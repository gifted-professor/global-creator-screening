from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any
from urllib import parse

from .feishu_api import FeishuApiError, FeishuOpenClient


@dataclass(frozen=True)
class ResolvedBitableView:
    source_url: str
    source_kind: str
    source_token: str
    app_token: str
    table_id: str
    view_id: str
    table_name: str = ""
    view_name: str = ""
    title: str = ""


def resolve_bitable_view_from_url(client: FeishuOpenClient, url: str) -> ResolvedBitableView:
    parsed = parse.urlparse(str(url or "").strip())
    if not parsed.scheme or not parsed.netloc:
        raise ValueError("飞书多维表格 URL 不合法。")
    segments = [item for item in parsed.path.split("/") if item]
    query = parse.parse_qs(parsed.query)
    table_id = str(query.get("table", [""])[0] or "").strip()
    view_id = str(query.get("view", [""])[0] or "").strip()
    if not table_id:
        raise ValueError("飞书多维表格 URL 缺少 table 参数。")
    if not view_id:
        raise ValueError("飞书多维表格 URL 缺少 view 参数。")

    if len(segments) >= 2 and segments[0] == "base":
        app_token = segments[1]
        return _populate_names(
            client,
            ResolvedBitableView(
                source_url=parsed.geturl(),
                source_kind="base",
                source_token=app_token,
                app_token=app_token,
                table_id=table_id,
                view_id=view_id,
            ),
        )

    if len(segments) >= 2 and segments[0] == "wiki":
        wiki_token = segments[1]
        node = client.resolve_wiki_node(wiki_token)
        obj_type = str(node.get("obj_type") or "").strip()
        if obj_type != "bitable":
            raise FeishuApiError(f"当前 wiki 节点不是 bitable，而是 {obj_type or 'unknown'}。")
        app_token = str(node.get("obj_token") or "").strip()
        if not app_token:
            raise FeishuApiError("wiki 节点解析成功，但没有拿到 bitable obj_token。")
        return _populate_names(
            client,
            ResolvedBitableView(
                source_url=parsed.geturl(),
                source_kind="wiki",
                source_token=wiki_token,
                app_token=app_token,
                table_id=table_id,
                view_id=view_id,
                title=str(node.get("title") or "").strip(),
            ),
        )

    raise ValueError("当前 URL 不是支持的飞书 base/wiki 多维表格链接。")


def export_bitable_view(
    client: FeishuOpenClient,
    *,
    url: str,
    output_path: str | Path,
    output_format: str = "json",
    page_size: int = 500,
) -> dict[str, Any]:
    resolved = resolve_bitable_view_from_url(client, url)
    records = _fetch_all_records(client, resolved, page_size=page_size)
    output = Path(output_path).expanduser()
    output.parent.mkdir(parents=True, exist_ok=True)

    normalized_format = str(output_format or "json").strip().lower()
    if normalized_format == "json":
        payload = {
            "ok": True,
            "sourceUrl": resolved.source_url,
            "sourceKind": resolved.source_kind,
            "sourceToken": resolved.source_token,
            "appToken": resolved.app_token,
            "title": resolved.title,
            "tableId": resolved.table_id,
            "tableName": resolved.table_name,
            "viewId": resolved.view_id,
            "viewName": resolved.view_name,
            "recordCount": len(records),
            "records": records,
        }
        output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    elif normalized_format == "xlsx":
        _write_records_to_xlsx(output, resolved, records)
    else:
        raise ValueError(f"不支持的输出格式: {output_format}")

    return {
        "ok": True,
        "sourceUrl": resolved.source_url,
        "sourceKind": resolved.source_kind,
        "title": resolved.title,
        "appToken": resolved.app_token,
        "tableId": resolved.table_id,
        "tableName": resolved.table_name,
        "viewId": resolved.view_id,
        "viewName": resolved.view_name,
        "recordCount": len(records),
        "outputPath": str(output),
        "outputFormat": normalized_format,
    }


def _populate_names(client: FeishuOpenClient, resolved: ResolvedBitableView) -> ResolvedBitableView:
    tables_payload = client.get_api_json(f"/bitable/v1/apps/{resolved.app_token}/tables")
    tables = tables_payload.get("data", {}).get("items") or []
    table_name = next(
        (str(item.get("name") or "") for item in tables if str(item.get("table_id") or "") == resolved.table_id),
        "",
    )

    views_payload = client.get_api_json(f"/bitable/v1/apps/{resolved.app_token}/tables/{resolved.table_id}/views")
    views = views_payload.get("data", {}).get("items") or []
    view_name = next(
        (str(item.get("view_name") or "") for item in views if str(item.get("view_id") or "") == resolved.view_id),
        "",
    )
    return ResolvedBitableView(
        source_url=resolved.source_url,
        source_kind=resolved.source_kind,
        source_token=resolved.source_token,
        app_token=resolved.app_token,
        table_id=resolved.table_id,
        view_id=resolved.view_id,
        table_name=table_name,
        view_name=view_name,
        title=resolved.title,
    )


def _fetch_all_records(
    client: FeishuOpenClient,
    resolved: ResolvedBitableView,
    *,
    page_size: int,
) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    page_token = ""
    while True:
        body: dict[str, Any] = {"view_id": resolved.view_id, "page_size": int(page_size)}
        if page_token:
            body["page_token"] = page_token
        payload = client.post_api_json(
            f"/bitable/v1/apps/{resolved.app_token}/tables/{resolved.table_id}/records/search",
            body=body,
        )
        data = payload.get("data", {}) or {}
        items = data.get("items") or []
        for item in items:
            if isinstance(item, dict):
                collected.append(
                    {
                        "record_id": str(item.get("record_id") or ""),
                        "fields": item.get("fields") or {},
                    }
                )
        has_more = bool(data.get("has_more"))
        next_page_token = str(data.get("page_token") or "").strip()
        if not has_more:
            break
        if not next_page_token:
            raise FeishuApiError("records/search 返回 has_more=true，但没有 page_token，无法继续分页。")
        page_token = next_page_token
    return collected


def _write_records_to_xlsx(output_path: Path, resolved: ResolvedBitableView, records: list[dict[str, Any]]) -> None:
    load_workbook = _require_openpyxl_workbook()
    workbook = load_workbook()
    sheet = workbook.active
    sheet.title = (resolved.table_name or "FeishuExport")[:31]

    field_order: list[str] = []
    seen: set[str] = set()
    for item in records:
        fields = item.get("fields") or {}
        if not isinstance(fields, dict):
            continue
        for key in fields.keys():
            normalized = str(key or "")
            if normalized and normalized not in seen:
                seen.add(normalized)
                field_order.append(normalized)

    headers = ["record_id", *field_order]
    for column_index, value in enumerate(headers, start=1):
        sheet.cell(1, column_index).value = value

    for row_index, item in enumerate(records, start=2):
        sheet.cell(row_index, 1).value = str(item.get("record_id") or "")
        fields = item.get("fields") or {}
        for column_index, key in enumerate(field_order, start=2):
            sheet.cell(row_index, column_index).value = _normalize_cell_value(fields.get(key))

    workbook.save(output_path)


def _normalize_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _require_openpyxl_workbook() -> Any:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("导出 xlsx 需要 openpyxl，请先安装。") from exc
    return Workbook
