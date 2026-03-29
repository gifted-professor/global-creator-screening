from __future__ import annotations

from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
import os
from pathlib import Path
import re
from typing import Any

from email_sync.config import Settings
from email_sync.db import Database
from email_sync.imap_sync import MailboxInfo, connect, discover_mailboxes, resolve_mailboxes, sync_mailboxes
from openpyxl import load_workbook
from workbook_template_parser import compile_workbook

from .bitable_export import resolve_bitable_view_from_url
from .email_project import load_email_project, resolve_email_env_file, resolve_email_project_root
from .feishu_api import FeishuApiError, FeishuOpenClient
from .repo_local_runtime import build_repo_local_workbook_runtime, safe_path_component, write_dashboard_html, write_json


CANONICAL_SECTION_BASIC = "A. 基本信息"
CANONICAL_SECTION_STEP_1 = "B. 步骤1：基础资质审核"
CANONICAL_SECTION_STEP_2 = "C. 步骤2：数据审核"
CANONICAL_SECTION_STEP_3 = "D. 步骤3：内容 / 视觉审核"
CANONICAL_SECTION_STEP_4 = "E. 步骤4：排除项审核"
CANONICAL_SECTION_STEP_6 = "F. 人工判断项 / 合规提醒"
CANONICAL_SECTION_STEP_7 = "G. 最终判定逻辑"

_SECTION_ALIASES = {
    "A. 基本信息": CANONICAL_SECTION_BASIC,
    "A. 项目基本信息": CANONICAL_SECTION_BASIC,
    "B. 步骤1：基础资质审核": CANONICAL_SECTION_STEP_1,
    "C. 步骤2：数据审核": CANONICAL_SECTION_STEP_2,
    "D. 步骤3：内容 / 视觉审核": CANONICAL_SECTION_STEP_3,
    "E. 步骤4：排除项审核": CANONICAL_SECTION_STEP_4,
    "E. 步骤4：排除项审核（不需要的封面清单）": CANONICAL_SECTION_STEP_4,
    "F. 人工判断项 / 合规提醒": CANONICAL_SECTION_STEP_6,
    "F. 人工判断项 / 合规提醒（当封面或数据出现什么情况时需要人工复核，如没有可不填）": CANONICAL_SECTION_STEP_6,
    "G. 最终判定逻辑": CANONICAL_SECTION_STEP_7,
}

_FIELD_ALIASES = {
    (CANONICAL_SECTION_BASIC, "项目名称（品牌名）"): "项目名称",
    (CANONICAL_SECTION_BASIC, "品牌 / 产品使用场景"): "品牌 / 产品",
}

_PLATFORM_TOKEN_MAP = {
    "tiktok": "tiktok",
    "tik tok": "tiktok",
    "instagram": "instagram",
    "youtube": "youtube",
    "yt": "youtube",
    "两者": "both",
    "both": "both",
}


@dataclass(frozen=True)
class TaskUploadEntry:
    record_id: str
    task_name: str
    employee_id: str
    owner_name: str
    owner_email: str
    owner_email_candidates: tuple[str, ...]
    responsible_name: str
    linked_bitable_url: str
    workbook_file_token: str
    workbook_file_name: str
    sending_list_file_token: str
    sending_list_file_name: str


@dataclass(frozen=True)
class EmployeeDirectoryEntry:
    record_id: str
    employee_id: str
    employee_name: str
    email: str
    imap_code: str


def resolve_task_upload_entry(
    *,
    client: FeishuOpenClient,
    task_upload_url: str,
    task_name: str,
) -> TaskUploadEntry:
    normalized_task_name = _normalize_lookup_key(task_name)
    if not normalized_task_name:
        raise ValueError("task_name 不能为空。")

    entries = _fetch_task_upload_entries(client, task_upload_url)
    matched_entries = [
        entry
        for entry in entries
        if _normalize_lookup_key(entry.task_name) == normalized_task_name
    ]
    if not matched_entries:
        available = ", ".join(sorted(entry.task_name for entry in entries if entry.task_name))
        raise ValueError(
            f"任务上传里找不到任务 {task_name!r}。"
            + (f" 当前可用任务：{available}" if available else "")
        )
    if len(matched_entries) > 1:
        record_ids = ", ".join(entry.record_id for entry in matched_entries)
        raise ValueError(
            f"任务名 {task_name!r} 匹配到多条任务上传记录：{record_ids}。"
            " 请先清理重复记录后再执行任务驱动流程。"
        )
    return matched_entries[0]


def download_task_upload_screening_assets(
    *,
    client: FeishuOpenClient,
    task_upload_url: str,
    task_name: str,
    download_dir: str | Path,
    download_template: bool = True,
    download_sending_list: bool = True,
) -> dict[str, Any]:
    entry = resolve_task_upload_entry(
        client=client,
        task_upload_url=task_upload_url,
        task_name=task_name,
    )
    download_root = Path(download_dir).expanduser()
    download_root.mkdir(parents=True, exist_ok=True)

    template_downloaded_path = ""
    sending_list_downloaded_path = ""

    if download_template:
        if not entry.workbook_file_token:
            raise ValueError(f"任务 {entry.task_name!r} 缺少 `需求上传（excel 格式）` 附件。")
        template_downloaded_path = str(_download_task_upload_workbook(client, entry, download_root))

    if download_sending_list:
        if not entry.sending_list_file_token:
            raise ValueError(f"任务 {entry.task_name!r} 缺少 `发信名单` 附件。")
        sending_list_downloaded_path = str(_download_task_upload_sending_list(client, entry, download_root))

    return {
        "recordId": entry.record_id,
        "taskName": entry.task_name,
        "linkedBitableUrl": entry.linked_bitable_url,
        "templateFileToken": entry.workbook_file_token,
        "templateFileName": entry.workbook_file_name,
        "templateDownloadedPath": template_downloaded_path,
        "sendingListFileToken": entry.sending_list_file_token,
        "sendingListFileName": entry.sending_list_file_name,
        "sendingListDownloadedPath": sending_list_downloaded_path,
    }


def inspect_task_upload_assignments(
    *,
    client: FeishuOpenClient,
    task_upload_url: str,
    employee_info_url: str,
    download_dir: str | Path,
    download_templates: bool = False,
    parse_templates: bool = False,
    parse_output_dir: str | Path | None = None,
    owner_email_overrides: dict[str, str] | None = None,
) -> dict[str, Any]:
    task_resolved = resolve_bitable_view_from_url(client, task_upload_url)
    if task_resolved.table_name and task_resolved.table_name != "任务上传":
        raise FeishuApiError(f"当前任务链接指向的不是任务上传表，而是 {task_resolved.table_name}。")

    employee_resolved = resolve_bitable_view_from_url(client, employee_info_url)
    if employee_resolved.table_name and "员工" not in employee_resolved.table_name:
        raise FeishuApiError(f"当前员工链接指向的不是员工信息表，而是 {employee_resolved.table_name}。")

    entries = _fetch_task_upload_entries(client, task_upload_url)
    employees = _fetch_employee_directory_entries(client, employee_info_url)
    employees_by_id = {
        _normalize_lookup_key(item.employee_id): item
        for item in employees
        if _normalize_lookup_key(item.employee_id)
    }
    employees_by_email = {
        _normalize_lookup_key(item.email): item
        for item in employees
        if _normalize_lookup_key(item.email)
    }

    should_download_templates = bool(download_templates or parse_templates)
    normalized_owner_email_overrides = {
        _normalize_lookup_key(key): str(value or "").strip()
        for key, value in (owner_email_overrides or {}).items()
        if _normalize_lookup_key(key) and str(value or "").strip()
    }
    download_root = Path(download_dir).expanduser()
    parse_root = (
        Path(parse_output_dir).expanduser()
        if parse_output_dir is not None
        else download_root / "parsed_outputs"
    )
    if should_download_templates:
        download_root.mkdir(parents=True, exist_ok=True)
    if parse_templates:
        parse_root.mkdir(parents=True, exist_ok=True)

    matched_count = 0
    downloaded_count = 0
    parsed_count = 0
    parse_failed_count = 0
    items: list[dict[str, Any]] = []
    for entry in entries:
        matched_by = ""
        employee: EmployeeDirectoryEntry | None = None
        employee_id_key = _normalize_lookup_key(entry.employee_id)
        owner_email_key = _normalize_lookup_key(entry.owner_email)
        task_key = _normalize_lookup_key(entry.task_name)
        preferred_owner_email = normalized_owner_email_overrides.get(task_key, "")
        preferred_owner_email_key = _normalize_lookup_key(preferred_owner_email)
        owner_email_candidates = list(entry.owner_email_candidates or ((entry.owner_email,) if entry.owner_email else ()))
        if preferred_owner_email_key:
            owner_email_candidates = [
                candidate
                for candidate in owner_email_candidates
                if _normalize_lookup_key(candidate) == preferred_owner_email_key
            ] + [
                candidate
                for candidate in owner_email_candidates
                if _normalize_lookup_key(candidate) != preferred_owner_email_key
            ]
            for candidate in owner_email_candidates:
                employee = employees_by_email.get(_normalize_lookup_key(candidate))
                if employee is not None:
                    matched_by = "owner_email_override"
                    break
        if employee is None and employee_id_key:
            employee = employees_by_id.get(employee_id_key)
            if employee is not None:
                matched_by = "employee_id"
        if employee is None:
            for candidate in owner_email_candidates:
                employee = employees_by_email.get(_normalize_lookup_key(candidate))
                if employee is not None:
                    matched_by = "owner_email"
                    break

        saved_template_path = ""
        parse_requested = False
        parse_result: dict[str, Any] = {
            "templateParsed": False,
            "templateParseOutputDir": "",
            "templateParseReportPath": "",
            "templateParseArtifacts": {},
            "templateParseWarnings": [],
            "templateParseStats": {},
            "templateParseError": "",
        }
        if should_download_templates and entry.workbook_file_token:
            saved_template_path = str(_download_task_upload_workbook(client, entry, download_root))
            downloaded_count += 1
            if parse_templates:
                parse_requested = True
                try:
                    parse_result = _parse_task_upload_workbook(
                        workbook_path=Path(saved_template_path),
                        output_root=parse_root,
                    )
                    parsed_count += 1
                except Exception as exc:  # noqa: BLE001
                    parse_failed_count += 1
                    parse_result["templateParseError"] = str(exc)
        elif parse_templates:
            parse_requested = True
            parse_failed_count += 1
            parse_result["templateParseError"] = "任务记录缺少可下载的模板文件，无法解析。"

        if employee is not None:
            matched_count += 1

        item = {
            "recordId": entry.record_id,
            "taskName": entry.task_name,
            "employeeId": entry.employee_id,
            "responsibleName": entry.responsible_name,
            "ownerName": entry.owner_name,
            "ownerEmail": entry.owner_email,
            "ownerEmailCandidates": list(entry.owner_email_candidates),
            "preferredOwnerEmail": preferred_owner_email,
            "linkedBitableUrl": entry.linked_bitable_url,
            "templateFileToken": entry.workbook_file_token,
            "templateFileName": entry.workbook_file_name,
            "templateDownloadedPath": saved_template_path,
            "sendingListFileToken": entry.sending_list_file_token,
            "sendingListFileName": entry.sending_list_file_name,
            "templateParseRequested": parse_requested,
            "employeeMatched": employee is not None,
            "matchedBy": matched_by,
            "employeeRecordId": employee.record_id if employee is not None else "",
            "employeeName": employee.employee_name if employee is not None else "",
            "employeeEmail": employee.email if employee is not None else "",
            "imapCode": employee.imap_code if employee is not None else "",
        }
        item.update(parse_result)
        items.append(item)

    return {
        "ok": True,
        "taskUploadUrl": task_upload_url,
        "employeeInfoUrl": employee_info_url,
        "taskTableName": task_resolved.table_name,
        "employeeTableName": employee_resolved.table_name,
        "recordCount": len(entries),
        "employeeCount": len(employees),
        "matchedCount": matched_count,
        "downloadedCount": downloaded_count,
        "parseTemplates": parse_templates,
        "parsedCount": parsed_count,
        "parseFailedCount": parse_failed_count,
        "downloadDir": str(download_root),
        "parseOutputDir": str(parse_root) if parse_templates else "",
        "items": items,
    }


def sync_task_upload_mailboxes(
    *,
    client: FeishuOpenClient,
    task_upload_url: str,
    employee_info_url: str,
    download_dir: str | Path,
    mail_data_dir: str | Path,
    task_names: list[str] | tuple[str, ...] | None = None,
    owner_email_overrides: dict[str, str] | None = None,
    folder_overrides: dict[str, str] | None = None,
    folder_prefixes: list[str] | tuple[str, ...] | None = None,
    limit: int | None = None,
    workers: int = 1,
    reset_state: bool = False,
    sent_since: str | None = None,
    imap_host: str = "imap.qq.com",
    imap_port: int = 993,
) -> dict[str, Any]:
    if limit is not None and limit <= 0:
        raise ValueError("--limit 必须是大于 0 的整数。")
    if workers <= 0:
        raise ValueError("--workers 必须是大于 0 的整数。")

    sent_since_date = date.fromisoformat(str(sent_since).strip()) if sent_since else None
    normalized_task_names = {
        _normalize_lookup_key(name)
        for name in (task_names or [])
        if _normalize_lookup_key(name)
    }
    normalized_folder_overrides = {
        _normalize_lookup_key(key): str(value or "").strip()
        for key, value in (folder_overrides or {}).items()
        if _normalize_lookup_key(key) and str(value or "").strip()
    }
    normalized_folder_prefixes = [
        str(prefix or "").strip()
        for prefix in (folder_prefixes or ["其他文件夹"])
        if str(prefix or "").strip()
    ]

    inspection = inspect_task_upload_assignments(
        client=client,
        task_upload_url=task_upload_url,
        employee_info_url=employee_info_url,
        download_dir=download_dir,
        download_templates=False,
        parse_templates=False,
        owner_email_overrides=owner_email_overrides,
    )

    mail_root = Path(mail_data_dir).expanduser()
    mail_root.mkdir(parents=True, exist_ok=True)

    selected_count = 0
    synced_count = 0
    failed_count = 0
    items: list[dict[str, Any]] = []
    for inspected in inspection["items"]:
        task_key = _normalize_lookup_key(inspected["taskName"])
        if normalized_task_names and task_key not in normalized_task_names:
            continue

        selected_count += 1
        requested_folder = normalized_folder_overrides.get(task_key, "")
        result_item = dict(inspected)
        result_item.update(
            {
                "mailSyncRequested": True,
                "mailSyncOk": False,
                "mailSyncError": "",
                "requestedFolder": requested_folder,
                "resolvedFolder": "",
                "mailDataDir": "",
                "mailDbPath": "",
                "mailRawDir": "",
                "mailFetchedCount": 0,
                "mailLastSeenUid": 0,
                "mailUidvalidity": None,
                "mailServerCount": None,
                "mailSkippedStateAdvance": False,
            }
        )

        try:
            employee_email = str(inspected.get("employeeEmail") or "").strip()
            imap_code = str(inspected.get("imapCode") or "").strip()
            if not inspected.get("employeeMatched"):
                raise ValueError("任务未匹配到员工信息，无法抓取邮件。")
            if not employee_email:
                raise ValueError("员工邮箱为空，无法抓取邮件。")
            if not imap_code:
                raise ValueError("员工 IMAP 码为空，无法抓取邮件。")

            task_data_dir = mail_root / _safe_path_component(inspected["taskName"] or inspected["recordId"] or "task-mail-sync")
            settings = Settings(
                account_email=employee_email,
                auth_code=imap_code,
                imap_host=str(imap_host or "imap.qq.com").strip() or "imap.qq.com",
                imap_port=int(imap_port),
                data_dir=task_data_dir,
                db_path=task_data_dir / "email_sync.db",
                raw_dir=task_data_dir / "raw",
                mail_folders=None,
            )
            settings.validate()
            settings.ensure_directories()

            discovered_mailboxes = _discover_mailboxes_for_settings(settings)
            resolved_folder = _resolve_task_mailbox_name(
                discovered_mailboxes,
                task_name=inspected["taskName"],
                explicit_folder=requested_folder,
                folder_prefixes=normalized_folder_prefixes,
            )

            db = Database(settings.db_path)
            try:
                db.init_schema()
                sync_results = sync_mailboxes(
                    settings,
                    db,
                    requested_folders=[resolved_folder],
                    limit=limit,
                    reset_state=reset_state,
                    workers=workers,
                    sent_since=sent_since_date,
                )
            finally:
                db.close()

            sync_result = sync_results[0] if sync_results else None
            result_item.update(
                {
                    "mailSyncOk": True,
                    "resolvedFolder": resolved_folder,
                    "mailDataDir": str(settings.data_dir),
                    "mailDbPath": str(settings.db_path),
                    "mailRawDir": str(settings.raw_dir),
                    "mailFetchedCount": sync_result.fetched if sync_result is not None else 0,
                    "mailLastSeenUid": sync_result.last_seen_uid if sync_result is not None else 0,
                    "mailUidvalidity": sync_result.uidvalidity if sync_result is not None else None,
                    "mailServerCount": sync_result.message_count_on_server if sync_result is not None else None,
                    "mailSkippedStateAdvance": bool(sync_result.skipped_state_advance) if sync_result is not None else False,
                }
            )
            synced_count += 1
        except Exception as exc:  # noqa: BLE001
            failed_count += 1
            result_item["mailSyncError"] = str(exc)

        items.append(result_item)

    return {
        "ok": True,
        "taskUploadUrl": task_upload_url,
        "employeeInfoUrl": employee_info_url,
        "taskTableName": inspection["taskTableName"],
        "employeeTableName": inspection["employeeTableName"],
        "recordCount": inspection["recordCount"],
        "selectedCount": selected_count,
        "syncedCount": synced_count,
        "failedCount": failed_count,
        "mailDataDir": str(mail_root),
        "imapHost": str(imap_host or "imap.qq.com").strip() or "imap.qq.com",
        "imapPort": int(imap_port),
        "sentSince": sent_since_date.isoformat() if sent_since_date is not None else "",
        "items": items,
    }


def sync_task_upload_view_to_email_project(
    *,
    client: FeishuOpenClient,
    task_upload_url: str,
    email_project_root: str | Path | None,
    email_env_file: str | Path | None,
    download_dir: str | Path,
    dashboard_output: str | Path | None = None,
    project_code_prefix: str = "P-FSH-",
    default_primary_category: str = "lifestyle",
    category_overrides: dict[str, str] | None = None,
) -> dict[str, Any]:
    if not str(email_project_root or "").strip():
        download_root = Path(download_dir).expanduser()
        download_root.mkdir(parents=True, exist_ok=True)
        runtime_root = download_root / "_repo_local"
        runtime_root.mkdir(parents=True, exist_ok=True)

        entries = _fetch_task_upload_entries(client, task_upload_url)
        imported_items: list[dict[str, Any]] = []
        latest_project_code = ""
        for entry in entries:
            saved_path = _download_task_upload_workbook(client, entry, download_root)
            project_code = _build_project_code(entry.task_name, prefix=project_code_prefix)
            primary_category = _resolve_primary_category(
                workbook_path=saved_path,
                task_name=entry.task_name,
                overrides=category_overrides or {},
                default_primary_category=default_primary_category,
            )
            item_runtime_root = runtime_root / safe_path_component(project_code or entry.task_name or entry.record_id)
            repo_local_summary = build_repo_local_workbook_runtime(
                workbook_path=saved_path,
                runtime_root=item_runtime_root,
                project_code=project_code,
                primary_category=primary_category,
                owner_name=entry.owner_name,
                task_name=entry.task_name,
                record_id=entry.record_id,
                linked_bitable_url=entry.linked_bitable_url,
            )
            latest_project_code = project_code
            imported_items.append(
                {
                    "recordId": entry.record_id,
                    "taskName": entry.task_name,
                    "projectCode": project_code,
                    "projectName": repo_local_summary["projectName"],
                    "primaryCategory": primary_category,
                    "ownerName": entry.owner_name,
                    "linkedBitableUrl": entry.linked_bitable_url,
                    "savedWorkbookPath": str(saved_path),
                    "compiledRowCount": repo_local_summary["compiledRowCount"],
                    "platforms": list(repo_local_summary["platforms"]),
                    "summaryJson": repo_local_summary["summaryJson"],
                    "projectStatePath": repo_local_summary["projectStatePath"],
                    "dashboardOutput": repo_local_summary["dashboardOutput"],
                    "templateParseArtifacts": dict(repo_local_summary["templateParseArtifacts"]),
                }
            )

        aggregate_dashboard_path = (
            Path(dashboard_output).expanduser().resolve()
            if dashboard_output is not None
            else runtime_root / "dashboard.html"
        )
        aggregate_summary_path = runtime_root / "summary.json"
        aggregate_project_state_path = runtime_root / "project_state_index.json"
        aggregate_summary = {
            "ok": True,
            "mode": "repo_local",
            "taskUploadUrl": task_upload_url,
            "emailProjectRoot": "",
            "emailEnvFile": str(email_env_file or ".env"),
            "dbPath": "",
            "dashboardOutput": str(aggregate_dashboard_path),
            "summaryJson": str(aggregate_summary_path),
            "projectStatePath": str(aggregate_project_state_path),
            "recordCount": len(entries),
            "importedCount": len(imported_items),
            "latestProjectCode": latest_project_code,
            "items": imported_items,
        }
        write_json(aggregate_project_state_path, {"items": imported_items, "latestProjectCode": latest_project_code})
        write_json(aggregate_summary_path, aggregate_summary)
        write_dashboard_html(
            aggregate_dashboard_path,
            {
                "projectCode": latest_project_code,
                "projectName": "task-upload-sync",
                "savedWorkbookPath": download_root,
                "summaryJson": str(aggregate_summary_path),
                "projectStatePath": str(aggregate_project_state_path),
                "templateParseArtifacts": {},
                "nextSteps": [
                    {
                        "label": "inspect_repo_local_sync_summary",
                        "description": "查看聚合 summary 和每个任务的 repo-local project-state。",
                        "path": str(aggregate_summary_path),
                    }
                ],
            },
        )
        return aggregate_summary

    resolved_project_root = resolve_email_project_root(email_project_root)
    resolved_env_file = resolve_email_env_file(resolved_project_root, email_env_file)
    modules = load_email_project(resolved_project_root)

    env_base_dir = resolved_env_file.parent if resolved_env_file.exists() else resolved_project_root
    with _pushd(env_base_dir):
        settings = modules.Settings.from_environment(str(resolved_env_file), require_credentials=False)
    settings.data_dir = _resolve_path_from(env_base_dir, settings.data_dir)
    settings.db_path = _resolve_path_from(env_base_dir, settings.db_path)
    settings.raw_dir = _resolve_path_from(env_base_dir, settings.raw_dir)
    settings.ensure_directories()

    output_path = (
        Path(dashboard_output).expanduser()
        if dashboard_output is not None
        else resolved_project_root / "exports" / "index.html"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    download_root = Path(download_dir).expanduser()
    download_root.mkdir(parents=True, exist_ok=True)

    entries = _fetch_task_upload_entries(client, task_upload_url)
    imported_items: list[dict[str, Any]] = []
    latest_project_code = ""
    for entry in entries:
        saved_path = _download_task_upload_workbook(client, entry, download_root)
        project_code = _build_project_code(entry.task_name, prefix=project_code_prefix)
        primary_category = _resolve_primary_category(
            workbook_path=saved_path,
            task_name=entry.task_name,
            overrides=category_overrides or {},
            default_primary_category=default_primary_category,
        )
        compile_result = _compile_workbook_for_import(
            modules,
            workbook_path=saved_path,
            project_code=project_code,
            primary_category=primary_category,
        )
        _import_compiled_rows(
            modules,
            db_path=settings.db_path,
            output_path=output_path,
            project_code=project_code,
            project_name=compile_result["project_name"],
            owner_name=entry.owner_name,
            compiled_rows=compile_result["compiled_rows"],
        )
        latest_project_code = project_code
        imported_items.append(
            {
                "recordId": entry.record_id,
                "taskName": entry.task_name,
                "projectCode": project_code,
                "projectName": compile_result["project_name"],
                "primaryCategory": primary_category,
                "ownerName": entry.owner_name,
                "linkedBitableUrl": entry.linked_bitable_url,
                "savedWorkbookPath": str(saved_path),
                "compiledRowCount": len(compile_result["compiled_rows"]),
                "platforms": [row.platform for row in compile_result["compiled_rows"]],
            }
        )

    if latest_project_code:
        modules.export_dashboard(settings.db_path, output_path, current_project_code=latest_project_code)

    return {
        "ok": True,
        "taskUploadUrl": task_upload_url,
        "emailProjectRoot": str(resolved_project_root),
        "emailEnvFile": str(resolved_env_file),
        "dbPath": str(settings.db_path),
        "dashboardOutput": str(output_path),
        "recordCount": len(entries),
        "importedCount": len(imported_items),
        "items": imported_items,
    }


def _fetch_task_upload_entries(client: FeishuOpenClient, task_upload_url: str) -> list[TaskUploadEntry]:
    resolved = resolve_bitable_view_from_url(client, task_upload_url)
    if resolved.table_name and resolved.table_name != "任务上传":
        raise FeishuApiError(f"当前链接指向的不是任务上传表，而是 {resolved.table_name}。")

    items = _fetch_all_records_for_view(client, resolved)
    results: list[TaskUploadEntry] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        fields = item.get("fields") or {}
        workbook = _extract_attachment_with_file_token(fields.get("需求上传（excel 格式）"))
        if not isinstance(workbook, dict):
            continue
        sending_list = _extract_attachment_with_file_token(fields.get("发信名单"))
        task_name = _extract_text_like(fields.get("任务名")) or str(item.get("record_id") or "")
        owner_email_candidates = _extract_email_values(fields.get("负责人邮箱"))
        owner_email = ",".join(owner_email_candidates)
        responsible_name = _extract_person_name(fields.get("负责人"))
        owner_name = owner_email or responsible_name or ""
        linked_bitable_url = _extract_mention_link(fields.get("达人管理表链接"))
        results.append(
            TaskUploadEntry(
                record_id=str(item.get("record_id") or ""),
                task_name=task_name,
                employee_id=_extract_text_like(fields.get("员工ID")),
                owner_name=owner_name,
                owner_email=owner_email,
                owner_email_candidates=owner_email_candidates,
                responsible_name=responsible_name,
                linked_bitable_url=linked_bitable_url,
                workbook_file_token=str(workbook.get("file_token") or "").strip(),
                workbook_file_name=str(workbook.get("name") or "").strip() or "screening.xlsx",
                sending_list_file_token=str((sending_list or {}).get("file_token") or "").strip(),
                sending_list_file_name=str((sending_list or {}).get("name") or "").strip() or "creator-source.xlsx",
            )
        )
    return results


def _fetch_employee_directory_entries(client: FeishuOpenClient, employee_info_url: str) -> list[EmployeeDirectoryEntry]:
    resolved = resolve_bitable_view_from_url(client, employee_info_url)
    items = _fetch_all_records_for_view(client, resolved)
    results: list[EmployeeDirectoryEntry] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        fields = item.get("fields") or {}
        results.append(
            EmployeeDirectoryEntry(
                record_id=str(item.get("record_id") or ""),
                employee_id=_first_non_empty(
                    _extract_text_like(fields.get("员工 ID")),
                    _extract_text_like(fields.get("员工ID")),
                ),
                employee_name=_first_non_empty(
                    _extract_person_name(fields.get("员工名")),
                    _extract_text_like(fields.get("员工名")),
                ),
                email=_first_non_empty(
                    _extract_email_value(fields.get("邮箱")),
                    _extract_text_like(fields.get("邮箱")),
                ),
                imap_code=_first_non_empty(
                    _extract_text_like(fields.get("imap 码")),
                    _extract_text_like(fields.get("IMAP 码")),
                    _extract_text_like(fields.get("IMAP码")),
                ),
            )
        )
    return results


def _fetch_all_records_for_view(client: FeishuOpenClient, resolved: Any, *, page_size: int = 500) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    page_token = ""
    while True:
        body: dict[str, Any] = {"view_id": resolved.view_id, "page_size": int(page_size)}
        if page_token:
            body["page_token"] = page_token
        payload = client.post_api_json(
            f"/bitable/v1/apps/{resolved.app_token}/tables/{resolved.table_id}/records/search",
            body=body,
        )
        data = payload.get("data", {}) or {}
        items = data.get("items") or []
        for item in items:
            if isinstance(item, dict):
                collected.append(item)
        if not bool(data.get("has_more")):
            break
        page_token = str(data.get("page_token") or "").strip()
        if not page_token:
            break
    return collected


def _download_task_upload_workbook(
    client: FeishuOpenClient,
    entry: TaskUploadEntry,
    download_root: Path,
) -> Path:
    return _download_task_upload_attachment(
        client,
        record_id=entry.record_id,
        attachment_label="需求上传（excel 格式）",
        file_token=entry.workbook_file_token,
        file_name=entry.workbook_file_name,
        download_root=download_root,
    )


def _download_task_upload_sending_list(
    client: FeishuOpenClient,
    entry: TaskUploadEntry,
    download_root: Path,
) -> Path:
    return _download_task_upload_attachment(
        client,
        record_id=entry.record_id,
        attachment_label="发信名单",
        file_token=entry.sending_list_file_token,
        file_name=entry.sending_list_file_name,
        download_root=download_root,
    )


def _download_task_upload_attachment(
    client: FeishuOpenClient,
    *,
    record_id: str,
    attachment_label: str,
    file_token: str,
    file_name: str,
    download_root: Path,
) -> Path:
    downloaded = client.download_file(file_token, desired_name=file_name)
    record_dir = download_root / record_id / attachment_label
    record_dir.mkdir(parents=True, exist_ok=True)
    return _write_unique_file(record_dir, downloaded.file_name, downloaded.content)


def _discover_mailboxes_for_settings(settings: Settings) -> list[MailboxInfo]:
    client = connect(settings)
    try:
        return discover_mailboxes(client)
    finally:
        try:
            client.close()
        except Exception:  # noqa: BLE001
            pass
        try:
            client.logout()
        except Exception:  # noqa: BLE001
            pass


def _resolve_task_mailbox_name(
    discovered_mailboxes: list[MailboxInfo],
    *,
    task_name: str,
    explicit_folder: str = "",
    folder_prefixes: list[str] | tuple[str, ...] | None = None,
) -> str:
    selectable = resolve_mailboxes(discovered_mailboxes, None)
    candidates: list[str] = []
    for raw_candidate in [explicit_folder, task_name]:
        candidate = str(raw_candidate or "").strip()
        if candidate and candidate not in candidates:
            candidates.append(candidate)
    for prefix in folder_prefixes or []:
        normalized_prefix = str(prefix or "").strip().strip("/")
        if normalized_prefix and task_name:
            candidate = f"{normalized_prefix}/{task_name}"
            if candidate not in candidates:
                candidates.append(candidate)

    for candidate in candidates:
        try:
            return resolve_mailboxes(discovered_mailboxes, [candidate])[0].display_name
        except ValueError:
            continue

    normalized_task = _normalize_lookup_key(task_name)
    fuzzy_matches = [
        mailbox.display_name
        for mailbox in selectable
        if _normalize_lookup_key(mailbox.display_name) == normalized_task
        or _normalize_lookup_key(mailbox.display_name).endswith(f"/{normalized_task}")
    ]
    if len(fuzzy_matches) == 1:
        return fuzzy_matches[0]

    available = ", ".join(mailbox.display_name for mailbox in selectable[:20])
    if len(selectable) > 20:
        available = f"{available}, ..."
    raise ValueError(f"找不到任务 {task_name!r} 对应的邮箱文件夹。可用文件夹：{available}")


def _parse_task_upload_workbook(*, workbook_path: Path, output_root: Path) -> dict[str, Any]:
    report = compile_workbook(workbook_path, output_root)
    workbook_output_dir = Path(str(report["output_dir"]))
    return {
        "templateParsed": True,
        "templateParseOutputDir": str(workbook_output_dir),
        "templateParseReportPath": str(workbook_output_dir / "compile_report.json"),
        "templateParseArtifacts": dict(report.get("artifacts") or {}),
        "templateParseWarnings": list(report.get("warnings") or []),
        "templateParseStats": dict(report.get("stats") or {}),
        "templateParseError": "",
    }


def _compile_workbook_for_import(
    modules: Any,
    *,
    workbook_path: Path,
    project_code: str,
    primary_category: str,
) -> dict[str, Any]:
    try:
        result = modules.compile_screening_workbook(
            workbook_path,
            project_code=project_code,
            primary_category=primary_category,
        )
        return {
            "project_name": result.project_name,
            "compiled_rows": tuple(result.compiled_rows),
        }
    except modules.ScreeningWorkbookImportError:
        return _compile_flexible_screening_workbook(
            modules,
            workbook_path=workbook_path,
            project_code=project_code,
            primary_category=primary_category,
        )


def _compile_flexible_screening_workbook(
    modules: Any,
    *,
    workbook_path: Path,
    project_code: str,
    primary_category: str,
) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet_name = "需求主表" if "需求主表" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    field_index: dict[tuple[str, str], dict[str, Any]] = {}
    current_section = ""
    for row_number in range(1, sheet.max_row + 1):
        field_name = _normalize_text(sheet.cell(row_number, 1).value)
        value = sheet.cell(row_number, 2).value
        note = _normalize_text(sheet.cell(row_number, 3).value)
        if row_number <= 3 or not field_name:
            continue
        section_name = _SECTION_ALIASES.get(field_name)
        if section_name:
            current_section = section_name
            continue
        canonical_field_name = _FIELD_ALIASES.get((current_section, field_name), field_name)
        field_index[(current_section, canonical_field_name)] = {
            "value": value,
            "note": note,
            "row_number": row_number,
        }

    project_name = _require_text(field_index, CANONICAL_SECTION_BASIC, "项目名称")
    brand_name = _require_text(field_index, CANONICAL_SECTION_BASIC, "品牌 / 产品")
    target_region = _require_text(field_index, CANONICAL_SECTION_STEP_1, "地区要求")
    target_language = _normalize_blank(_optional_text(field_index, CANONICAL_SECTION_STEP_1, "语言要求")) or "不限"
    median_views_min = _require_non_negative_int(field_index, CANONICAL_SECTION_STEP_2, "中位数播放量阈值")
    follower_min = _optional_non_negative_int(field_index, CANONICAL_SECTION_STEP_2, "粉丝数阈值（可选）", default=0)
    platforms = _parse_platform_scope_extended(_require_text(field_index, CANONICAL_SECTION_BASIC, "适用平台"))
    visual_constraints = _build_visual_constraints(field_index)

    compiled_rows = []
    for platform in platforms:
        compiled_rows.append(
            modules.parse_requirement_row(
                {
                    "requirement_code": f"SWB-{platform.upper()}",
                    "brand_name": brand_name,
                    "platform": platform,
                    "target_region": target_region,
                    "target_language": target_language,
                    "primary_category": primary_category,
                    "follower_min": follower_min,
                    "median_views_min": median_views_min,
                    "engagement_rate_min": 0.0,
                    "target_cpm_cap": 0.0,
                    "visual_constraints": visual_constraints,
                },
                row_number=len(compiled_rows) + 1,
            )
        )
    return {
        "project_name": project_name,
        "compiled_rows": tuple(compiled_rows),
    }


def _import_compiled_rows(
    modules: Any,
    *,
    db_path: Path,
    output_path: Path,
    project_code: str,
    project_name: str,
    owner_name: str,
    compiled_rows: tuple[Any, ...],
) -> None:
    db = modules.Database(db_path)
    try:
        db.init_schema()
        modules.init_influencer_schema(db)
        modules.init_creator_ops_schema(db)
        project_id = modules.ensure_project(
            db,
            project_code=project_code,
            project_name=project_name,
            owner_name=owner_name,
        )
        modules.import_requirements(
            db,
            project_id=project_id,
            rows=compiled_rows,
            source_kind="screening_workbook",
        )
        modules.rebuild_project_home_read_model(db)
        modules.rebuild_project_workbench_read_model(db)
    finally:
        db.close()
    modules.export_dashboard(db_path, output_path, current_project_code=project_code)


def _resolve_primary_category(
    *,
    workbook_path: Path,
    task_name: str,
    overrides: dict[str, str],
    default_primary_category: str,
) -> str:
    override = overrides.get(task_name.casefold())
    if override:
        return override
    workbook = load_workbook(workbook_path, data_only=True)
    sheet_name = "需求主表" if "需求主表" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    text_parts = [task_name]
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        for column_index in range(1, 3):
            value = _normalize_text(sheet.cell(row_number, column_index).value)
            if value:
                text_parts.append(value)
    haystack = " ".join(text_parts).casefold()
    if any(keyword in haystack for keyword in ("tapo", "智能家居", "smart home", "home security")):
        return "smart_home"
    if any(keyword in haystack for keyword in ("lifestyle", "街头", "采访", "剧情", "年轻人")):
        return "lifestyle"
    if any(keyword in haystack for keyword in ("电影", "影迷", "潮玩", "开箱", "盲盒", "coser", "diy")):
        return "lifestyle"
    return default_primary_category


def _build_project_code(task_name: str, *, prefix: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", task_name.casefold())
    normalized = normalized.strip("-").upper()
    return f"{prefix}{normalized or 'TASK'}"


def _safe_path_component(value: str) -> str:
    normalized = re.sub(r"[\\/]+", "-", str(value or "").strip())
    normalized = re.sub(r"\s+", "_", normalized)
    normalized = re.sub(r"[^\w.-]+", "-", normalized, flags=re.UNICODE)
    normalized = normalized.strip("._-")
    return normalized or "task-mail-sync"


def _extract_text_like(value: Any) -> str:
    if isinstance(value, dict):
        nested = value.get("value")
        if nested is not None:
            return _extract_text_like(nested)
        for key in ("text", "link", "name"):
            candidate = _normalize_text(value.get(key))
            if candidate:
                return candidate
        return ""
    if isinstance(value, list):
        parts = []
        for item in value:
            text = _extract_text_like(item)
            if text:
                parts.append(text)
        return "".join(parts).strip().strip(",")
    return _normalize_text(value)


def _extract_email_value(value: Any) -> str:
    values = _extract_email_values(value)
    if values:
        return values[0]
    if isinstance(value, dict):
        nested = value.get("value")
        if nested is not None:
            return _extract_email_value(nested)
    text = _extract_text_like(value)
    if text.startswith("mailto:"):
        return text[len("mailto:") :]
    return text


def _extract_email_values(value: Any) -> tuple[str, ...]:
    results: list[str] = []
    seen: set[str] = set()

    def add_candidate(raw: Any) -> None:
        text = _normalize_text(raw)
        if text.startswith("mailto:"):
            text = text[len("mailto:") :]
        for item in re.split(r"[,;，\\s]+", text):
            candidate = item.strip()
            if "@" not in candidate:
                continue
            key = candidate.casefold()
            if key in seen:
                continue
            seen.add(key)
            results.append(candidate)

    def walk(raw: Any) -> None:
        if isinstance(raw, dict):
            nested = raw.get("value")
            if nested is not None:
                walk(nested)
            for key in ("text", "link", "name", "email"):
                candidate = raw.get(key)
                if candidate is not None:
                    add_candidate(candidate)
            return
        if isinstance(raw, list):
            for item in raw:
                walk(item)
            return
        add_candidate(raw)

    walk(value)
    return tuple(results)


def _extract_attachment_with_file_token(value: Any) -> dict[str, Any] | None:
    if not isinstance(value, list):
        return None
    for item in value:
        if isinstance(item, dict) and str(item.get("file_token") or "").strip():
            return item
    return None


def _extract_person_name(value: Any) -> str:
    if isinstance(value, list):
        person = value[0] if value else None
        if isinstance(person, dict):
            return _normalize_text(person.get("name") or person.get("en_name"))
    return _extract_text_like(value)


def _extract_mention_link(value: Any) -> str:
    if isinstance(value, dict):
        return _normalize_text(value.get("link") or value.get("url"))
    return ""


def _require_text(field_index: dict[tuple[str, str], dict[str, Any]], section: str, field_name: str) -> str:
    value = _normalize_blank(_optional_text(field_index, section, field_name))
    if not value:
        raise ValueError(f"缺少必填字段: {section} / {field_name}")
    return value


def _optional_text(field_index: dict[tuple[str, str], dict[str, Any]], section: str, field_name: str) -> str:
    entry = field_index.get((section, field_name))
    return _normalize_text(entry["value"] if entry else "")


def _require_non_negative_int(field_index: dict[tuple[str, str], dict[str, Any]], section: str, field_name: str) -> int:
    value = _require_text(field_index, section, field_name)
    return _parse_non_negative_int(value, field_name)


def _optional_non_negative_int(
    field_index: dict[tuple[str, str], dict[str, Any]],
    section: str,
    field_name: str,
    *,
    default: int,
) -> int:
    value = _normalize_blank(_optional_text(field_index, section, field_name))
    if not value:
        return default
    return _parse_non_negative_int(value, field_name)


def _parse_non_negative_int(value: Any, field_name: str) -> int:
    text = _normalize_blank(_normalize_text(value))
    if not text:
        raise ValueError(f"{field_name} 缺少数值。")
    return int(float(text))


def _parse_platform_scope_extended(value: str) -> tuple[str, ...]:
    normalized = _normalize_blank(value)
    if not normalized:
        raise ValueError("适用平台不能为空。")
    if normalized in {"两者", "both"}:
        return ("tiktok", "instagram")

    tokens = [
        token.strip().lower()
        for token in re.split(r"[,/，]+", normalized)
        if token.strip()
    ]
    results: list[str] = []
    for token in tokens:
        mapped = _PLATFORM_TOKEN_MAP.get(token)
        if mapped == "both":
            for platform in ("tiktok", "instagram"):
                if platform not in results:
                    results.append(platform)
            continue
        if not mapped:
            raise ValueError(f"适用平台不合法: {value}")
        if mapped not in results:
            results.append(mapped)
    return tuple(results)


def _build_visual_constraints(field_index: dict[tuple[str, str], dict[str, Any]]) -> str:
    sections = []
    for section_name in (
        CANONICAL_SECTION_STEP_3,
        CANONICAL_SECTION_STEP_4,
        CANONICAL_SECTION_STEP_6,
        CANONICAL_SECTION_STEP_7,
    ):
        items = []
        for (entry_section, field_name), entry in field_index.items():
            if entry_section != section_name:
                continue
            value = _normalize_blank(_normalize_text(entry.get("value")))
            if not value:
                continue
            line = f"{field_name}: {value}"
            note = _normalize_blank(_normalize_text(entry.get("note")))
            if note:
                line += f"（说明：{note}）"
            items.append(line)
        if items:
            sections.append(f"{section_name}\n" + "\n".join(f"- {item}" for item in items))
    return "\n\n".join(sections)


def _normalize_blank(value: str) -> str:
    normalized = _normalize_text(value)
    if normalized in {"", "/", "／", "无", "None", "none"}:
        return ""
    return normalized


def _normalize_lookup_key(value: str) -> str:
    return _normalize_text(value).casefold()


def _first_non_empty(*values: str) -> str:
    for value in values:
        normalized = _normalize_text(value)
        if normalized:
            return normalized
    return ""


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _write_unique_file(directory: Path, file_name: str, content: bytes) -> Path:
    candidate = directory / Path(file_name).name
    if not candidate.exists():
        candidate.write_bytes(content)
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    counter = 1
    while True:
        next_candidate = candidate.with_name(f"{stem}-{counter}{suffix}")
        if not next_candidate.exists():
            next_candidate.write_bytes(content)
            return next_candidate
        counter += 1


def _resolve_path_from(base_dir: Path, path_value: Path) -> Path:
    path = Path(path_value).expanduser()
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


@contextmanager
def _pushd(path: Path) -> Any:
    previous = Path.cwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(previous)
