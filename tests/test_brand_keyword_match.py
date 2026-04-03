from __future__ import annotations

import csv
import json
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from email_sync.brand_keyword_match import match_brand_keyword, split_shared_email_candidates
from email_sync.db import Database


def _addresses(*pairs: tuple[str, str]) -> str:
    return json.dumps([{"name": name, "address": address} for name, address in pairs], ensure_ascii=False)


class BrandKeywordMatchTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="brand_keyword_match_test_"))
        self.db_path = self.temp_dir / "email_sync.db"
        self.input_path = self.temp_dir / "creators.xlsx"
        self.output_prefix = self.temp_dir / "exports" / "miniso_fast_path"

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _make_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["Platform", "@username", "URL", "nickname", "Region", "Email"])
        sheet.append(["Instagram", "alice", "https://instagram.com/alice", "Alice", "US", "alice@mgmt.com"])
        sheet.append(["Instagram", "alice", "https://instagram.com/alice", "Alice Duplicate", "US", "alt@agency.com"])
        sheet.append(["Instagram", "bob", "https://instagram.com/bob", "Bob", "UK", "shared@agency.com"])
        sheet.append(["Instagram", "cara", "https://instagram.com/cara", "Cara", "UK", "shared@agency.com"])
        workbook.save(self.input_path)

    def _make_custom_country_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["Creator", "Country", "IGlink", "Email"])
        sheet.append(["Alice", "US", "https://instagram.com/alice", "alice@mgmt.com"])
        workbook.save(self.input_path)

    def _make_handle_email_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["博主用户名", "地区", "平台", "邮箱"])
        sheet.append(["alice", "US", "Instagram", "alice@mgmt.com"])
        workbook.save(self.input_path)

    def _make_four_column_sending_list_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["地区", "博主用户名", "邮箱", "主页链接"])
        sheet.append(["US", "alice", "alice@mgmt.com", "https://instagram.com/alice"])
        workbook.save(self.input_path)

    def _make_formula_sending_list_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["地区", "博主用户名", "邮箱", "主页链接"])
        sheet["A2"] = "US"
        sheet["B2"] = '=HYPERLINK("https://instagram.com/alice","alice")'
        sheet["C2"] = '=HYPERLINK("mailto:alice@mgmt.com","alice@mgmt.com")'
        sheet["D2"] = "https://instagram.com/alice"
        workbook.save(self.input_path)

    def _seed_messages(self) -> Database:
        db = Database(self.db_path)
        db.init_schema()
        rows = [
            (
                "operator@example.com",
                "其他文件夹/MINISO",
                1,
                1,
                "<m1>",
                "MINISO outreach for Alice and agency roster",
                None,
                None,
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Brand Manager", "brand@miniso.com")),
                _addresses(("Alice Manager", "alice@mgmt.com"), ("Shared Manager", "shared@agency.com")),
                "[]",
                "[]",
                "[]",
                "[]",
                "MINISO wants to review Alice and a shared roster.",
                "",
                "MINISO wants to review Alice and a shared roster.",
                "{}",
                "raw/1.eml",
                "sha1",
                100,
                0,
                0,
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
            ),
            (
                "operator@example.com",
                "INBOX",
                2,
                1,
                "<m2>",
                "Re: MINISO outreach for Alice duplicate",
                "<m1>",
                "<m1>",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Alternate Manager", "alt@agency.com")),
                _addresses(("Brand Manager", "brand@miniso.com")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Following up on MINISO for Alice duplicate.",
                "",
                "Following up on MINISO for Alice duplicate.",
                "{}",
                "raw/2.eml",
                "sha2",
                100,
                0,
                0,
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
            ),
            (
                "operator@example.com",
                "INBOX",
                3,
                1,
                "<m3>",
                "Other brand outreach",
                None,
                None,
                "2026-03-12T10:00:00+00:00",
                "2026-03-12T10:00:00+00:00",
                "2026-03-12T10:00:00+00:00",
                "2026-03-12T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Brand Manager", "brand@example.com")),
                _addresses(("Other Manager", "other@agency.com")),
                "[]",
                "[]",
                "[]",
                "[]",
                "This should not match the keyword.",
                "",
                "This should not match the keyword.",
                "{}",
                "raw/3.eml",
                "sha3",
                100,
                0,
                0,
                "2026-03-12T10:00:00+00:00",
                "2026-03-12T10:00:00+00:00",
            ),
        ]
        db.conn.executemany(
            """
            INSERT INTO messages (
                account_email, folder_name, uid, uidvalidity, message_id, subject, in_reply_to, references_header,
                sent_at, sent_at_raw, internal_date, internal_date_raw, flags_json, size_bytes,
                from_json, to_json, cc_json, bcc_json, reply_to_json, sender_json,
                body_text, body_html, snippet, headers_json, raw_path, raw_sha256, raw_size_bytes,
                has_attachments, attachment_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        db.conn.commit()
        return db

    def test_match_brand_keyword_outputs_deduped_and_shared_email_artifacts(self) -> None:
        self._make_workbook()
        db = self._seed_messages()
        try:
            result = match_brand_keyword(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="MINISO",
                include_from=True,
            )
        finally:
            db.close()

        self.assertEqual(result["message_hit_count"], 2)
        self.assertEqual(result["matched_email_count"], 4)
        self.assertEqual(result["email_direct_match_row_count"], 4)
        self.assertEqual(result["profile_deduped_row_count"], 3)
        self.assertEqual(result["unique_email_row_count"], 1)
        self.assertEqual(result["shared_email_row_count"], 2)
        self.assertEqual(result["shared_email_group_count"], 1)

        with self.output_prefix.with_suffix(".csv").open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(len(rows), 4)

        with self.output_prefix.with_name(f"{self.output_prefix.name}_deduped").with_suffix(".csv").open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as handle:
            deduped_rows = list(csv.DictReader(handle))
        self.assertEqual(len(deduped_rows), 3)
        self.assertEqual(deduped_rows[0]["matched_email"], "alt@agency.com")
        self.assertEqual(deduped_rows[0]["Region"], "US")

        shared_xlsx = Path(result["shared_xlsx_path"])
        workbook = load_workbook(shared_xlsx, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        self.assertEqual(len(rows) - 1, 2)
        headers = list(rows[0])
        matched_email_index = headers.index("matched_email")
        distinct_count_index = headers.index("shared_email_distinct_profile_count")
        region_index = headers.index("Region")
        self.assertTrue(all(row[matched_email_index] == "shared@agency.com" for row in rows[1:]))
        self.assertTrue(all(row[distinct_count_index] == 2 for row in rows[1:]))
        self.assertTrue(all(row[region_index] == "UK" for row in rows[1:]))

    def test_split_shared_email_can_reprocess_deduped_output(self) -> None:
        self._make_workbook()
        db = self._seed_messages()
        try:
            result = match_brand_keyword(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="MINISO",
                include_from=True,
            )
        finally:
            db.close()

        split_prefix = self.temp_dir / "exports" / "manual_split"
        split_result = split_shared_email_candidates(
            input_path=Path(result["deduped_xlsx_path"]),
            output_prefix=split_prefix,
        )
        self.assertEqual(split_result["unique_email_row_count"], 1)
        self.assertEqual(split_result["shared_email_row_count"], 2)
        self.assertEqual(split_result["shared_email_group_count"], 1)

    def test_match_brand_keyword_maps_country_column_to_region_for_custom_headers(self) -> None:
        self._make_custom_country_workbook()
        db = self._seed_messages()
        try:
            result = match_brand_keyword(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="MINISO",
                include_from=True,
            )
        finally:
            db.close()

        self.assertEqual(result["source_kind"], "custom_columns")
        deduped_path = Path(result["deduped_xlsx_path"])
        workbook = load_workbook(deduped_path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        headers = list(rows[0])
        region_index = headers.index("Region")
        self.assertEqual(rows[1][region_index], "US")

    def test_match_brand_keyword_accepts_blog_username_and_email_columns(self) -> None:
        self._make_handle_email_workbook()
        db = self._seed_messages()
        try:
            result = match_brand_keyword(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="MINISO",
                include_from=True,
            )
        finally:
            db.close()

        self.assertEqual(result["source_kind"], "custom_columns")
        deduped_path = Path(result["deduped_xlsx_path"])
        workbook = load_workbook(deduped_path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        headers = list(rows[0])
        username_index = headers.index("@username")
        email_index = headers.index("Email")
        self.assertEqual(rows[1][username_index], "alice")
        self.assertEqual(rows[1][email_index], "alice@mgmt.com")

    def test_match_brand_keyword_accepts_four_column_sending_list(self) -> None:
        self._make_four_column_sending_list_workbook()
        db = self._seed_messages()
        try:
            result = match_brand_keyword(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="MINISO",
                include_from=True,
            )
        finally:
            db.close()

        self.assertEqual(result["source_kind"], "custom_columns")
        deduped_path = Path(result["deduped_xlsx_path"])
        workbook = load_workbook(deduped_path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        headers = list(rows[0])
        username_index = headers.index("@username")
        url_index = headers.index("URL")
        email_index = headers.index("Email")
        self.assertEqual(rows[1][username_index], "alice")
        self.assertEqual(rows[1][url_index], "https://instagram.com/alice")
        self.assertEqual(rows[1][email_index], "alice@mgmt.com")

    def test_match_brand_keyword_accepts_hyperlink_formula_sending_list(self) -> None:
        self._make_formula_sending_list_workbook()
        db = self._seed_messages()
        try:
            result = match_brand_keyword(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="MINISO",
                include_from=True,
            )
        finally:
            db.close()

        self.assertEqual(result["source_kind"], "custom_columns")
        self.assertEqual(result["email_direct_match_row_count"], 1)
        deduped_path = Path(result["deduped_xlsx_path"])
        workbook = load_workbook(deduped_path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        headers = list(rows[0])
        username_index = headers.index("@username")
        email_index = headers.index("Email")
        matched_email_index = headers.index("matched_email")
        self.assertEqual(rows[1][username_index], "alice")
        self.assertEqual(rows[1][email_index], "mailto:alice@mgmt.com")
        self.assertEqual(rows[1][matched_email_index], "alice@mgmt.com")


if __name__ == "__main__":
    unittest.main()
