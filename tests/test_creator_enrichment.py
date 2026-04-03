from __future__ import annotations

import csv
import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

import email_sync.creator_enrichment as creator_enrichment
from email_sync.creator_enrichment import enrich_creator_workbook
from email_sync.db import Database
from email_sync.relation_index import rebuild_relation_index


def _addresses(*pairs: tuple[str, str]) -> str:
    return json.dumps([{"name": name, "address": address} for name, address in pairs], ensure_ascii=False)


class CreatorEnrichmentTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="creator_enrichment_test_"))
        self.db_path = self.temp_dir / "email_sync.db"
        self.input_path = self.temp_dir / "creators.xlsx"
        self.output_prefix = self.temp_dir / "exports" / "达人邮件可获取信息_v1"

    def _make_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TikTok"
        sheet.append(
            [
                "Project",
                "Member Email",
                "Likes Time",
                "nickname",
                "@username",
                "Region",
                "Platform",
                "Followers",
                "Avg. Views",
                "Avg. Likes",
                "is SuperLike",
                "Tags",
                "Email",
                "Email Send Status",
                "YouTube Description Email Button",
                "Last Post",
                "Posts (7d)>",
                "Posts (30d)>",
                "URL",
            ]
        )
        sheet.append(
            [
                "Test",
                "",
                "",
                "Creator X",
                "creatorx",
                "US",
                "TIKTOK",
                10000,
                2000,
                100,
                "",
                "",
                "creator@example.com",
                "",
                "",
                "",
                0,
                0,
                "https://tiktok.com/@creatorx",
            ]
        )
        sheet.append(
            [
                "Test",
                "",
                "",
                "Cass And Home",
                "cass-and-home",
                "US",
                "YOUTUBE",
                50000,
                5000,
                300,
                "",
                "",
                "",
                "",
                "",
                "",
                0,
                0,
                "https://youtube.com/@cass-and-home",
            ]
        )
        workbook.save(self.input_path)

    def _make_sending_list_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SendingList"
        sheet.append(["国家", "Creator", "邮箱地址", "TTlink", "YTlink"])
        sheet.append(["US", "Creator X", "creator@example.com", "https://www.tiktok.com/@creatorx", ""])
        sheet.append(["US", "Cass And Home", "", "", "https://www.youtube.com/@cass-and-home"])
        workbook.save(self.input_path)

    def _make_four_column_sending_list_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SendingList"
        sheet.append(["地区", "邮箱", "博主用户名", "主页链接"])
        sheet.append(["US", "creator@example.com", "@creatorx", "https://www.tiktok.com/@creatorx"])
        sheet.append(["US", "", "cass-and-home", "https://www.youtube.com/@cass-and-home"])
        workbook.save(self.input_path)

    def _make_formula_sending_list_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SendingList"
        sheet.append(["地区", "博主用户名", "邮箱", "主页链接"])
        sheet["A2"] = "US"
        sheet["B2"] = '=HYPERLINK("https://www.tiktok.com/@creatorx","creatorx")'
        sheet["C2"] = '=HYPERLINK("mailto:creator@example.com","creator@example.com")'
        sheet["D2"] = "https://www.tiktok.com/@creatorx"
        workbook.save(self.input_path)

    def test_low_level_header_detection_handles_partial_read_only_rows(self) -> None:
        class FakeSheet:
            def __init__(self) -> None:
                self.title = "SendingList"
                self.max_row = 2
                self.max_column = 4
                self._rows = [
                    ["地区", "邮箱", "博主用户名", "主页链接"],
                    ["US", "creator@example.com", "@creatorx", "https://www.tiktok.com/@creatorx"],
                ]

            def cell(self, row_number: int, column_number: int) -> SimpleNamespace:
                value = ""
                if 1 <= row_number <= len(self._rows):
                    row = self._rows[row_number - 1]
                    if 1 <= column_number <= len(row):
                        value = row[column_number - 1]
                return SimpleNamespace(value=value)

            def iter_rows(self, min_row: int = 1, max_row: int | None = None, values_only: bool = False):
                end_row = max_row or self.max_row
                for row_number in range(min_row, end_row + 1):
                    row = self._rows[row_number - 1]
                    if row_number == 1:
                        row = row[:1]
                    if values_only:
                        yield tuple(row)
                    else:
                        yield tuple(SimpleNamespace(value=value) for value in row)

        class FakeWorkbook:
            def __init__(self) -> None:
                self.worksheets = [FakeSheet()]

            def close(self) -> None:
                return None

        original_loader = creator_enrichment._load_workbook
        creator_enrichment._load_workbook = lambda: (None, lambda **_: FakeWorkbook())
        try:
            headers = creator_enrichment._source_headers(self.input_path)
            rows = list(creator_enrichment._iter_sheet_rows(self.input_path, headers))
        finally:
            creator_enrichment._load_workbook = original_loader

        self.assertEqual(headers, ["地区", "邮箱", "博主用户名", "主页链接"])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["地区"], "US")
        self.assertEqual(rows[0]["邮箱"], "creator@example.com")
        self.assertEqual(rows[0]["博主用户名"], "@creatorx")
        self.assertEqual(rows[0]["主页链接"], "https://www.tiktok.com/@creatorx")

    def _seed_messages(self) -> Database:
        db = Database(self.db_path)
        db.init_schema()
        rows = [
            (
                "william@amagency.biz",
                "Sent Messages",
                1,
                1,
                "<m1>",
                "Paid Collaboration with Brand - @creatorx",
                None,
                None,
                "2025-01-01T10:00:00+00:00",
                "2025-01-01T10:00:00+00:00",
                "2025-01-01T10:00:00+00:00",
                "2025-01-01T10:00:00+00:00",
                "[]",
                100,
                _addresses(("William", "william@amagency.biz")),
                _addresses(("Creator X", "creator@example.com")),
                "[]",
                "[]",
                "[]",
                "[]",
                "We can offer $200 for one video.",
                "",
                "We can offer $200 for one video.",
                "{}",
                "raw/1.eml",
                "sha1",
                100,
                0,
                0,
                "2025-01-01T10:00:00+00:00",
                "2025-01-01T10:00:00+00:00",
            ),
            (
                "william@amagency.biz",
                "INBOX",
                2,
                1,
                "<m2>",
                "Re: Paid Collaboration with Brand - @creatorx",
                "<m1>",
                "<m1>",
                "2025-01-02T10:00:00+00:00",
                "2025-01-02T10:00:00+00:00",
                "2025-01-02T10:00:00+00:00",
                "2025-01-02T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Creator X", "creator@example.com")),
                _addresses(("William", "william@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "My rate for a TikTok and repost is $600.",
                "",
                "My rate for a TikTok and repost is $600.",
                "{}",
                "raw/2.eml",
                "sha2",
                100,
                0,
                0,
                "2025-01-02T10:00:00+00:00",
                "2025-01-02T10:00:00+00:00",
            ),
            (
                "william@amagency.biz",
                "INBOX",
                3,
                1,
                "<m3>",
                "Re: Exciting Collaboration with Anker | We Love Your Content! @cass-and-home",
                None,
                None,
                "2025-01-03T10:00:00+00:00",
                "2025-01-03T10:00:00+00:00",
                "2025-01-03T10:00:00+00:00",
                "2025-01-03T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Cass", "cass@undercurrent.net")),
                _addresses(("William", "william@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "The dock supports 8K output and 4K quality, but let's discuss rates separately for @cass-and-home.",
                "",
                "The dock supports 8K output and 4K quality.",
                "{}",
                "raw/3.eml",
                "sha3",
                100,
                0,
                0,
                "2025-01-03T10:00:00+00:00",
                "2025-01-03T10:00:00+00:00",
            ),
        ]
        db.conn.executemany(
            """
            INSERT INTO messages (
                account_email, folder_name, uid, uidvalidity, message_id, subject, in_reply_to, references_header,
                sent_at, sent_at_raw, internal_date, internal_date_raw, flags_json, size_bytes,
                from_json, to_json, cc_json, bcc_json, reply_to_json, sender_json,
                body_text, body_html, snippet, headers_json, raw_path, raw_sha256, raw_size_bytes,
                has_attachments, attachment_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        db.conn.commit()
        rebuild_relation_index(db)
        return db

    def test_enrich_creator_workbook_outputs_last_mail_and_quote(self) -> None:
        self._make_workbook()
        db = self._seed_messages()
        try:
            result = enrich_creator_workbook(db, self.input_path, self.output_prefix)
        finally:
            db.close()

        self.assertEqual(result["rows"], 2)
        self.assertEqual(result["matched_rows"], 2)

        with self.output_prefix.with_suffix(".csv").open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        creator_row = rows[0]
        self.assertEqual(creator_row["match_rule"], "email_exact")
        self.assertEqual(creator_row["last_mail_subject"], "Re: Paid Collaboration with Brand - @creatorx")
        self.assertEqual(creator_row["latest_quote_amount"], "600.0")

        handle_row = rows[1]
        self.assertEqual(handle_row["derived_handle"], "cassandhome")
        self.assertEqual(handle_row["latest_quote_amount"], "")

    def test_enrich_creator_workbook_accepts_sending_list_workbook(self) -> None:
        self._make_sending_list_workbook()
        db = self._seed_messages()
        try:
            result = enrich_creator_workbook(db, self.input_path, self.output_prefix)
        finally:
            db.close()

        self.assertEqual(result["source_kind"], "sending_list")
        self.assertEqual(result["rows"], 2)
        self.assertEqual(result["matched_rows"], 2)

        with self.output_prefix.with_suffix(".csv").open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        creator_row = rows[0]
        self.assertEqual(creator_row["Platform"], "TikTok")
        self.assertEqual(creator_row["@username"], "creatorx")
        self.assertEqual(creator_row["match_rule"], "email_exact")
        self.assertEqual(creator_row["latest_quote_amount"], "600.0")

        handle_row = rows[1]
        self.assertEqual(handle_row["Platform"], "YouTube")
        self.assertEqual(handle_row["@username"], "cassandhome")
        self.assertEqual(handle_row["match_confidence"], "high")

    def test_real_duet_sending_list_headers_are_all_detected(self) -> None:
        duet_path = Path(
            "temp/runs/task_upload_to_final_export/20260402_145244_duet_8864d6e5/children/01_Duet1/upstream/downloads/reczxcYykT/发信名单/duet 冷名单.xlsx"
        )
        if not duet_path.exists():
            self.skipTest(f"missing fixture workbook: {duet_path}")

        headers = creator_enrichment._source_headers(duet_path)
        rows = list(creator_enrichment._iter_sending_list_rows(duet_path))

        self.assertEqual(headers[:4], ["地区", "邮箱", "博主用户名", "主页链接"])
        self.assertGreater(len(rows), 1000)
        self.assertIn("Platform", rows[0])
        self.assertIn("@username", rows[0])

    def test_enrich_creator_workbook_accepts_four_column_sending_list_workbook(self) -> None:
        self._make_four_column_sending_list_workbook()
        db = self._seed_messages()
        try:
            result = enrich_creator_workbook(db, self.input_path, self.output_prefix)
        finally:
            db.close()

        self.assertEqual(result["source_kind"], "sending_list")
        self.assertEqual(result["rows"], 2)
        self.assertEqual(result["matched_rows"], 2)

        with self.output_prefix.with_suffix(".csv").open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        creator_row = rows[0]
        self.assertEqual(creator_row["Platform"], "TikTok")
        self.assertEqual(creator_row["@username"], "creatorx")
        self.assertEqual(creator_row["match_rule"], "email_exact")

        handle_row = rows[1]
        self.assertEqual(handle_row["Platform"], "YouTube")
        self.assertEqual(handle_row["@username"], "cassandhome")
        self.assertEqual(handle_row["match_confidence"], "high")

    def test_iter_sending_list_rows_extracts_hyperlink_formula_targets(self) -> None:
        self._make_formula_sending_list_workbook()

        headers = creator_enrichment._source_headers(self.input_path)
        rows = list(creator_enrichment._iter_sending_list_rows(self.input_path))

        self.assertEqual(headers[:4], ["地区", "博主用户名", "邮箱", "主页链接"])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Platform"], "TikTok")
        self.assertEqual(rows[0]["@username"], "creatorx")
        self.assertEqual(rows[0]["Email"], "mailto:creator@example.com")
        self.assertEqual(rows[0]["URL"], "https://www.tiktok.com/@creatorx")


if __name__ == "__main__":
    unittest.main()
