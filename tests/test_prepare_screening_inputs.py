from __future__ import annotations

import json
import tempfile
from pathlib import Path
import unittest
from unittest.mock import patch

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_TEMPLATE = REPO_ROOT / "tests" / "fixtures" / "template_parser" / "11.xlsx"

IMPORT_ERROR: Exception | None = None
prepare_screening_inputs = None

try:
    from scripts.prepare_screening_inputs import prepare_screening_inputs
except Exception as exc:  # pragma: no cover - dependency availability differs by runtime
    IMPORT_ERROR = exc


def build_creator_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"
    sheet.append(["Platform", "@username", "nickname", "Region", "URL"])
    sheet.append(["Instagram", "@creatoralpha", "Alpha", "US", "https://www.instagram.com/creatoralpha/"])
    sheet.append(["TikTok", "@creatorbeta", "Beta", "US", "https://www.tiktok.com/@creatorbeta"])
    workbook.save(path)


def build_creator_workbook_with_tiktok_search_url(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"
    sheet.append(["Platform", "@username", "nickname", "Region", "URL"])
    sheet.append([
        "TikTok",
        "httpswww.tiktok.comsearchqtinozacht1773382255532",
        "tinozach",
        "US",
        "https://www.tiktok.com/search?q=tinozach&t=1773382255532",
    ])
    workbook.save(path)


def build_sending_list_workbook(path: Path) -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "TikTokPool"
    first_sheet.append(["Country", "Creator", "邮箱地址", "IGlink"])
    first_sheet.append(["US", "Creator TikTok", "tiktok@example.com", "https://www.tiktok.com/@creatorbeta"])
    first_sheet.append(["US", "Creator Insta", "insta@example.com", "https://www.instagram.com/creatoralpha/"])

    second_sheet = workbook.create_sheet("Mixed")
    second_sheet.append(["国家", "Creator", "邮箱地址", "IGlink", "TTlink", "YTlink"])
    second_sheet.append([
        "US",
        "Creator Mixed",
        "mixed@example.com",
        "https://www.instagram.com/creatormixed/",
        "@creatormixed",
        "https://www.youtube.com/@creatormixed",
    ])
    workbook.save(path)


def build_four_column_sending_list_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"
    sheet.append(["地区", "博主用户名", "邮箱", "主页链接"])
    sheet.append(["US", "@creatoralpha", "insta@example.com", "https://www.instagram.com/creatoralpha/"])
    sheet.append(["US", "creatorbeta", "tiktok@example.com", "https://www.tiktok.com/@creatorbeta"])
    workbook.save(path)


def build_keep_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append([
        "Project",
        "nickname",
        "@username",
        "Region",
        "Platform",
        "URL",
        "Email",
        "creator_dedupe_key",
        "llm_review_decision",
    ])
    sheet.append([
        "MINISO",
        "Creator Alpha",
        "@creatoralpha",
        "US",
        "Instagram",
        "https://www.instagram.com/creatoralpha/",
        "insta@example.com",
        "instagram:creatoralpha",
        "match_all",
    ])
    sheet.append([
        "MINISO",
        "Creator Beta",
        "@creatorbeta",
        "US",
        "TikTok",
        "https://www.tiktok.com/@creatorbeta",
        "tiktok@example.com",
        "tiktok:creatorbeta",
        "match_all",
    ])
    sheet.append([
        "MINISO",
        "Creator Gamma",
        "@creatoryt",
        "US",
        "YouTube",
        "https://www.youtube.com/@creatoryt",
        "yt@example.com",
        "youtube:creatoryt",
        "match_all",
    ])
    workbook.save(path)


def build_keep_workbook_with_missing_platform(path: Path) -> None:
    build_keep_workbook(path)
    from openpyxl import load_workbook

    loaded = load_workbook(path)
    sheet = loaded[loaded.sheetnames[0]]
    sheet.append([
        "MINISO",
        "samandcitra90day",
        "samandcitra90day",
        "",
        None,
        "Sam&Citra 90 day fiancè (@samandcitra90day) | TikTok",
        "booking.samandcitra@outlook.com",
        "tiktok:samandcitra90day",
        "match_all",
    ])
    loaded.save(path)


def build_mail_thread_funnel_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "threads"
    sheet.append([
        "Platform",
        "thread_key",
        "subject",
        "latest_external_from",
        "latest_external_sent_at",
        "latest_external_clean_body",
        "latest_external_full_body",
        "resolution_stage_final",
        "resolution_confidence_final",
        "final_id_final",
        "llm_handle",
        "raw_path",
    ])
    sheet.append([
        "TikTok",
        "thread-1",
        "Re: Paid collaboration with SKG",
        "manager@example.com",
        "2026-04-02T19:31:15+02:00",
        "Hello Lilith, interested.",
        "Hello Lilith\n\n> Hi @ livio.official ,",
        "regex_pass1",
        "",
        "livio.official",
        "",
        "raw/thread-1.eml",
    ])
    sheet.append([
        "TikTok",
        "thread-2",
        "Automatic Reply: Paid collaboration with SKG",
        "auto.reply@example.com",
        "2026-04-03T06:00:00+08:00",
        "",
        "Thank you for your email. I am currently out of office and will get back to you soon.",
        "llm",
        "high",
        "auto.reply.user",
        "auto.reply.user",
        "raw/thread-2.eml",
    ])
    sheet.append([
        "TikTok",
        "thread-3",
        "Re: Paid collaboration with SKG",
        "team@example.com",
        "2026-04-03T07:15:00+08:00",
        "",
        "Hallo Lilith,\n\nvielen Dank.\n\n> Hi @ maggy_valentine ,",
        "regex_pass1",
        "",
        "maggy_valentine",
        "",
        "raw/thread-3.eml",
    ])
    sheet.append([
        "TikTok",
        "thread-4",
        "Re: Paid collaboration with SKG",
        "agent@example.com",
        "2026-04-03T08:15:00+08:00",
        "Could work. Sharing rates below.",
        "Hello Astrid\n\ncreator details ...",
        "llm",
        "high",
        "high.confidence.creator",
        "high.confidence.creator",
        "raw/thread-4.eml",
    ])
    sheet.append([
        "TikTok",
        "thread-5",
        "Re: Paid collaboration with SKG",
        "agent2@example.com",
        "2026-04-03T08:30:00+08:00",
        "Maybe relevant",
        "Hello Astrid\n\ncreator details ...",
        "llm",
        "medium",
        "medium.confidence.creator",
        "medium.confidence.creator",
        "raw/thread-5.eml",
    ])
    sheet.append([
        "TikTok",
        "thread-6",
        "Re: Paid collaboration with SKG",
        "agent3@example.com",
        "2026-04-03T08:45:00+08:00",
        "Maybe relevant",
        "Hello Astrid\n\ncreator details ...",
        "weak_rule",
        "high",
        "weak.rule.creator",
        "llm.upgraded.creator",
        "raw/thread-6.eml",
    ])
    sheet.append([
        "TikTok",
        "thread-7",
        "Re: Paid collaboration with OTHERBRAND",
        "brand@example.com",
        "2026-04-03T09:00:00+08:00",
        "Hello Astrid, interested.",
        "Hello Astrid\n\nWe are excited about OTHERBRAND.\n\n> Hi @ wrong.brand.creator ,",
        "regex_pass1",
        "",
        "wrong.brand.creator",
        "",
        "raw/thread-7.eml",
    ])
    workbook.save(path)


@unittest.skipIf(prepare_screening_inputs is None, f"screening deps unavailable: {IMPORT_ERROR}")
class PrepareScreeningInputsTests(unittest.TestCase):
    def test_prepare_screening_inputs_persists_rulespec_and_upload_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "creator_upload.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"
            summary_json = tmp_path / "summary.json"

            build_creator_workbook(creator_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
                summary_json=summary_json,
            )

            self.assertEqual(summary["rulespec"]["source"], "template_workbook", summary)
            self.assertEqual(summary["env_file_raw"], ".env", summary)
            self.assertEqual(summary["env_file"], summary["resolved_inputs"]["env_file"]["path"], summary)
            self.assertIn("resolved_config_sources", summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["instagram"], 1, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 1, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["youtube"], 0, summary)

            active_rulespec_path = Path(summary["active_rulespec_path"])
            active_visual_prompts_path = Path(summary["active_visual_prompts_path"])
            self.assertTrue(active_rulespec_path.exists(), active_rulespec_path)
            self.assertTrue(active_visual_prompts_path.exists(), active_visual_prompts_path)
            self.assertTrue(Path(summary["rulespec"]["visual_prompts_json_path"]).exists(), summary)
            self.assertTrue(Path(summary["rulespec"]["runtime_prompt_artifacts_json_path"]).exists(), summary)

            instagram_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["instagram"])
            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            self.assertTrue(instagram_metadata_path.exists(), instagram_metadata_path)
            self.assertTrue(tiktok_metadata_path.exists(), tiktok_metadata_path)

            instagram_metadata = json.loads(instagram_metadata_path.read_text(encoding="utf-8"))
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))
            self.assertIn("creatoralpha", instagram_metadata, instagram_metadata)
            self.assertIn("creatorbeta", tiktok_metadata, tiktok_metadata)
            prompt_artifacts = json.loads(Path(summary["rulespec"]["runtime_prompt_artifacts_json_path"]).read_text(encoding="utf-8"))
            self.assertIn("instagram", prompt_artifacts["platforms"], prompt_artifacts)
            self.assertEqual(
                prompt_artifacts["platforms"]["instagram"]["visual_review"]["prompt_source"],
                "platform_prompt",
                prompt_artifacts,
            )
            self.assertIn(
                "达人：{{username}}",
                prompt_artifacts["platforms"]["instagram"]["visual_review"]["preview_prompt"],
                prompt_artifacts,
            )
            self.assertEqual(
                prompt_artifacts["platforms"]["instagram"]["positioning_card_analysis"]["prompt_source"],
                "generic_brand_fit",
                prompt_artifacts,
            )

            self.assertTrue(summary_json.exists(), summary_json)

    def test_prepare_screening_inputs_prefers_tiktok_search_query_over_bad_handle(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "creator_upload_search.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"

            build_creator_workbook_with_tiktok_search_url(creator_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))
            self.assertIn("tinozach", tiktok_metadata, tiktok_metadata)
            self.assertNotIn("httpswww.tiktok.comsearchqtinozacht1773382255532", tiktok_metadata, tiktok_metadata)
            self.assertEqual(
                tiktok_metadata["tinozach"]["url"],
                "https://www.tiktok.com/@tinozach",
                tiktok_metadata["tinozach"],
            )

    def test_prepare_screening_inputs_normalizes_sending_list_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            sending_list_workbook = tmp_path / "sending_list.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"

            build_sending_list_workbook(sending_list_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=sending_list_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["upload"]["parsed_source_kind"], "sending_list", summary)
            self.assertTrue(summary["upload"]["normalized_upload_source_path"], summary)
            self.assertTrue(Path(summary["upload"]["normalized_upload_source_path"]).exists(), summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["instagram"], 2, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 2, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["youtube"], 1, summary)

            instagram_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["instagram"])
            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            youtube_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["youtube"])
            instagram_metadata = json.loads(instagram_metadata_path.read_text(encoding="utf-8"))
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))
            youtube_metadata = json.loads(youtube_metadata_path.read_text(encoding="utf-8"))

            self.assertEqual(instagram_metadata["creatoralpha"]["email"], "insta@example.com", instagram_metadata)
            self.assertEqual(tiktok_metadata["creatormixed"]["email"], "mixed@example.com", tiktok_metadata)
            self.assertEqual(youtube_metadata["creatormixed"]["region"], "US", youtube_metadata)

    def test_prepare_screening_inputs_accepts_four_column_sending_list_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            sending_list_workbook = tmp_path / "sending_list_4col.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"

            build_four_column_sending_list_workbook(sending_list_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=sending_list_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["upload"]["parsed_source_kind"], "sending_list", summary)
            instagram_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["instagram"])
            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            instagram_metadata = json.loads(instagram_metadata_path.read_text(encoding="utf-8"))
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))

            self.assertIn("creatoralpha", instagram_metadata, instagram_metadata)
            self.assertIn("creatorbeta", tiktok_metadata, tiktok_metadata)
            self.assertEqual(instagram_metadata["creatoralpha"]["email"], "insta@example.com", instagram_metadata)
            self.assertEqual(tiktok_metadata["creatorbeta"]["region"], "US", tiktok_metadata)

    def test_prepare_screening_inputs_accepts_keep_workbook_and_reports_row_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            keep_workbook = tmp_path / "miniso_llm_reviewed_keep.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            summary_json = tmp_path / "summary.json"

            build_keep_workbook(keep_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=keep_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
                summary_json=summary_json,
            )

            self.assertEqual(summary["creator_workbook"], str(keep_workbook), summary)
            self.assertEqual(summary["parsed_source_kind"], "keep_list", summary)
            self.assertEqual(summary["input_row_count"], 3, summary)
            self.assertEqual(summary["upload"]["parsed_source_kind"], "keep_list", summary)
            self.assertEqual(summary["upload"]["input_row_count"], 3, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["instagram"], 1, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 1, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["youtube"], 1, summary)
            self.assertTrue(summary_json.exists(), summary_json)

    def test_prepare_screening_inputs_infers_platform_and_canonical_url_for_keep_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            keep_workbook = tmp_path / "miniso_llm_reviewed_keep_with_gap.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"

            build_keep_workbook_with_missing_platform(keep_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=keep_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["upload"]["parsed_source_kind"], "keep_list", summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 2, summary)
            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))
            self.assertIn("samandcitra90day", tiktok_metadata, tiktok_metadata)

    def test_prepare_screening_inputs_accepts_mail_thread_funnel_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "mail_thread_funnel.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"

            build_mail_thread_funnel_workbook(creator_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["upload"]["parsed_source_kind"], "mail_thread_funnel", summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 5, summary)
            self.assertTrue(summary["upload"]["normalized_upload_source_path"])
            self.assertEqual(summary["upload"]["normalized_upload_summary"]["autoReplySkippedCount"], 1, summary)
            self.assertEqual(summary["upload"]["normalized_upload_summary"]["evidenceFallbackCount"], 1, summary)
            self.assertEqual(summary["upload"]["normalized_upload_summary"]["manualReviewSkippedCount"], 1, summary)
            self.assertEqual(summary["upload"]["normalized_upload_summary"]["llmHighAcceptedCount"], 2, summary)
            self.assertEqual(summary["upload"]["normalized_upload_summary"]["llmNonHighSkippedCount"], 1, summary)
            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))
            self.assertIn("livio.official", tiktok_metadata, tiktok_metadata)
            self.assertIn("maggy_valentine", tiktok_metadata, tiktok_metadata)
            self.assertIn("high.confidence.creator", tiktok_metadata, tiktok_metadata)
            self.assertIn("llm.upgraded.creator", tiktok_metadata, tiktok_metadata)
            self.assertNotIn("auto.reply.user", tiktok_metadata, tiktok_metadata)
            self.assertNotIn("medium.confidence.creator", tiktok_metadata, tiktok_metadata)
            self.assertIn("wrong.brand.creator", tiktok_metadata, tiktok_metadata)
            self.assertEqual(
                tiktok_metadata["livio.official"]["instagram_url"],
                "https://www.instagram.com/livio.official/",
            )
            self.assertEqual(
                tiktok_metadata["livio.official"]["youtube_url"],
                "https://www.youtube.com/@livio.official",
            )
            self.assertEqual(
                tiktok_metadata["livio.official"]["mail_resolution_stage"],
                "regex_pass1",
            )
            self.assertEqual(
                tiktok_metadata["maggy_valentine"]["mail_evidence"],
                "Hallo Lilith,\n\nvielen Dank.\n\n> Hi @ maggy_valentine ,",
            )
            self.assertEqual(
                tiktok_metadata["high.confidence.creator"]["mail_resolution_stage"],
                "llm",
            )
            self.assertEqual(
                tiktok_metadata["high.confidence.creator"]["mail_resolution_confidence"],
                "high",
            )
            self.assertEqual(
                tiktok_metadata["high.confidence.creator"]["mail_apify_gate"],
                "ready_for_apify",
            )
            self.assertEqual(
                tiktok_metadata["livio.official"]["latest_external_sent_at"],
                "2026-04-03T01:31:15+08:00",
            )
            self.assertEqual(
                tiktok_metadata["llm.upgraded.creator"]["mail_resolution_stage"],
                "llm",
            )
            self.assertEqual(
                tiktok_metadata["llm.upgraded.creator"]["mail_resolution_confidence"],
                "high",
            )

    def test_prepare_screening_inputs_respects_platform_column_for_mail_thread_funnel(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "mail_thread_funnel.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"

            build_mail_thread_funnel_workbook(creator_workbook)

            from openpyxl import load_workbook

            workbook = load_workbook(creator_workbook)
            sheet = workbook.active
            platform_col = 1
            sheet.cell(2, platform_col).value = "Instagram"
            sheet.cell(4, platform_col).value = "TikTok"
            sheet.cell(5, platform_col).value = "YouTube"
            sheet.cell(7, platform_col).value = "Instagram"
            sheet.cell(8, platform_col).value = "TikTok"
            workbook.save(creator_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["upload"]["parsed_source_kind"], "mail_thread_funnel", summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["instagram"], 2, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 2, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["youtube"], 1, summary)

            instagram_metadata = json.loads(Path(summary["upload"]["upload_metadata_paths"]["instagram"]).read_text(encoding="utf-8"))
            tiktok_metadata = json.loads(Path(summary["upload"]["upload_metadata_paths"]["tiktok"]).read_text(encoding="utf-8"))
            youtube_metadata = json.loads(Path(summary["upload"]["upload_metadata_paths"]["youtube"]).read_text(encoding="utf-8"))

            self.assertIn("livio.official", instagram_metadata, instagram_metadata)
            self.assertIn("llm.upgraded.creator", instagram_metadata, instagram_metadata)
            self.assertIn("maggy_valentine", tiktok_metadata, tiktok_metadata)
            self.assertIn("wrong.brand.creator", tiktok_metadata, tiktok_metadata)
            self.assertIn("high.confidence.creator", youtube_metadata, youtube_metadata)
            self.assertEqual(
                instagram_metadata["livio.official"]["url"],
                "https://www.instagram.com/livio.official/",
            )
            self.assertEqual(
                youtube_metadata["high.confidence.creator"]["url"],
                "https://www.youtube.com/@high.confidence.creator",
            )

    def test_prepare_screening_inputs_filters_brand_mismatch_for_mail_thread_funnel(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "mail_thread_funnel.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            template_output_dir = tmp_path / "parsed_outputs"

            build_mail_thread_funnel_workbook(creator_workbook)

            summary = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                task_name="SKG",
                template_output_dir=template_output_dir,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["upload"]["parsed_source_kind"], "mail_thread_funnel", summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 4, summary)
            self.assertEqual(summary["upload"]["normalized_upload_summary"]["brandFilteredSkippedCount"], 1, summary)
            tiktok_metadata_path = Path(summary["upload"]["upload_metadata_paths"]["tiktok"])
            tiktok_metadata = json.loads(tiktok_metadata_path.read_text(encoding="utf-8"))
            self.assertNotIn("wrong.brand.creator", tiktok_metadata, tiktok_metadata)

    def test_prepare_screening_inputs_can_source_task_upload_assets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "downloaded_sending_list.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"
            summary_json = tmp_path / "summary.json"

            build_sending_list_workbook(creator_workbook)

            fake_task_source = {
                "recordId": "rec-task-miniso",
                "taskName": "MINISO",
                "taskUploadUrl": "https://example.com/wiki/task-upload",
                "downloadDir": str(tmp_path / "downloads"),
                "templateFileToken": "file-template-001",
                "templateFileName": "miniso-template.xlsx",
                "templateDownloadedPath": str(FIXTURE_TEMPLATE),
                "sendingListFileToken": "file-sending-001",
                "sendingListFileName": "miniso-sending-list.xlsx",
                "sendingListDownloadedPath": str(creator_workbook),
            }

            with patch(
                "scripts.prepare_screening_inputs.resolve_task_upload_source_files",
                return_value=fake_task_source,
            ) as mocked_resolver:
                task_download_dir = tmp_path / "downloads"
                template_output_dir = tmp_path / "parsed_outputs"
                summary = prepare_screening_inputs(
                    task_name="MINISO",
                    task_upload_url="https://example.com/wiki/task-upload",
                    task_download_dir=task_download_dir,
                    template_output_dir=template_output_dir,
                    screening_data_dir=screening_data_dir,
                    config_dir=config_dir,
                    temp_dir=temp_dir,
                    summary_json=summary_json,
                )

            mocked_resolver.assert_called_once()
            self.assertEqual(summary["taskSource"]["taskName"], "MINISO", summary)
            self.assertEqual(summary["rulespec"]["source"], "task_upload_template", summary)
            self.assertEqual(summary["upload"]["parsed_source_kind"], "sending_list", summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["instagram"], 2, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["tiktok"], 2, summary)
            self.assertEqual(summary["upload"]["metadata_count_by_platform"]["youtube"], 1, summary)
            self.assertEqual(summary["resolved_inputs"]["task_download_dir"]["path"], str(task_download_dir.resolve()), summary)
            self.assertEqual(summary["resolved_inputs"]["template_output_dir"]["path"], str(template_output_dir.resolve()), summary)
            self.assertEqual(summary["resolved_config_sources"]["task_download_dir"], "cli", summary)
            self.assertEqual(summary["resolved_config_sources"]["template_output_dir"], "cli", summary)
            self.assertTrue(summary_json.exists(), summary_json)

    def test_prepare_screening_inputs_skips_task_upload_resolution_when_local_inputs_are_complete(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "keep.xlsx"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"

            build_keep_workbook(creator_workbook)

            with patch("scripts.prepare_screening_inputs.resolve_task_upload_source_files") as mocked_resolver:
                summary = prepare_screening_inputs(
                    creator_workbook=creator_workbook,
                    template_workbook=FIXTURE_TEMPLATE,
                    task_name="MINISO",
                    screening_data_dir=screening_data_dir,
                    config_dir=config_dir,
                    temp_dir=temp_dir,
                )

            mocked_resolver.assert_not_called()
            self.assertEqual(summary["preflight"]["creator_input_mode"], "creator_workbook", summary)
            self.assertEqual(summary["preflight"]["template_input_mode"], "template_workbook", summary)
            self.assertEqual(summary["rulespec"]["source"], "template_workbook", summary)

    def test_prepare_screening_inputs_clears_stale_active_visual_prompts_for_rulespec_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            creator_workbook = tmp_path / "creator_upload.xlsx"
            rulespec_json = tmp_path / "rulespec.json"
            screening_data_dir = tmp_path / "screening_data"
            config_dir = tmp_path / "config"
            temp_dir = tmp_path / "temp"

            build_creator_workbook(creator_workbook)
            rulespec_json.write_text(
                json.dumps(
                    {
                        "platform_overrides": {
                            "instagram": {"visual_review_cover_limit": 9},
                            "tiktok": {"visual_review_cover_limit": 9},
                        }
                    },
                    ensure_ascii=False,
                    indent=2,
                )
                + "\n",
                encoding="utf-8",
            )

            seeded = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                template_workbook=FIXTURE_TEMPLATE,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )
            self.assertTrue(Path(seeded["active_visual_prompts_path"]).exists(), seeded)

            summary = prepare_screening_inputs(
                creator_workbook=creator_workbook,
                rulespec_json=rulespec_json,
                screening_data_dir=screening_data_dir,
                config_dir=config_dir,
                temp_dir=temp_dir,
            )

            self.assertEqual(summary["rulespec"]["source"], "rulespec_json", summary)
            self.assertEqual(summary["rulespec"]["visual_prompts_json_path"], "", summary)
            self.assertFalse(Path(summary["active_visual_prompts_path"]).exists(), summary)
            self.assertTrue(Path(summary["rulespec"]["runtime_prompt_artifacts_json_path"]).exists(), summary)


if __name__ == "__main__":
    unittest.main()
