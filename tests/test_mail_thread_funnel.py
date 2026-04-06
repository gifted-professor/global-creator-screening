from __future__ import annotations

import json
import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from email_sync.db import Database
from email_sync.mail_thread_funnel import build_mail_thread_funnel_keep_workbook


def _addresses(*pairs: tuple[str, str]) -> str:
    return json.dumps([{"name": name, "address": address} for name, address in pairs], ensure_ascii=False)


class MailThreadFunnelTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="mail_thread_funnel_test_"))
        self.db_path = self.temp_dir / "email_sync.db"
        self.input_path = self.temp_dir / "sending_list.xlsx"
        self.output_prefix = self.temp_dir / "exports" / "skg_mail_thread_funnel"

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _make_sending_list_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["Platform", "@username", "URL", "nickname", "Region", "Email"])
        sheet.append(["TikTok", "creatoralpha", "https://www.tiktok.com/@creatoralpha", "Alpha", "US", "creatoralpha@mgmt.com"])
        sheet.append(["TikTok", "creatorbeta", "https://www.tiktok.com/@creatorbeta", "Beta", "US", "creatorbeta@mgmt.com"])
        workbook.save(self.input_path)

    def _seed_messages(self) -> Database:
        db = Database(self.db_path)
        db.init_schema()
        rows = [
            (
                "partnerships@amagency.biz",
                "INBOX",
                1,
                1,
                "<m1>",
                "Re: Paid collab with SKG",
                None,
                None,
                "2026-04-02T12:00:00+00:00",
                "2026-04-02T12:00:00+00:00",
                "2026-04-02T12:00:00+00:00",
                "2026-04-02T12:00:00+00:00",
                "[]",
                100,
                _addresses(("Creator Alpha", "creatoralpha@mgmt.com")),
                _addresses(("Astrid", "astrid@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Thanks Astrid, happy to collaborate with SKG.",
                "",
                "Thanks Astrid, happy to collaborate with SKG.",
                "{}",
                "raw/1.eml",
                "sha1",
                100,
                0,
                0,
                "2026-04-02T12:00:00+00:00",
                "2026-04-02T12:00:00+00:00",
            ),
            (
                "partnerships@amagency.biz",
                "INBOX",
                2,
                1,
                "<m2>",
                "Re: Paid collab with SKG",
                None,
                None,
                "2026-04-02T13:00:00+00:00",
                "2026-04-02T13:00:00+00:00",
                "2026-04-02T13:00:00+00:00",
                "2026-04-02T13:00:00+00:00",
                "[]",
                100,
                _addresses(("Manager", "team@agency.example")),
                _addresses(("Astrid", "astrid@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Hallo Astrid,\n\nvielen Dank.\n\n> Hi @ downwithopc0 ,\n\nThis SKG campaign sounds great.",
                "",
                "Hallo Astrid,\n\nvielen Dank.\n\n> Hi @ downwithopc0 ,\n\nThis SKG campaign sounds great.",
                "{}",
                "raw/2.eml",
                "sha2",
                100,
                0,
                0,
                "2026-04-02T13:00:00+00:00",
                "2026-04-02T13:00:00+00:00",
            ),
            (
                "partnerships@amagency.biz",
                "INBOX",
                3,
                1,
                "<m3>",
                "Automatic reply: Paid collab with SKG",
                None,
                None,
                "2026-04-02T14:00:00+00:00",
                "2026-04-02T14:00:00+00:00",
                "2026-04-02T14:00:00+00:00",
                "2026-04-02T14:00:00+00:00",
                "[]",
                100,
                _addresses(("Auto Reply", "auto@example.com")),
                _addresses(("Astrid", "astrid@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Thank you for your email. I am currently out of office.",
                "",
                "Thank you for your email. I am currently out of office.",
                "{}",
                "raw/3.eml",
                "sha3",
                100,
                0,
                0,
                "2026-04-02T14:00:00+00:00",
                "2026-04-02T14:00:00+00:00",
            ),
            (
                "partnerships@amagency.biz",
                "INBOX",
                4,
                1,
                "<m4>",
                "Re: Paid collab with SKG",
                None,
                None,
                "2026-04-02T15:00:00+00:00",
                "2026-04-02T15:00:00+00:00",
                "2026-04-02T15:00:00+00:00",
                "2026-04-02T15:00:00+00:00",
                "[]",
                100,
                _addresses(("Manager", "manager@example.com")),
                _addresses(("Astrid", "astrid@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Hello Astrid,\n\nWe can move ahead on the SKG campaign and share rates shortly.",
                "",
                "Hello Astrid,\n\nWe can move ahead on the SKG campaign and share rates shortly.",
                "{}",
                "raw/4.eml",
                "sha4",
                100,
                0,
                0,
                "2026-04-02T15:00:00+00:00",
                "2026-04-02T15:00:00+00:00",
            ),
        ]
        db.conn.executemany(
            """
            INSERT INTO messages (
                account_email, folder_name, uid, uidvalidity, message_id, subject, in_reply_to, references_header,
                sent_at, sent_at_raw, internal_date, internal_date_raw, flags_json, size_bytes,
                from_json, to_json, cc_json, bcc_json, reply_to_json, sender_json,
                body_text, body_html, snippet, headers_json, raw_path, raw_sha256, raw_size_bytes,
                has_attachments, attachment_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        db.conn.commit()
        return db

    def test_build_mail_thread_funnel_keep_workbook_runs_pass0_regex_and_llm(self) -> None:
        self._make_sending_list_workbook()
        db = self._seed_messages()

        def fake_llm_runner(records):
            results = []
            for record in records:
                results.append(
                    {
                        **record,
                        "llm_handle": "llmresolvedcreator",
                        "resolution_confidence_final": "high",
                        "llm_evidence": "Manager confirms creator handle llmresolvedcreator.",
                        "llm_reason": "explicit manager confirmation",
                    }
                )
            return results

        try:
            result = build_mail_thread_funnel_keep_workbook(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
                keyword="SKG",
                sent_since=date(2026, 4, 1),
                llm_runner=fake_llm_runner,
            )
        finally:
            db.close()

        self.assertEqual(result["message_hit_count"], 4)
        self.assertEqual(result["external_message_count"], 4)
        self.assertEqual(result["pass0_sending_list_email_count"], 1)
        self.assertEqual(result["regex_pass1_count"], 1)
        self.assertEqual(result["regex_pass2_count"], 0)
        self.assertEqual(result["llm_high_count"], 1)
        self.assertEqual(result["filtered_auto_reply_count"], 1)
        self.assertEqual(result["manual_row_count"], 0)
        self.assertEqual(result["keep_row_count"], 3)

        review_book = load_workbook(result["review_xlsx_path"], read_only=True, data_only=True)
        keep_book = load_workbook(result["keep_xlsx_path"], read_only=True, data_only=True)
        try:
            review_rows = list(review_book.active.iter_rows(values_only=True))
            keep_rows = list(keep_book.active.iter_rows(values_only=True))
        finally:
            review_book.close()
            keep_book.close()

        review_headers = list(review_rows[0])
        keep_headers = list(keep_rows[0])
        review_thread_index = review_headers.index("thread_key")
        review_stage_index = review_headers.index("resolution_stage_final")
        review_id_index = review_headers.index("final_id_final")
        keep_thread_index = keep_headers.index("thread_key")
        keep_platform_index = keep_headers.index("Platform")
        keep_stage_index = keep_headers.index("resolution_stage_final")
        keep_id_index = keep_headers.index("final_id_final")

        review_stages = [row[review_stage_index] for row in review_rows[1:]]
        self.assertIn("pass0_sending_list_email", review_stages)
        self.assertIn("regex_pass1", review_stages)
        self.assertIn("filtered_auto_reply", review_stages)
        self.assertIn("llm", review_stages)

        keep_pairs = {(row[keep_stage_index], row[keep_id_index]) for row in keep_rows[1:]}
        self.assertIn(("pass0_sending_list_email", "creatoralpha"), keep_pairs)
        self.assertIn(("regex_pass1", "downwithopc0"), keep_pairs)
        self.assertIn(("llm", "llmresolvedcreator"), keep_pairs)
        keep_platform_by_id = {row[keep_id_index]: row[keep_platform_index] for row in keep_rows[1:]}
        self.assertEqual(keep_platform_by_id["creatoralpha"], "TikTok")
        self.assertTrue(all(str(row[review_thread_index]).startswith("mid:<m") for row in review_rows[1:]))
        self.assertTrue(all(str(row[keep_thread_index]).startswith("mid:<m") for row in keep_rows[1:]))


if __name__ == "__main__":
    unittest.main()
