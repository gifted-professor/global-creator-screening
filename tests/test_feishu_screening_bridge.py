from __future__ import annotations

import json
import tempfile
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import patch

from email_sync.imap_sync import MailboxInfo, SyncResult
from openpyxl import Workbook

from feishu_screening_bridge import (
    download_bitable_attachments,
    download_task_upload_screening_assets,
    extract_file_token,
    import_screening_workbook_from_feishu,
    inspect_task_upload_assignments,
    resolve_bitable_view_from_url,
    sync_task_upload_mailboxes,
    sync_task_upload_view_to_email_project,
)
from feishu_screening_bridge.bitable_export import export_bitable_view
from feishu_screening_bridge.feishu_api import FeishuOpenClient
from feishu_screening_bridge.local_env import get_preferred_value, load_local_env


EMAIL_PROJECT_ROOT = Path("/Users/a1234/Desktop/Coding/网红/email")


class _FakeUrlopenResponse:
    def __init__(self, *, url: str, status: int, headers: dict[str, str], body: bytes) -> None:
        self._url = url
        self.status = status
        self.headers = headers
        self._body = body

    def __enter__(self) -> "_FakeUrlopenResponse":
        return self

    def __exit__(self, exc_type: object, exc: object, tb: object) -> bool:
        return False

    def read(self) -> bytes:
        return self._body

    def geturl(self) -> str:
        return self._url


class FeishuScreeningBridgeTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temp_dir.name)
        self.env_path = self.base_path / ".env"
        self.output_path = self.base_path / "index.html"
        self.download_dir = self.base_path / "downloads"
        self.db_path = self.base_path / "email_sync.db"
        self.workbook_bytes = self._build_workbook_bytes()
        self.request_log: list[tuple[str, str]] = []
        self.env_path.write_text(
            f"DATA_DIR={self.base_path / 'data'}\nDB_PATH={self.db_path}\n",
            encoding="utf-8",
        )
        self.feishu_base_url = "https://unit-test.feishu.mock/open-apis"

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_import_screening_workbook_from_feishu_downloads_writes_and_reexports_dashboard(self) -> None:
        with patch("urllib.request.urlopen", side_effect=self._fake_urlopen):
            result = import_screening_workbook_from_feishu(
                email_project_root=EMAIL_PROJECT_ROOT,
                email_env_file=self.env_path,
                feishu_app_id="cli_test",
                feishu_app_secret="secret_test",
                file_token_or_url="boxcn-test-workbook",
                project_code="P-FEISHU-001",
                primary_category="smart_home",
                owner_name="pm.alice",
                dashboard_output=self.output_path,
                download_dir=self.download_dir,
                feishu_base_url=self.feishu_base_url,
            )

        self.assertTrue(result["ok"])
        self.assertEqual(result["fileToken"], "boxcn-test-workbook")
        self.assertEqual(result["importResult"]["projectCode"], "P-FEISHU-001")
        self.assertEqual(result["importResult"]["projectName"], "Tapo")
        self.assertTrue(self.db_path.exists())
        self.assertTrue(self.output_path.exists())

        saved_workbook_path = Path(result["savedWorkbookPath"])
        self.assertTrue(saved_workbook_path.exists())
        self.assertEqual(saved_workbook_path.read_bytes(), self.workbook_bytes)

        html = self.output_path.read_text(encoding="utf-8")
        self.assertIn("直接导入", html)
        self.assertIn("提交人工更新", html)

        self.assertEqual(
            self.request_log,
            [
                ("POST", "https://unit-test.feishu.mock/open-apis/auth/v3/tenant_access_token/internal"),
                ("GET", "https://unit-test.feishu.mock/open-apis/drive/v1/files/boxcn-test-workbook/download"),
            ],
        )

    def test_extract_file_token_supports_feishu_file_url(self) -> None:
        self.assertEqual(
            extract_file_token("https://example.feishu.cn/file/boxcn-test-workbook"),
            "boxcn-test-workbook",
        )

    def test_resolve_bitable_view_from_base_url(self) -> None:
        client = _FakeBitableClient()
        resolved = resolve_bitable_view_from_url(
            client,
            "https://bcnorxdfy50v.feishu.cn/base/P42ub2bX3aZY7jszJpKcUtConWg?table=tbl7mANMzT4kjuqC&view=vewWIN9jul",
        )

        self.assertEqual(resolved.source_kind, "base")
        self.assertEqual(resolved.app_token, "P42ub2bX3aZY7jszJpKcUtConWg")
        self.assertEqual(resolved.table_name, "达人管理")
        self.assertEqual(resolved.view_name, "总视图")

    def test_export_bitable_view_from_wiki_url_writes_json(self) -> None:
        client = _FakeBitableClient()
        output_path = self.base_path / "exports" / "wiki_export.json"

        result = export_bitable_view(
            client,
            url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblQho4xE6SrOtmw&view=vewHoxVXaC",
            output_path=output_path,
            output_format="json",
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["sourceKind"], "wiki")
        self.assertEqual(result["tableName"], "员工信息表")
        self.assertEqual(result["viewName"], "表格")
        payload = json.loads(output_path.read_text(encoding="utf-8"))
        self.assertEqual(payload["recordCount"], 2)
        self.assertEqual(payload["records"][0]["fields"]["员工名"], "张三")

    def test_download_bitable_attachments_downloads_real_attachment_items(self) -> None:
        client = _FakeAttachmentBitableClient()
        output_dir = self.base_path / "attachments"

        result = download_bitable_attachments(
            client,
            url="https://bcnorxdfy50v.feishu.cn/base/P42ub2bX3aZY7jszJpKcUtConWg?table=tbl7mANMzT4kjuqC&view=vewWIN9jul",
            output_dir=output_dir,
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["attachmentCount"], 2)
        self.assertEqual(result["downloadedCount"], 2)
        saved_paths = [Path(item["savedPath"]) for item in result["items"]]
        for saved_path in saved_paths:
            self.assertTrue(saved_path.exists())
        self.assertEqual(saved_paths[0].read_bytes(), b"file-a")
        self.assertEqual(saved_paths[1].read_bytes(), b"file-b")

    def test_sync_task_upload_view_to_email_project_imports_new_template_variant(self) -> None:
        email_env_path = self.base_path / "email.env"
        email_data_dir = self.base_path / "email_data"
        email_db_path = email_data_dir / "email_sync.db"
        dashboard_output = self.base_path / "exports" / "index.html"
        email_env_path.write_text(
            f"DATA_DIR={email_data_dir}\nDB_PATH={email_db_path}\n",
            encoding="utf-8",
        )

        result = sync_task_upload_view_to_email_project(
            client=_FakeTaskUploadClient(self._build_new_template_workbook_bytes()),
            task_upload_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblYvtOYLoGWCRna&view=vewNwYvkQL",
            email_project_root=EMAIL_PROJECT_ROOT,
            email_env_file=email_env_path,
            download_dir=self.base_path / "downloads",
            dashboard_output=dashboard_output,
            project_code_prefix="P-FSH-",
            default_primary_category="lifestyle",
            category_overrides={"duet": "lifestyle"},
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["importedCount"], 1)
        item = result["items"][0]
        self.assertEqual(item["projectCode"], "P-FSH-DUET")
        self.assertEqual(item["primaryCategory"], "lifestyle")
        self.assertEqual(item["platforms"], ["instagram", "youtube"])
        self.assertTrue(email_db_path.exists())
        self.assertTrue(dashboard_output.exists())

    def test_inspect_task_upload_assignments_matches_employee_and_downloads_template(self) -> None:
        download_dir = self.base_path / "downloads"
        result = inspect_task_upload_assignments(
            client=_FakeInspectionClient(self._build_new_template_workbook_bytes()),
            task_upload_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblYvtOYLoGWCRna&view=vewNwYvkQL",
            employee_info_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblQho4xE6SrOtmw&view=vewHoxVXaC",
            download_dir=download_dir,
            download_templates=True,
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["taskTableName"], "任务上传")
        self.assertEqual(result["employeeTableName"], "员工信息表")
        self.assertEqual(result["recordCount"], 1)
        self.assertEqual(result["matchedCount"], 1)
        self.assertEqual(result["downloadedCount"], 1)
        item = result["items"][0]
        self.assertEqual(item["taskName"], "duet")
        self.assertEqual(item["matchedBy"], "employee_id")
        self.assertEqual(item["ownerEmail"], "yvette@amagency.biz")
        self.assertEqual(item["employeeEmail"], "yvette@amagency.biz")
        self.assertEqual(item["imapCode"], "imap-yvette-123")
        self.assertEqual(item["sendingListFileToken"], "boxcn-duet-sending-list")
        self.assertEqual(item["sendingListFileName"], "duet-发信名单.xlsx")
        self.assertTrue(Path(item["templateDownloadedPath"]).exists())
        self.assertFalse(item["templateParseRequested"])

    def test_inspect_task_upload_assignments_parses_downloaded_template(self) -> None:
        download_dir = self.base_path / "downloads"
        parse_output_dir = self.base_path / "parsed"
        result = inspect_task_upload_assignments(
            client=_FakeInspectionClient(self._build_new_template_workbook_bytes()),
            task_upload_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblYvtOYLoGWCRna&view=vewNwYvkQL",
            employee_info_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblQho4xE6SrOtmw&view=vewHoxVXaC",
            download_dir=download_dir,
            parse_templates=True,
            parse_output_dir=parse_output_dir,
        )

        self.assertTrue(result["ok"])
        self.assertTrue(result["parseTemplates"])
        self.assertEqual(result["downloadedCount"], 1)
        self.assertEqual(result["parsedCount"], 1)
        self.assertEqual(result["parseFailedCount"], 0)
        self.assertEqual(Path(result["parseOutputDir"]).resolve(), parse_output_dir.resolve())

        item = result["items"][0]
        self.assertTrue(item["templateParseRequested"])
        self.assertTrue(item["templateParsed"])
        self.assertEqual(item["templateParseError"], "")
        self.assertTrue(Path(item["templateDownloadedPath"]).exists())

        report_path = Path(item["templateParseReportPath"])
        self.assertTrue(report_path.exists(), report_path)
        self.assertEqual(report_path.name, "compile_report.json")

        artifact_paths = item["templateParseArtifacts"]
        self.assertEqual(
            sorted(artifact_paths.keys()),
            [
                "rulespec_json",
                "structured_requirement_json",
                "visual_prompts_json",
                "visual_reuse_spec_json",
            ],
            artifact_paths,
        )
        for raw_path in artifact_paths.values():
            self.assertTrue(Path(raw_path).exists(), raw_path)

    def test_download_task_upload_screening_assets_by_task_name(self) -> None:
        download_dir = self.base_path / "downloads"
        result = download_task_upload_screening_assets(
            client=_FakeTaskUploadClient(self._build_new_template_workbook_bytes()),
            task_upload_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblYvtOYLoGWCRna&view=vewNwYvkQL",
            task_name="duet",
            download_dir=download_dir,
        )

        self.assertEqual(result["taskName"], "duet")
        self.assertEqual(result["recordId"], "rec-task-001")
        self.assertEqual(result["sendingListFileToken"], "boxcn-duet-sending-list")
        self.assertEqual(result["sendingListFileName"], "duet-发信名单.xlsx")
        self.assertTrue(Path(result["templateDownloadedPath"]).exists(), result)
        self.assertTrue(Path(result["sendingListDownloadedPath"]).exists(), result)

    def test_sync_task_upload_mailboxes_resolves_prefixed_folder_and_runs_sync(self) -> None:
        mail_data_dir = self.base_path / "task-mail-data"

        class _FakeImapClient:
            def close(self) -> None:
                return None

            def logout(self) -> None:
                return None

        with (
            patch("feishu_screening_bridge.task_upload_sync.connect", return_value=_FakeImapClient()),
            patch(
                "feishu_screening_bridge.task_upload_sync.discover_mailboxes",
                return_value=[
                    MailboxInfo(display_name="INBOX", imap_name="INBOX", delimiter="/", flags=["\\HasNoChildren"]),
                    MailboxInfo(display_name="其他文件夹", imap_name="&UXZO1mWHTvZZOQ-", delimiter="/", flags=["\\NoSelect", "\\HasChildren"]),
                    MailboxInfo(display_name="其他文件夹/duet", imap_name="其他文件夹/duet", delimiter="/", flags=["\\HasNoChildren"]),
                ],
            ),
            patch(
                "feishu_screening_bridge.task_upload_sync.sync_mailboxes",
                return_value=[
                    SyncResult(
                        folder_name="其他文件夹/duet",
                        fetched=3,
                        skipped_state_advance=True,
                        last_seen_uid=0,
                        uidvalidity=9527,
                        message_count_on_server=18,
                    )
                ],
            ) as sync_mock,
        ):
            result = sync_task_upload_mailboxes(
                client=_FakeInspectionClient(self._build_new_template_workbook_bytes()),
                task_upload_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblYvtOYLoGWCRna&view=vewNwYvkQL",
                employee_info_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblQho4xE6SrOtmw&view=vewHoxVXaC",
                download_dir=self.base_path / "downloads",
                mail_data_dir=mail_data_dir,
                folder_prefixes=["其他文件夹"],
                sent_since="2026-03-01",
                limit=3,
            )

        self.assertTrue(result["ok"])
        self.assertEqual(result["selectedCount"], 1)
        self.assertEqual(result["syncedCount"], 1)
        self.assertEqual(result["failedCount"], 0)
        item = result["items"][0]
        self.assertTrue(item["mailSyncOk"])
        self.assertEqual(item["resolvedFolder"], "其他文件夹/duet")
        self.assertEqual(item["mailFetchedCount"], 3)
        self.assertTrue(Path(item["mailDbPath"]).exists())
        self.assertEqual(result["sentSince"], "2026-03-01")

        _, kwargs = sync_mock.call_args
        self.assertEqual(kwargs["requested_folders"], ["其他文件夹/duet"])
        self.assertEqual(kwargs["limit"], 3)
        self.assertEqual(kwargs["sent_since"], date(2026, 3, 1))

    def test_sync_task_upload_mail_cli_defaults_to_recent_three_months_when_omitted(self) -> None:
        mail_data_dir = self.base_path / "task-mail-data-default-window"
        env_path = self.base_path / "mail.env"
        env_path.write_text(
            "\n".join(
                [
                    "FEISHU_APP_ID=test_app",
                    "FEISHU_APP_SECRET=test_secret",
                    "TASK_UPLOAD_URL=https://example.com/task",
                    "EMPLOYEE_INFO_URL=https://example.com/employee",
                ]
            ),
            encoding="utf-8",
        )

        parser = __import__("feishu_screening_bridge.__main__", fromlist=["_build_parser"])._build_parser()
        args = parser.parse_args(
            [
                "sync-task-upload-mail",
                "--env-file",
                str(env_path),
                "--task-name",
                "duet",
                "--download-dir",
                str(self.base_path / "downloads"),
                "--mail-data-dir",
                str(mail_data_dir),
            ]
        )

        with (
            patch("feishu_screening_bridge.__main__.FeishuOpenClient", return_value=object()),
            patch("feishu_screening_bridge.__main__.resolve_sync_sent_since", return_value=date(2025, 12, 27)),
            patch(
                "feishu_screening_bridge.__main__.sync_task_upload_mailboxes",
                return_value={
                    "selectedCount": 1,
                    "syncedCount": 1,
                    "failedCount": 0,
                    "mailDataDir": str(mail_data_dir),
                    "items": [],
                },
            ) as sync_mock,
        ):
            __import__("feishu_screening_bridge.__main__", fromlist=["_cmd_sync_task_upload_mail"])._cmd_sync_task_upload_mail(args)

        _, kwargs = sync_mock.call_args
        self.assertEqual(kwargs["sent_since"], "2025-12-27")

    def test_inspect_task_upload_assignments_can_override_owner_email_match(self) -> None:
        class _FakeOverrideClient(_FakeInspectionClient):
            def post_api_json(self, url_path: str, *, body: dict[str, Any] | None = None, headers: dict[str, str] | None = None) -> dict[str, Any]:
                if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblYvtOYLoGWCRna/records/search":
                    return {
                        "code": 0,
                        "msg": "success",
                        "data": {
                            "has_more": False,
                            "items": [
                                {
                                    "record_id": "rec-task-002",
                                    "fields": {
                                        "任务名": [{"text": "MINISO", "type": "text"}],
                                        "员工ID": {"type": 1, "value": [{"text": "ou_primary", "type": "text"}]},
                                        "负责人邮箱": {
                                            "type": 1,
                                            "value": [
                                                {"link": "mailto:chenjunren@amagency.biz", "text": "chenjunren@amagency.biz", "type": "url"},
                                                {"text": ",", "type": "text"},
                                                {"link": "mailto:eden@amagency.biz", "text": "eden@amagency.biz", "type": "url"},
                                            ],
                                        },
                                        "负责人": [{"name": "陈俊仁"}],
                                        "发信名单": [{"file_token": "boxcn-miniso-sending-list", "name": "陈俊仁的总表.xlsx"}],
                                        "达人管理表链接": {"link": "https://example.com/base/miniso", "type": "mention"},
                                        "需求上传（excel 格式）": [{"file_token": "boxcn-miniso-file", "name": "miniso.xlsx"}],
                                    },
                                }
                            ],
                        },
                    }
                if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblQho4xE6SrOtmw/records/search":
                    return {
                        "code": 0,
                        "msg": "success",
                        "data": {
                            "has_more": False,
                            "items": [
                                {
                                    "record_id": "rec-employee-primary",
                                    "fields": {
                                        "imap 码": [{"text": "imap-chen-123", "type": "text"}],
                                        "员工 ID": [{"text": "ou_primary", "type": "text"}],
                                        "员工名": [{"name": "陈俊仁"}],
                                        "邮箱": [{"link": "mailto:chenjunren@amagency.biz", "text": "chenjunren@amagency.biz", "type": "url"}],
                                    },
                                },
                                {
                                    "record_id": "rec-employee-eden",
                                    "fields": {
                                        "imap 码": [{"text": "imap-eden-456", "type": "text"}],
                                        "员工 ID": [{"text": "ou_eden", "type": "text"}],
                                        "员工名": [{"name": "Eden"}],
                                        "邮箱": [{"link": "mailto:eden@amagency.biz", "text": "eden@amagency.biz", "type": "url"}],
                                    },
                                },
                            ],
                        },
                    }
                return super().post_api_json(url_path, body=body, headers=headers)

        result = inspect_task_upload_assignments(
            client=_FakeOverrideClient(self._build_new_template_workbook_bytes()),
            task_upload_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblYvtOYLoGWCRna&view=vewNwYvkQL",
            employee_info_url="https://bcnorxdfy50v.feishu.cn/wiki/S0bbwTnlZiJlVMk1Q04ctPXBnje?table=tblQho4xE6SrOtmw&view=vewHoxVXaC",
            download_dir=self.base_path / "downloads",
            owner_email_overrides={"MINISO": "eden@amagency.biz"},
        )

        self.assertTrue(result["ok"])
        item = result["items"][0]
        self.assertEqual(item["preferredOwnerEmail"], "eden@amagency.biz")
        self.assertEqual(item["matchedBy"], "owner_email_override")
        self.assertEqual(item["employeeEmail"], "eden@amagency.biz")
        self.assertEqual(item["imapCode"], "imap-eden-456")

    def test_local_env_loader_parses_key_values(self) -> None:
        env_path = self.base_path / ".env.local"
        env_path.write_text(
            "FEISHU_APP_ID=test_app\nPROJECT_CODE=P-001\nPRIMARY_CATEGORY=smart_home\n",
            encoding="utf-8",
        )

        values = load_local_env(env_path)

        self.assertEqual(values["FEISHU_APP_ID"], "test_app")
        self.assertEqual(get_preferred_value("", values, "PROJECT_CODE"), "P-001")
        self.assertEqual(get_preferred_value("override", values, "PROJECT_CODE"), "override")

    def test_import_uses_env_file_directory_for_relative_data_dir(self) -> None:
        relative_env_path = self.base_path / "relative.env"
        relative_env_path.write_text("DATA_DIR=./relative-data\n", encoding="utf-8")

        with patch("urllib.request.urlopen", side_effect=self._fake_urlopen):
            result = import_screening_workbook_from_feishu(
                email_project_root=EMAIL_PROJECT_ROOT,
                email_env_file=relative_env_path,
                feishu_app_id="cli_test",
                feishu_app_secret="secret_test",
                file_token_or_url="boxcn-test-workbook",
                project_code="P-FEISHU-002",
                primary_category="smart_home",
                dashboard_output=self.output_path,
                download_dir=self.download_dir,
                feishu_base_url=self.feishu_base_url,
            )

        expected_db_path = self.base_path / "relative-data" / "email_sync.db"
        self.assertEqual(Path(result["dbPath"]).resolve(), expected_db_path.resolve())
        self.assertTrue(expected_db_path.resolve().exists())

    def _build_workbook_bytes(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "需求主表"
        rows = [
            ("字段", "内容", "说明"),
            ("", "", ""),
            ("", "", ""),
            ("A. 基本信息", "", ""),
            ("项目名称", "Tapo", ""),
            ("品牌 / 产品", "Tapo 智能家居 / 宠物 / 户外生活场景", ""),
            ("适用平台", "两者", ""),
            ("B. 步骤1：基础资质审核", "", ""),
            ("地区要求", "美国", ""),
            ("语言要求", "", ""),
            ("C. 步骤2：数据审核", "", ""),
            ("粉丝数阈值（可选）", "", ""),
            ("中位数播放量阈值", 10000, ""),
            ("D. 步骤3：内容 / 视觉审核", "", ""),
            ("多人互动", "需要", ""),
            ("E. 步骤4：排除项审核", "", ""),
            ("不符合时处理", "排除", ""),
            ("F. 人工判断项 / 合规提醒", "", ""),
            ("合规提醒", "FTC disclosure", ""),
            ("G. 最终判定逻辑", "", ""),
            ("最终判定", "满足前序条件则推进", ""),
        ]
        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row_index, column_index).value = value
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _build_new_template_workbook_bytes(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "需求主表"
        rows = [
            ("字段", "内容", "说明"),
            ("", "", ""),
            ("", "", ""),
            ("A. 项目基本信息", "", ""),
            ("项目名称（品牌名）", "duet", ""),
            ("品牌 / 产品使用场景", "duet 街头采访场景", ""),
            ("适用平台", "Instagram,YouTube", ""),
            ("B. 步骤1：基础资质审核", "", ""),
            ("地区要求", "美国", ""),
            ("语言要求", "英语", ""),
            ("C. 步骤2：数据审核", "", ""),
            ("粉丝数阈值（可选）", 5000, ""),
            ("中位数播放量阈值", 15000, ""),
            ("D. 步骤3：内容 / 视觉审核", "", ""),
            ("多人互动", "需要", "最好有真实路人反馈"),
            ("E. 步骤4：排除项审核（不需要的封面清单）", "", ""),
            ("不符合时处理", "排除", ""),
            ("F. 人工判断项 / 合规提醒（当封面或数据出现什么情况时需要人工复核，如没有可不填）", "", ""),
            ("合规提醒", "需要人工复核敏感内容", ""),
            ("G. 最终判定逻辑", "", ""),
            ("最终判定", "满足前序条件则推进", ""),
        ]
        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row_index, column_index).value = value
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _fake_urlopen(self, req: Any, timeout: float | None = None) -> _FakeUrlopenResponse:
        method = str(getattr(req, "method", None) or req.get_method())
        full_url = str(req.full_url)
        self.request_log.append((method, full_url))

        if full_url == "https://unit-test.feishu.mock/open-apis/auth/v3/tenant_access_token/internal":
            return _FakeUrlopenResponse(
                url=full_url,
                status=200,
                headers={"Content-Type": "application/json; charset=utf-8"},
                body=json.dumps(
                    {
                        "code": 0,
                        "msg": "success",
                        "tenant_access_token": "tenant-access-token",
                        "expire": 7140,
                    },
                    ensure_ascii=False,
                ).encode("utf-8"),
            )
        if full_url == "https://unit-test.feishu.mock/open-apis/drive/v1/files/boxcn-test-workbook/download":
            return _FakeUrlopenResponse(
                url=full_url,
                status=200,
                headers={
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": "attachment; filename*=UTF-8''feishu-screening.xlsx",
                },
                body=self.workbook_bytes,
            )
        raise AssertionError(f"unexpected urlopen request: method={method} url={full_url} timeout={timeout}")


if __name__ == "__main__":
    unittest.main()


class _FakeBitableClient(FeishuOpenClient):
    def __init__(self) -> None:
        super().__init__(app_id="test_app", app_secret="test_secret", base_url="https://unit-test.feishu.mock/open-apis")

    def get_tenant_access_token(self) -> str:
        return "tenant-access-token"

    def get_api_json(self, url_path: str, *, headers: dict[str, str] | None = None) -> dict[str, Any]:
        if url_path == "/wiki/v2/spaces/get_node?token=S0bbwTnlZiJlVMk1Q04ctPXBnje":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "node": {
                        "obj_type": "bitable",
                        "obj_token": "WVxtbuOkdaoqbxscPfJcG3oEnMd",
                        "title": "AI 系统任务设置",
                    }
                },
            }
        if url_path == "/bitable/v1/apps/P42ub2bX3aZY7jszJpKcUtConWg/tables":
            return {
                "code": 0,
                "msg": "success",
                "data": {"items": [{"table_id": "tbl7mANMzT4kjuqC", "name": "达人管理"}]},
            }
        if url_path == "/bitable/v1/apps/P42ub2bX3aZY7jszJpKcUtConWg/tables/tbl7mANMzT4kjuqC/views":
            return {
                "code": 0,
                "msg": "success",
                "data": {"items": [{"view_id": "vewWIN9jul", "view_name": "总视图"}]},
            }
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables":
            return {
                "code": 0,
                "msg": "success",
                "data": {"items": [{"table_id": "tblQho4xE6SrOtmw", "name": "员工信息表"}]},
            }
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblQho4xE6SrOtmw/views":
            return {
                "code": 0,
                "msg": "success",
                "data": {"items": [{"view_id": "vewHoxVXaC", "view_name": "表格"}]},
            }
        raise AssertionError(f"unexpected get_api_json call: {url_path}")

    def post_api_json(
        self,
        url_path: str,
        *,
        body: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblYvtOYLoGWCRna/records/search":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "has_more": False,
                    "items": [
                        {
                            "record_id": "rec-task-001",
                            "fields": {
                                "任务名": [{"text": "duet", "type": "text"}],
                                "员工ID": {"type": 1, "value": [{"text": "ou_fbb11a48bd394428", "type": "text"}]},
                                "负责人邮箱": {"type": 1, "value": [{"link": "mailto:yvette@amagency.biz", "text": "yvette@amagency.biz", "type": "url"}]},
                                "负责人": [{"name": "Yvette"}],
                                "达人管理表链接": {"link": "https://example.com/base/foo", "type": "mention"},
                                "需求上传（excel 格式）": [{"file_token": "boxcn-duet-file", "name": "红人筛号需求-duet.xlsx"}],
                            },
                        }
                    ],
                },
            }
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblQho4xE6SrOtmw/records/search":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "has_more": False,
                    "items": [
                        {"record_id": "rec001", "fields": {"员工名": "张三", "邮箱": "zhangsan@example.com"}},
                        {"record_id": "rec002", "fields": {"员工名": "李四", "邮箱": "lisi@example.com"}},
                    ],
                },
            }
        raise AssertionError(f"unexpected post_api_json call: {url_path} body={body}")


class _FakeAttachmentBitableClient(_FakeBitableClient):
    def post_api_json(
        self,
        url_path: str,
        *,
        body: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        if url_path == "/bitable/v1/apps/P42ub2bX3aZY7jszJpKcUtConWg/tables/tbl7mANMzT4kjuqC/records/search":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "has_more": False,
                    "items": [
                        {
                            "record_id": "rec001",
                            "fields": {
                                "附件列": [
                                    {"file_token": "boxcn-file-a", "name": "a.txt"},
                                    {"file_token": "boxcn-file-b", "name": "b.txt"},
                                ]
                            },
                        }
                    ],
                },
            }
        return super().post_api_json(url_path, body=body, headers=headers)

    def download_file(self, file_token_or_url: str, *, desired_name: str | None = None):  # type: ignore[override]
        from feishu_screening_bridge.feishu_api import DownloadedFeishuFile

        content_map = {
            "boxcn-file-a": b"file-a",
            "boxcn-file-b": b"file-b",
        }
        token = str(file_token_or_url)
        return DownloadedFeishuFile(
            file_token=token,
            file_name=str(desired_name or token),
            content_type="text/plain",
            content=content_map[token],
            source_url=f"https://unit-test.feishu.mock/download/{token}",
        )


class _FakeTaskUploadClient(_FakeBitableClient):
    def __init__(self, workbook_bytes: bytes) -> None:
        super().__init__()
        self.workbook_bytes = workbook_bytes

    def get_api_json(self, url_path: str, *, headers: dict[str, str] | None = None) -> dict[str, Any]:
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables":
            return {
                "code": 0,
                "msg": "success",
                "data": {"items": [{"table_id": "tblYvtOYLoGWCRna", "name": "任务上传"}]},
            }
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblYvtOYLoGWCRna/views":
            return {
                "code": 0,
                "msg": "success",
                "data": {"items": [{"view_id": "vewNwYvkQL", "view_name": "表格"}]},
            }
        return super().get_api_json(url_path, headers=headers)

    def post_api_json(
        self,
        url_path: str,
        *,
        body: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblYvtOYLoGWCRna/records/search":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "has_more": False,
                    "items": [
                        {
                            "record_id": "rec-task-001",
                            "fields": {
                                "任务名": [{"text": "duet", "type": "text"}],
                                "负责人邮箱": {"type": 1, "value": [{"link": "mailto:yvette@amagency.biz", "text": "yvette@amagency.biz", "type": "url"}]},
                                "负责人": [{"name": "Yvette"}],
                                "发信名单": [{"file_token": "boxcn-duet-sending-list", "name": "duet-发信名单.xlsx"}],
                                "达人管理表链接": {"link": "https://example.com/base/foo", "type": "mention"},
                                "需求上传（excel 格式）": [{"file_token": "boxcn-duet-file", "name": "红人筛号需求-duet.xlsx"}],
                            },
                        }
                    ],
                    "total": 1,
                },
            }
        return super().post_api_json(url_path, body=body, headers=headers)

    def download_file(self, file_token_or_url: str, *, desired_name: str | None = None):  # type: ignore[override]
        from feishu_screening_bridge.feishu_api import DownloadedFeishuFile

        return DownloadedFeishuFile(
            file_token=str(file_token_or_url),
            file_name=str(desired_name or "screening.xlsx"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=self.workbook_bytes,
            source_url=f"https://unit-test.feishu.mock/download/{file_token_or_url}",
        )


class _FakeInspectionClient(_FakeTaskUploadClient):
    def get_api_json(self, url_path: str, *, headers: dict[str, str] | None = None) -> dict[str, Any]:
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "items": [
                        {"table_id": "tblQho4xE6SrOtmw", "name": "员工信息表"},
                        {"table_id": "tblYvtOYLoGWCRna", "name": "任务上传"},
                    ]
                },
            }
        return super().get_api_json(url_path, headers=headers)

    def post_api_json(
        self,
        url_path: str,
        *,
        body: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblYvtOYLoGWCRna/records/search":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "has_more": False,
                    "items": [
                        {
                            "record_id": "rec-task-001",
                            "fields": {
                                "任务名": [{"text": "duet", "type": "text"}],
                                "员工ID": {"type": 1, "value": [{"text": "ou_fbb11a48bd394428", "type": "text"}]},
                                "负责人邮箱": {"type": 1, "value": [{"link": "mailto:yvette@amagency.biz", "text": "yvette@amagency.biz", "type": "url"}]},
                                "负责人": [{"name": "Yvette"}],
                                "发信名单": [{"file_token": "boxcn-duet-sending-list", "name": "duet-发信名单.xlsx"}],
                                "达人管理表链接": {"link": "https://example.com/base/foo", "type": "mention"},
                                "需求上传（excel 格式）": [{"file_token": "boxcn-duet-file", "name": "红人筛号需求-duet.xlsx"}],
                            },
                        }
                    ],
                },
            }
        if url_path == "/bitable/v1/apps/WVxtbuOkdaoqbxscPfJcG3oEnMd/tables/tblQho4xE6SrOtmw/records/search":
            return {
                "code": 0,
                "msg": "success",
                "data": {
                    "has_more": False,
                    "items": [
                        {
                            "record_id": "rec-employee-001",
                            "fields": {
                                "imap 码": [{"text": "imap-yvette-123", "type": "text"}],
                                "员工 ID": [{"text": "ou_fbb11a48bd394428", "type": "text"}],
                                "员工名": [{"en_name": "Yvette", "name": "Yvette"}],
                                "邮箱": [{"link": "mailto:yvette@amagency.biz", "text": "yvette@amagency.biz", "type": "url"}],
                            },
                        }
                    ],
                },
            }
        return super().post_api_json(url_path, body=body, headers=headers)
