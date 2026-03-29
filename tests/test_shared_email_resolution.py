from __future__ import annotations

import json
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import requests
from openpyxl import Workbook, load_workbook

from email_sync.db import Database
from email_sync.shared_email_resolution import (
    resolve_shared_email_candidates,
    run_shared_email_final_review,
)


class SharedEmailResolutionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="shared_email_resolution_test_"))
        self.db_path = self.temp_dir / "email_sync.db"
        self.shared_input_path = self.temp_dir / "shared_email.xlsx"
        self.output_prefix = self.temp_dir / "exports" / "shared_email_resolution"

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _make_shared_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        headers = [
            "Platform",
            "@username",
            "URL",
            "nickname",
            "Region",
            "Email",
            "matched_email",
            "brand_message_id",
            "brand_message_sent_at",
            "brand_message_subject",
            "brand_message_snippet",
            "profile_dedupe_key",
            "sheet_name",
            "source_row_number",
        ]
        sheet.append(headers)
        sheet.append(
            [
                "Instagram",
                "bob",
                "https://instagram.com/bob",
                "Bob",
                "UK",
                "shared@agency.com",
                "shared@agency.com",
                1,
                "2026-03-10T10:00:00+00:00",
                "MINISO shared roster",
                "Please review @cara for the MINISO campaign.",
                "instagram:cara_bob:bob",
                "results",
                2,
            ]
        )
        sheet.append(
            [
                "Instagram",
                "cara",
                "https://instagram.com/cara",
                "Cara",
                "UK",
                "shared@agency.com",
                "shared@agency.com",
                1,
                "2026-03-10T10:00:00+00:00",
                "MINISO shared roster",
                "Please review @cara for the MINISO campaign.",
                "instagram:cara",
                "results",
                3,
            ]
        )
        sheet.append(
            [
                "Instagram",
                "dean",
                "https://instagram.com/dean",
                "Dean",
                "US",
                "team@agency.com",
                "team@agency.com",
                2,
                "2026-03-11T10:00:00+00:00",
                "MINISO team shortlist",
                "Please review this MINISO group.",
                "instagram:dean",
                "results",
                4,
            ]
        )
        sheet.append(
            [
                "Instagram",
                "eric",
                "https://instagram.com/eric",
                "Eric",
                "CA",
                "team@agency.com",
                "team@agency.com",
                2,
                "2026-03-11T10:00:00+00:00",
                "MINISO team shortlist",
                "Please review this MINISO group.",
                "instagram:eric",
                "results",
                5,
            ]
        )
        workbook.save(self.shared_input_path)

    def _make_auto_keep_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        headers = [
            "Platform",
            "@username",
            "URL",
            "nickname",
            "Region",
            "Email",
            "profile_dedupe_key",
        ]
        sheet.append(headers)
        sheet.append(
            [
                "Instagram",
                "solo",
                "https://instagram.com/solo",
                "Solo",
                "US",
                "solo@creator.com",
                "instagram:solo",
            ]
        )
        workbook.save(path)

    def _seed_messages(self) -> Database:
        db = Database(self.db_path)
        db.init_schema()
        rows = [
            (
                "operator@example.com",
                "INBOX",
                1,
                1,
                "<m1>",
                "MINISO shared roster",
                None,
                None,
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "[]",
                100,
                "[]",
                "[]",
                "[]",
                "[]",
                "[]",
                "[]",
                "Please review @cara for the MINISO campaign.",
                "",
                "Please review @cara for the MINISO campaign.",
                "{}",
                "raw/1.eml",
                "sha1",
                100,
                0,
                0,
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
            ),
            (
                "operator@example.com",
                "INBOX",
                2,
                1,
                "<m2>",
                "MINISO team shortlist",
                None,
                None,
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "[]",
                100,
                "[]",
                "[]",
                "[]",
                "[]",
                "[]",
                "[]",
                "Please review this MINISO group.",
                "",
                "Please review this MINISO group.",
                "{}",
                "raw/2.eml",
                "sha2",
                100,
                0,
                0,
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
            ),
        ]
        db.conn.executemany(
            """
            INSERT INTO messages (
                account_email, folder_name, uid, uidvalidity, message_id, subject, in_reply_to, references_header,
                sent_at, sent_at_raw, internal_date, internal_date_raw, flags_json, size_bytes,
                from_json, to_json, cc_json, bcc_json, reply_to_json, sender_json,
                body_text, body_html, snippet, headers_json, raw_path, raw_sha256, raw_size_bytes,
                has_attachments, attachment_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        db.conn.commit()
        return db

    def test_resolve_shared_email_candidates_splits_resolved_and_unresolved_groups(self) -> None:
        self._make_shared_workbook()
        db = self._seed_messages()
        try:
            result = resolve_shared_email_candidates(
                db=db,
                input_path=self.shared_input_path,
                output_prefix=self.output_prefix,
            )
        finally:
            db.close()

        self.assertEqual(result["resolved_group_count"], 1)
        self.assertEqual(result["resolved_row_count"], 1)
        self.assertEqual(result["unresolved_group_count"], 1)
        self.assertEqual(result["unresolved_row_count"], 2)
        self.assertEqual(result["llm_candidate_group_count"], 1)

        resolved_wb = load_workbook(result["resolved_xlsx_path"], read_only=True, data_only=True)
        try:
            resolved_rows = list(resolved_wb.active.iter_rows(values_only=True))
        finally:
            resolved_wb.close()
        self.assertEqual(len(resolved_rows) - 1, 1)
        headers = list(resolved_rows[0])
        nickname_index = headers.index("nickname")
        self.assertEqual(resolved_rows[1][nickname_index], "Cara")

        unresolved_wb = load_workbook(result["unresolved_xlsx_path"], read_only=True, data_only=True)
        try:
            unresolved_rows = list(unresolved_wb.active.iter_rows(values_only=True))
        finally:
            unresolved_wb.close()
        self.assertEqual(len(unresolved_rows) - 1, 2)

        llm_candidates = [
            json.loads(line)
            for line in Path(result["llm_candidates_jsonl_path"]).read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(len(llm_candidates), 1)
        self.assertEqual(llm_candidates[0]["shared_email"], "team@agency.com")

    def test_llm_final_review_merges_auto_keep_with_llm_resolved_rows(self) -> None:
        self._make_shared_workbook()
        db = self._seed_messages()
        try:
            resolve_shared_email_candidates(
                db=db,
                input_path=self.shared_input_path,
                output_prefix=self.output_prefix,
            )
        finally:
            db.close()

        auto_keep_path = self.temp_dir / "auto_keep.xlsx"
        self._make_auto_keep_workbook(auto_keep_path)

        with mock.patch(
            "email_sync.shared_email_resolution._invoke_shared_email_llm_review",
            return_value={
                "decision": "match_one",
                "matched_candidate_keys": ["instagram:dean"],
                "confidence": "high",
                "reason": "The message context points to Dean.",
                "raw_text": "{}",
            },
        ):
            result = run_shared_email_final_review(
                input_prefix=self.output_prefix,
                env_path=str(self.temp_dir / ".env"),
                auto_keep_paths=[auto_keep_path],
                base_url="https://example.com/v1",
                api_key="sk-test",
                model="gpt-test",
                wire_api="responses",
            )

        self.assertEqual(result["review_group_count"], 1)
        self.assertEqual(result["llm_resolved_row_count"], 1)
        self.assertEqual(result["manual_row_count"], 0)
        self.assertEqual(result["final_keep_row_count"], 2)

        final_keep_wb = load_workbook(result["final_keep_xlsx_path"], read_only=True, data_only=True)
        try:
            final_rows = list(final_keep_wb.active.iter_rows(values_only=True))
        finally:
            final_keep_wb.close()
        headers = list(final_rows[0])
        nickname_index = headers.index("nickname")
        region_index = headers.index("Region")
        nicknames = {row[nickname_index] for row in final_rows[1:]}
        regions = {row[region_index] for row in final_rows[1:]}
        self.assertEqual(nicknames, {"Solo", "Dean"})
        self.assertEqual(regions, {"US"})

    def test_llm_final_review_falls_back_to_secondary_candidate_after_retryable_failure(self) -> None:
        self._make_shared_workbook()
        db = self._seed_messages()
        try:
            resolve_shared_email_candidates(
                db=db,
                input_path=self.shared_input_path,
                output_prefix=self.output_prefix,
            )
        finally:
            db.close()

        env_path = self.temp_dir / ".env"
        env_path.write_text(
            "\n".join(
                [
                    "OPENAI_PROVIDER_NAME=Primary",
                    "OPENAI_SECONDARY_API_KEY=sk-secondary",
                    "OPENAI_SECONDARY_BASE_URL=https://secondary.example/v1",
                    "OPENAI_SECONDARY_MODEL=qwen-max",
                    "OPENAI_SECONDARY_WIRE_API=responses",
                    "OPENAI_SECONDARY_PROVIDER_NAME=Secondary",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        with mock.patch(
            "email_sync.shared_email_resolution._invoke_shared_email_llm_review",
            side_effect=[
                requests.exceptions.SSLError("EOF occurred in violation of protocol"),
                {
                    "decision": "match_one",
                    "matched_candidate_keys": ["instagram:dean"],
                    "confidence": "high",
                    "reason": "Secondary candidate resolved the group.",
                    "raw_text": "{}",
                },
            ],
        ):
            result = run_shared_email_final_review(
                input_prefix=self.output_prefix,
                env_path=str(env_path),
                auto_keep_paths=[],
                base_url="https://primary.example/v1",
                api_key="sk-primary",
                model="gpt-5.4",
                wire_api="responses",
            )

        self.assertEqual(result["selected_provider"], "Secondary")
        self.assertEqual(result["selected_model"], "qwen-max")
        self.assertEqual(result["selected_wire_api"], "responses")
        self.assertEqual(result["retryable_failure_count"], 1)
        self.assertEqual(len(result["absorbed_failures"]), 1)
        self.assertTrue(any(item["provider"] == "Primary" for item in result["provider_attempts"]))
        self.assertTrue(any(item["provider"] == "Secondary" for item in result["provider_attempts"]))


if __name__ == "__main__":
    unittest.main()
