from __future__ import annotations

import json
import os
import shutil
import tempfile
import unittest
from unittest import mock
from pathlib import Path

from openpyxl import Workbook, load_workbook

from email_sync.db import Database
from email_sync.llm_review import (
    _resolve_llm_review_config,
    prepare_llm_review_candidates,
    resolve_llm_review_config_chain,
    run_and_apply_llm_review,
)
from email_sync.relation_index import rebuild_relation_index


def _addresses(*items: tuple[str, str]) -> str:
    return json.dumps([{"name": name, "address": address} for name, address in items], ensure_ascii=False)


class LlmReviewPrepTests(unittest.TestCase):
    ENV_KEYS = {
        "OPENAI_API_KEY",
        "OPENAI_BASE_URL",
        "OPENAI_MODEL",
        "OPENAI_VISION_MODEL",
        "OPENAI_WIRE_API",
        "OPENAI_PROVIDER_NAME",
        "OPENAI_REASONING_EFFORT",
        "OPENAI_SECONDARY_API_KEY",
        "OPENAI_SECONDARY_BASE_URL",
        "OPENAI_SECONDARY_MODEL",
        "OPENAI_SECONDARY_WIRE_API",
        "OPENAI_SECONDARY_PROVIDER_NAME",
        "OPENAI_SECONDARY_REASONING_EFFORT",
        "OPENAI_TERTIARY_API_KEY",
        "OPENAI_TERTIARY_BASE_URL",
        "OPENAI_TERTIARY_MODEL",
        "OPENAI_TERTIARY_WIRE_API",
        "OPENAI_TERTIARY_PROVIDER_NAME",
        "OPENAI_TERTIARY_REASONING_EFFORT",
        "LLM_API_KEY",
        "LLM_API_BASE",
        "LLM_MODEL",
        "LLM_TIMEOUT_SECONDS",
        "VISION_MODEL",
    }

    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="llm_review_test_"))
        self.input_path = self.temp_dir / "high_confidence.xlsx"
        self.output_prefix = self.temp_dir / "exports" / "测试达人库_MINISO_匹配结果_高置信_按我们去重"
        self.db_path = self.temp_dir / "email_sync.db"
        self.env_path = self.temp_dir / ".env"
        self.original_env = {key: os.environ.get(key) for key in self.ENV_KEYS}
        for key in self.ENV_KEYS:
            os.environ.pop(key, None)

    def tearDown(self) -> None:
        for key, value in self.original_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _make_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        headers = [
            "nickname",
            "@username",
            "Platform",
            "Email",
            "URL",
            "sheet_name",
            "source_row_number",
            "derived_handle",
            "creator_emails",
            "matched_contact_email",
            "matched_contact_name",
            "match_rule",
            "match_confidence",
            "last_mail_message_id",
            "last_mail_time",
            "last_mail_subject",
            "last_mail_snippet",
            "last_mail_raw_path",
            "evidence_thread_key",
        ]
        sheet.append(headers)
        sheet.append(
            [
                "Jess Prime",
                "justjess",
                "YOUTUBE",
                "jess@creator.com",
                "https://youtube.com/@justjess",
                "results",
                2,
                "justjess",
                "jess@creator.com",
                "mgmt@shared.com",
                "Talent Manager",
                "email_exact",
                "high",
                2,
                "2026-03-10T10:00:00+00:00",
                "Re: MINISO collaboration",
                "Hi Eden, I manage Jess and wanted to share rates.",
                "raw/2.eml",
                "mid:<m1>",
            ]
        )
        sheet.append(
            [
                "Jess Prime Duplicate",
                "justjess",
                "YOUTUBE",
                "jess@creator.com",
                "https://youtube.com/@justjess",
                "results",
                3,
                "justjess",
                "jess@creator.com",
                "mgmt@shared.com",
                "Talent Manager",
                "email_exact",
                "high",
                2,
                "2026-03-10T10:00:00+00:00",
                "Re: MINISO collaboration",
                "Hi Eden, I manage Jess and wanted to share rates.",
                "raw/2.eml",
                "mid:<m1>",
            ]
        )
        sheet.append(
            [
                "Kat Shine",
                "katshine",
                "YOUTUBE",
                "kat@creator.com",
                "https://youtube.com/@katshine",
                "results",
                4,
                "katshine",
                "kat@creator.com",
                "mgmt@shared.com",
                "Talent Manager",
                "email_exact",
                "high",
                2,
                "2026-03-10T10:00:00+00:00",
                "Re: MINISO collaboration",
                "Hi Eden, I manage Jess and wanted to share rates.",
                "raw/2.eml",
                "mid:<m1>",
            ]
        )
        sheet.append(
            [
                "Solo Creator",
                "solocreator",
                "TIKTOK",
                "solo@creator.com",
                "https://tiktok.com/@solocreator",
                "results",
                5,
                "solocreator",
                "solo@creator.com",
                "solo@creator.com",
                "Solo Creator",
                "email_exact",
                "high",
                3,
                "2026-03-11T10:00:00+00:00",
                "MINISO outreach",
                "Hi there, sharing this across the broader campaign team.",
                "raw/5.eml",
                "mid:<m5>",
            ]
        )
        workbook.save(self.input_path)

    def _seed_messages(self) -> Database:
        db = Database(self.db_path)
        db.init_schema()
        rows = [
            (
                "william@amagency.biz",
                "INBOX",
                1,
                1,
                "<m1>",
                "MINISO collaboration",
                None,
                None,
                "2026-03-09T10:00:00+00:00",
                "2026-03-09T10:00:00+00:00",
                "2026-03-09T10:00:00+00:00",
                "2026-03-09T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Eden", "chenjunren@amagency.biz")),
                _addresses(("Talent Manager", "mgmt@shared.com")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Hi team, can you share Jess's rate for MINISO?",
                "",
                "Hi team, can you share Jess's rate for MINISO?",
                "{}",
                "raw/1.eml",
                "sha1",
                100,
                0,
                0,
                "2026-03-09T10:00:00+00:00",
                "2026-03-09T10:00:00+00:00",
            ),
            (
                "william@amagency.biz",
                "INBOX",
                2,
                1,
                "<m2>",
                "Re: MINISO collaboration",
                "<m1>",
                "<m1>",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Talent Manager", "mgmt@shared.com")),
                _addresses(("Eden", "chenjunren@amagency.biz")),
                "[]",
                "[]",
                "[]",
                "[]",
                "Hi Eden, I manage Jess and wanted to share rates.",
                "",
                "Hi Eden, I manage Jess and wanted to share rates.",
                "{}",
                "raw/2.eml",
                "sha2",
                100,
                0,
                0,
                "2026-03-10T10:00:00+00:00",
                "2026-03-10T10:00:00+00:00",
            ),
            (
                "william@amagency.biz",
                "INBOX",
                5,
                1,
                "<m5>",
                "MINISO outreach",
                None,
                None,
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
                "[]",
                100,
                _addresses(("Eden", "chenjunren@amagency.biz")),
                _addresses(("Solo Creator", "solo@creator.com")),
                _addresses(("Team A", "team-a@example.com"), ("Team B", "team-b@example.com"), ("Team C", "team-c@example.com")),
                "[]",
                "[]",
                "[]",
                "Hi there, sharing this across the broader campaign team.",
                "",
                "Hi there, sharing this across the broader campaign team.",
                "{}",
                "raw/5.eml",
                "sha5",
                100,
                0,
                0,
                "2026-03-11T10:00:00+00:00",
                "2026-03-11T10:00:00+00:00",
            ),
        ]
        db.conn.executemany(
            """
            INSERT INTO messages (
                account_email, folder_name, uid, uidvalidity, message_id, subject, in_reply_to, references_header,
                sent_at, sent_at_raw, internal_date, internal_date_raw, flags_json, size_bytes,
                from_json, to_json, cc_json, bcc_json, reply_to_json, sender_json,
                body_text, body_html, snippet, headers_json, raw_path, raw_sha256, raw_size_bytes,
                has_attachments, attachment_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        db.conn.commit()
        rebuild_relation_index(db)
        return db

    def _sheet_rows(self, path: Path) -> tuple[list[str], list[dict[str, object]]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        headers = [str(item) if item is not None else "" for item in rows[0]]
        payload = []
        for row in rows[1:]:
            payload.append({header: row[index] for index, header in enumerate(headers)})
        return headers, payload

    def test_prepare_llm_review_candidates_emits_production_artifacts(self) -> None:
        self._make_workbook()
        db = self._seed_messages()
        try:
            result = prepare_llm_review_candidates(
                db=db,
                input_path=self.input_path,
                output_prefix=self.output_prefix,
            )
        finally:
            db.close()

        self.assertEqual(result["source_row_count"], 4)
        self.assertEqual(result["prep_row_count"], 4)
        self.assertEqual(result["deduped_row_count"], 3)
        self.assertEqual(result["llm_candidate_group_count"], 2)
        self.assertTrue(Path(result["prep_xlsx_path"]).exists())
        self.assertTrue(Path(result["deduped_xlsx_path"]).exists())
        self.assertTrue(Path(result["llm_candidates_jsonl_path"]).exists())

        prep_headers, prep_rows = self._sheet_rows(Path(result["prep_xlsx_path"]))
        self.assertIn("creator_dedupe_key", prep_headers)
        self.assertIn("same_last_mail_row_count", prep_headers)
        self.assertIn("shared_contact_suspected", prep_headers)
        self.assertIn("manager_suspected", prep_headers)
        self.assertIn("mass_cc_suspected", prep_headers)

        prep_by_nickname = {str(row["nickname"]): row for row in prep_rows}
        self.assertEqual(prep_by_nickname["Jess Prime"]["creator_dedupe_key"], "youtube:handle:justjess")
        self.assertEqual(prep_by_nickname["Jess Prime"]["same_last_mail_row_count"], 3)
        self.assertEqual(prep_by_nickname["Jess Prime"]["shared_contact_suspected"], 1)
        self.assertEqual(prep_by_nickname["Jess Prime"]["manager_suspected"], 1)
        self.assertEqual(prep_by_nickname["Solo Creator"]["mass_cc_suspected"], 1)

        _, deduped_rows = self._sheet_rows(Path(result["deduped_xlsx_path"]))
        deduped_by_nickname = {str(row["nickname"]): row for row in deduped_rows}
        self.assertNotIn("Jess Prime Duplicate", deduped_by_nickname)
        self.assertEqual(deduped_by_nickname["Jess Prime"]["same_last_mail_row_count"], 2)
        self.assertEqual(deduped_by_nickname["Kat Shine"]["shared_contact_suspected"], 1)

        with Path(result["llm_candidates_jsonl_path"]).open("r", encoding="utf-8") as handle:
            candidate_records = [json.loads(line) for line in handle if line.strip()]

        self.assertEqual(len(candidate_records), 2)
        by_group_key = {record["group_key"]: record for record in candidate_records}
        self.assertIn("last_mail_message_id:2", by_group_key)
        self.assertIn("last_mail_message_id:3", by_group_key)

        shared_group = by_group_key["last_mail_message_id:2"]
        self.assertEqual(shared_group["same_last_mail_row_count"], 2)
        self.assertEqual(len(shared_group["candidate_rows"]), 2)
        self.assertEqual(
            shared_group["candidate_creator_dedupe_keys"],
            ["youtube:handle:justjess", "youtube:handle:katshine"],
        )
        self.assertEqual(shared_group["matched_contact_emails"], ["mgmt@shared.com"])
        self.assertEqual(shared_group["candidate_rows"][0]["creator_dedupe_key"], "youtube:handle:justjess")

        mass_cc_group = by_group_key["last_mail_message_id:3"]
        self.assertEqual(mass_cc_group["mass_cc_suspected"], 1)
        self.assertEqual(len(mass_cc_group["candidate_rows"]), 1)
        self.assertEqual(mass_cc_group["candidate_rows"][0]["creator_dedupe_key"], "tiktok:handle:solocreator")

    def test_resolve_llm_review_config_prefers_openai_surface(self) -> None:
        self.env_path.write_text(
            "\n".join(
                [
                    "OPENAI_API_KEY=sk-openai",
                    "OPENAI_BASE_URL=https://openai-compatible.example/v1",
                    "OPENAI_MODEL=qwen-max",
                    "OPENAI_WIRE_API=responses",
                    "OPENAI_PROVIDER_NAME=Qwen",
                    "OPENAI_REASONING_EFFORT=high",
                    "LLM_API_KEY=sk-legacy",
                    "LLM_API_BASE=https://legacy.example/v1",
                    "LLM_MODEL=legacy-chat",
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        config = _resolve_llm_review_config(str(self.env_path))
        self.assertEqual(config.api_key, "sk-openai")
        self.assertEqual(config.base_url, "https://openai-compatible.example/v1")
        self.assertEqual(config.model, "qwen-max")
        self.assertEqual(config.wire_api, "responses")
        self.assertEqual(config.provider_name, "Qwen")
        self.assertEqual(config.reasoning_effort, "high")

    def test_resolve_llm_review_config_falls_back_to_legacy_surface(self) -> None:
        self.env_path.write_text(
            "\n".join(
                [
                    "OPENAI_BASE_URL=https://api.openai.com/v1",
                    "LLM_API_KEY=sk-legacy",
                    "LLM_API_BASE=https://legacy.example/v1",
                    "LLM_MODEL=legacy-chat",
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        config = _resolve_llm_review_config(str(self.env_path))
        self.assertEqual(config.api_key, "sk-legacy")
        self.assertEqual(config.base_url, "https://legacy.example/v1")
        self.assertEqual(config.model, "legacy-chat")
        self.assertEqual(config.wire_api, "chat_completions")

    def test_resolve_llm_review_config_chain_includes_secondary_and_tertiary_candidates(self) -> None:
        self.env_path.write_text(
            "\n".join(
                [
                    "OPENAI_API_KEY=sk-primary",
                    "OPENAI_BASE_URL=https://primary.example/v1",
                    "OPENAI_MODEL=gpt-5.4",
                    "OPENAI_PROVIDER_NAME=Primary",
                    "OPENAI_SECONDARY_API_KEY=sk-secondary",
                    "OPENAI_SECONDARY_BASE_URL=https://secondary.example/v1",
                    "OPENAI_SECONDARY_MODEL=qwen-max",
                    "OPENAI_SECONDARY_PROVIDER_NAME=Secondary",
                    "OPENAI_TERTIARY_API_KEY=sk-tertiary",
                    "OPENAI_TERTIARY_BASE_URL=https://tertiary.example/v1",
                    "OPENAI_TERTIARY_MODEL=gemini-2.5-pro",
                    "OPENAI_TERTIARY_PROVIDER_NAME=Tertiary",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        configs = resolve_llm_review_config_chain(str(self.env_path))

        self.assertEqual([config.candidate_stage for config in configs], ["primary", "secondary", "tertiary"])
        self.assertEqual([config.provider_name for config in configs], ["Primary", "Secondary", "Tertiary"])
        self.assertEqual([config.model for config in configs], ["gpt-5.4", "qwen-max", "gemini-2.5-pro"])

    def test_run_and_apply_llm_review_supports_responses_wire_api(self) -> None:
        self._make_workbook()
        db = self._seed_messages()
        try:
            prepare_llm_review_candidates(db=db, input_path=self.input_path, output_prefix=self.output_prefix)
        finally:
            db.close()

        self.env_path.write_text(
            "\n".join(
                [
                    "OPENAI_API_KEY=sk-openai",
                    "OPENAI_BASE_URL=https://openai-compatible.example/v1",
                    "OPENAI_MODEL=qwen-max",
                    "OPENAI_WIRE_API=responses",
                    "OPENAI_PROVIDER_NAME=Qwen",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        def fake_post(url, headers, json, timeout):  # noqa: ANN001
            self.assertEqual(url, "https://openai-compatible.example/v1/responses")
            self.assertEqual(json["model"], "qwen-max")
            self.assertIn("input", json)
            content_text = json["input"][1]["content"][0]["text"]
            if "last_mail_message_id:2" in content_text:
                payload = {
                    "output_text": json_module.dumps(
                        {
                            "decision": "match_some",
                            "matched_creator_dedupe_keys": ["youtube:handle:justjess"],
                            "sender_role": "manager",
                            "confidence": "high",
                            "reason": "Only Jess should stay.",
                        },
                        ensure_ascii=False,
                    )
                }
            else:
                payload = {
                    "output_text": json_module.dumps(
                        {
                            "decision": "uncertain",
                            "matched_creator_dedupe_keys": [],
                            "sender_role": "unclear",
                            "confidence": "low",
                            "reason": "Mass cc but weak evidence.",
                        },
                        ensure_ascii=False,
                    )
                }
            response = mock.Mock()
            response.status_code = 200
            response.json.return_value = payload
            return response

        import json as json_module

        with mock.patch("email_sync.llm_review.requests.post", side_effect=fake_post):
            result = run_and_apply_llm_review(input_prefix=self.output_prefix, env_path=str(self.env_path))

        self.assertEqual(result["review_group_count"], 2)
        self.assertEqual(result["keep_row_count"], 2)
        self.assertTrue(Path(result["llm_review_jsonl_path"]).exists())
        self.assertTrue(Path(result["llm_reviewed_xlsx_path"]).exists())
        self.assertTrue(Path(result["llm_reviewed_keep_xlsx_path"]).exists())

        with Path(result["llm_review_jsonl_path"]).open("r", encoding="utf-8") as handle:
            review_records = [json.loads(line) for line in handle if line.strip()]
        by_group = {record["group_key"]: record for record in review_records}
        self.assertEqual(by_group["last_mail_message_id:2"]["decision"], "match_some")
        self.assertEqual(by_group["last_mail_message_id:2"]["matched_creator_dedupe_keys"], ["youtube:handle:justjess"])
        self.assertEqual(by_group["last_mail_message_id:3"]["decision"], "uncertain")

        reviewed_headers, reviewed_rows = self._sheet_rows(Path(result["llm_reviewed_xlsx_path"]))
        self.assertIn("llm_review_decision", reviewed_headers)
        by_nickname = {str(row["nickname"]): row for row in reviewed_rows}
        self.assertEqual(by_nickname["Jess Prime"]["llm_review_keep"], "yes")
        self.assertEqual(by_nickname["Kat Shine"]["llm_review_keep"], "no")
        self.assertEqual(by_nickname["Solo Creator"]["llm_review_keep"], "yes")
        self.assertEqual(by_nickname["Solo Creator"]["llm_review_decision"], "uncertain")

        _, keep_rows = self._sheet_rows(Path(result["llm_reviewed_keep_xlsx_path"]))
        keep_nicknames = {str(row["nickname"]) for row in keep_rows}
        self.assertEqual(keep_nicknames, {"Jess Prime", "Solo Creator"})

    def test_run_and_apply_llm_review_supports_chat_completions_wire_api(self) -> None:
        self._make_workbook()
        db = self._seed_messages()
        try:
            prepare_llm_review_candidates(db=db, input_path=self.input_path, output_prefix=self.output_prefix)
        finally:
            db.close()

        self.env_path.write_text(
            "\n".join(
                [
                    "OPENAI_API_KEY=sk-openai",
                    "OPENAI_BASE_URL=https://chat.example/v1",
                    "OPENAI_MODEL=qwen-chat",
                    "OPENAI_WIRE_API=chat_completions",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        def fake_post(url, headers, json, timeout):  # noqa: ANN001
            self.assertEqual(url, "https://chat.example/v1/chat/completions")
            self.assertIn("messages", json)
            response = mock.Mock()
            response.status_code = 200
            response.json.return_value = {
                "choices": [
                    {
                        "message": {
                            "content": json_module.dumps(
                                {
                                    "decision": "match_all",
                                    "matched_creator_dedupe_keys": [],
                                    "sender_role": "manager",
                                    "confidence": "medium",
                                    "reason": "Keep the full group.",
                                },
                                ensure_ascii=False,
                            )
                        }
                    }
                ]
            }
            return response

        import json as json_module

        with mock.patch("email_sync.llm_review.requests.post", side_effect=fake_post):
            result = run_and_apply_llm_review(input_prefix=self.output_prefix, env_path=str(self.env_path))

        self.assertEqual(result["review_group_count"], 2)
        _, keep_rows = self._sheet_rows(Path(result["llm_reviewed_keep_xlsx_path"]))
        keep_nicknames = {str(row["nickname"]) for row in keep_rows}
        self.assertEqual(keep_nicknames, {"Jess Prime", "Kat Shine", "Solo Creator"})


if __name__ == "__main__":
    unittest.main()
