from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

from email_sync.db import Database
import scripts.run_parsed_field_mail_funnel as funnel


class ParsedFieldMailFunnelTests(unittest.TestCase):
    def test_llm_high_rows_preserve_inferred_platform_in_keep_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "email_sync.db"
            db = Database(db_path)
            db.init_schema()
            try:
                sent_at = "2026-04-08T10:00:00+08:00"
                db.conn.execute(
                    """
                    INSERT INTO messages (
                        account_email, folder_name, uid, uidvalidity, message_id, subject, in_reply_to, references_header,
                        sent_at, sent_at_raw, internal_date, internal_date_raw, flags_json, size_bytes,
                        from_json, to_json, cc_json, bcc_json, reply_to_json, sender_json,
                        body_text, body_html, snippet, headers_json, raw_path, raw_sha256, raw_size_bytes,
                        has_attachments, attachment_count, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, '', '', ?, ?, ?, ?, '[]', 0, ?, ?, '[]', '[]', '[]', '[]', ?, '', ?, '{}', ?, '', 0, 0, 0, ?, ?)
                    """,
                    (
                        "partnerships@amagency.biz",
                        "其他文件夹/达人回信",
                        1,
                        1,
                        "<msg-llm-1>",
                        "MINISO collab",
                        sent_at,
                        sent_at,
                        sent_at,
                        sent_at,
                        '[{"address":"creator@example.com","name":"Alpha"}]',
                        '[{"address":"partnerships@amagency.biz","name":"AM"}]',
                        "Hi team, my Instagram: @alpha_creator",
                        "MINISO campaign",
                        "alpha-last.eml",
                        sent_at,
                        sent_at,
                    ),
                )
                message_row_id = int(db.conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                db.conn.execute(
                    """
                    INSERT INTO message_index (
                        message_row_id, normalized_subject, direction, thread_key, thread_root_message_id,
                        thread_parent_message_id, thread_depth, sent_sort_at, external_contact_count, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        message_row_id,
                        "miniso collab",
                        "inbound",
                        "thread-llm-alpha",
                        "<msg-llm-1>",
                        "",
                        0,
                        sent_at,
                        1,
                        sent_at,
                    ),
                )
                db.conn.commit()
            finally:
                db.close()

            output_prefix = root / "miniso_funnel"

            def fake_llm_review(rows, **_kwargs):
                reviewed = []
                for row in rows:
                    reviewed.append(
                        {
                            **row,
                            "llm_handle": "alpha_creator",
                            "llm_evidence": "body mentioned Instagram handle explicitly",
                            "resolution_confidence_final": "high",
                        }
                    )
                return reviewed

            with patch.object(funnel, "_run_default_llm_review", side_effect=fake_llm_review):
                result = funnel.run(
                    SimpleNamespace(
                        env_file=".env",
                        db_path=str(db_path),
                        input_workbook="",
                        keyword="MINISO",
                        local_date="2026-04-08",
                        timezone="Asia/Shanghai",
                        output_prefix=str(output_prefix),
                        base_url="",
                        api_key="",
                        model="",
                        wire_api="",
                        llm_max_workers=1,
                        llm_limit=0,
                    )
                )

            self.assertEqual(result["keep_row_count"], 1)
            keep_wb = load_workbook(result["keep_xlsx_path"], read_only=True)
            try:
                ws = keep_wb.active
                rows = list(ws.iter_rows(values_only=True))
            finally:
                keep_wb.close()
            self.assertEqual(rows[0][6], "Platform")
            self.assertEqual(rows[1][6], "Instagram")
            self.assertEqual(rows[1][10], "alpha_creator")


if __name__ == "__main__":
    unittest.main()
