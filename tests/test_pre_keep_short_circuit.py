from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from email_sync.db import Database
from email_sync.pre_keep_short_circuit import route_pre_keep_workbook


class PreKeepShortCircuitTests(unittest.TestCase):
    def test_route_pre_keep_workbook_splits_screened_and_heavy_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "email_sync.db"
            keep_path = root / "keep.xlsx"
            routed_keep_path = root / "keep_full_screening_only.xlsx"
            mail_only_path = root / "keep_pre_keep_mail_only.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Platform",
                    "@username",
                    "URL",
                    "matched_contact_email",
                    "subject",
                    "thread_key",
                ]
            )
            sheet.append(["Instagram", "alpha", "https://www.instagram.com/alpha", "alpha@example.com", "Re: Alpha outreach", "mid:<alpha-root>"])
            sheet.append(["TikTok", "beta", "https://www.tiktok.com/@beta", "beta@example.com", "Re: Beta outreach", ""])
            sheet.append(["Instagram", "gamma", "https://www.instagram.com/gamma", "gamma@example.com", "Re: Gamma outreach", ""])
            workbook.save(keep_path)

            db = Database(db_path)
            db.init_schema()
            db.conn.execute(
                """
                INSERT INTO thread_assignments (
                    thread_key, owner_scope, creator_id, platform, brand,
                    matched_contact_email, normalized_subject, source_stage, source_run_id,
                    last_mail_message_id, last_mail_sent_at, mail_update_revision, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "mid:<alpha-root>",
                    "ou_owner",
                    "alpha",
                    "instagram",
                    "MINISO",
                    "alpha@example.com",
                    "alpha outreach",
                    "final_keep",
                    "run-1",
                    "101",
                    "2026-04-05T10:00:00+08:00",
                    3,
                    "2026-04-05T10:00:00+08:00",
                    "2026-04-05T10:00:00+08:00",
                ),
            )
            db.conn.commit()
            db.close()

            result = route_pre_keep_workbook(
                keep_workbook_path=keep_path,
                routed_keep_workbook_path=routed_keep_path,
                mail_only_workbook_path=mail_only_path,
                db_path=db_path,
                owner_scope="ou_owner",
                brand="MINISO",
                existing_index={
                    "ou_owner::alpha::instagram": {
                        "record_id": "rec_alpha",
                        "fields": {"ai 是否通过": "是"},
                    },
                    "ou_owner::beta::tiktok": {
                        "record_id": "rec_beta",
                        "fields": {"ai 是否通过": ""},
                    },
                },
                owner_scope_enabled=True,
            )

            stats = result["stats"]
            self.assertEqual(stats["input_row_count"], 3)
            self.assertEqual(stats["mail_only_count"], 1)
            self.assertEqual(stats["full_screening_count"], 2)
            self.assertEqual(stats["known_thread_hit_count"], 1)
            self.assertEqual(stats["thread_assignment_cache_hit_count"], 1)
            self.assertEqual(stats["existing_screened_count"], 1)
            self.assertEqual(stats["existing_unscreened_count"], 1)
            self.assertEqual(stats["new_creator_count"], 1)
            self.assertEqual(stats["partial_refresh_count"], 0)

            routed_rows = list(load_workbook(routed_keep_path, read_only=True).active.iter_rows(values_only=True))
            mail_only_rows = list(load_workbook(mail_only_path, read_only=True).active.iter_rows(values_only=True))
            self.assertEqual(len(routed_rows), 3)
            self.assertEqual(len(mail_only_rows), 2)
            self.assertEqual({row[1] for row in routed_rows[1:]}, {"beta", "gamma"})
            self.assertEqual(mail_only_rows[1][1], "alpha")


if __name__ == "__main__":
    unittest.main()
