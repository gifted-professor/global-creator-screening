#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import zlib
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .visual_prompt_adapter import build_visual_prompt_artifacts


VERSION = "v1-flex-visual-reuse"
MAIN_SHEET_NAME = "需求主表"
DEFAULT_OUTPUT_DIRNAME = "compiled_outputs"
PLATFORM_ORDER = ["tiktok", "instagram", "youtube"]

EMPTY_MARKERS = {
    "",
    "/",
    "-",
    "无",
    "none",
    "null",
    "n/a",
    "na",
    "暂不填写",
    "暂时先不填",
}
YES_VALUES = {"是", "需要", "yes", "true", "1", "y"}
NO_VALUES = {"否", "不需要", "no", "false", "0", "n"}
FAIL_ACTION_MAP = {"直接不通过": "reject", "转人工": "manual_review"}
RELATION_MAP = {"同时满足": "all", "任一满足": "any"}
MANUAL_ACTION_MAP = {"转人工": "manual_review", "仅提醒": "note_only"}
OUTPUT_STATUS_MAP = {"通过": "pass", "不通过": "reject", "转人工": "manual_review"}

SENSITIVE_TERMS = [
    "黑人",
    "白人",
    "黄种人",
    "亚裔",
    "拉丁裔",
    "中老年",
    "老年",
    "老人",
    "种族",
    "民族",
    "年龄",
    "race",
    "ethnic",
    "black",
    "white",
]
IGNORE_ROW_TERMS = [
    "上传需求时请删掉此行",
    "以上为举例",
    "只保留封面能稳定看出来的项",
    "自动化特征清单",
    "需要的特征清单",
]

STEP_NAME_MAP = {
    "step_1_qualification": ["步骤1：基础资质审核", "步骤1:基础资质审核", "步骤1：基础资质"],
    "step_2_data": ["步骤2：数据审核", "步骤2:数据审核"],
    "step_3_visual": ["步骤3：内容 / 视觉审核", "步骤3:内容/视觉审核", "步骤3：内容/视觉审核"],
    "step_4_exclusions": ["步骤4：排除项审核", "步骤4:排除项审核"],
}

VISUAL_FEATURE_DEFS = {
    "multi_person_interaction": {
        "label": "多人互动",
        "aliases": ["多人互动"],
        "reuse_status": "runtime_supported",
    },
    "speaking_led": {
        "label": "Speaking-led",
        "aliases": ["speaking-led", "镜头口播", "口播", "镜头前开口说话"],
        "reuse_status": "runtime_supported",
    },
    "real_life_scene": {
        "label": "真实生活场景",
        "aliases": ["真实生活场景"],
        "reuse_status": "runtime_supported",
    },
    "kid_interaction": {
        "label": "孩子互动",
        "aliases": ["孩子互动"],
        "reuse_status": "runtime_supported",
    },
    "product_display": {
        "label": "产品展示",
        "aliases": ["产品展示", "潮玩产品展示"],
        "reuse_status": "runtime_supported",
    },
    "outdoor_yard": {
        "label": "户外庭院",
        "aliases": ["户外庭院"],
        "reuse_status": "runtime_supported",
    },
    "urban_street_scene": {
        "label": "户外城市和街景",
        "aliases": ["户外城市和街景", "城市街景", "街景"],
        "reuse_status": "template_only",
    },
    "pet_interaction": {
        "label": "宠物互动",
        "aliases": ["宠物互动"],
        "reuse_status": "runtime_supported",
    },
    "blind_box_unboxing": {
        "label": "盲盒开箱",
        "aliases": ["盲盒开箱"],
        "reuse_status": "template_only",
    },
    "couple_interaction": {
        "label": "情侣互动",
        "aliases": ["情侣互动"],
        "reuse_status": "template_only",
    },
    "coser": {
        "label": "coser",
        "aliases": ["coser", "变装"],
        "reuse_status": "template_only",
    },
}

RUNTIME_RULE_PLATFORMS = {
    "region_gate": {"instagram"},
    "language_gate": {"instagram"},
    "view_count_mean": {"tiktok"},
    "view_count_median": {"tiktok"},
    "follower_count": {"tiktok", "instagram"},
    "activity_recency_days": {"tiktok", "instagram"},
    "text_keyword_blocklist": {"tiktok", "instagram"},
    "relationship_keyword_block": {"instagram"},
    "content_keyword_ratio": {"tiktok", "instagram"},
    "multi_dance_ratio": {"tiktok", "instagram"},
    "selfie_or_couple_ratio": {"tiktok", "instagram"},
    "green_screen": {"tiktok", "instagram"},
    "visual_feature_group": {"tiktok", "instagram"},
}


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(text: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(text)).lower()


def is_empty_marker(text: Any) -> bool:
    return compact(text) in {compact(item) for item in EMPTY_MARKERS}


def clean_text(text: Any) -> str | None:
    value = normalize_text(text)
    if not value or is_empty_marker(value):
        return None
    return value


def split_list_text(text: Any) -> list[str]:
    raw = normalize_text(text)
    if not raw or is_empty_marker(raw):
        return []
    normalized = raw.replace("；", ",").replace("，", ",").replace("\n", ",").replace("/", ",")
    items = [item.strip() for item in normalized.split(",")]
    return [item for item in items if item and not is_empty_marker(item)]


def parse_int(text: Any) -> int | None:
    raw = normalize_text(text)
    if not raw or is_empty_marker(raw):
        return None
    match = re.search(r"-?\d+", raw.replace(",", ""))
    if not match:
        raise ValueError(f"无法解析整数: {raw}")
    return int(match.group(0))


def parse_boolish(text: Any) -> bool | None:
    raw = normalize_text(text)
    if not raw or is_empty_marker(raw):
        return False
    lowered = raw.strip().lower()
    if lowered in YES_VALUES:
        return True
    if lowered in NO_VALUES:
        return False
    if raw in YES_VALUES:
        return True
    if raw in NO_VALUES:
        return False
    return None


def parse_choice(text: Any, mapping: dict[str, str]) -> str | None:
    raw = normalize_text(text)
    if not raw or is_empty_marker(raw):
        return None
    return mapping.get(raw)


def parse_platform_scope(text: Any) -> list[str]:
    raw = normalize_text(text)
    if not raw:
        return []
    if raw == "两者":
        return ["tiktok", "instagram"]
    normalized = raw.replace("／", "/").replace("，", ",").replace("、", ",").replace("/", ",")
    items = [item.strip() for item in normalized.split(",") if item.strip()]
    mapped: list[str] = []
    for item in items:
        lowered = item.lower()
        if lowered == "tiktok":
            mapped.append("tiktok")
        elif lowered == "instagram":
            mapped.append("instagram")
        elif lowered == "youtube":
            mapped.append("youtube")
    ordered = [platform for platform in PLATFORM_ORDER if platform in mapped]
    return ordered


def normalize_step_name(text: str) -> str | None:
    value = normalize_text(text)
    if not value:
        return None
    for step_key, aliases in STEP_NAME_MAP.items():
        if any(compact(value) == compact(alias) for alias in aliases):
            return step_key
    return None


def parse_step_list(text: Any) -> list[str]:
    values = split_list_text(text)
    result: list[str] = []
    for item in values:
        step_name = normalize_step_name(item)
        if step_name:
            result.append(step_name)
    return result


def slugify_custom_key(text: str, prefix: str) -> str:
    ascii_part = re.sub(r"[^a-zA-Z0-9]+", "_", normalize_text(text).lower()).strip("_")
    if ascii_part:
        return f"{prefix}_{ascii_part}"
    checksum = zlib.crc32(normalize_text(text).encode("utf-8")) & 0xFFFFFFFF
    return f"{prefix}_{checksum:08x}"


def contains_sensitive_terms(*parts: Any) -> bool:
    merged = " ".join(normalize_text(item) for item in parts if normalize_text(item))
    lowered = merged.lower()
    return any(term in merged or term in lowered for term in SENSITIVE_TERMS)


def is_ignored_row(label: str) -> bool:
    return any(term in label for term in IGNORE_ROW_TERMS)


def detect_section(label: str) -> str | None:
    raw = normalize_text(label)
    if not raw:
        return None
    if re.match(r"^A\.", raw, re.IGNORECASE) and ("基本信息" in raw or "项目信息" in raw):
        return "basic_info"
    if re.match(r"^B\.", raw, re.IGNORECASE) and ("步骤1" in raw or "资质审核" in raw):
        return "qualification"
    if re.match(r"^C\.", raw, re.IGNORECASE) and ("步骤2" in raw or "数据审核" in raw):
        return "data_audit"
    if re.match(r"^D\.", raw, re.IGNORECASE) and ("步骤3" in raw or "视觉审核" in raw):
        return "visual_audit"
    if re.match(r"^E\.", raw, re.IGNORECASE) and ("步骤4" in raw or "排除项审核" in raw):
        return "exclusions"
    if re.match(r"^F\.", raw, re.IGNORECASE) and "人工判断项" in raw:
        return "manual_review"
    if re.match(r"^G\.", raw, re.IGNORECASE) and "最终判定逻辑" in raw:
        return "final_logic"
    return None


def iter_main_rows(worksheet) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(1, worksheet.max_row + 1):
        label = normalize_text(worksheet.cell(row=row_index, column=1).value)
        value = normalize_text(worksheet.cell(row=row_index, column=2).value)
        note = normalize_text(worksheet.cell(row=row_index, column=3).value)
        if not any((label, value, note)):
            continue
        rows.append(
            {
                "row": row_index,
                "label": label,
                "value": value,
                "note": note,
                "value_cell": f"B{row_index}",
                "note_cell": f"C{row_index}",
            }
        )
    return rows


def split_sections(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    sections = {
        "basic_info": [],
        "qualification": [],
        "data_audit": [],
        "visual_audit": [],
        "exclusions": [],
        "manual_review": [],
        "final_logic": [],
    }
    current_section: str | None = None
    for row in rows:
        section = detect_section(row["label"])
        if section:
            current_section = section
            continue
        if current_section:
            sections[current_section].append(row)
    return sections


def match_visual_feature(label: str) -> tuple[str, dict[str, Any]] | None:
    current = compact(label)
    for feature_key, feature_info in VISUAL_FEATURE_DEFS.items():
        for alias in feature_info["aliases"]:
            alias_compact = compact(alias)
            if alias_compact == current or alias_compact in current or current in alias_compact:
                return feature_key, feature_info
    return None


def parse_basic_info(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "project_name": None,
        "brand_product": None,
        "platform_scope": [],
        "audit_goal": None,
        "reference_accounts": [],
        "counterexample_accounts": [],
        "notes": None,
        "extra_rows": [],
    }
    warnings: list[str] = []

    for row in rows:
        label = row["label"]
        value = row["value"]
        if "项目名称" in label:
            payload["project_name"] = clean_text(value)
        elif "品牌" in label and ("产品" in label or "使用场景" in label):
            payload["brand_product"] = clean_text(value)
        elif "适用平台" in label:
            payload["platform_scope"] = parse_platform_scope(value)
            if not payload["platform_scope"] and clean_text(value):
                warnings.append(f"{row['value_cell']} 适用平台未识别：{value}")
        elif "审核目标" in label:
            payload["audit_goal"] = clean_text(value)
        elif "参考账号" in label:
            payload["reference_accounts"] = split_list_text(value)
        elif "反例账号" in label:
            payload["counterexample_accounts"] = split_list_text(value)
        elif label == "备注":
            payload["notes"] = clean_text(value)
        else:
            payload["extra_rows"].append(deepcopy(row))
    return payload, warnings


def parse_qualification(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "region_requirement": [],
        "language_requirement": [],
        "check_scope": None,
        "fail_action": None,
        "notes": None,
        "extra_rows": [],
    }
    warnings: list[str] = []
    for row in rows:
        label = row["label"]
        value = row["value"]
        if "地区要求" in label:
            payload["region_requirement"] = split_list_text(value)
        elif "语言要求" in label:
            payload["language_requirement"] = split_list_text(value)
        elif "检查位置" in label:
            payload["check_scope"] = "profile_metadata" if "主页" in value else "all_content" if "内容" in value else None
            if clean_text(value) and payload["check_scope"] is None:
                warnings.append(f"{row['value_cell']} 检查位置未识别：{value}")
        elif "不符合时处理" in label:
            payload["fail_action"] = parse_choice(value, FAIL_ACTION_MAP)
        elif "补充说明" in label:
            payload["notes"] = clean_text(value)
        else:
            payload["extra_rows"].append(deepcopy(row))
    return payload, warnings


def parse_data_audit(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "sample_size": None,
        "mean_view_threshold": None,
        "median_view_threshold": None,
        "follower_threshold": None,
        "recent_active_days": None,
        "judgement_relation": None,
        "fail_action": None,
        "notes": None,
        "extra_rows": [],
    }
    warnings: list[str] = []
    for row in rows:
        label = row["label"]
        value = row["value"]
        try:
            if "取样视频数" in label:
                payload["sample_size"] = parse_int(value)
            elif "平均播放量阈值" in label:
                payload["mean_view_threshold"] = parse_int(value)
            elif "中位数播放量阈值" in label:
                payload["median_view_threshold"] = parse_int(value)
            elif "粉丝数阈值" in label:
                payload["follower_threshold"] = parse_int(value)
            elif "最近活跃要求" in label:
                payload["recent_active_days"] = parse_int(value)
            elif "判定关系" in label:
                payload["judgement_relation"] = parse_choice(value, RELATION_MAP)
            elif "不符合时处理" in label:
                payload["fail_action"] = parse_choice(value, FAIL_ACTION_MAP)
            elif "补充说明" in label:
                payload["notes"] = clean_text(value)
            else:
                payload["extra_rows"].append(deepcopy(row))
        except ValueError:
            warnings.append(f"{row['value_cell']} 数值无法解析：{value}")
    return payload, warnings


def parse_visual_feature_row(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    label = row["label"]
    value = row["value"]
    note = row["note"]
    if not label or label == "无" or is_ignored_row(label):
        return None, None
    enabled = parse_boolish(value)
    if enabled is None:
        return None, f"{row['value_cell']} 视觉项开关未识别：{value}"
    feature_match = match_visual_feature(label)
    if feature_match:
        feature_key, feature_info = feature_match
        return (
            {
                "key": feature_key,
                "label": feature_info["label"],
                "source_label": label,
                "source_cell": row["value_cell"],
                "enabled": enabled,
                "note": clean_text(note),
                "reuse_status": feature_info["reuse_status"],
            },
            None,
        )
    custom_key = slugify_custom_key(label, "visual")
    return (
        {
            "key": custom_key,
            "label": label,
            "source_label": label,
            "source_cell": row["value_cell"],
            "enabled": enabled,
            "note": clean_text(note),
            "reuse_status": "template_only",
        },
        None,
    )


def parse_visual_audit(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "cover_count": None,
        "min_hit_features": None,
        "notes": None,
        "features": [],
        "extra_rows": [],
    }
    warnings: list[str] = []
    for row in rows:
        label = row["label"]
        value = row["value"]
        try:
            if "查看封面数量" in label:
                payload["cover_count"] = parse_int(value)
            elif "至少命中几类特征" in label:
                payload["min_hit_features"] = parse_int(value)
            elif "补充说明" in label:
                payload["notes"] = clean_text(value)
                if contains_sensitive_terms(value):
                    warnings.append(f"{row['value_cell']} 视觉说明含敏感属性，只会作为合规提醒。")
            elif "特征清单" in label:
                continue
            else:
                feature, warning = parse_visual_feature_row(row)
                if feature:
                    payload["features"].append(feature)
                elif warning:
                    warnings.append(warning)
                else:
                    payload["extra_rows"].append(deepcopy(row))
        except ValueError:
            warnings.append(f"{row['value_cell']} 视觉字段数值无法解析：{value}")
    return payload, warnings


def normalize_custom_visual_risk(label: str) -> tuple[str, str]:
    if "未成年" in label or "小孩" in label:
        return "minor_presence", "未成年/小孩出镜"
    return slugify_custom_key(label, "visual_risk"), label


def parse_exclusions(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "text_keyword_blocklist": [],
        "relationship_keyword_blocklist": [],
        "beauty_ratio_threshold": None,
        "multi_dance_ratio_threshold": None,
        "selfie_couple_ratio_threshold": None,
        "green_screen_direct_reject": False,
        "other_exclusion_items": None,
        "notes": None,
        "custom_visual_risks": [],
        "compliance_notes": [],
        "extra_rows": [],
    }
    warnings: list[str] = []

    for row in rows:
        label = row["label"]
        value = row["value"]
        note = row["note"]
        try:
            if "文本关键词排除" in label:
                payload["text_keyword_blocklist"] = split_list_text(value)
            elif "关系词排除" in label:
                payload["relationship_keyword_blocklist"] = split_list_text(value)
            elif "美妆内容占比阈值" in label:
                payload["beauty_ratio_threshold"] = parse_int(value)
            elif "多人跳舞占比阈值" in label:
                payload["multi_dance_ratio_threshold"] = parse_int(value)
            elif "自拍" in label and "占比阈值" in label:
                payload["selfie_couple_ratio_threshold"] = parse_int(value)
            elif "绿幕" in label:
                parsed = parse_boolish(value)
                payload["green_screen_direct_reject"] = bool(parsed)
            elif "其他排除项" in label:
                payload["other_exclusion_items"] = clean_text(value)
            elif "补充说明" in label:
                payload["notes"] = clean_text(value)
            elif is_ignored_row(label) or label == "无":
                continue
            else:
                sensitive = contains_sensitive_terms(label, value)
                toggle = parse_boolish(value)
                if sensitive:
                    payload["compliance_notes"].append(
                        {
                            "source_cell": row["value_cell"],
                            "label": label,
                            "value": clean_text(value),
                            "note": clean_text(note),
                            "policy": "never_compile_to_automation",
                        }
                    )
                    warnings.append(f"{row['value_cell']} 排除项涉及敏感属性，已降级为合规提醒。")
                    continue
                if toggle is True:
                    risk_key, normalized_label = normalize_custom_visual_risk(label)
                    payload["custom_visual_risks"].append(
                        {
                            "key": risk_key,
                            "label": normalized_label,
                            "source_label": label,
                            "source_cell": row["value_cell"],
                            "operator": "must_not_appear",
                            "reuse_status": "template_only",
                            "note": clean_text(note),
                        }
                    )
                    continue
                payload["extra_rows"].append(deepcopy(row))
        except ValueError:
            warnings.append(f"{row['value_cell']} 排除项数值无法解析：{value}")
    return payload, warnings


def parse_manual_review(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "comment_emoji_abnormal": False,
        "persona_or_niche_manual": False,
        "full_video_review": False,
        "other_subjective_items": None,
        "protected_attribute_notice": None,
        "notes": None,
        "extra_items": [],
        "compliance_notes": [],
    }
    warnings: list[str] = []
    for row in rows:
        label = row["label"]
        value = row["value"]
        note = row["note"]
        parsed_bool = parse_boolish(value)
        if label == "无":
            continue
        if "评论区 emoji" in label:
            payload["comment_emoji_abnormal"] = bool(parsed_bool)
        elif "鲜明人设" in label or "垂直 niche" in label:
            payload["persona_or_niche_manual"] = bool(parsed_bool)
        elif "完整视频内容判断" in label:
            payload["full_video_review"] = bool(parsed_bool)
        elif "其他主观判断项" in label:
            payload["other_subjective_items"] = clean_text(value)
        elif "合规提醒" in label:
            compliance_value = clean_text(value) or clean_text(note)
            if not compliance_value:
                normalized_label = clean_text(label)
                if normalized_label and normalized_label != "合规提醒":
                    compliance_value = normalized_label
            if compliance_value:
                payload["compliance_notes"].append(
                    {
                        "source_cell": row["value_cell"],
                        "label": label,
                        "value": compliance_value,
                        "note": clean_text(note),
                        "policy": "never_compile_to_automation",
                    }
                )
        elif "受保护属性相关判断" in label:
            payload["protected_attribute_notice"] = clean_text(value)
            if payload["protected_attribute_notice"]:
                payload["compliance_notes"].append(
                    {
                        "source_cell": row["value_cell"],
                        "label": label,
                        "value": payload["protected_attribute_notice"],
                        "policy": "never_compile_to_automation",
                    }
                )
        elif "补充说明" in label:
            payload["notes"] = clean_text(value)
        else:
            raw_text = clean_text(value) or clean_text(note)
            if not raw_text:
                normalized_label = clean_text(label)
                if normalized_label and normalized_label not in {"合规提醒", "补充说明"}:
                    raw_text = normalized_label
            if raw_text:
                payload["extra_items"].append(
                    {
                        "source_cell": row["value_cell"],
                        "label": label,
                        "value": clean_text(value),
                        "note": clean_text(note),
                    }
                )
                warnings.append(f"{row['value_cell']} 存在未标准化人工项，已保留为 extra_items。")
    return payload, warnings


def parse_final_logic(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[str]]:
    payload = {
        "must_pass_steps": [],
        "reject_if_any_steps": [],
        "manual_hit_action": None,
        "success_output": None,
        "failure_output": None,
        "notes": None,
        "extra_rows": [],
    }
    warnings: list[str] = []
    for row in rows:
        label = row["label"]
        value = row["value"]
        if "必须满足哪些步骤" in label:
            payload["must_pass_steps"] = parse_step_list(value)
            if clean_text(value) and not payload["must_pass_steps"]:
                warnings.append(f"{row['value_cell']} 最终逻辑步骤未识别：{value}")
        elif "任一触发即不通过的项" in label:
            payload["reject_if_any_steps"] = parse_step_list(value)
            if clean_text(value) and not payload["reject_if_any_steps"]:
                warnings.append(f"{row['value_cell']} 最终逻辑步骤未识别：{value}")
        elif "人工判断项命中时如何处理" in label:
            payload["manual_hit_action"] = parse_choice(value, MANUAL_ACTION_MAP)
        elif "满足条件时输出" in label:
            payload["success_output"] = parse_choice(value, OUTPUT_STATUS_MAP)
        elif "不满足时输出" in label:
            payload["failure_output"] = parse_choice(value, OUTPUT_STATUS_MAP)
        elif "补充说明" in label:
            payload["notes"] = clean_text(value)
        else:
            payload["extra_rows"].append(deepcopy(row))
    return payload, warnings


def intersect_platforms(requested: list[str], supported_key: str) -> list[str]:
    supported = RUNTIME_RULE_PLATFORMS[supported_key]
    return [platform for platform in PLATFORM_ORDER if platform in requested and platform in supported]


def build_manual_review_items(manual_review: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if manual_review.get("comment_emoji_abnormal"):
        items.append({"key": "comment_emoji_abnormal", "label": "评论区 emoji 异常"})
    if manual_review.get("persona_or_niche_manual"):
        items.append({"key": "persona_or_niche_manual", "label": "鲜明人设 / 垂直 niche"})
    if manual_review.get("full_video_review"):
        items.append({"key": "full_video_review", "label": "完整视频内容判断"})
    if manual_review.get("other_subjective_items"):
        items.append(
            {
                "key": "other_subjective_items",
                "label": "其他主观判断项",
                "value": manual_review["other_subjective_items"],
            }
        )
    items.extend(deepcopy(manual_review.get("extra_items") or []))
    return items


def build_rulespec(structured: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    requested_platforms = structured["basic_info"].get("platform_scope") or []
    rules: list[dict[str, Any]] = []
    warnings: list[str] = []

    qualification = structured["qualification"]
    if qualification.get("region_requirement"):
        platforms = intersect_platforms(requested_platforms, "region_gate")
        if platforms:
            rules.append(
                {
                    "id": "qualification_region_gate",
                    "step": "step_1_qualification",
                    "rule_type": "region_gate",
                    "platforms": platforms,
                    "operator": "contains_any",
                    "values": qualification["region_requirement"],
                }
            )
        else:
            warnings.append("地区要求当前没有可自动执行的平台支持。")

    if qualification.get("language_requirement"):
        platforms = intersect_platforms(requested_platforms, "language_gate")
        if platforms:
            rules.append(
                {
                    "id": "qualification_language_gate",
                    "step": "step_1_qualification",
                    "rule_type": "language_gate",
                    "platforms": platforms,
                    "operator": "contains_any",
                    "values": qualification["language_requirement"],
                }
            )
        else:
            warnings.append("语言要求当前没有可自动执行的平台支持。")

    data_audit = structured["data_audit"]
    sample_size = data_audit.get("sample_size") or 50
    for field_name, rule_id, support_key, threshold_key in [
        ("mean_view_threshold", "data_mean_view_threshold", "view_count_mean", "threshold"),
        ("median_view_threshold", "data_median_view_threshold", "view_count_median", "threshold"),
        ("follower_threshold", "data_follower_threshold", "follower_count", "threshold"),
        ("recent_active_days", "data_recent_active_days", "activity_recency_days", "threshold_days"),
    ]:
        value = data_audit.get(field_name)
        if value is None:
            continue
        platforms = intersect_platforms(requested_platforms, support_key)
        if not platforms:
            warnings.append(f"{field_name} 当前没有可自动执行的平台支持。")
            continue
        rule = {
            "id": rule_id,
            "step": "step_2_data",
            "rule_type": "metric_threshold",
            "metric_key": support_key,
            "platforms": platforms,
            threshold_key: value,
            "operator": "<=" if support_key == "activity_recency_days" else ">",
        }
        if support_key in {"view_count_mean", "view_count_median"}:
            rule["window"] = {"size": sample_size, "unit": "posts"}
        rules.append(rule)

    visual_audit = structured["visual_audit"]
    runtime_features = [
        {
            "key": feature["key"],
            "label": feature["label"],
            "source_label": feature["source_label"],
            "source_cell": feature["source_cell"],
            "note": feature.get("note"),
        }
        for feature in visual_audit.get("features") or []
        if feature.get("enabled") and feature.get("reuse_status") == "runtime_supported"
    ]
    runtime_platforms = intersect_platforms(requested_platforms, "visual_feature_group")
    if runtime_features and runtime_platforms:
        rules.append(
            {
                "id": "visual_feature_group",
                "step": "step_3_visual",
                "rule_type": "visual_feature_group",
                "platforms": runtime_platforms,
                "cover_count": visual_audit.get("cover_count") or 18,
                "min_hit_features": visual_audit.get("min_hit_features") or 1,
                "features": runtime_features,
            }
        )
    elif any(feature.get("enabled") for feature in visual_audit.get("features") or []):
        warnings.append("视觉正向特征已解析，但当前没有稳定 runtime 平台支持；已保留到 visual_reuse_spec。")

    exclusions = structured["exclusions"]
    if exclusions.get("text_keyword_blocklist"):
        platforms = intersect_platforms(requested_platforms, "text_keyword_blocklist")
        if platforms:
            rules.append(
                {
                    "id": "exclusion_text_keywords",
                    "step": "step_4_exclusions",
                    "rule_type": "text_keyword_blocklist",
                    "platforms": platforms,
                    "operator": "contains_any",
                    "values": exclusions["text_keyword_blocklist"],
                }
            )

    if exclusions.get("relationship_keyword_blocklist"):
        platforms = intersect_platforms(requested_platforms, "relationship_keyword_block")
        if platforms:
            rules.append(
                {
                    "id": "exclusion_relationship_keywords",
                    "step": "step_4_exclusions",
                    "rule_type": "relationship_keyword_block",
                    "platforms": platforms,
                    "operator": "contains_any",
                    "values": exclusions["relationship_keyword_blocklist"],
                }
            )

    if exclusions.get("beauty_ratio_threshold") is not None:
        platforms = intersect_platforms(requested_platforms, "content_keyword_ratio")
        if platforms:
            keyword_values = exclusions.get("text_keyword_blocklist") or ["beauty", "makeup", "skincare"]
            rules.append(
                {
                    "id": "exclusion_beauty_ratio",
                    "step": "step_4_exclusions",
                    "rule_type": "content_keyword_ratio",
                    "platforms": platforms,
                    "operator": "ratio_gt",
                    "threshold_percent": exclusions["beauty_ratio_threshold"],
                    "keyword_values": keyword_values,
                }
            )

    if exclusions.get("multi_dance_ratio_threshold") is not None:
        platforms = intersect_platforms(requested_platforms, "multi_dance_ratio")
        if platforms:
            rules.append(
                {
                    "id": "exclusion_multi_dance_ratio",
                    "step": "step_4_exclusions",
                    "rule_type": "multi_dance_ratio",
                    "platforms": platforms,
                    "operator": "ratio_gt",
                    "threshold_percent": exclusions["multi_dance_ratio_threshold"],
                }
            )

    if exclusions.get("selfie_couple_ratio_threshold") is not None:
        platforms = intersect_platforms(requested_platforms, "selfie_or_couple_ratio")
        if platforms:
            rules.append(
                {
                    "id": "exclusion_selfie_couple_ratio",
                    "step": "step_4_exclusions",
                    "rule_type": "selfie_or_couple_ratio",
                    "platforms": platforms,
                    "operator": "ratio_gt",
                    "threshold_percent": exclusions["selfie_couple_ratio_threshold"],
                }
            )

    if exclusions.get("green_screen_direct_reject"):
        platforms = intersect_platforms(requested_platforms, "green_screen")
        if platforms:
            rules.append(
                {
                    "id": "exclusion_green_screen",
                    "step": "step_4_exclusions",
                    "rule_type": "green_screen",
                    "platforms": platforms,
                    "operator": "must_not_appear",
                }
            )

    compiled_steps = {rule["step"] for rule in rules}
    final_logic = structured["final_logic"]
    for step in final_logic.get("must_pass_steps") or []:
        if step not in compiled_steps:
            warnings.append(f"{step} 被声明为 must_pass，但当前没有稳定自动规则，需人工兜底。")
    for step in final_logic.get("reject_if_any_steps") or []:
        if step not in compiled_steps:
            warnings.append(f"{step} 被声明为 reject_if_any，但当前没有稳定自动规则，需人工兜底。")

    rulespec = {
        "version": VERSION,
        "compile_mode": "workbook_flexible_visual_reuse",
        "source_layout": "flex_sectioned_workbook",
        "goal": structured["basic_info"].get("audit_goal"),
        "basic_info": deepcopy(structured["basic_info"]),
        "platform_scope": requested_platforms,
        "rules": rules,
        "manual_review_items": build_manual_review_items(structured["manual_review"]),
        "compliance_notes": (
            deepcopy(structured["manual_review"].get("compliance_notes") or [])
            + deepcopy(structured["exclusions"].get("compliance_notes") or [])
        ),
        "final_logic": {
            "must_pass_steps": final_logic.get("must_pass_steps") or [],
            "reject_if_any_steps": final_logic.get("reject_if_any_steps") or [],
            "manual_hit_action": final_logic.get("manual_hit_action"),
            "success_output": final_logic.get("success_output"),
            "failure_output": final_logic.get("failure_output"),
        },
    }
    return rulespec, warnings


def build_visual_reuse_spec(structured: dict[str, Any]) -> dict[str, Any]:
    requested_platforms = structured["basic_info"].get("platform_scope") or []
    positive_features = [
        {
            "key": feature["key"],
            "label": feature["label"],
            "source_label": feature["source_label"],
            "source_cell": feature["source_cell"],
            "note": feature.get("note"),
            "reuse_status": feature.get("reuse_status"),
        }
        for feature in structured["visual_audit"].get("features") or []
        if feature.get("enabled")
    ]

    negative_features: list[dict[str, Any]] = []
    exclusions = structured["exclusions"]
    if exclusions.get("green_screen_direct_reject"):
        negative_features.append(
            {
                "key": "green_screen",
                "label": "绿幕",
                "operator": "must_not_appear",
                "reuse_status": "runtime_supported",
            }
        )
    if exclusions.get("multi_dance_ratio_threshold") is not None:
        negative_features.append(
            {
                "key": "multi_dance_ratio",
                "label": "多人跳舞占比",
                "operator": "ratio_gt",
                "threshold_percent": exclusions["multi_dance_ratio_threshold"],
                "reuse_status": "runtime_supported",
            }
        )
    if exclusions.get("selfie_couple_ratio_threshold") is not None:
        negative_features.append(
            {
                "key": "selfie_or_couple_ratio",
                "label": "自拍 / 情侣出镜占比",
                "operator": "ratio_gt",
                "threshold_percent": exclusions["selfie_couple_ratio_threshold"],
                "reuse_status": "runtime_supported",
            }
        )
    negative_features.extend(deepcopy(exclusions.get("custom_visual_risks") or []))

    runtime_positive = [item["label"] for item in positive_features if item.get("reuse_status") == "runtime_supported"]
    template_only_positive = [item["label"] for item in positive_features if item.get("reuse_status") != "runtime_supported"]
    template_only_negative = [item["label"] for item in negative_features if item.get("reuse_status") != "runtime_supported"]

    return {
        "version": VERSION,
        "goal": structured["basic_info"].get("audit_goal"),
        "basic_info": deepcopy(structured["basic_info"]),
        "requested_platforms": requested_platforms,
        "visual_scope": {
            "cover_count": structured["visual_audit"].get("cover_count") or 18,
            "min_hit_features": structured["visual_audit"].get("min_hit_features") or 1,
            "positive_features": positive_features,
            "negative_features": negative_features,
        },
        "manual_review_items": build_manual_review_items(structured["manual_review"]),
        "reuse_prompt_hints": {
            "runtime_supported_positive": runtime_positive,
            "template_only_positive": template_only_positive,
            "template_only_negative": template_only_negative,
        },
        "final_logic": deepcopy(structured["final_logic"]),
        "compliance_notes": (
            deepcopy(structured["manual_review"].get("compliance_notes") or [])
            + deepcopy(structured["exclusions"].get("compliance_notes") or [])
        ),
    }


def build_structured_requirement(workbook_path: Path) -> tuple[dict[str, Any], list[str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    if MAIN_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"工作簿缺少 {MAIN_SHEET_NAME}")
    rows = iter_main_rows(workbook[MAIN_SHEET_NAME])
    sections = split_sections(rows)

    warnings: list[str] = []
    basic_info, current = parse_basic_info(sections["basic_info"])
    warnings.extend(current)
    qualification, current = parse_qualification(sections["qualification"])
    warnings.extend(current)
    data_audit, current = parse_data_audit(sections["data_audit"])
    warnings.extend(current)
    visual_audit, current = parse_visual_audit(sections["visual_audit"])
    warnings.extend(current)
    exclusions, current = parse_exclusions(sections["exclusions"])
    warnings.extend(current)
    manual_review, current = parse_manual_review(sections["manual_review"])
    warnings.extend(current)
    final_logic, current = parse_final_logic(sections["final_logic"])
    warnings.extend(current)

    structured = {
        "version": VERSION,
        "workbook_path": str(workbook_path),
        "parsed_at": utc_now_iso(),
        "basic_info": basic_info,
        "qualification": qualification,
        "data_audit": data_audit,
        "visual_audit": visual_audit,
        "exclusions": exclusions,
        "manual_review": manual_review,
        "final_logic": final_logic,
    }
    return structured, warnings


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def compile_workbook(workbook_path: Path, output_root: Path) -> dict[str, Any]:
    structured, parse_warnings = build_structured_requirement(workbook_path)
    rulespec, rulespec_warnings = build_rulespec(structured)
    visual_reuse_spec = build_visual_reuse_spec(structured)
    visual_prompts = build_visual_prompt_artifacts(visual_reuse_spec)
    warnings = parse_warnings + rulespec_warnings

    workbook_output_dir = output_root / workbook_path.stem
    ensure_dir(workbook_output_dir)

    structured_path = workbook_output_dir / "structured_requirement.json"
    rulespec_path = workbook_output_dir / "rulespec.json"
    visual_path = workbook_output_dir / "visual_reuse_spec.json"
    prompts_path = workbook_output_dir / "visual_prompts.json"
    report_path = workbook_output_dir / "compile_report.json"

    write_json(structured_path, structured)
    write_json(rulespec_path, rulespec)
    write_json(visual_path, visual_reuse_spec)
    write_json(prompts_path, visual_prompts)

    report = {
        "success": True,
        "compiled_at": utc_now_iso(),
        "workbook_path": str(workbook_path),
        "output_dir": str(workbook_output_dir),
        "warnings": warnings,
        "artifacts": {
            "structured_requirement_json": str(structured_path),
            "rulespec_json": str(rulespec_path),
            "visual_reuse_spec_json": str(visual_path),
            "visual_prompts_json": str(prompts_path),
        },
        "stats": {
            "rule_count": len(rulespec.get("rules") or []),
            "manual_review_item_count": len(rulespec.get("manual_review_items") or []),
            "compliance_note_count": len(rulespec.get("compliance_notes") or []),
            "positive_visual_feature_count": len(visual_reuse_spec["visual_scope"]["positive_features"]),
            "negative_visual_feature_count": len(visual_reuse_spec["visual_scope"]["negative_features"]),
            "visual_prompt_platform_count": len(visual_prompts),
        },
    }
    write_json(report_path, report)
    return report


def collect_workbooks(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    return sorted(path for path in input_path.glob("*.xlsx") if not path.name.startswith(".~"))


def main() -> None:
    parser = argparse.ArgumentParser(description="Compile requirement workbooks into rulespec + visual reuse artifacts.")
    parser.add_argument("--input", default=".", help="单个 Excel 文件路径，或包含 Excel 的目录")
    parser.add_argument("--output-dir", help="输出目录，默认写到输入目录下的 compiled_outputs")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    workbooks = collect_workbooks(input_path)
    if not workbooks:
        raise SystemExit("没有找到可编译的 .xlsx 文件。")

    default_output_root = input_path.parent / DEFAULT_OUTPUT_DIRNAME if input_path.is_file() else input_path / DEFAULT_OUTPUT_DIRNAME
    output_root = Path(args.output_dir).resolve() if args.output_dir else default_output_root
    ensure_dir(output_root)

    reports = [compile_workbook(workbook_path, output_root) for workbook_path in workbooks]
    print(json.dumps({"success": True, "compiled": reports}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
