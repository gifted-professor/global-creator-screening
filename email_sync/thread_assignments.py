from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .db import Database
from .relation_index import normalize_subject


_REPLY_PREFIX_PATTERN = re.compile(r"^\s*@?")
_HANDLE_URL_PATTERN = re.compile(r"(?:@|/)([A-Za-z0-9._-]{2,80})/?$")
_ALLOWED_CONFIDENCES = {"high", "medium"}


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_email(value: Any) -> str:
    return _clean_text(value).lower()


def _normalize_owner_scope(value: Any) -> str:
    return _clean_text(value).casefold()


def _normalize_platform(value: Any) -> str:
    text = _clean_text(value).casefold()
    if text in {"ig", "ins", "instagram"}:
        return "instagram"
    if text in {"tt", "tiktok", "douyin"}:
        return "tiktok"
    if text in {"yt", "youtube"}:
        return "youtube"
    return text


def _extract_handle(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    match = _HANDLE_URL_PATTERN.search(text)
    if match:
        return _REPLY_PREFIX_PATTERN.sub("", match.group(1)).strip().casefold()
    return _REPLY_PREFIX_PATTERN.sub("", text).strip().casefold()


def _first_non_blank(*values: Any) -> str:
    for value in values:
        cleaned = _clean_text(value)
        if cleaned:
            return cleaned
    return ""


def _iter_workbook_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_clean_text(item) for item in rows[0]]
        results: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            row: dict[str, Any] = {}
            has_value = False
            for header, cell in zip(headers, raw_row):
                if not header:
                    continue
                row[header] = cell
                if cell not in (None, ""):
                    has_value = True
            if has_value:
                results.append(row)
        return results
    finally:
        workbook.close()


def _extract_confidence(row: dict[str, Any]) -> str:
    for key in (
        "resolution_confidence_final",
        "resolution_confidence",
        "match_confidence",
    ):
        confidence = _clean_text(row.get(key)).casefold()
        if confidence:
            return confidence
    return ""


def _is_persistable_keep_row(row: dict[str, Any]) -> bool:
    confidence = _extract_confidence(row)
    if confidence:
        return confidence in _ALLOWED_CONFIDENCES
    keep_status = _clean_text(row.get("resolution_keep")).casefold()
    if keep_status:
        return keep_status in {"1", "true", "yes", "y", "keep"}
    return True


def _build_assignment_record(
    row: dict[str, Any],
    *,
    brand: str,
    owner_scope: str,
    source_run_id: str,
    default_source_stage: str,
) -> dict[str, Any] | None:
    thread_key = _first_non_blank(row.get("thread_key"), row.get("evidence_thread_key"))
    creator_id = _extract_handle(
        _first_non_blank(
            row.get("final_id_final"),
            row.get("@username"),
            row.get("creator_id"),
            row.get("identifier"),
            row.get("URL"),
        )
    )
    platform = _normalize_platform(_first_non_blank(row.get("Platform"), row.get("platform")))
    if not thread_key or not creator_id or not platform:
        return None
    if not _is_persistable_keep_row(row):
        return None
    now = _utc_now()
    last_mail_message_id = _clean_text(_first_non_blank(row.get("last_mail_message_id"), row.get("evidence_message_id")))
    last_mail_sent_at = _first_non_blank(
        row.get("latest_external_sent_at"),
        row.get("last_mail_time"),
        row.get("evidence_sent_at"),
        row.get("brand_message_sent_at"),
    )
    source_stage = _first_non_blank(
        row.get("resolution_stage_final"),
        row.get("resolution_stage"),
        row.get("resolution_method"),
        default_source_stage,
    )
    return {
        "thread_key": thread_key,
        "owner_scope": _normalize_owner_scope(owner_scope),
        "creator_id": creator_id,
        "platform": platform,
        "brand": _clean_text(brand),
        "matched_contact_email": _normalize_email(_first_non_blank(row.get("matched_contact_email"), row.get("matched_email"))),
        "normalized_subject": normalize_subject(_first_non_blank(row.get("subject"), row.get("evidence_subject"), row.get("last_mail_subject"))),
        "source_stage": _clean_text(source_stage),
        "source_run_id": _clean_text(source_run_id),
        "last_mail_message_id": last_mail_message_id,
        "last_mail_sent_at": _clean_text(last_mail_sent_at),
        "created_at": now,
        "updated_at": now,
    }


def persist_thread_assignments_from_keep_workbook(
    *,
    db: Database,
    keep_workbook_path: str | Path,
    brand: str,
    owner_scope: str,
    source_run_id: str = "",
    default_source_stage: str = "final_keep",
) -> dict[str, int | str]:
    path = Path(str(keep_workbook_path)).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"找不到 keep workbook: {path}")

    db.init_schema()
    rows = _iter_workbook_rows(path)
    upserted = 0
    skipped = 0

    for row in rows:
        record = _build_assignment_record(
            row,
            brand=brand,
            owner_scope=owner_scope,
            source_run_id=source_run_id,
            default_source_stage=default_source_stage,
        )
        if record is None:
            skipped += 1
            continue
        existing = db.conn.execute(
            """
            SELECT last_mail_message_id, last_mail_sent_at, mail_update_revision
            FROM thread_assignments
            WHERE thread_key = ? AND owner_scope = ?
            """,
            (record["thread_key"], record["owner_scope"]),
        ).fetchone()
        revision = int(existing["mail_update_revision"]) if existing is not None else 0
        existing_message_id = _clean_text(existing["last_mail_message_id"]) if existing is not None else ""
        existing_sent_at = _clean_text(existing["last_mail_sent_at"]) if existing is not None else ""
        if record["last_mail_message_id"]:
            if record["last_mail_message_id"] != existing_message_id:
                revision += 1
        elif record["last_mail_sent_at"] and record["last_mail_sent_at"] > existing_sent_at:
            revision += 1
        record["mail_update_revision"] = revision
        db.conn.execute(
            """
            INSERT INTO thread_assignments (
                thread_key,
                owner_scope,
                creator_id,
                platform,
                brand,
                matched_contact_email,
                normalized_subject,
                source_stage,
                source_run_id,
                last_mail_message_id,
                last_mail_sent_at,
                mail_update_revision,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(thread_key, owner_scope)
            DO UPDATE SET
                creator_id = excluded.creator_id,
                platform = excluded.platform,
                brand = excluded.brand,
                matched_contact_email = excluded.matched_contact_email,
                normalized_subject = CASE
                    WHEN excluded.normalized_subject != '' THEN excluded.normalized_subject
                    ELSE thread_assignments.normalized_subject
                END,
                source_stage = excluded.source_stage,
                source_run_id = excluded.source_run_id,
                last_mail_message_id = CASE
                    WHEN excluded.last_mail_message_id != '' THEN excluded.last_mail_message_id
                    ELSE thread_assignments.last_mail_message_id
                END,
                last_mail_sent_at = CASE
                    WHEN excluded.last_mail_sent_at != '' THEN excluded.last_mail_sent_at
                    ELSE thread_assignments.last_mail_sent_at
                END,
                mail_update_revision = excluded.mail_update_revision,
                updated_at = excluded.updated_at
            """,
            (
                record["thread_key"],
                record["owner_scope"],
                record["creator_id"],
                record["platform"],
                record["brand"],
                record["matched_contact_email"],
                record["normalized_subject"],
                record["source_stage"],
                record["source_run_id"],
                record["last_mail_message_id"],
                record["last_mail_sent_at"],
                record["mail_update_revision"],
                record["created_at"],
                record["updated_at"],
            ),
        )
        upserted += 1

    db.conn.commit()
    return {
        "keep_workbook_path": str(path),
        "source_row_count": len(rows),
        "upserted_count": upserted,
        "skipped_count": skipped,
    }


def lookup_thread_assignment(
    *,
    db_path: str | Path,
    owner_scope: str,
    creator_id: str,
    platform: str,
    brand: str = "",
    matched_contact_email: str = "",
    subject: str = "",
) -> dict[str, Any] | None:
    path = Path(str(db_path)).expanduser().resolve()
    if not path.exists():
        return None
    normalized_owner_scope = _normalize_owner_scope(owner_scope)
    normalized_creator_id = _extract_handle(creator_id)
    normalized_platform = _normalize_platform(platform)
    if not normalized_owner_scope or not normalized_creator_id or not normalized_platform:
        return None

    conn = sqlite3.connect(str(path))
    conn.row_factory = sqlite3.Row
    try:
        query = """
            SELECT *
            FROM thread_assignments
            WHERE owner_scope = ? AND creator_id = ? AND platform = ?
        """
        params: list[Any] = [normalized_owner_scope, normalized_creator_id, normalized_platform]
        normalized_brand = _clean_text(brand)
        if normalized_brand:
            query += " AND brand = ?"
            params.append(normalized_brand)
        query += " ORDER BY datetime(updated_at) DESC, thread_key ASC"
        rows = list(conn.execute(query, params).fetchall())
    except sqlite3.OperationalError:
        return None
    finally:
        conn.close()

    if not rows:
        return None

    normalized_contact = _normalize_email(matched_contact_email)
    normalized_subject = normalize_subject(subject)
    best_match: dict[str, Any] | None = None

    for row in rows:
        row_contact = _normalize_email(row["matched_contact_email"])
        row_subject = _clean_text(row["normalized_subject"])
        contact_match = bool(normalized_contact and row_contact and normalized_contact == row_contact)
        subject_match = bool(normalized_subject and row_subject and normalized_subject == row_subject)
        if normalized_contact and row_contact and not contact_match:
            continue
        if normalized_subject and row_subject and not subject_match:
            continue
        if contact_match or subject_match:
            best_match = dict(row)
            best_match["match_reason"] = "contact+subject" if contact_match and subject_match else ("contact" if contact_match else "subject")
            break
    return best_match
