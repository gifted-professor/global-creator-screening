from __future__ import annotations

import csv
import json
import os
import re
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

import requests
from openpyxl import Workbook, load_workbook

from .db import Database
from .relation_index import rebuild_relation_index


REVIEW_CANDIDATE_FIELDS = (
    "candidate_id",
    "sheet_name",
    "source_row_number",
    "nickname",
    "@username",
    "Platform",
    "Email",
    "matched_contact_email",
    "URL",
    "match_confidence",
    "last_mail_message_id",
    "last_mail_subject",
    "last_mail_snippet",
    "last_mail_raw_path",
    "evidence_thread_key",
)

REVIEW_RESULT_HEADERS = [
    "review_candidate_id",
    "review_group_key",
    "review_decision",
    "review_selected",
    "review_selected_candidate_ids",
    "review_reason",
    "review_confidence",
    "review_model",
    "review_reviewed_at",
]

RETRYABLE_STATUS_CODES = {408, 429, 500, 502, 503, 504}


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


@dataclass(frozen=True)
class ReviewGroupEvidence:
    message_row_id: int
    thread_key: str
    subject: str
    sent_at: str
    snippet: str
    body_text: str
    raw_path: str
    direction: str
    thread_message_count: int
    thread_last_sent_at: str
    thread_messages: tuple[dict[str, Any], ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "message_row_id": self.message_row_id,
            "thread_key": self.thread_key,
            "subject": self.subject,
            "sent_at": self.sent_at,
            "snippet": self.snippet,
            "body_text": self.body_text,
            "raw_path": self.raw_path,
            "direction": self.direction,
            "thread_message_count": self.thread_message_count,
            "thread_last_sent_at": self.thread_last_sent_at,
            "thread_messages": list(self.thread_messages),
        }


@dataclass(frozen=True)
class ReviewGroup:
    group_key: str
    key_source: str
    key_value: str
    rows: tuple[dict[str, Any], ...]
    evidence: ReviewGroupEvidence

    @property
    def candidate_count(self) -> int:
        return len(self.rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "group_key": self.group_key,
            "key_source": self.key_source,
            "key_value": self.key_value,
            "candidate_count": self.candidate_count,
            "rows": list(self.rows),
            "evidence": self.evidence.to_dict(),
        }


@dataclass(frozen=True)
class DuplicateReviewConfig:
    base_url: str
    api_key: str
    model: str
    timeout_seconds: int


def _ensure_relation_index(db: Database) -> None:
    db.init_schema()
    message_count = int(db.conn.execute("SELECT COUNT(*) FROM messages").fetchone()[0])
    indexed_count = int(db.conn.execute("SELECT COUNT(*) FROM message_index").fetchone()[0])
    if message_count == 0:
        raise RuntimeError("本地邮件库为空，先运行 python3 -m email_sync sync")
    if indexed_count != message_count:
        rebuild_relation_index(db)


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _source_headers(input_path: Path) -> list[str]:
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        ordered_headers: list[str] = []
        seen: set[str] = set()
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(iterator, ())
            for cell in header_row:
                header = _stringify(cell)
                if not header or header in seen:
                    continue
                seen.add(header)
                ordered_headers.append(header)
        return ordered_headers
    finally:
        workbook.close()


def _iter_rows(input_path: Path, source_headers: Sequence[str]) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [_stringify(cell) for cell in header_row]
            header_positions = {header: index for index, header in enumerate(headers) if header}
            for row_number, values in enumerate(rows, start=2):
                row: dict[str, Any] = {
                    "sheet_name": sheet.title,
                    "source_row_number": row_number,
                }
                has_value = False
                for header in source_headers:
                    index = header_positions.get(header)
                    value = values[index] if index is not None and index < len(values) else ""
                    row[header] = value
                    if not has_value and _stringify(value):
                        has_value = True
                if has_value:
                    yield row
    finally:
        workbook.close()


def _build_group_key(row: dict[str, Any]) -> tuple[str, str, str]:
    message_id = _stringify(row.get("last_mail_message_id"))
    if message_id:
        return (f"last_mail_message_id:{message_id}", "last_mail_message_id", message_id)
    raw_path = _stringify(row.get("last_mail_raw_path"))
    if raw_path:
        return (f"last_mail_raw_path:{raw_path}", "last_mail_raw_path", raw_path)
    raise RuntimeError(
        f"source row {row.get('source_row_number')} 缺少 last_mail_message_id 和 last_mail_raw_path，无法建立 duplicate review group。"
    )


def _candidate_payload(row: dict[str, Any]) -> dict[str, Any]:
    candidate_id = f"{_stringify(row.get('sheet_name'))}:{int(row.get('source_row_number') or 0)}"
    payload = {field: _stringify(row.get(field)) for field in REVIEW_CANDIDATE_FIELDS}
    payload["candidate_id"] = candidate_id
    payload["source_row_number"] = int(row.get("source_row_number") or 0)
    return payload


def _fetch_message_evidence(db: Database, key_source: str, key_value: str) -> ReviewGroupEvidence:
    if key_source == "last_mail_message_id":
        row = db.conn.execute(
            """
            SELECT
                m.id,
                COALESCE(mi.thread_key, '') AS thread_key,
                COALESCE(mi.sent_sort_at, m.sent_at, m.internal_date, m.created_at, '') AS sent_sort_at,
                COALESCE(m.subject, '') AS subject,
                COALESCE(m.snippet, '') AS snippet,
                COALESCE(m.body_text, '') AS body_text,
                COALESCE(m.raw_path, '') AS raw_path,
                COALESCE(mi.direction, '') AS direction
            FROM messages m
            LEFT JOIN message_index mi ON mi.message_row_id = m.id
            WHERE m.id = ?
            """,
            (int(key_value),),
        ).fetchone()
    else:
        row = db.conn.execute(
            """
            SELECT
                m.id,
                COALESCE(mi.thread_key, '') AS thread_key,
                COALESCE(mi.sent_sort_at, m.sent_at, m.internal_date, m.created_at, '') AS sent_sort_at,
                COALESCE(m.subject, '') AS subject,
                COALESCE(m.snippet, '') AS snippet,
                COALESCE(m.body_text, '') AS body_text,
                COALESCE(m.raw_path, '') AS raw_path,
                COALESCE(mi.direction, '') AS direction
            FROM messages m
            LEFT JOIN message_index mi ON mi.message_row_id = m.id
            WHERE m.raw_path = ?
            """,
            (key_value,),
        ).fetchone()

    if row is None:
        raise RuntimeError(f"找不到 duplicate review evidence: {key_source}={key_value}")

    thread_key = _stringify(row["thread_key"])
    thread_messages: list[dict[str, Any]] = []
    thread_message_count = 1
    thread_last_sent_at = _stringify(row["sent_sort_at"])
    if thread_key:
        thread_rows = db.conn.execute(
            """
            SELECT
                m.id,
                COALESCE(mi.sent_sort_at, m.sent_at, m.internal_date, m.created_at, '') AS sent_sort_at,
                COALESCE(m.subject, '') AS subject,
                COALESCE(m.snippet, '') AS snippet,
                COALESCE(m.body_text, '') AS body_text,
                COALESCE(m.raw_path, '') AS raw_path,
                COALESCE(mi.direction, '') AS direction
            FROM message_index mi
            JOIN messages m ON m.id = mi.message_row_id
            WHERE mi.thread_key = ?
            ORDER BY mi.sent_sort_at, m.id
            """,
            (thread_key,),
        ).fetchall()
        thread_messages = [
            {
                "message_row_id": int(thread_row["id"]),
                "sent_at": _stringify(thread_row["sent_sort_at"]),
                "subject": _stringify(thread_row["subject"]),
                "snippet": _stringify(thread_row["snippet"]),
                "body_text": _stringify(thread_row["body_text"]),
                "raw_path": _stringify(thread_row["raw_path"]),
                "direction": _stringify(thread_row["direction"]),
            }
            for thread_row in thread_rows
        ]
        if thread_messages:
            thread_message_count = len(thread_messages)
            thread_last_sent_at = thread_messages[-1]["sent_at"]

    return ReviewGroupEvidence(
        message_row_id=int(row["id"]),
        thread_key=thread_key,
        subject=_stringify(row["subject"]),
        sent_at=_stringify(row["sent_sort_at"]),
        snippet=_stringify(row["snippet"]),
        body_text=_stringify(row["body_text"]),
        raw_path=_stringify(row["raw_path"]),
        direction=_stringify(row["direction"]),
        thread_message_count=thread_message_count,
        thread_last_sent_at=thread_last_sent_at,
        thread_messages=tuple(thread_messages),
    )


def _build_duplicate_groups(db: Database, input_path: Path) -> tuple[list[ReviewGroup], dict[str, Any]]:
    source_headers = _source_headers(input_path)
    grouped_rows: dict[str, dict[str, Any]] = {}
    total_rows = 0

    for row in _iter_rows(input_path, source_headers):
        total_rows += 1
        group_key, key_source, key_value = _build_group_key(row)
        entry = grouped_rows.setdefault(
            group_key,
            {"key_source": key_source, "key_value": key_value, "rows": []},
        )
        entry["rows"].append(_candidate_payload(row))

    counts = Counter(len(entry["rows"]) for entry in grouped_rows.values())
    duplicate_groups: list[ReviewGroup] = []
    duplicate_rows = 0
    singleton_groups = 0

    for group_key, entry in grouped_rows.items():
        rows = tuple(entry["rows"])
        if len(rows) == 1:
            singleton_groups += 1
            continue
        evidence = _fetch_message_evidence(db, str(entry["key_source"]), str(entry["key_value"]))
        duplicate_groups.append(
            ReviewGroup(
                group_key=group_key,
                key_source=str(entry["key_source"]),
                key_value=str(entry["key_value"]),
                rows=rows,
                evidence=evidence,
            )
        )
        duplicate_rows += len(rows)

    duplicate_groups.sort(key=lambda item: (-item.candidate_count, item.group_key))

    stats = {
        "total_rows": total_rows,
        "group_count": len(grouped_rows),
        "singleton_group_count": singleton_groups,
        "duplicate_group_count": len(duplicate_groups),
        "duplicate_row_count": duplicate_rows,
        "group_size_histogram": {str(size): count for size, count in sorted(counts.items())},
    }
    return duplicate_groups, stats


def prepare_duplicate_review(
    db: Database,
    input_path: Path,
    output_prefix: Path,
    sample_limit: int = 3,
    group_keys: Optional[Sequence[str]] = None,
) -> dict[str, Any]:
    if not input_path.exists():
        raise FileNotFoundError(f"找不到 enrichment workbook: {input_path}")
    if sample_limit <= 0:
        raise ValueError("--sample-limit 必须是大于 0 的整数。")

    _ensure_relation_index(db)
    duplicate_groups, stats = _build_duplicate_groups(db, input_path)
    selected_groups = duplicate_groups
    selection_mode = "top_duplicate_groups"
    if group_keys:
        normalized = {str(key).strip() for key in group_keys if str(key).strip()}
        if not normalized:
            raise ValueError("传入了空的 group key。")
        selected_groups = [group for group in duplicate_groups if group.group_key in normalized]
        missing = sorted(normalized - {group.group_key for group in selected_groups})
        if missing:
            raise RuntimeError(f"找不到指定 group key: {', '.join(missing)}")
        selection_mode = "explicit_group_keys"
    else:
        selected_groups = duplicate_groups[:sample_limit]

    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    groups_path = output_prefix.with_name(f"{output_prefix.name}_groups.json")
    summary_path = output_prefix.with_name(f"{output_prefix.name}_summary.json")

    groups_payload = {
        "input_path": str(input_path),
        "duplicate_group_count": stats["duplicate_group_count"],
        "selected_group_count": len(selected_groups),
        "groups": [group.to_dict() for group in selected_groups],
    }
    groups_path.write_text(json.dumps(groups_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary = {
        "input_path": str(input_path),
        "db_path": str(db.db_path),
        "selection_mode": selection_mode,
        "sample_limit": sample_limit,
        "selected_group_count": len(selected_groups),
        "selected_group_keys": [group.group_key for group in selected_groups],
        "stats": stats,
        "groups_json_path": str(groups_path),
        "summary_json_path": str(summary_path),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary


def _load_env_file(env_path: str) -> dict[str, str]:
    path = Path(env_path).expanduser()
    candidates = [path]
    if path.name == ".env":
        local_path = path.with_name(".env.local")
        if local_path.exists():
            candidates.append(local_path)

    values: dict[str, str] = {}
    for candidate in candidates:
        if not candidate.exists():
            continue
        for raw_line in candidate.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def _resolve_env_value(key: str, env_values: dict[str, str], default: str = "") -> str:
    value = os.environ.get(key)
    if value is not None and str(value).strip():
        return str(value).strip()
    return str(env_values.get(key, default) or "").strip()


def _resolve_review_config(
    env_path: str,
    *,
    base_url: Optional[str] = None,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
) -> DuplicateReviewConfig:
    env_values = _load_env_file(env_path)
    resolved_base_url = (
        str(base_url or "").strip()
        or _resolve_env_value("OPENAI_BASE_URL", env_values)
        or _resolve_env_value("LLM_API_BASE", env_values)
        or "https://api.openai.com/v1"
    )
    resolved_api_key = (
        str(api_key or "").strip()
        or _resolve_env_value("OPENAI_API_KEY", env_values)
        or _resolve_env_value("LLM_API_KEY", env_values)
    )
    resolved_model = (
        str(model or "").strip()
        or _resolve_env_value("OPENAI_VISION_MODEL", env_values)
        or _resolve_env_value("LLM_MODEL", env_values)
        or _resolve_env_value("VISION_MODEL", env_values)
        or "gpt-5.4"
    )
    timeout_text = (
        _resolve_env_value("LLM_TIMEOUT_SECONDS", env_values)
        or _resolve_env_value("VISION_REQUEST_TIMEOUT", env_values)
        or "60"
    )
    try:
        timeout_seconds = max(5, int(timeout_text))
    except Exception:
        timeout_seconds = 60

    if not resolved_base_url:
        raise RuntimeError("缺少 LLM base url，请设置 OPENAI_BASE_URL 或 LLM_API_BASE。")
    if not resolved_api_key:
        raise RuntimeError("缺少 LLM API key，请设置 OPENAI_API_KEY 或 LLM_API_KEY。")

    return DuplicateReviewConfig(
        base_url=resolved_base_url.rstrip("/"),
        api_key=resolved_api_key,
        model=resolved_model,
        timeout_seconds=timeout_seconds,
    )


def _truncate_text(value: str, limit: int) -> str:
    text = _stringify(value)
    if len(text) <= limit:
        return text
    return f"{text[: max(0, limit - 3)]}..."


def _build_group_prompt_payload(group: dict[str, Any]) -> dict[str, Any]:
    evidence = dict(group.get("evidence") or {})
    thread_messages = list(evidence.get("thread_messages") or [])
    compact_messages = [
        {
            "message_row_id": item.get("message_row_id"),
            "sent_at": item.get("sent_at"),
            "subject": _truncate_text(_stringify(item.get("subject")), 200),
            "snippet": _truncate_text(_stringify(item.get("snippet")), 500),
            "body_text": _truncate_text(_stringify(item.get("body_text")), 1200),
            "direction": item.get("direction"),
        }
        for item in thread_messages[-4:]
    ]

    return {
        "group_key": group.get("group_key"),
        "key_source": group.get("key_source"),
        "key_value": group.get("key_value"),
        "candidate_count": group.get("candidate_count"),
        "decision_options": ["match_one", "match_some", "reject_group", "uncertain"],
        "instructions": [
            "Judge which creator candidate(s) this mail thread actually refers to.",
            "Use the mail content first, then use candidate identity fields to disambiguate.",
            "If exactly one candidate clearly matches, use match_one.",
            "If several candidates clearly belong to the same mail, use match_some.",
            "If none fit, use reject_group.",
            "If the evidence is too weak, use uncertain.",
            "Return JSON only.",
        ],
        "evidence": {
            "message_row_id": evidence.get("message_row_id"),
            "thread_key": evidence.get("thread_key"),
            "subject": _truncate_text(_stringify(evidence.get("subject")), 240),
            "sent_at": evidence.get("sent_at"),
            "snippet": _truncate_text(_stringify(evidence.get("snippet")), 700),
            "body_text": _truncate_text(_stringify(evidence.get("body_text")), 2000),
            "thread_message_count": evidence.get("thread_message_count"),
            "thread_last_sent_at": evidence.get("thread_last_sent_at"),
            "thread_messages_tail": compact_messages,
        },
        "candidates": [
            {
                "candidate_id": row.get("candidate_id"),
                "sheet_name": row.get("sheet_name"),
                "source_row_number": row.get("source_row_number"),
                "nickname": row.get("nickname"),
                "username": row.get("@username"),
                "platform": row.get("Platform"),
                "email": row.get("Email"),
                "matched_contact_email": row.get("matched_contact_email"),
                "url": row.get("URL"),
            }
            for row in list(group.get("rows") or [])
        ],
        "output_schema": {
            "decision": "match_one | match_some | reject_group | uncertain",
            "selected_candidate_ids": ["candidate ids from the input candidate list"],
            "reason": "brief justification",
            "confidence": "high | medium | low",
        },
    }


def _extract_response_text(payload: Any) -> str:
    if isinstance(payload, str):
        return payload.strip()
    if not isinstance(payload, dict):
        return str(payload)

    output_text = payload.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    choices = payload.get("choices")
    if isinstance(choices, list):
        for choice in choices:
            message = (choice or {}).get("message") or {}
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                return content.strip()
            if isinstance(content, list):
                parts = []
                for item in content:
                    text_value = (item or {}).get("text")
                    if isinstance(text_value, str) and text_value.strip():
                        parts.append(text_value.strip())
                if parts:
                    return "\n".join(parts)

    output = payload.get("output")
    if isinstance(output, list):
        parts = []
        for item in output:
            for content_item in (item or {}).get("content") or []:
                text_value = content_item.get("text")
                if isinstance(text_value, str) and text_value.strip():
                    parts.append(text_value.strip())
        if parts:
            return "\n".join(parts)

    return json.dumps(payload, ensure_ascii=False)


def _parse_review_response(raw_text: str, valid_candidate_ids: set[str]) -> dict[str, Any]:
    cleaned = str(raw_text or "").strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```[a-zA-Z0-9_]*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```$", "", cleaned).strip()

    payload: dict[str, Any] | None = None
    try:
        maybe_payload = json.loads(cleaned)
        if isinstance(maybe_payload, dict):
            payload = maybe_payload
    except json.JSONDecodeError:
        start_idx = cleaned.find("{")
        end_idx = cleaned.rfind("}")
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            snippet = cleaned[start_idx : end_idx + 1]
            try:
                maybe_payload = json.loads(snippet)
                if isinstance(maybe_payload, dict):
                    payload = maybe_payload
            except json.JSONDecodeError:
                payload = None

    if payload is None:
        return {
            "decision": "uncertain",
            "selected_candidate_ids": [],
            "reason": cleaned or "模型未返回可解析 JSON",
            "confidence": "low",
            "raw_text": raw_text,
        }

    decision = _stringify(payload.get("decision")).lower()
    if decision not in {"match_one", "match_some", "reject_group", "uncertain"}:
        decision = "uncertain"

    selected_candidate_ids = payload.get("selected_candidate_ids")
    if not isinstance(selected_candidate_ids, list):
        selected_candidate_ids = []
    normalized_ids = []
    seen = set()
    for item in selected_candidate_ids:
        candidate_id = _stringify(item)
        if not candidate_id or candidate_id not in valid_candidate_ids or candidate_id in seen:
            continue
        seen.add(candidate_id)
        normalized_ids.append(candidate_id)

    if decision == "match_one":
        if len(normalized_ids) != 1:
            decision = "uncertain"
    elif decision == "match_some":
        if len(normalized_ids) < 2:
            decision = "uncertain"
    if decision in {"reject_group", "uncertain"}:
        normalized_ids = []

    confidence = _stringify(payload.get("confidence")).lower()
    if confidence not in {"high", "medium", "low"}:
        confidence = "medium" if decision in {"match_one", "match_some"} else "low"

    return {
        "decision": decision,
        "selected_candidate_ids": normalized_ids,
        "reason": _stringify(payload.get("reason")) or cleaned or "模型未返回原因",
        "confidence": confidence,
        "raw_text": raw_text,
    }


def _build_chat_messages(group: dict[str, Any]) -> list[dict[str, str]]:
    payload = _build_group_prompt_payload(group)
    return [
        {
            "role": "system",
            "content": (
                "You adjudicate duplicate creator-to-email matches. "
                "Return JSON only. Use one of: match_one, match_some, reject_group, uncertain."
            ),
        },
        {
            "role": "user",
            "content": json.dumps(payload, ensure_ascii=False, indent=2),
        },
    ]


def _invoke_duplicate_review_llm(config: DuplicateReviewConfig, group: dict[str, Any]) -> dict[str, Any]:
    url = f"{config.base_url}/chat/completions"
    body = {
        "model": config.model,
        "messages": _build_chat_messages(group),
        "temperature": 0,
    }
    headers = {
        "Authorization": f"Bearer {config.api_key}",
        "Content-Type": "application/json",
    }

    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=config.timeout_seconds)
            if response.status_code >= 400:
                if response.status_code in RETRYABLE_STATUS_CODES and attempt < 3:
                    time.sleep(min(4.0, 1.5 * attempt))
                    continue
                raise RuntimeError(f"LLM HTTP {response.status_code}: {response.text[:400]}")
            payload = response.json()
            raw_text = _extract_response_text(payload)
            valid_candidate_ids = {row.get("candidate_id") for row in list(group.get("rows") or [])}
            parsed = _parse_review_response(raw_text, {item for item in valid_candidate_ids if item})
            parsed["provider_payload"] = payload
            return parsed
        except requests.exceptions.RequestException as exc:
            last_error = exc
            if attempt < 3:
                time.sleep(min(4.0, 1.5 * attempt))
                continue
            break
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            break

    raise RuntimeError(str(last_error) if last_error else "duplicate review LLM call failed")


def _write_annotated_exports(
    input_path: Path,
    output_prefix: Path,
    annotations_by_candidate_id: dict[str, dict[str, Any]],
) -> tuple[str, str]:
    source_headers = _source_headers(input_path)
    output_headers = list(source_headers) + REVIEW_RESULT_HEADERS
    csv_path = output_prefix.with_name(f"{output_prefix.name}_annotated.csv")
    xlsx_path = output_prefix.with_name(f"{output_prefix.name}_annotated.xlsx")

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="results")
    worksheet.append(output_headers)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=output_headers)
        writer.writeheader()
        for row in _iter_rows(input_path, source_headers):
            candidate_id = f"{_stringify(row.get('sheet_name'))}:{int(row.get('source_row_number') or 0)}"
            annotation = annotations_by_candidate_id.get(candidate_id, {})
            output_row = {header: row.get(header, "") for header in source_headers}
            output_row.update({header: annotation.get(header, "") for header in REVIEW_RESULT_HEADERS})
            writer.writerow(output_row)
            worksheet.append([output_row.get(header, "") for header in output_headers])

    workbook.save(xlsx_path)
    return (str(csv_path), str(xlsx_path))


def review_duplicate_groups(
    db: Database,
    input_path: Path,
    output_prefix: Path,
    *,
    env_path: str = ".env",
    sample_limit: int = 3,
    group_keys: Optional[Sequence[str]] = None,
    base_url: Optional[str] = None,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
) -> dict[str, Any]:
    prepare_summary = prepare_duplicate_review(
        db=db,
        input_path=input_path,
        output_prefix=output_prefix,
        sample_limit=sample_limit,
        group_keys=group_keys,
    )
    groups_payload = json.loads(Path(prepare_summary["groups_json_path"]).read_text(encoding="utf-8"))
    config = _resolve_review_config(env_path, base_url=base_url, api_key=api_key, model=model)

    group_results = []
    annotations_by_candidate_id: dict[str, dict[str, Any]] = {}

    for group in list(groups_payload.get("groups") or []):
        result = _invoke_duplicate_review_llm(config, group)
        reviewed_at = _utc_now()
        selected_ids = list(result.get("selected_candidate_ids") or [])
        selected_set = set(selected_ids)
        group_results.append(
            {
                "group_key": group.get("group_key"),
                "key_source": group.get("key_source"),
                "key_value": group.get("key_value"),
                "candidate_count": group.get("candidate_count"),
                "decision": result.get("decision"),
                "selected_candidate_ids": selected_ids,
                "reason": result.get("reason"),
                "confidence": result.get("confidence"),
                "model": config.model,
                "reviewed_at": reviewed_at,
                "raw_text": result.get("raw_text"),
                "group": group,
            }
        )
        for row in list(group.get("rows") or []):
            candidate_id = _stringify(row.get("candidate_id"))
            annotations_by_candidate_id[candidate_id] = {
                "review_candidate_id": candidate_id,
                "review_group_key": _stringify(group.get("group_key")),
                "review_decision": _stringify(result.get("decision")),
                "review_selected": "yes" if candidate_id in selected_set else "no",
                "review_selected_candidate_ids": " | ".join(selected_ids),
                "review_reason": _stringify(result.get("reason")),
                "review_confidence": _stringify(result.get("confidence")),
                "review_model": config.model,
                "review_reviewed_at": reviewed_at,
            }

    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    audit_path = output_prefix.with_name(f"{output_prefix.name}_audit.json")
    summary_path = output_prefix.with_name(f"{output_prefix.name}_review_summary.json")
    audit_path.write_text(json.dumps({"groups": group_results}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    annotated_csv_path, annotated_xlsx_path = _write_annotated_exports(input_path, output_prefix, annotations_by_candidate_id)
    summary = {
        "input_path": str(input_path),
        "db_path": str(db.db_path),
        "env_path": env_path,
        "selection_mode": prepare_summary["selection_mode"],
        "sample_limit": sample_limit,
        "selected_group_count": prepare_summary["selected_group_count"],
        "selected_group_keys": prepare_summary["selected_group_keys"],
        "model": config.model,
        "base_url": config.base_url,
        "stats": prepare_summary["stats"],
        "audit_json_path": str(audit_path),
        "annotated_csv_path": annotated_csv_path,
        "annotated_xlsx_path": annotated_xlsx_path,
        "prepare_groups_json_path": prepare_summary["groups_json_path"],
        "prepare_summary_json_path": prepare_summary["summary_json_path"],
        "review_summary_json_path": str(summary_path),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary
