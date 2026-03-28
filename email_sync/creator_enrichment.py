from __future__ import annotations

import csv
import html
import json
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence

from .db import Database
from .relation_index import rebuild_relation_index


EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
MENTION_HANDLE_PATTERN = re.compile(r"(?<![\w@])@([a-z0-9._-]{3,50})", re.IGNORECASE)
YOUTUBE_HANDLE_PATTERN = re.compile(r"youtube\.com/@([a-z0-9._-]{3,50})", re.IGNORECASE)
TIKTOK_HANDLE_PATTERN = re.compile(r"tiktok\.com/@([a-z0-9._-]{3,50})", re.IGNORECASE)
INSTAGRAM_HANDLE_PATTERN = re.compile(r"instagram\.com/([a-z0-9._-]{3,50})", re.IGNORECASE)
URL_PATTERN = re.compile(r"https?://\S+", re.IGNORECASE)
AMOUNT_PATTERNS = (
    re.compile(r"(?<![\w])\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)"),
    re.compile(r"(?<![\w])usd\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)", re.IGNORECASE),
    re.compile(r"(?<![\w])([0-9]{1,3}(?:\.\d+)?)\s*k\b", re.IGNORECASE),
)
COLLAB_KEYWORDS = (
    "collab",
    "collaboration",
    "campaign",
    "partnership",
    "partner",
    "offer",
    "paid",
    "sponsor",
    "sponsored",
    "合作",
    "报价",
    "payment",
)
QUOTE_POSITIVE_TERMS = (
    "my rate",
    "quoted rate",
    "my quote",
    "my fee",
    "flat fee",
    "i charge",
    "our rate",
    "our budget",
    "rate for",
    "rate would be",
    "expected cost",
    "looking for",
    "my standard rate",
    "budget for this collaboration",
    "i don't work for",
    "i dont work for",
    "not quite at your quoted rate",
)
QUOTE_NEGATIVE_TERMS = (
    "valued at",
    "worth",
    "worth around",
    "retail",
    "discount",
    "coupon",
    "save ",
    "% off",
    "zip code",
    "phone number",
    "tracking number",
    "shopping cart",
    "sample",
    "free sample",
    "gift",
    "gmv",
    "commission",
    "4k quality",
    "8k",
    "waterproof",
    "thunderbolt",
    "resolution",
)
K_AMOUNT_FINANCE_TERMS = (
    "rate",
    "fee",
    "budget",
    "offer",
    "quoted",
    "quote",
    "cost",
    "price",
    "pay",
    "payment",
    "counteroffer",
)
K_AMOUNT_TECH_TERMS = (
    "4k",
    "8k",
    "quality",
    "resolution",
    "waterproof",
    "thunderbolt",
    "output",
    "display",
    "fps",
)
OUTBOUND_BUDGET_TERMS = (
    "we're offering",
    "we are offering",
    "we can offer",
    "our current budget",
    "our total budget",
    "we’re able to offer",
    "we are able to offer",
    "we can adjust",
    "we'd like to offer",
    "we would like to offer",
)
APPENDED_HEADERS = [
    "sheet_name",
    "source_row_number",
    "derived_handle",
    "creator_emails",
    "matched_contact_email",
    "matched_contact_name",
    "match_rule",
    "match_confidence",
    "evidence_message_id",
    "evidence_subject",
    "evidence_sent_at",
    "evidence_thread_key",
    "evidence_raw_path",
    "last_mail_message_id",
    "last_mail_time",
    "last_mail_subject",
    "last_mail_snippet",
    "last_mail_raw_path",
    "thread_message_count",
    "thread_last_sent_at",
    "latest_quote_amount",
    "latest_quote_currency",
    "latest_quote_source",
    "latest_quote_message_id",
    "latest_quote_time",
    "latest_quote_text",
]
CANONICAL_SOURCE_HEADERS = [
    "Platform",
    "@username",
    "URL",
    "nickname",
    "Region",
    "Email",
]
SENDING_LIST_COUNTRY_ALIASES = ("country", "国家", "region", "地区")
SENDING_LIST_CREATOR_ALIASES = ("creator", "nickname", "达人", "红人", "博主")
SENDING_LIST_EMAIL_ALIASES = ("邮箱地址", "邮箱", "email", "emailaddress", "mail")
SENDING_LIST_GENERIC_LINK_ALIASES = ("link", "url", "主页链接", "账号链接", "profilelink", "profileurl")
SENDING_LIST_PLATFORM_LINK_ALIASES = {
    "instagram": ("iglink", "igurl", "instagramlink", "instagramurl", "inslink", "insurl"),
    "tiktok": ("ttlink", "tturl", "tiktoklink", "tiktokurl", "douyinlink"),
    "youtube": ("ytlink", "yturl", "youtubelink", "youtubeurl", "channelurl", "channellink"),
}


@dataclass(frozen=True)
class ThreadRef:
    thread_key: str
    example_subject: str
    last_sent_at: str
    message_count: int
    contact_email: str
    contact_name: str
    contact_message_count: int


@dataclass(frozen=True)
class MessageSnapshot:
    row_id: int
    thread_key: str
    sent_sort_at: str
    subject: str
    snippet: str
    body_text: str
    raw_path: str
    direction: str
    participants: frozenset[str]
    handle_tokens: frozenset[str]


@dataclass(frozen=True)
class MatchResult:
    matched_contact_email: str
    matched_contact_name: str
    match_rule: str
    match_confidence: str
    evidence_message: MessageSnapshot


@dataclass(frozen=True)
class QuoteResult:
    amount: float
    currency: str
    source: str
    message_id: int
    sent_at: str
    text: str
    score: int


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _timestamp(value: Optional[str]) -> float:
    parsed = _parse_iso_datetime(value)
    if parsed is None:
        return float("-inf")
    return parsed.timestamp()


def _normalize_email(value: Optional[str]) -> str:
    if not value:
        return ""
    return value.strip().lower()


def _normalize_handle(value: Optional[str]) -> str:
    if not value:
        return ""
    text = html.unescape(str(value)).strip()
    if not text:
        return ""
    handles = _extract_handles_from_text(text)
    if handles:
        return max(handles, key=len)
    text = text.lstrip("@").strip().lower()
    return re.sub(r"[^a-z0-9._]", "", text)


def _normalize_name(value: Optional[str]) -> str:
    if not value:
        return ""
    text = html.unescape(str(value)).strip().lower()
    return "".join(char for char in text if char.isalnum())


def _clean_text(value: Optional[str]) -> str:
    if not value:
        return ""
    text = html.unescape(str(value)).replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _strip_urls(value: str) -> str:
    return URL_PATTERN.sub(" ", value)


def _extract_handles_from_text(value: Optional[str]) -> set[str]:
    text = _clean_text(value).lower()
    if not text:
        return set()

    handles: set[str] = set()
    for pattern in (MENTION_HANDLE_PATTERN, YOUTUBE_HANDLE_PATTERN, TIKTOK_HANDLE_PATTERN, INSTAGRAM_HANDLE_PATTERN):
        for match in pattern.finditer(text):
            handle = re.sub(r"[^a-z0-9._]", "", match.group(1).lower())
            if len(handle) >= 3:
                handles.add(handle)
    return handles


def _extract_emails(value: Optional[str]) -> list[str]:
    if not value:
        return []
    seen: set[str] = set()
    result: list[str] = []
    for match in EMAIL_PATTERN.finditer(str(value)):
        email = _normalize_email(match.group(0))
        if not email or email in seen:
            continue
        seen.add(email)
        result.append(email)
    return result


def _load_addresses(raw_value: str) -> list[dict[str, str]]:
    try:
        items = json.loads(raw_value or "[]")
    except json.JSONDecodeError:
        return []

    result: list[dict[str, str]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "") or "").strip()
        address = _normalize_email(item.get("address"))
        if not name and not address:
            continue
        result.append({"name": name, "address": address})
    return result


def _message_participants(row: sqlite3.Row) -> frozenset[str]:  # type: ignore[name-defined]
    participants: set[str] = set()
    for key in ("from_json", "to_json", "cc_json", "reply_to_json", "bcc_json", "sender_json"):
        for item in _load_addresses(row[key] or "[]"):
            if item["address"]:
                participants.add(item["address"])
    return frozenset(participants)


def _collab_score(subject: Optional[str]) -> int:
    lowered = _clean_text(subject).lower()
    if not lowered:
        return 0
    return sum(1 for keyword in COLLAB_KEYWORDS if keyword in lowered)


def _handle_variants_from_email(email: str) -> set[str]:
    local_part = email.split("@", 1)[0].lower()
    local_part = re.sub(r"[^a-z0-9._]", "", local_part)
    if not local_part:
        return set()

    variants = {local_part}
    for suffix in ("mgmt", "management", "business", "biz", "official", "collab", "team", "media", "contact", "mail"):
        if local_part.endswith(suffix) and len(local_part) > len(suffix) + 2:
            variants.add(local_part[: -len(suffix)])
    for prefix in ("collabwith", "contact", "team", "official", "the"):
        if local_part.startswith(prefix) and len(local_part) > len(prefix) + 2:
            variants.add(local_part[len(prefix) :])
    variants.update({variant.replace(".", "").replace("_", "") for variant in variants if "." in variant or "_" in variant})
    return {_normalize_handle(variant) for variant in variants if _normalize_handle(variant)}


def _pick_best_thread_ref(refs: Sequence[ThreadRef]) -> ThreadRef:
    return max(
        refs,
        key=lambda item: (_collab_score(item.example_subject), _timestamp(item.last_sent_at), item.message_count, item.contact_message_count),
    )


def _pick_best_message(messages: Sequence[MessageSnapshot]) -> MessageSnapshot:
    return max(messages, key=lambda item: (_collab_score(item.subject), _timestamp(item.sent_sort_at), item.row_id))


def _iter_amount_matches(text: str) -> Iterator[tuple[float, str, int, int, str]]:
    seen_spans: set[tuple[int, int]] = set()
    for pattern in AMOUNT_PATTERNS:
        for match in pattern.finditer(text):
            span = match.span()
            if span in seen_spans:
                continue
            seen_spans.add(span)
            raw_amount = match.group(1).replace(",", "")
            try:
                amount = float(raw_amount)
            except ValueError:
                continue
            currency = "USD"
            source_type = "body"
            if "k" in match.group(0).lower():
                amount = amount * 1000
                source_type = "k_amount"
            yield amount, currency, span[0], span[1], source_type


def _quote_score(message: MessageSnapshot, context: str, source: str) -> int:
    lowered = context.lower()
    score = 0
    if message.direction == "inbound":
        score += 2
    if source == "subject":
        score -= 1
    if any(term in lowered for term in QUOTE_POSITIVE_TERMS):
        score += 5
    if any(term in lowered for term in OUTBOUND_BUDGET_TERMS):
        score -= 3
    if any(term in lowered for term in QUOTE_NEGATIVE_TERMS):
        score -= 4
    score += _collab_score(message.subject)
    return score


def _extract_quote_candidates(message: MessageSnapshot) -> list[QuoteResult]:
    candidates: list[QuoteResult] = []
    cleaned_body = _clean_text(message.body_text)
    cleaned_subject = _clean_text(message.subject)

    for text, source in ((cleaned_body, "body"), (cleaned_subject, "subject")):
        if not text:
            continue
        text_no_urls = _strip_urls(text)
        for amount, currency, start, end, source_type in _iter_amount_matches(text_no_urls):
            context = text_no_urls[max(0, start - 60) : min(len(text_no_urls), end + 60)]
            lowered_context = context.lower()
            if source_type == "k_amount" and not any(term in lowered_context for term in K_AMOUNT_FINANCE_TERMS):
                continue
            if source_type == "k_amount" and any(term in lowered_context for term in K_AMOUNT_TECH_TERMS):
                continue
            score = _quote_score(message, context, source)
            if amount > 20000 and "rate" not in lowered_context and "fee" not in lowered_context:
                score -= 2
            if amount < 30 and "rate" not in lowered_context and "fee" not in lowered_context:
                score -= 2
            candidates.append(
                QuoteResult(
                    amount=amount,
                    currency=currency,
                    source=source,
                    message_id=message.row_id,
                    sent_at=message.sent_sort_at,
                    text=context[:220],
                    score=score,
                )
            )
    return candidates


class MailIndex:
    def __init__(self, db: Database) -> None:
        self.db = db
        self.threads_by_email: dict[str, list[ThreadRef]] = defaultdict(list)
        self.threads_by_name: dict[str, list[ThreadRef]] = defaultdict(list)
        self.threads_by_handle: dict[str, list[ThreadRef]] = defaultdict(list)
        self.thread_contacts: dict[str, list[ThreadRef]] = defaultdict(list)
        self.messages_by_handle: dict[str, list[MessageSnapshot]] = defaultdict(list)
        self.thread_messages: dict[str, list[MessageSnapshot]] = defaultdict(list)
        self.thread_last_message: dict[str, MessageSnapshot] = {}
        self.thread_quote_cache: dict[str, Optional[QuoteResult]] = {}
        self._load_thread_refs()
        self._load_messages()

    def _load_thread_refs(self) -> None:
        rows = self.db.conn.execute(
            """
            SELECT
                t.thread_key,
                t.example_subject,
                t.last_sent_at,
                t.message_count,
                c.email_normalized,
                c.display_name,
                tc.message_count AS contact_message_count
            FROM thread_contacts tc
            JOIN contacts c ON c.id = tc.contact_id
            JOIN threads t ON t.thread_key = tc.thread_key
            ORDER BY datetime(t.last_sent_at) DESC, t.thread_key, c.email_normalized
            """
        ).fetchall()

        for row in rows:
            ref = ThreadRef(
                thread_key=str(row["thread_key"]),
                example_subject=str(row["example_subject"] or ""),
                last_sent_at=str(row["last_sent_at"] or ""),
                message_count=int(row["message_count"] or 0),
                contact_email=str(row["email_normalized"] or ""),
                contact_name=str(row["display_name"] or ""),
                contact_message_count=int(row["contact_message_count"] or 0),
            )
            self.threads_by_email[ref.contact_email].append(ref)
            normalized_name = _normalize_name(ref.contact_name)
            if normalized_name:
                self.threads_by_name[normalized_name].append(ref)
            for handle in _handle_variants_from_email(ref.contact_email):
                self.threads_by_handle[handle].append(ref)
            self.thread_contacts[ref.thread_key].append(ref)

    def _load_messages(self) -> None:
        rows = self.db.conn.execute(
            """
            SELECT
                m.id,
                m.subject,
                m.snippet,
                m.body_text,
                m.raw_path,
                m.from_json,
                m.to_json,
                m.cc_json,
                m.bcc_json,
                m.reply_to_json,
                m.sender_json,
                mi.thread_key,
                mi.direction,
                COALESCE(mi.sent_sort_at, m.sent_at, m.internal_date, m.created_at) AS sent_sort_at
            FROM messages m
            JOIN message_index mi ON mi.message_row_id = m.id
            ORDER BY datetime(COALESCE(mi.sent_sort_at, m.sent_at, m.internal_date, m.created_at)), m.id
            """
        ).fetchall()

        for row in rows:
            combined_text = " ".join(
                part
                for part in (
                    row["subject"] or "",
                    row["snippet"] or "",
                    (row["body_text"] or "")[:4000],
                )
                if part
            )
            snapshot = MessageSnapshot(
                row_id=int(row["id"]),
                thread_key=str(row["thread_key"]),
                sent_sort_at=str(row["sent_sort_at"] or ""),
                subject=str(row["subject"] or ""),
                snippet=str(row["snippet"] or ""),
                body_text=str(row["body_text"] or ""),
                raw_path=str(row["raw_path"] or ""),
                direction=str(row["direction"] or ""),
                participants=_message_participants(row),
                handle_tokens=frozenset(_extract_handles_from_text(combined_text)),
            )
            self.thread_messages[snapshot.thread_key].append(snapshot)
            self.thread_last_message[snapshot.thread_key] = snapshot
            for handle in snapshot.handle_tokens:
                self.messages_by_handle[handle].append(snapshot)

    def _primary_contact_for_thread(self, thread_key: str, preferred_name: str = "", preferred_handle: str = "") -> tuple[str, str]:
        refs = self.thread_contacts.get(thread_key, [])
        if not refs:
            return "", ""

        normalized_name = _normalize_name(preferred_name)
        if normalized_name:
            for ref in refs:
                if _normalize_name(ref.contact_name) == normalized_name:
                    return ref.contact_email, ref.contact_name

        if preferred_handle:
            for ref in refs:
                if preferred_handle in _handle_variants_from_email(ref.contact_email):
                    return ref.contact_email, ref.contact_name

        best_ref = max(refs, key=lambda item: (item.contact_message_count, _timestamp(item.last_sent_at), item.contact_email))
        return best_ref.contact_email, best_ref.contact_name

    def _evidence_for_thread(self, thread_key: str, emails: Iterable[str] = (), handle: str = "") -> Optional[MessageSnapshot]:
        target_emails = {_normalize_email(item) for item in emails if _normalize_email(item)}
        handle = _normalize_handle(handle)
        messages = self.thread_messages.get(thread_key, [])
        if not messages:
            return None

        for message in reversed(messages):
            if target_emails and target_emails & set(message.participants):
                return message
            if handle and handle in message.handle_tokens:
                return message
        return messages[-1]

    def match_by_email(self, emails: Sequence[str]) -> Optional[MatchResult]:
        refs: list[ThreadRef] = []
        for email in emails:
            refs.extend(self.threads_by_email.get(_normalize_email(email), []))
        if not refs:
            return None

        best_ref = _pick_best_thread_ref(refs)
        evidence = self._evidence_for_thread(best_ref.thread_key, emails=emails)
        if evidence is None:
            return None
        return MatchResult(
            matched_contact_email=best_ref.contact_email,
            matched_contact_name=best_ref.contact_name,
            match_rule="email_exact",
            match_confidence="high",
            evidence_message=evidence,
        )

    def match_by_handle(self, handle: str, nickname: str = "") -> Optional[MatchResult]:
        normalized_handle = _normalize_handle(handle)
        if not normalized_handle:
            return None

        handle_messages = self.messages_by_handle.get(normalized_handle, [])
        handle_refs = self.threads_by_handle.get(normalized_handle, [])
        unique_threads = {message.thread_key for message in handle_messages}
        unique_threads.update(ref.thread_key for ref in handle_refs)
        if not unique_threads:
            return None

        if handle_messages:
            best_message = _pick_best_message(handle_messages)
            matched_email, matched_name = self._primary_contact_for_thread(best_message.thread_key, preferred_name=nickname, preferred_handle=normalized_handle)
            confidence = "high" if len({message.thread_key for message in handle_messages}) == 1 else "medium"
            rule = "handle_exact_unique" if confidence == "high" else "handle_exact"
            return MatchResult(
                matched_contact_email=matched_email,
                matched_contact_name=matched_name,
                match_rule=rule,
                match_confidence=confidence,
                evidence_message=best_message,
            )

        best_ref = _pick_best_thread_ref(handle_refs)
        evidence = self._evidence_for_thread(best_ref.thread_key, handle=normalized_handle)
        if evidence is None:
            return None
        confidence = "high" if len(unique_threads) == 1 else "medium"
        rule = "handle_contact_unique" if confidence == "high" else "handle_contact"
        return MatchResult(
            matched_contact_email=best_ref.contact_email,
            matched_contact_name=best_ref.contact_name,
            match_rule=rule,
            match_confidence=confidence,
            evidence_message=evidence,
        )

    def match_by_nickname(self, nickname: str) -> Optional[MatchResult]:
        normalized_name = _normalize_name(nickname)
        if not normalized_name:
            return None

        refs = self.threads_by_name.get(normalized_name, [])
        if not refs:
            return None
        best_ref = _pick_best_thread_ref(refs)
        evidence = self._evidence_for_thread(best_ref.thread_key)
        if evidence is None:
            return None

        unique_threads = {ref.thread_key for ref in refs}
        unique_emails = {ref.contact_email for ref in refs}
        confidence = "high" if len(unique_threads) == 1 and len(unique_emails) == 1 else "medium"
        rule = "nickname_exact_unique" if confidence == "high" else "nickname_exact"
        return MatchResult(
            matched_contact_email=best_ref.contact_email,
            matched_contact_name=best_ref.contact_name,
            match_rule=rule,
            match_confidence=confidence,
            evidence_message=evidence,
        )

    def get_last_mail(self, thread_key: str) -> Optional[MessageSnapshot]:
        return self.thread_last_message.get(thread_key)

    def get_quote(self, thread_key: str) -> Optional[QuoteResult]:
        if thread_key in self.thread_quote_cache:
            return self.thread_quote_cache[thread_key]

        messages = self.thread_messages.get(thread_key, [])
        best: Optional[QuoteResult] = None
        for message in messages:
            for candidate in _extract_quote_candidates(message):
                if best is None:
                    best = candidate
                    continue
                left = (candidate.score, _timestamp(candidate.sent_at), candidate.amount)
                right = (best.score, _timestamp(best.sent_at), best.amount)
                if left > right:
                    best = candidate

        self.thread_quote_cache[thread_key] = best
        return best

    def get_thread_summary(self, thread_key: str) -> tuple[int, str]:
        messages = self.thread_messages.get(thread_key, [])
        if not messages:
            return 0, ""
        return len(messages), messages[-1].sent_sort_at


def _load_workbook() -> Any:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("缺少 openpyxl，请先安装后再运行 enrich-creators。") from exc
    return Workbook, load_workbook


def _source_headers(input_path: Path) -> list[str]:
    _, load_workbook = _load_workbook()
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        ordered_headers: list[str] = []
        seen: set[str] = set()
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(iterator, ())
            for cell in header_row:
                header = _clean_text(cell)
                if not header or header in seen:
                    continue
                seen.add(header)
                ordered_headers.append(header)
        return ordered_headers
    finally:
        workbook.close()


def _iter_sheet_rows(input_path: Path, source_headers: Sequence[str]) -> Iterator[dict[str, Any]]:
    _, load_workbook = _load_workbook()
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [_clean_text(cell) for cell in header_row]
            header_positions = {header: index for index, header in enumerate(headers) if header}
            for row_number, values in enumerate(rows, start=2):
                row: dict[str, Any] = {
                    "sheet_name": sheet.title,
                    "source_row_number": row_number,
                }
                for header in source_headers:
                    index = header_positions.get(header)
                    row[header] = values[index] if index is not None and index < len(values) else ""
                yield row
    finally:
        workbook.close()


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _normalize_source_column_name(name: Any) -> str:
    return re.sub(r"[\s_\-./（）()]+", "", str(name or "").strip().lower())


def _resolve_source_column(columns: Sequence[Any], aliases: Sequence[str]) -> Any:
    normalized_aliases = {_normalize_source_column_name(alias) for alias in aliases if str(alias or "").strip()}
    for column in columns:
        if _normalize_source_column_name(column) in normalized_aliases:
            return column
    return None


def _clean_source_cell(value: Any) -> str:
    return _stringify(value)


def _infer_platform_from_value(value: Any) -> str:
    text = _clean_source_cell(value).lower()
    if not text:
        return ""
    if "instagram.com" in text:
        return "instagram"
    if "tiktok.com" in text:
        return "tiktok"
    if "youtube.com" in text or "youtu.be/" in text:
        return "youtube"
    return ""


def _platform_label(platform: str) -> str:
    mapping = {
        "instagram": "Instagram",
        "tiktok": "TikTok",
        "youtube": "YouTube",
    }
    return mapping.get(platform, platform)


def _resolve_sending_list_link_columns(columns: Sequence[Any], rows: Sequence[dict[str, Any]]) -> list[tuple[Any, str]]:
    resolved: list[tuple[Any, str]] = []
    normalized_generic_aliases = {
        _normalize_source_column_name(alias)
        for alias in SENDING_LIST_GENERIC_LINK_ALIASES
    }
    normalized_platform_aliases = {
        platform: {
            _normalize_source_column_name(alias)
            for alias in aliases
        }
        for platform, aliases in SENDING_LIST_PLATFORM_LINK_ALIASES.items()
    }

    for column in columns:
        normalized_column = _normalize_source_column_name(column)
        explicit_platform = ""
        for platform, aliases in normalized_platform_aliases.items():
            if normalized_column in aliases:
                explicit_platform = platform
                break

        inferred_platform = ""
        if not explicit_platform:
            checked = 0
            hits: dict[str, int] = {}
            for row in rows:
                platform = _infer_platform_from_value(row.get(column))
                if platform:
                    hits[platform] = hits.get(platform, 0) + 1
                if _clean_source_cell(row.get(column)):
                    checked += 1
                if checked >= 20:
                    break
            if hits:
                inferred_platform = max(hits.items(), key=lambda item: item[1])[0]

        if explicit_platform or inferred_platform or normalized_column in normalized_generic_aliases:
            resolved.append((column, explicit_platform or inferred_platform))
    return resolved


def _has_canonical_creator_columns(headers: Sequence[str]) -> bool:
    normalized = {_normalize_source_column_name(header) for header in headers}
    required = {_normalize_source_column_name(header) for header in ("Platform", "@username", "URL")}
    return required.issubset(normalized)


def _iter_sending_list_rows(input_path: Path) -> Iterator[dict[str, Any]]:
    _, load_workbook = _load_workbook()
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        records_by_key: dict[tuple[str, str], dict[str, Any]] = {}
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [_clean_text(cell) for cell in header_row]
            if not any(headers):
                continue

            parsed_rows: list[dict[str, Any]] = []
            for row_number, values in enumerate(rows, start=2):
                row_dict: dict[str, Any] = {"sheet_name": sheet.title, "source_row_number": row_number}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = values[index] if index < len(values) else ""
                parsed_rows.append(row_dict)

            columns = [header for header in headers if header]
            country_column = _resolve_source_column(columns, SENDING_LIST_COUNTRY_ALIASES)
            creator_column = _resolve_source_column(columns, SENDING_LIST_CREATOR_ALIASES)
            email_column = _resolve_source_column(columns, SENDING_LIST_EMAIL_ALIASES)
            link_columns = _resolve_sending_list_link_columns(columns, parsed_rows)
            if not link_columns:
                continue

            for row_dict in parsed_rows:
                nickname = _clean_source_cell(row_dict.get(creator_column)) if creator_column else ""
                region = _clean_source_cell(row_dict.get(country_column)) if country_column else ""
                email = _clean_source_cell(row_dict.get(email_column)) if email_column else ""
                row_seen_keys: set[tuple[str, str]] = set()

                for link_column, default_platform in link_columns:
                    raw_link_value = _clean_source_cell(row_dict.get(link_column))
                    if not raw_link_value:
                        continue
                    platform = _infer_platform_from_value(raw_link_value) or default_platform
                    if not platform:
                        continue
                    identifier = _normalize_handle(raw_link_value) or _normalize_handle(nickname)
                    if not identifier:
                        continue
                    record_key = (platform, identifier)
                    if record_key in row_seen_keys:
                        continue
                    row_seen_keys.add(record_key)

                    existing = records_by_key.get(record_key)
                    if existing is None:
                        records_by_key[record_key] = {
                            "Platform": _platform_label(platform),
                            "@username": identifier,
                            "URL": raw_link_value,
                            "nickname": nickname,
                            "Region": region,
                            "Email": email,
                            "sheet_name": row_dict["sheet_name"],
                            "source_row_number": row_dict["source_row_number"],
                        }
                        continue
                    if not existing.get("URL") and raw_link_value:
                        existing["URL"] = raw_link_value
                    if not existing.get("nickname") and nickname:
                        existing["nickname"] = nickname
                    if not existing.get("Region") and region:
                        existing["Region"] = region
                    if not existing.get("Email") and email:
                        existing["Email"] = email

        for row in records_by_key.values():
            yield row
    finally:
        workbook.close()


def _build_output_row(source_row: dict[str, Any], match: Optional[MatchResult], index: MailIndex) -> dict[str, Any]:
    output = dict(source_row)
    output["sheet_name"] = source_row.get("sheet_name", "")
    output["source_row_number"] = source_row.get("source_row_number", "")

    creator_email_text = _stringify(source_row.get("Email", ""))
    creator_emails = _extract_emails(creator_email_text)
    username = _stringify(source_row.get("@username", ""))
    url = _stringify(source_row.get("URL", ""))
    nickname = _stringify(source_row.get("nickname", ""))
    derived_handle = _normalize_handle(url) or _normalize_handle(username)

    output["derived_handle"] = derived_handle
    output["creator_emails"] = " | ".join(creator_emails)
    output["matched_contact_email"] = ""
    output["matched_contact_name"] = ""
    output["match_rule"] = ""
    output["match_confidence"] = ""
    output["evidence_message_id"] = ""
    output["evidence_subject"] = ""
    output["evidence_sent_at"] = ""
    output["evidence_thread_key"] = ""
    output["evidence_raw_path"] = ""
    output["last_mail_message_id"] = ""
    output["last_mail_time"] = ""
    output["last_mail_subject"] = ""
    output["last_mail_snippet"] = ""
    output["last_mail_raw_path"] = ""
    output["thread_message_count"] = ""
    output["thread_last_sent_at"] = ""
    output["latest_quote_amount"] = ""
    output["latest_quote_currency"] = ""
    output["latest_quote_source"] = ""
    output["latest_quote_message_id"] = ""
    output["latest_quote_time"] = ""
    output["latest_quote_text"] = ""

    if match is None:
        return output

    output["matched_contact_email"] = match.matched_contact_email
    output["matched_contact_name"] = match.matched_contact_name
    output["match_rule"] = match.match_rule
    output["match_confidence"] = match.match_confidence
    output["evidence_message_id"] = match.evidence_message.row_id
    output["evidence_subject"] = match.evidence_message.subject
    output["evidence_sent_at"] = match.evidence_message.sent_sort_at
    output["evidence_thread_key"] = match.evidence_message.thread_key
    output["evidence_raw_path"] = match.evidence_message.raw_path

    last_mail = index.get_last_mail(match.evidence_message.thread_key)
    if last_mail is not None:
        output["last_mail_message_id"] = last_mail.row_id
        output["last_mail_time"] = last_mail.sent_sort_at
        output["last_mail_subject"] = last_mail.subject
        output["last_mail_snippet"] = last_mail.snippet
        output["last_mail_raw_path"] = last_mail.raw_path

    thread_message_count, thread_last_sent_at = index.get_thread_summary(match.evidence_message.thread_key)
    output["thread_message_count"] = thread_message_count
    output["thread_last_sent_at"] = thread_last_sent_at

    quote = index.get_quote(match.evidence_message.thread_key)
    if quote is not None:
        output["latest_quote_amount"] = round(quote.amount, 2)
        output["latest_quote_currency"] = quote.currency
        output["latest_quote_source"] = quote.source
        output["latest_quote_message_id"] = quote.message_id
        output["latest_quote_time"] = quote.sent_at
        output["latest_quote_text"] = quote.text

    return output


def _select_match(index: MailIndex, row: dict[str, Any]) -> Optional[MatchResult]:
    creator_emails = _extract_emails(_stringify(row.get("Email", "")))
    if creator_emails:
        email_match = index.match_by_email(creator_emails)
        if email_match is not None:
            return email_match

    nickname = _stringify(row.get("nickname", ""))
    handle = _normalize_handle(_stringify(row.get("URL", ""))) or _normalize_handle(_stringify(row.get("@username", "")))
    handle_match = index.match_by_handle(handle, nickname=nickname) if handle else None
    name_match = index.match_by_nickname(nickname) if nickname else None

    if handle_match is not None and handle_match.match_confidence == "high":
        return handle_match
    if name_match is not None and name_match.match_confidence == "high":
        return name_match
    return handle_match or name_match


def _write_csv(path: Path, headers: Sequence[str], rows: Iterable[dict[str, Any]]) -> int:
    count = 0
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers))
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})
            count += 1
    return count


def _write_xlsx(path: Path, headers: Sequence[str], rows: Iterable[dict[str, Any]]) -> int:
    Workbook, _ = _load_workbook()
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="results")
    worksheet.append(list(headers))
    count = 0
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
        count += 1
    workbook.save(path)
    return count


def _row_values(row: dict[str, Any], headers: Sequence[str]) -> list[Any]:
    return [row.get(header, "") for header in headers]


def _ensure_relation_index(db: Database) -> None:
    message_count = int(db.conn.execute("SELECT COUNT(*) FROM messages").fetchone()[0])
    indexed_count = int(db.conn.execute("SELECT COUNT(*) FROM message_index").fetchone()[0])
    if message_count == 0:
        raise RuntimeError("本地邮件库为空，先运行 python3 -m email_sync sync")
    if indexed_count != message_count:
        rebuild_relation_index(db)


def enrich_creator_workbook(
    db: Database,
    input_path: Path,
    output_prefix: Path,
) -> dict[str, Any]:
    if not input_path.exists():
        raise FileNotFoundError(f"找不到达人库文件: {input_path}")

    db.init_schema()
    _ensure_relation_index(db)
    index = MailIndex(db)

    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    source_headers = _source_headers(input_path)
    source_kind = "canonical_upload"
    source_rows: Iterable[dict[str, Any]]
    if _has_canonical_creator_columns(source_headers):
        source_rows = _iter_sheet_rows(input_path, source_headers)
    else:
        source_kind = "sending_list"
        source_headers = list(CANONICAL_SOURCE_HEADERS)
        source_rows = _iter_sending_list_rows(input_path)
    output_headers = list(source_headers) + APPENDED_HEADERS

    csv_path = output_prefix.with_suffix(".csv")
    xlsx_path = output_prefix.with_suffix(".xlsx")
    high_csv_path = output_prefix.with_name(f"{output_prefix.name}_高置信").with_suffix(".csv")
    high_xlsx_path = output_prefix.with_name(f"{output_prefix.name}_高置信").with_suffix(".xlsx")

    Workbook, _ = _load_workbook()
    all_workbook = Workbook(write_only=True)
    all_sheet = all_workbook.create_sheet(title="results")
    all_sheet.append(list(output_headers))
    high_workbook = Workbook(write_only=True)
    high_sheet = high_workbook.create_sheet(title="results")
    high_sheet.append(list(output_headers))

    rows = 0
    matched_rows = 0
    high_confidence_rows = 0

    with csv_path.open("w", encoding="utf-8-sig", newline="") as all_csv_handle, high_csv_path.open(
        "w", encoding="utf-8-sig", newline=""
    ) as high_csv_handle:
        all_writer = csv.DictWriter(all_csv_handle, fieldnames=list(output_headers))
        high_writer = csv.DictWriter(high_csv_handle, fieldnames=list(output_headers))
        all_writer.writeheader()
        high_writer.writeheader()

        for source_row in source_rows:
            match = _select_match(index, source_row)
            if match is not None:
                matched_rows += 1
            output_row = _build_output_row(source_row, match, index)
            all_writer.writerow({header: output_row.get(header, "") for header in output_headers})
            all_sheet.append(_row_values(output_row, output_headers))
            rows += 1

            if output_row["match_confidence"] == "high":
                high_writer.writerow({header: output_row.get(header, "") for header in output_headers})
                high_sheet.append(_row_values(output_row, output_headers))
                high_confidence_rows += 1

    all_workbook.save(xlsx_path)
    high_workbook.save(high_xlsx_path)

    return {
        "source_kind": source_kind,
        "rows": rows,
        "matched_rows": matched_rows,
        "high_confidence_rows": high_confidence_rows,
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "high_csv_path": str(high_csv_path),
        "high_xlsx_path": str(high_xlsx_path),
    }
