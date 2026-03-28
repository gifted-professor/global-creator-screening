from __future__ import annotations

import json
import os
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

import requests
from openpyxl import Workbook, load_workbook

from .creator_enrichment import _extract_emails, _load_addresses, _normalize_email, _normalize_handle, _normalize_name
from .creator_review import _ensure_relation_index, _fetch_message_evidence
from .db import Database


PRODUCTION_HELPER_HEADERS = [
    "creator_dedupe_key",
    "same_last_mail_row_count",
    "shared_contact_suspected",
    "manager_suspected",
    "mass_cc_suspected",
]

LLM_REVIEW_HEADERS = [
    "llm_review_status",
    "llm_review_group_key",
    "llm_review_decision",
    "llm_review_keep",
    "llm_review_matched_creator_dedupe_keys",
    "llm_review_reason",
    "llm_review_confidence",
    "llm_review_sender_role",
    "llm_review_provider",
    "llm_review_model",
    "llm_review_wire_api",
    "llm_review_reviewed_at",
]

MANAGER_HINT_TERMS = (
    "manager",
    "management",
    "mgmt",
    "agent",
    "agency",
    "assistant",
    "partnership",
    "partnerships",
    "talent",
    "coordinator",
    "booking",
    "customer experience",
    "customer service",
    "representing",
    "team",
)

RETRYABLE_STATUS_CODES = {408, 429, 500, 502, 503, 504}


@dataclass(frozen=True)
class LlmReviewConfig:
    base_url: str
    api_key: str
    model: str
    wire_api: str
    timeout_seconds: int
    provider_name: str
    reasoning_effort: str


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _json_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _source_headers(input_path: Path) -> list[str]:
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        ordered_headers: list[str] = []
        seen: set[str] = set()
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(iterator, ())
            for cell in header_row:
                header = _stringify(cell)
                if not header or header in seen:
                    continue
                seen.add(header)
                ordered_headers.append(header)
        return ordered_headers
    finally:
        workbook.close()


def _iter_rows(input_path: Path, source_headers: Sequence[str]) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [_stringify(cell) for cell in header_row]
            header_positions = {header: index for index, header in enumerate(headers) if header}
            for row_number, values in enumerate(rows, start=2):
                row: dict[str, Any] = {
                    "sheet_name": sheet.title,
                    "source_row_number": row_number,
                }
                has_value = False
                for header in source_headers:
                    index = header_positions.get(header)
                    value = values[index] if index is not None and index < len(values) else ""
                    row[header] = value
                    if not has_value and _stringify(value):
                        has_value = True
                if has_value:
                    yield row
    finally:
        workbook.close()


def _build_group_key(row: dict[str, Any]) -> tuple[str, str, str]:
    message_id = _stringify(row.get("last_mail_message_id"))
    if message_id:
        return (f"last_mail_message_id:{message_id}", "last_mail_message_id", message_id)
    raw_path = _stringify(row.get("last_mail_raw_path"))
    if raw_path:
        return (f"last_mail_raw_path:{raw_path}", "last_mail_raw_path", raw_path)
    return ("", "", "")


def _creator_dedupe_key(row: dict[str, Any]) -> str:
    platform = _stringify(row.get("Platform")).strip().lower() or "unknown"
    handle = (
        _normalize_handle(row.get("derived_handle"))
        or _normalize_handle(row.get("URL"))
        or _normalize_handle(row.get("@username"))
    )
    creator_emails = _extract_emails(_stringify(row.get("creator_emails"))) or _extract_emails(_stringify(row.get("Email")))
    nickname = _normalize_name(row.get("nickname"))
    if handle:
        return f"{platform}:handle:{handle}"
    if creator_emails:
        return f"{platform}:email:{creator_emails[0]}"
    if nickname:
        return f"{platform}:name:{nickname}"
    sheet_name = _stringify(row.get("sheet_name")) or "sheet"
    source_row = _stringify(row.get("source_row_number")) or "0"
    return f"fallback:{sheet_name}:{source_row}"


def _row_score(row: dict[str, Any]) -> tuple[Any, ...]:
    populated = sum(1 for value in row.values() if _stringify(value))
    quote_score = 1 if _stringify(row.get("latest_quote_amount")) else 0
    match_score = 1 if _stringify(row.get("matched_contact_email")) else 0
    last_mail_time = _stringify(row.get("last_mail_time"))
    try:
        last_mail_dt = datetime.fromisoformat(last_mail_time).timestamp() if last_mail_time else float("-inf")
    except ValueError:
        last_mail_dt = float("-inf")
    source_row_number = int(_stringify(row.get("source_row_number")) or "0")
    return (populated, quote_score, match_score, last_mail_dt, -source_row_number)


def _write_workbook(path: Path, headers: Sequence[str], rows: Iterable[dict[str, Any]]) -> None:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="results")
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def _fetch_message_recipient_metrics(db: Database, key_source: str, key_value: str) -> dict[str, int]:
    if not key_source or not key_value:
        return {
            "to_count": 0,
            "cc_count": 0,
            "bcc_count": 0,
            "external_recipient_count": 0,
            "mass_cc_suspected": 0,
        }

    if key_source == "last_mail_message_id":
        row = db.conn.execute(
            """
            SELECT from_json, sender_json, to_json, cc_json, bcc_json
            FROM messages
            WHERE id = ?
            """,
            (int(key_value),),
        ).fetchone()
    else:
        row = db.conn.execute(
            """
            SELECT from_json, sender_json, to_json, cc_json, bcc_json
            FROM messages
            WHERE raw_path = ?
            """,
            (key_value,),
        ).fetchone()

    if row is None:
        return {
            "to_count": 0,
            "cc_count": 0,
            "bcc_count": 0,
            "external_recipient_count": 0,
            "mass_cc_suspected": 0,
        }

    sender_emails = {
        _normalize_email(item["address"])
        for key in ("from_json", "sender_json")
        for item in _load_addresses(row[key] or "[]")
        if _normalize_email(item["address"])
    }
    to_items = _load_addresses(row["to_json"] or "[]")
    cc_items = _load_addresses(row["cc_json"] or "[]")
    bcc_items = _load_addresses(row["bcc_json"] or "[]")
    external_recipients = {
        _normalize_email(item["address"])
        for item in [*to_items, *cc_items, *bcc_items]
        if _normalize_email(item["address"]) and _normalize_email(item["address"]) not in sender_emails
    }
    mass_cc_suspected = 1 if len(cc_items) >= 2 or len(external_recipients) >= 4 else 0
    return {
        "to_count": len(to_items),
        "cc_count": len(cc_items),
        "bcc_count": len(bcc_items),
        "external_recipient_count": len(external_recipients),
        "mass_cc_suspected": mass_cc_suspected,
    }


def _manager_suspected(row: dict[str, Any]) -> int:
    lowered = " ".join(
        part.lower()
        for part in (
            _stringify(row.get("matched_contact_name")),
            _stringify(row.get("matched_contact_email")),
            _stringify(row.get("last_mail_subject")),
            _stringify(row.get("last_mail_snippet")),
        )
        if part
    )
    return 1 if any(term in lowered for term in MANAGER_HINT_TERMS) else 0


def _annotate_rows(db: Database, rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    annotated = [dict(row) for row in rows]
    by_contact: dict[str, set[str]] = defaultdict(set)
    group_counts: Counter[str] = Counter()
    recipient_metrics_cache: dict[str, dict[str, int]] = {}

    for row in annotated:
        dedupe_key = _creator_dedupe_key(row)
        row["creator_dedupe_key"] = dedupe_key
        matched_contact_email = _normalize_email(row.get("matched_contact_email"))
        if matched_contact_email:
            by_contact[matched_contact_email].add(dedupe_key)
        group_key, _, _ = _build_group_key(row)
        if group_key:
            group_counts[group_key] += 1

    for row in annotated:
        group_key, key_source, key_value = _build_group_key(row)
        if group_key and group_key not in recipient_metrics_cache:
            recipient_metrics_cache[group_key] = _fetch_message_recipient_metrics(db, key_source, key_value)
        metrics = recipient_metrics_cache.get(group_key, {})
        matched_contact_email = _normalize_email(row.get("matched_contact_email"))
        row["same_last_mail_row_count"] = group_counts.get(group_key, 0)
        row["shared_contact_suspected"] = 1 if matched_contact_email and len(by_contact[matched_contact_email]) >= 2 else 0
        row["manager_suspected"] = _manager_suspected(row)
        row["mass_cc_suspected"] = int(metrics.get("mass_cc_suspected", 0))
    return annotated


def _ordinary_creator_dedupe(rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    best_rows: dict[str, dict[str, Any]] = {}
    for row in rows:
        dedupe_key = _stringify(row.get("creator_dedupe_key"))
        existing = best_rows.get(dedupe_key)
        if existing is None or _row_score(row) > _row_score(existing):
            best_rows[dedupe_key] = dict(row)
    return list(best_rows.values())


def _candidate_row_payload(row: dict[str, Any], headers: Sequence[str]) -> dict[str, Any]:
    return {header: _json_safe(row.get(header, "")) for header in headers}


def _iter_candidate_groups(
    db: Database,
    rows: Sequence[dict[str, Any]],
    headers: Sequence[str],
) -> list[dict[str, Any]]:
    grouped_rows: dict[str, dict[str, Any]] = {}
    for row in rows:
        group_key, key_source, key_value = _build_group_key(row)
        if not group_key:
            continue
        entry = grouped_rows.setdefault(
            group_key,
            {"key_source": key_source, "key_value": key_value, "rows": []},
        )
        entry["rows"].append(dict(row))

    candidate_groups: list[dict[str, Any]] = []
    for group_key, entry in grouped_rows.items():
        candidate_rows = list(entry["rows"])
        same_last_mail_row_count = max(int(row.get("same_last_mail_row_count") or 0) for row in candidate_rows)
        mass_cc_suspected = max(int(row.get("mass_cc_suspected") or 0) for row in candidate_rows)
        if same_last_mail_row_count < 2 and mass_cc_suspected < 1:
            continue

        evidence = _fetch_message_evidence(db, str(entry["key_source"]), str(entry["key_value"])).to_dict()
        candidate_groups.append(
            {
                "group_key": group_key,
                "key_source": entry["key_source"],
                "key_value": entry["key_value"],
                "same_last_mail_row_count": same_last_mail_row_count,
                "mass_cc_suspected": mass_cc_suspected,
                "matched_contact_emails": sorted(
                    {
                        _normalize_email(row.get("matched_contact_email"))
                        for row in candidate_rows
                        if _normalize_email(row.get("matched_contact_email"))
                    }
                ),
                "candidate_creator_dedupe_keys": [row["creator_dedupe_key"] for row in candidate_rows],
                "candidate_rows": [_candidate_row_payload(row, headers) for row in candidate_rows],
                "representative_message": {
                    "message_row_id": evidence.get("message_row_id"),
                    "thread_key": evidence.get("thread_key"),
                    "subject": evidence.get("subject"),
                    "sent_at": evidence.get("sent_at"),
                    "snippet": evidence.get("snippet"),
                    "body_text": evidence.get("body_text"),
                    "raw_path": evidence.get("raw_path"),
                    "direction": evidence.get("direction"),
                    "thread_message_count": evidence.get("thread_message_count"),
                    "thread_last_sent_at": evidence.get("thread_last_sent_at"),
                },
                "thread_messages": list(evidence.get("thread_messages") or []),
            }
        )

    candidate_groups.sort(
        key=lambda item: (
            -int(item.get("same_last_mail_row_count") or 0),
            -int(item.get("mass_cc_suspected") or 0),
            str(item.get("group_key") or ""),
        )
    )
    return candidate_groups


def prepare_llm_review_candidates(
    db: Database,
    input_path: Path,
    output_prefix: Path,
) -> dict[str, Any]:
    if not input_path.exists():
        raise FileNotFoundError(f"找不到高置信 workbook: {input_path}")

    db.init_schema()
    _ensure_relation_index(db)

    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    source_headers = _source_headers(input_path)
    input_rows = list(_iter_rows(input_path, source_headers))

    prep_rows = _annotate_rows(db, input_rows)
    prep_headers = list(source_headers) + [header for header in PRODUCTION_HELPER_HEADERS if header not in source_headers]
    prep_path = output_prefix.with_suffix(".xlsx")
    _write_workbook(prep_path, prep_headers, prep_rows)

    deduped_rows = _ordinary_creator_dedupe(prep_rows)
    deduped_rows = _annotate_rows(db, deduped_rows)
    deduped_path = output_prefix.with_name(f"{output_prefix.name}_去重").with_suffix(".xlsx")
    _write_workbook(deduped_path, prep_headers, deduped_rows)

    candidate_groups = _iter_candidate_groups(db, deduped_rows, prep_headers)
    candidates_path = output_prefix.with_name(f"{output_prefix.name}_llm_candidates").with_suffix(".jsonl")
    with candidates_path.open("w", encoding="utf-8") as handle:
        for record in candidate_groups:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")

    return {
        "input_path": str(input_path),
        "db_path": str(db.db_path),
        "prep_xlsx_path": str(prep_path),
        "deduped_xlsx_path": str(deduped_path),
        "llm_candidates_jsonl_path": str(candidates_path),
        "source_row_count": len(input_rows),
        "prep_row_count": len(prep_rows),
        "deduped_row_count": len(deduped_rows),
        "llm_candidate_group_count": len(candidate_groups),
    }


def _load_env_file(env_path: str) -> dict[str, str]:
    path = Path(env_path).expanduser()
    candidates = [path]
    if path.name == ".env":
        local_path = path.with_name(".env.local")
        if local_path.exists():
            candidates.append(local_path)

    values: dict[str, str] = {}
    for candidate in candidates:
        if not candidate.exists():
            continue
        for raw_line in candidate.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def _resolve_env_value(key: str, env_values: dict[str, str], default: str = "") -> str:
    value = os.environ.get(key)
    if value is not None and str(value).strip():
        return str(value).strip()
    return str(env_values.get(key, default) or "").strip()


def _resolve_llm_review_config(
    env_path: str,
    *,
    base_url: Optional[str] = None,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
    wire_api: Optional[str] = None,
) -> LlmReviewConfig:
    env_values = _load_env_file(env_path)

    openai_api_key = _resolve_env_value("OPENAI_API_KEY", env_values)
    legacy_api_key = _resolve_env_value("LLM_API_KEY", env_values)
    use_legacy_surface = not openai_api_key and bool(legacy_api_key)

    resolved_api_key = str(api_key or "").strip() or openai_api_key or legacy_api_key
    resolved_base_url = (
        str(base_url or "").strip()
        or ("" if use_legacy_surface else _resolve_env_value("OPENAI_BASE_URL", env_values))
        or _resolve_env_value("LLM_API_BASE", env_values)
        or "https://api.openai.com/v1"
    )
    resolved_model = (
        str(model or "").strip()
        or ("" if use_legacy_surface else _resolve_env_value("OPENAI_MODEL", env_values))
        or _resolve_env_value("OPENAI_VISION_MODEL", env_values)
        or _resolve_env_value("LLM_MODEL", env_values)
        or _resolve_env_value("VISION_MODEL", env_values)
        or "gpt-5.4"
    )
    resolved_wire_api = (
        str(wire_api or "").strip()
        or _resolve_env_value("OPENAI_WIRE_API", env_values)
        or "chat_completions"
    ).strip().lower()
    if resolved_wire_api not in {"responses", "chat_completions"}:
        raise RuntimeError("OPENAI_WIRE_API 只支持 responses 或 chat_completions。")

    timeout_text = (
        _resolve_env_value("LLM_TIMEOUT_SECONDS", env_values)
        or _resolve_env_value("VISION_REQUEST_TIMEOUT", env_values)
        or "60"
    )
    try:
        timeout_seconds = max(5, int(timeout_text))
    except Exception:
        timeout_seconds = 60

    provider_name = (
        _resolve_env_value("OPENAI_PROVIDER_NAME", env_values)
        or ("legacy-llm" if use_legacy_surface else "openai-compatible")
    )
    reasoning_effort = _resolve_env_value("OPENAI_REASONING_EFFORT", env_values)

    if not resolved_base_url:
        raise RuntimeError("缺少 LLM base url，请设置 OPENAI_BASE_URL 或 LLM_API_BASE。")
    if not resolved_api_key:
        raise RuntimeError("缺少 LLM API key，请设置 OPENAI_API_KEY 或 LLM_API_KEY。")

    return LlmReviewConfig(
        base_url=resolved_base_url.rstrip("/"),
        api_key=resolved_api_key,
        model=resolved_model,
        wire_api=resolved_wire_api,
        timeout_seconds=timeout_seconds,
        provider_name=provider_name,
        reasoning_effort=reasoning_effort,
    )


def _truncate_text(value: str, limit: int) -> str:
    text = _stringify(value)
    if len(text) <= limit:
        return text
    return f"{text[: max(0, limit - 3)]}..."


def _build_group_prompt_payload(group: dict[str, Any]) -> dict[str, Any]:
    representative_message = dict(group.get("representative_message") or {})
    thread_messages = list(group.get("thread_messages") or [])
    compact_messages = [
        {
            "message_row_id": item.get("message_row_id"),
            "sent_at": item.get("sent_at"),
            "subject": _truncate_text(_stringify(item.get("subject")), 240),
            "snippet": _truncate_text(_stringify(item.get("snippet")), 600),
            "body_text": _truncate_text(_stringify(item.get("body_text")), 1200),
            "direction": item.get("direction"),
        }
        for item in thread_messages[-4:]
    ]

    return {
        "group_key": group.get("group_key"),
        "same_last_mail_row_count": group.get("same_last_mail_row_count"),
        "mass_cc_suspected": group.get("mass_cc_suspected"),
        "matched_contact_emails": list(group.get("matched_contact_emails") or []),
        "instructions": [
            "Judge which creator candidate(s) this email thread truly belongs to.",
            "Only choose from candidate_rows.matched_creator_dedupe_key values.",
            "Shared manager, agency, auto-reply, customer-service, and broad CC situations can be non-exclusive.",
            "Use match_all when every candidate in the group should remain.",
            "Use match_some when only part of the candidate set should remain.",
            "Use reject_group when none of the candidates should remain tied to this email evidence.",
            "Use uncertain when the evidence is too weak and the rows should stay for manual follow-up.",
            "Return one JSON object only.",
        ],
        "representative_message": {
            "message_row_id": representative_message.get("message_row_id"),
            "thread_key": representative_message.get("thread_key"),
            "subject": _truncate_text(_stringify(representative_message.get("subject")), 240),
            "sent_at": representative_message.get("sent_at"),
            "snippet": _truncate_text(_stringify(representative_message.get("snippet")), 900),
            "body_text": _truncate_text(_stringify(representative_message.get("body_text")), 2500),
            "direction": representative_message.get("direction"),
            "thread_message_count": representative_message.get("thread_message_count"),
            "thread_last_sent_at": representative_message.get("thread_last_sent_at"),
        },
        "thread_messages": compact_messages,
        "candidate_rows": [
            {
                "matched_creator_dedupe_key": row.get("creator_dedupe_key"),
                "nickname": row.get("nickname"),
                "username": row.get("@username"),
                "platform": row.get("Platform"),
                "email": row.get("Email"),
                "creator_emails": row.get("creator_emails"),
                "matched_contact_email": row.get("matched_contact_email"),
                "matched_contact_name": row.get("matched_contact_name"),
                "url": row.get("URL"),
                "shared_contact_suspected": row.get("shared_contact_suspected"),
                "manager_suspected": row.get("manager_suspected"),
                "mass_cc_suspected": row.get("mass_cc_suspected"),
            }
            for row in list(group.get("candidate_rows") or [])
        ],
        "output_schema": {
            "decision": "match_all | match_some | reject_group | uncertain",
            "matched_creator_dedupe_keys": ["creator_dedupe_key values from candidate_rows"],
            "sender_role": "creator | manager | agency | auto_reply | support | unclear",
            "confidence": "high | medium | low",
            "reason": "brief justification",
        },
    }


def _extract_response_text(payload: Any) -> str:
    if isinstance(payload, str):
        return payload.strip()
    if not isinstance(payload, dict):
        return str(payload)

    output_text = payload.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    choices = payload.get("choices")
    if isinstance(choices, list):
        for choice in choices:
            message = (choice or {}).get("message") or {}
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                return content.strip()
            if isinstance(content, list):
                parts = []
                for item in content:
                    text_value = (item or {}).get("text")
                    if isinstance(text_value, str) and text_value.strip():
                        parts.append(text_value.strip())
                if parts:
                    return "\n".join(parts)

    output = payload.get("output")
    if isinstance(output, list):
        parts = []
        for item in output:
            for content_item in (item or {}).get("content") or []:
                text_value = content_item.get("text")
                if isinstance(text_value, str) and text_value.strip():
                    parts.append(text_value.strip())
            if parts:
                return "\n".join(parts)

    return json.dumps(payload, ensure_ascii=False)


def _parse_llm_review_response(raw_text: str, valid_keys: set[str]) -> dict[str, Any]:
    cleaned = str(raw_text or "").strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```[a-zA-Z0-9_]*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```$", "", cleaned).strip()

    payload: dict[str, Any] | None = None
    try:
        maybe_payload = json.loads(cleaned)
        if isinstance(maybe_payload, dict):
            payload = maybe_payload
    except json.JSONDecodeError:
        start_idx = cleaned.find("{")
        end_idx = cleaned.rfind("}")
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            snippet = cleaned[start_idx : end_idx + 1]
            try:
                maybe_payload = json.loads(snippet)
                if isinstance(maybe_payload, dict):
                    payload = maybe_payload
            except json.JSONDecodeError:
                payload = None

    if payload is None:
        return {
            "decision": "uncertain",
            "matched_creator_dedupe_keys": [],
            "sender_role": "unclear",
            "confidence": "low",
            "reason": cleaned or "模型未返回可解析 JSON",
            "raw_text": raw_text,
        }

    decision = _stringify(payload.get("decision")).lower()
    if decision not in {"match_all", "match_some", "reject_group", "uncertain"}:
        decision = "uncertain"

    selected = payload.get("matched_creator_dedupe_keys")
    if not isinstance(selected, list):
        selected = []
    normalized: list[str] = []
    seen: set[str] = set()
    for item in selected:
        candidate_key = _stringify(item)
        if not candidate_key or candidate_key not in valid_keys or candidate_key in seen:
            continue
        seen.add(candidate_key)
        normalized.append(candidate_key)

    if decision == "match_all":
        normalized = sorted(valid_keys)
    elif decision == "match_some":
        if not normalized:
            decision = "uncertain"
    elif decision in {"reject_group", "uncertain"}:
        normalized = []

    sender_role = _stringify(payload.get("sender_role")).lower() or "unclear"
    confidence = _stringify(payload.get("confidence")).lower()
    if confidence not in {"high", "medium", "low"}:
        confidence = "medium" if decision in {"match_all", "match_some"} else "low"

    return {
        "decision": decision,
        "matched_creator_dedupe_keys": normalized,
        "sender_role": sender_role,
        "confidence": confidence,
        "reason": _stringify(payload.get("reason")) or cleaned or "模型未返回原因",
        "raw_text": raw_text,
    }


def _build_chat_messages(group: dict[str, Any]) -> list[dict[str, str]]:
    payload = _build_group_prompt_payload(group)
    return [
        {
            "role": "system",
            "content": (
                "You adjudicate duplicate creator-to-email matches at group level. "
                "Return JSON only. Use one of: match_all, match_some, reject_group, uncertain."
            ),
        },
        {
            "role": "user",
            "content": json.dumps(payload, ensure_ascii=False, indent=2),
        },
    ]


def _invoke_llm_review_http(config: LlmReviewConfig, group: dict[str, Any]) -> dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {config.api_key}",
        "Content-Type": "application/json",
    }
    chat_messages = _build_chat_messages(group)
    if config.wire_api == "responses":
        url = f"{config.base_url}/responses"
        body: dict[str, Any] = {
            "model": config.model,
            "input": [
                {
                    "role": message["role"],
                    "content": [{"type": "input_text", "text": message["content"]}],
                }
                for message in chat_messages
            ],
        }
        if config.reasoning_effort:
            body["reasoning"] = {"effort": config.reasoning_effort}
    else:
        url = f"{config.base_url}/chat/completions"
        body = {
            "model": config.model,
            "messages": chat_messages,
            "temperature": 0,
        }

    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=config.timeout_seconds)
            if response.status_code >= 400:
                if response.status_code in RETRYABLE_STATUS_CODES and attempt < 3:
                    time.sleep(min(4.0, 1.5 * attempt))
                    continue
                raise RuntimeError(f"LLM HTTP {response.status_code}: {response.text[:400]}")
            payload = response.json()
            raw_text = _extract_response_text(payload)
            valid_keys = set(group.get("candidate_creator_dedupe_keys") or [])
            parsed = _parse_llm_review_response(raw_text, valid_keys)
            parsed["provider_payload"] = payload
            return parsed
        except requests.exceptions.RequestException as exc:
            last_error = exc
            if attempt < 3:
                time.sleep(min(4.0, 1.5 * attempt))
                continue
            break
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            break

    raise RuntimeError(str(last_error) if last_error else "llm review request failed")


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            text = line.strip()
            if not text:
                continue
            payload = json.loads(text)
            if isinstance(payload, dict):
                records.append(payload)
    return records


def review_candidate_record(config: LlmReviewConfig, candidate_record: dict[str, Any]) -> dict[str, Any]:
    result = _invoke_llm_review_http(config, candidate_record)
    matched_creator_dedupe_keys = list(result.get("matched_creator_dedupe_keys") or [])
    return {
        "group_key": candidate_record.get("group_key"),
        "decision": result.get("decision"),
        "matched_creator_dedupe_keys": matched_creator_dedupe_keys,
        "sender_role": result.get("sender_role"),
        "confidence": result.get("confidence"),
        "reason": result.get("reason"),
        "raw_text": result.get("raw_text"),
        "provider_name": config.provider_name,
        "model": config.model,
        "wire_api": config.wire_api,
        "reviewed_at": _utc_now(),
    }


def run_llm_review(
    *,
    input_prefix: Path,
    env_path: str = ".env",
    base_url: Optional[str] = None,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
    wire_api: Optional[str] = None,
) -> dict[str, Any]:
    candidates_path = input_prefix.with_name(f"{input_prefix.name}_llm_candidates").with_suffix(".jsonl")
    if not candidates_path.exists():
        raise FileNotFoundError(f"找不到 llm candidate 文件: {candidates_path}")

    candidate_records = _load_jsonl(candidates_path)
    config = _resolve_llm_review_config(
        env_path,
        base_url=base_url,
        api_key=api_key,
        model=model,
        wire_api=wire_api,
    )
    review_records = [review_candidate_record(config, record) for record in candidate_records]

    review_path = input_prefix.with_name(f"{input_prefix.name}_llm_review").with_suffix(".jsonl")
    with review_path.open("w", encoding="utf-8") as handle:
        for record in review_records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")

    return {
        "input_prefix": str(input_prefix),
        "llm_candidates_jsonl_path": str(candidates_path),
        "llm_review_jsonl_path": str(review_path),
        "review_group_count": len(review_records),
        "provider_name": config.provider_name,
        "model": config.model,
        "wire_api": config.wire_api,
    }


def _apply_group_decision(row: dict[str, Any], review_record: Optional[dict[str, Any]]) -> tuple[str, str, str]:
    if review_record is None:
        return ("auto_keep", "yes", "")

    decision = _stringify(review_record.get("decision"))
    matched_keys = set(review_record.get("matched_creator_dedupe_keys") or [])
    row_key = _stringify(row.get("creator_dedupe_key"))
    if decision == "match_all":
        return ("reviewed", "yes", decision)
    if decision == "match_some":
        return ("reviewed", "yes" if row_key in matched_keys else "no", decision)
    if decision == "reject_group":
        return ("reviewed", "no", decision)
    if decision == "uncertain":
        return ("reviewed", "yes", decision)
    return ("review_error", "yes", decision)


def apply_llm_review(*, input_prefix: Path) -> dict[str, Any]:
    deduped_path = input_prefix.with_name(f"{input_prefix.name}_去重").with_suffix(".xlsx")
    review_path = input_prefix.with_name(f"{input_prefix.name}_llm_review").with_suffix(".jsonl")
    if not deduped_path.exists():
        raise FileNotFoundError(f"找不到去重 workbook: {deduped_path}")
    if not review_path.exists():
        raise FileNotFoundError(f"找不到 llm review 文件: {review_path}")

    review_records = _load_jsonl(review_path)
    review_by_group_key = {str(record.get("group_key")): record for record in review_records}

    source_headers = _source_headers(deduped_path)
    output_headers = list(source_headers) + [header for header in LLM_REVIEW_HEADERS if header not in source_headers]
    reviewed_rows: list[dict[str, Any]] = []
    keep_rows: list[dict[str, Any]] = []

    for row in _iter_rows(deduped_path, source_headers):
        group_key, _, _ = _build_group_key(row)
        dirty_group = int(row.get("same_last_mail_row_count") or 0) >= 2 or int(row.get("mass_cc_suspected") or 0) >= 1
        review_record = review_by_group_key.get(group_key) if dirty_group and group_key else None
        status, keep, decision = _apply_group_decision(row, review_record)
        annotated = dict(row)
        annotated.update(
            {
                "llm_review_status": status,
                "llm_review_group_key": group_key,
                "llm_review_decision": decision,
                "llm_review_keep": keep,
                "llm_review_matched_creator_dedupe_keys": " | ".join(review_record.get("matched_creator_dedupe_keys", []))
                if review_record
                else "",
                "llm_review_reason": _stringify(review_record.get("reason")) if review_record else "",
                "llm_review_confidence": _stringify(review_record.get("confidence")) if review_record else "",
                "llm_review_sender_role": _stringify(review_record.get("sender_role")) if review_record else "",
                "llm_review_provider": _stringify(review_record.get("provider_name")) if review_record else "",
                "llm_review_model": _stringify(review_record.get("model")) if review_record else "",
                "llm_review_wire_api": _stringify(review_record.get("wire_api")) if review_record else "",
                "llm_review_reviewed_at": _stringify(review_record.get("reviewed_at")) if review_record else "",
            }
        )
        reviewed_rows.append(annotated)
        if keep == "yes":
            keep_rows.append(annotated)

    reviewed_path = input_prefix.with_name(f"{input_prefix.name}_llm_reviewed").with_suffix(".xlsx")
    keep_path = input_prefix.with_name(f"{input_prefix.name}_llm_reviewed_keep").with_suffix(".xlsx")
    _write_workbook(reviewed_path, output_headers, reviewed_rows)
    _write_workbook(keep_path, output_headers, keep_rows)

    return {
        "input_prefix": str(input_prefix),
        "llm_review_jsonl_path": str(review_path),
        "llm_reviewed_xlsx_path": str(reviewed_path),
        "llm_reviewed_keep_xlsx_path": str(keep_path),
        "reviewed_row_count": len(reviewed_rows),
        "keep_row_count": len(keep_rows),
    }


def run_and_apply_llm_review(
    *,
    input_prefix: Path,
    env_path: str = ".env",
    base_url: Optional[str] = None,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
    wire_api: Optional[str] = None,
) -> dict[str, Any]:
    review_summary = run_llm_review(
        input_prefix=input_prefix,
        env_path=env_path,
        base_url=base_url,
        api_key=api_key,
        model=model,
        wire_api=wire_api,
    )
    apply_summary = apply_llm_review(input_prefix=input_prefix)
    return {
        **review_summary,
        **apply_summary,
    }
