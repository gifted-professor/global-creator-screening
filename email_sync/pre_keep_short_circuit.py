from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook, load_workbook

from .known_thread_update import process_known_thread_updates
from .thread_assignments import lookup_thread_assignment


_HANDLE_URL_PATTERN = re.compile(r"(?:@|/)([A-Za-z0-9._-]{2,80})/?$")


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_platform(value: Any) -> str:
    text = _clean_text(value).casefold()
    if text in {"ig", "ins", "instagram"}:
        return "instagram"
    if text in {"tt", "tiktok", "douyin"}:
        return "tiktok"
    if text in {"yt", "youtube"}:
        return "youtube"
    return text


def _extract_creator_id(row: Mapping[str, Any]) -> str:
    for value in (
        row.get("final_id_final"),
        row.get("@username"),
        row.get("达人ID"),
        row.get("creator_id"),
        row.get("identifier"),
        row.get("URL"),
        row.get("主页链接"),
    ):
        text = _clean_text(value)
        if not text:
            continue
        match = _HANDLE_URL_PATTERN.search(text)
        if match:
            return _clean_text(match.group(1)).lstrip("@").casefold()
        return text.lstrip("@").casefold()
    return ""


def _first_non_blank(*values: Any) -> str:
    for value in values:
        cleaned = _clean_text(value)
        if cleaned:
            return cleaned
    return ""


def _iter_workbook_rows(workbook_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not raw_rows:
        return [], []
    headers = [_clean_text(item) for item in raw_rows[0]]
    rows: list[dict[str, Any]] = []
    for raw_row in raw_rows[1:]:
        row: dict[str, Any] = {}
        has_value = False
        for header, value in zip(headers, raw_row):
            if not header:
                continue
            row[header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            rows.append(row)
    return headers, rows


def _write_workbook_rows(workbook_path: Path, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(headers))
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)


def route_pre_keep_workbook(
    *,
    keep_workbook_path: str | Path,
    routed_keep_workbook_path: str | Path,
    mail_only_workbook_path: str | Path,
    db_path: str | Path,
    owner_scope: str,
    brand: str,
    existing_index: Mapping[str, dict[str, Any]],
    owner_scope_enabled: bool,
) -> dict[str, Any]:
    source_path = Path(str(keep_workbook_path)).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"找不到 keep workbook: {source_path}")

    headers, rows = _iter_workbook_rows(source_path)
    prepared_candidates: list[dict[str, Any]] = []
    thread_assignment_cache_hit_count = 0
    for row in rows:
        updated_row = dict(row)
        creator_id = _extract_creator_id(updated_row)
        platform = _normalize_platform(_first_non_blank(updated_row.get("Platform"), updated_row.get("platform")))
        assignment: dict[str, Any] = {}
        if _clean_text(owner_scope) and creator_id and platform:
            assignment = dict(
                lookup_thread_assignment(
                    db_path=Path(str(db_path)).expanduser().resolve(),
                    owner_scope=owner_scope,
                    creator_id=creator_id,
                    platform=platform,
                    brand=brand,
                matched_contact_email=_first_non_blank(updated_row.get("matched_contact_email"), updated_row.get("matched_email")),
                subject=_first_non_blank(
                    updated_row.get("subject"),
                    updated_row.get("brand_message_subject"),
                        updated_row.get("evidence_subject"),
                        updated_row.get("last_mail_subject"),
                    ),
                )
                or {}
            )
        resolution = {"status": "cache_miss"}
        if assignment:
            resolution = {
                "status": "cache_hit",
                "thread_key": _clean_text(assignment.get("thread_key")),
                "match_reason": _clean_text(assignment.get("match_reason")),
            }
            thread_assignment_cache_hit_count += 1
            updated_row["evidence_thread_key"] = _clean_text(assignment.get("thread_key"))
            if not _clean_text(updated_row.get("matched_contact_email")) and _clean_text(assignment.get("matched_contact_email")):
                updated_row["matched_contact_email"] = _clean_text(assignment.get("matched_contact_email"))
            if not _clean_text(updated_row.get("last_mail_message_id")) and _clean_text(assignment.get("last_mail_message_id")):
                updated_row["last_mail_message_id"] = _clean_text(assignment.get("last_mail_message_id"))
            if not _clean_text(updated_row.get("last_mail_time")) and _clean_text(assignment.get("last_mail_sent_at")):
                updated_row["last_mail_time"] = _clean_text(assignment.get("last_mail_sent_at"))
            if not _clean_text(updated_row.get("last_mail_subject")) and _clean_text(assignment.get("normalized_subject")):
                updated_row["last_mail_subject"] = _clean_text(assignment.get("normalized_subject"))
            if assignment.get("mail_update_revision") not in (None, ""):
                updated_row["mail_update_revision"] = assignment.get("mail_update_revision")
        prepared_candidates.append(
            {
                "keep_row": updated_row,
                "owner_scope": owner_scope,
                "creator_id": creator_id,
                "platform": platform,
                "thread_key": _clean_text(updated_row.get("evidence_thread_key")) or _clean_text(assignment.get("thread_key")),
                "thread_assignment_resolution": resolution,
            }
        )

    routing = process_known_thread_updates(
        prepared_candidates,
        existing_index=existing_index,
        owner_scope_enabled=bool(owner_scope_enabled),
    )
    stats = dict(routing.get("stats") or {})
    stats["thread_assignment_cache_hit_count"] = thread_assignment_cache_hit_count
    stats["mail_only_count"] = len(routing.get("mail_only_candidates") or [])
    stats["partial_refresh_count"] = 0
    stats["full_screening_count"] = len(routing.get("full_screening_candidates") or [])
    stats["input_row_count"] = len(rows)

    mail_only_rows = [dict((candidate or {}).get("keep_row") or {}) for candidate in (routing.get("mail_only_candidates") or [])]
    heavy_rows = [dict((candidate or {}).get("keep_row") or {}) for candidate in (routing.get("full_screening_candidates") or [])]

    routed_keep_path = Path(str(routed_keep_workbook_path)).expanduser().resolve()
    mail_only_path = Path(str(mail_only_workbook_path)).expanduser().resolve()
    _write_workbook_rows(routed_keep_path, headers, heavy_rows)
    _write_workbook_rows(mail_only_path, headers, mail_only_rows)

    return {
        "routed_keep_workbook": str(routed_keep_path),
        "mail_only_workbook": str(mail_only_path),
        "stats": stats,
    }
