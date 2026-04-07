from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.final_export_merge import _build_quote_text, _clean_text, _format_date  # type: ignore


UPLOAD_COLUMNS = [
    "达人ID",
    "平台",
    "主页链接",
    "当前网红报价",
    "达人最后一次回复邮件时间",
    "full body",
]

WORKBOOK_COLUMNS = UPLOAD_COLUMNS + [
    "latest_external_from",
    "subject",
    "resolution_stage_final",
    "resolution_confidence_final",
    "thread_key",
    "raw_path",
    "brand_keyword",
    "原达人ID候选",
]


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a manual-pool mail-only workbook/payload from funnel manual-tail workbook.")
    parser.add_argument("--manual-tail-workbook", required=True, help="Manual-tail workbook path.")
    parser.add_argument(
        "--task-owner-payload-json",
        required=True,
        help="Existing payload JSON path used to source task_owner metadata.",
    )
    parser.add_argument("--task-name", required=True, help="Task name, for example SKG.")
    parser.add_argument(
        "--local-date",
        required=True,
        help="Local date in YYYY-MM-DD. Used to build synthetic manual IDs.",
    )
    parser.add_argument("--output-prefix", required=True, help="Output prefix.")
    return parser.parse_args()


def _load_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))
        return headers, [dict(zip(headers, row)) for row in rows]
    finally:
        wb.close()


def _build_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(WORKBOOK_COLUMNS)
    for row in rows:
        ws.append([row.get(column, "") for column in WORKBOOK_COLUMNS])
    wb.save(path)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _build_manual_creator_id(task_name: str, local_day: date, index: int) -> str:
    return f"{task_name}{local_day.month}/{local_day.day}转人工{index}"


def _build_profile_url(_creator_id: str, platform: str) -> str:
    if _clean_text(platform) == "转人工":
        return ""
    return ""


def run(args: argparse.Namespace) -> dict[str, Any]:
    manual_tail_workbook = Path(args.manual_tail_workbook).expanduser().resolve()
    task_owner_payload_json = Path(args.task_owner_payload_json).expanduser().resolve()
    output_prefix = Path(args.output_prefix).expanduser().resolve()
    local_day = date.fromisoformat(str(args.local_date).strip())
    task_name = _clean_text(args.task_name)

    _headers, source_rows = _load_rows(manual_tail_workbook)
    task_owner_payload = json.loads(task_owner_payload_json.read_text(encoding="utf-8"))
    task_owner = dict(task_owner_payload.get("task_owner") or {})

    workbook_rows: list[dict[str, Any]] = []
    payload_rows: list[dict[str, Any]] = []
    quote_count = 0

    for index, row in enumerate(source_rows, start=1):
        full_body = _clean_text(row.get("latest_external_full_body"))
        quote_text = _build_quote_text(row, mail_body_text=full_body)
        if quote_text:
            quote_count += 1
        last_mail_time = _format_date(row.get("latest_external_sent_at"))
        creator_id = _build_manual_creator_id(task_name, local_day, index)

        workbook_row = {
            "达人ID": creator_id,
            "平台": "转人工",
            "主页链接": _build_profile_url(creator_id, "转人工"),
            "当前网红报价": quote_text,
            "达人最后一次回复邮件时间": last_mail_time,
            "full body": full_body,
            "latest_external_from": _clean_text(row.get("latest_external_from")),
            "subject": _clean_text(row.get("subject")),
            "resolution_stage_final": _clean_text(row.get("resolution_stage_final")),
            "resolution_confidence_final": _clean_text(row.get("resolution_confidence_final")),
            "thread_key": _clean_text(row.get("thread_key")),
            "raw_path": _clean_text(row.get("raw_path")),
            "brand_keyword": _clean_text(row.get("brand_keyword")) or task_name,
            "原达人ID候选": _clean_text(row.get("final_id_final")),
        }
        workbook_rows.append(workbook_row)

        payload_row = {column: workbook_row.get(column, "") for column in UPLOAD_COLUMNS}
        payload_row["__feishu_update_mode"] = "create_or_mail_only_update"
        payload_rows.append(payload_row)

    workbook_path = output_prefix.with_name(f"{output_prefix.name}_with_mail_fields").with_suffix(".xlsx")
    payload_path = output_prefix.with_name(f"{output_prefix.name}_upload_payload").with_suffix(".json")
    summary_path = output_prefix.with_name(f"{output_prefix.name}_build_summary").with_suffix(".json")

    payload = {
        "task_owner": task_owner,
        "columns": list(UPLOAD_COLUMNS),
        "source_row_count": len(source_rows),
        "row_count": len(payload_rows),
        "skipped_row_count": 0,
        "rows": payload_rows,
        "skipped_rows": [],
    }
    summary = {
        "manual_tail_workbook": str(manual_tail_workbook),
        "task_owner_payload_json": str(task_owner_payload_json),
        "task_name": task_name,
        "local_date": local_day.isoformat(),
        "source_row_count": len(source_rows),
        "output_row_count": len(payload_rows),
        "quote_count": quote_count,
        "manual_platform_value": "转人工",
        "manual_id_pattern_preview": _build_manual_creator_id(task_name, local_day, 1),
        "workbook_path": str(workbook_path),
        "payload_path": str(payload_path),
    }

    _build_workbook(workbook_path, workbook_rows)
    _write_json(payload_path, payload)
    _write_json(summary_path, summary)

    summary["summary_path"] = str(summary_path)
    return summary


def main() -> int:
    args = _parse_args()
    result = run(args)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
