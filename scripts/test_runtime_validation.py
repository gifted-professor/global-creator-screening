import json
import os
import sys
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


os.environ.setdefault("SCREENING_DATA_DIR", tempfile.mkdtemp(prefix="screening-data-"))

import backend.app as backend_app


def make_upload_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Upload"
    sheet.append(["Platform", "@username", "nickname", "Region", "Language", "Followers", "URL"])
    sheet.append(["Instagram", "@creatoralpha", "Alpha", "US", "en", 100000, ""])
    sheet.append(["TikTok", "@creatorbeta", "Beta", "US", "en", 200000, ""])
    sheet.append(["YouTube", "@creatorgamma", "Gamma", "US", "en", 300000, ""])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def workbook_headers(workbook_bytes, sheet_name=None):
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    sheet = workbook[sheet_name or workbook.sheetnames[0]]
    return [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]


class DummyApifyResponse:
    def __init__(self, payload, status_code=200, text=""):
        self._payload = payload
        self.status_code = status_code
        self.text = text

    def json(self):
        return self._payload


def main():
    backend_app.ensure_runtime_dirs()
    client = backend_app.app.test_client()

    upload_buffer = make_upload_workbook()
    response = client.post(
        "/api/upload",
        data={"file": (upload_buffer, "sample_upload.xlsx")},
        content_type="multipart/form-data",
    )
    payload = response.get_json()
    assert response.status_code == 200, payload
    assert payload["success"] is True, payload
    assert payload["stats"]["Instagram"] == 1, payload
    assert payload["stats"]["TikTok"] == 1, payload
    assert payload["stats"]["YouTube"] == 1, payload

    compiled = client.post(
        "/api/rulespec/compile",
        json={"sop_text": "Instagram 仅保留美国；最近30天有更新；TikTok 平均播放量 10000；中位数 8000；视觉看 9 张封面。"},
    )
    compiled_payload = compiled.get_json()
    assert compiled.status_code == 200, compiled_payload
    assert compiled_payload["success"] is True, compiled_payload
    assert compiled_payload["rule_spec"]["platform_overrides"]["instagram"]["allowed_regions"] == ["US"], compiled_payload
    assert compiled_payload["rule_spec"]["platform_overrides"]["shared"]["active_days_max"] == 30, compiled_payload

    instagram_raw = [
        {
            "username": "creatoralpha",
            "url": "https://www.instagram.com/creatoralpha/",
            "biography": "US creator",
            "latestPosts": [
                {
                    "timestamp": "2026-03-20T00:00:00Z",
                    "displayUrl": "https://example.com/ig-1.jpg",
                }
            ],
        }
    ]
    tiktok_raw = [
        {
            "authorMeta": {
                "name": "creatorbeta",
                "profileUrl": "https://www.tiktok.com/@creatorbeta",
            },
            "createTimeISO": "2026-03-24T00:00:00Z",
            "playCount": 20000,
            "videoMeta": {"coverUrl": "https://example.com/tt-1.jpg"},
        },
        {
            "authorMeta": {
                "name": "creatorbeta",
                "profileUrl": "https://www.tiktok.com/@creatorbeta",
            },
            "createTimeISO": "2026-03-21T00:00:00Z",
            "playCount": 15000,
            "videoMeta": {"coverUrl": "https://example.com/tt-2.jpg"},
        },
    ]
    youtube_raw = [
        {
            "channelName": "creatorgamma",
            "channelUrl": "https://www.youtube.com/@creatorgamma",
            "date": "2026-03-23T00:00:00Z",
            "thumbnailUrl": "https://example.com/yt-1.jpg",
            "isPaidContent": False,
        }
    ]

    backend_app.write_json_file(backend_app.get_raw_data_path("instagram"), instagram_raw)
    backend_app.write_json_file(backend_app.get_raw_data_path("tiktok"), tiktok_raw)
    backend_app.write_json_file(backend_app.get_raw_data_path("youtube"), youtube_raw)

    active_rulespec = backend_app.load_active_rulespec()
    for platform, raw_items, expected in (
        ("instagram", instagram_raw, ["https://www.instagram.com/creatoralpha/"]),
        ("tiktok", tiktok_raw, ["https://www.tiktok.com/@creatorbeta"]),
        ("youtube", youtube_raw, ["https://www.youtube.com/@creatorgamma"]),
    ):
        filtered = backend_app.screening.filter_scraped_items(
            platform,
            raw_items,
            expected_profiles=expected,
            upload_metadata_lookup=backend_app.load_upload_metadata(platform),
            active_rulespec=active_rulespec,
        )
        backend_app.save_profile_reviews(platform, filtered["profile_reviews"])
        assert filtered["profile_reviews"], filtered

    for platform in ("instagram", "tiktok", "youtube"):
        prescreen = client.get(f"/api/download/{platform}/prescreen-review")
        assert prescreen.status_code == 200, prescreen.get_json()
        image_review = client.get(f"/api/download/{platform}/image-review")
        assert image_review.status_code == 200, image_review.get_json()
        test_info = client.get(f"/api/download/{platform}/test-info")
        assert test_info.status_code == 200, test_info.get_json()
        test_info_json = client.get(f"/api/download/{platform}/test-info-json")
        assert test_info_json.status_code == 200, test_info_json.get_json()
        final_review = client.post(f"/api/download/{platform}/final-review", json={})
        assert final_review.status_code == 200, final_review.get_json()

    health = client.get("/api/health")
    health_payload = health.get_json()
    assert health.status_code == 200, health_payload
    assert health_payload["status"] == "ok", health_payload

    dashboard = client.get("/apify/balance")
    assert dashboard.status_code == 200, dashboard.status_code
    dashboard_html = dashboard.get_data(as_text=True)
    assert "Apify 余额控制台" in dashboard_html, dashboard_html[:200]
    assert "/api/apify/balance" in dashboard_html, dashboard_html[:200]

    original_home = os.environ.get("HOME")
    original_token_env = {
        "APIFY_TOKEN": os.environ.get("APIFY_TOKEN"),
        "APIFY_API_TOKEN": os.environ.get("APIFY_API_TOKEN"),
        "APIFY_BACKUP_TOKENS": os.environ.get("APIFY_BACKUP_TOKENS"),
        "APIFY_FREE_TOKENS": os.environ.get("APIFY_FREE_TOKENS"),
    }
    auth_home = Path(tempfile.mkdtemp(prefix="apify-home-"))
    auth_dir = auth_home / ".apify"
    auth_dir.mkdir(parents=True, exist_ok=True)
    (auth_dir / "auth.json").write_text(json.dumps({"token": "tok-auth-9999"}), encoding="utf-8")

    try:
        os.environ["HOME"] = str(auth_home)
        for key in original_token_env:
            os.environ.pop(key, None)
        assert backend_app.get_apify_token_pool() == ["tok-auth-9999"]
    finally:
        if original_home is None:
            os.environ.pop("HOME", None)
        else:
            os.environ["HOME"] = original_home
        for key, value in original_token_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value

    os.environ["APIFY_TOKEN"] = "tok-primary-1234"
    os.environ["APIFY_BACKUP_TOKENS"] = "tok-backup-5678"
    original_apify_request = backend_app.apify_request

    def fake_apify_request(method, url, *, token, params=None, json_payload=None):
        assert method == "GET", method
        assert url.endswith("/users/me/limits"), url
        if token == "tok-primary-1234":
            return DummyApifyResponse({
                "data": {
                    "limits": {"maxMonthlyUsageUsd": 25},
                    "current": {"monthlyUsageUsd": 6.5},
                    "monthlyUsageCycle": {
                        "startedAt": "2026-03-01T00:00:00.000Z",
                        "endsAt": "2026-03-31T23:59:59.000Z",
                    },
                }
            })
        if token == "tok-backup-5678":
            return DummyApifyResponse({
                "data": {
                    "limits": {"maxMonthlyUsageUsd": 10},
                    "current": {"monthlyUsageUsd": 1},
                    "monthlyUsageCycle": {
                        "startedAt": "2026-03-01T00:00:00.000Z",
                        "endsAt": "2026-03-31T23:59:59.000Z",
                    },
                }
            })
        return DummyApifyResponse({"error": {"message": "unknown token"}}, status_code=401)

    try:
        backend_app.apify_request = fake_apify_request
        live_balance = client.post("/api/apify/balance/refresh")
        balance = client.get("/api/apify/balance")
        dashboard_filled = client.get("/apify/balance")
    finally:
        backend_app.apify_request = original_apify_request

    live_balance_payload = live_balance.get_json()
    assert live_balance.status_code == 200, live_balance_payload
    assert live_balance_payload["data_source"] == "live", live_balance_payload

    balance_payload = balance.get_json()
    assert balance.status_code == 200, balance_payload
    assert balance_payload["success"] is True, balance_payload
    assert balance_payload["data_source"] == "cache", balance_payload
    assert balance_payload["token_pool_size"] == 2, balance_payload
    assert balance_payload["summary"]["max_monthly_usage_usd_total"] == 35.0, balance_payload
    assert balance_payload["summary"]["monthly_usage_usd_total"] == 7.5, balance_payload
    assert balance_payload["summary"]["remaining_monthly_usage_usd_total"] == 27.5, balance_payload
    assert len(balance_payload["tokens"]) == 2, balance_payload
    assert balance_payload["refresh_interval_seconds"] == backend_app.APIFY_BALANCE_REFRESH_INTERVAL_SECONDS, balance_payload

    token_state = backend_app.load_apify_token_state()
    assert "tokens" in token_state, token_state
    assert "tok-primary-1234" in token_state["tokens"], token_state
    cache_state = backend_app.load_apify_balance_cache()
    assert cache_state.get("payload", {}).get("success") is True, cache_state
    assert cache_state.get("status_code") == 200, cache_state

    dashboard_filled_html = dashboard_filled.get_data(as_text=True)
    assert dashboard_filled.status_code == 200, dashboard_filled.status_code
    assert "35.000000 USD" in dashboard_filled_html, dashboard_filled_html[:400]
    assert "tok-...1234" in dashboard_filled_html, dashboard_filled_html[:400]
    assert "tok-...5678" in dashboard_filled_html, dashboard_filled_html[:400]
    assert "/api/apify/balance/refresh" in dashboard_filled_html, dashboard_filled_html[:400]
    assert "后端轮询" in dashboard_filled_html, dashboard_filled_html[:400]

    try:
        backend_app.apify_request = fake_apify_request
        budget_pick = backend_app.select_apify_token_for_batch(
            ["tok-primary-1234", "tok-backup-5678"],
            required_budget_usd=12.0,
        )
        budget_miss = backend_app.select_apify_token_for_batch(
            ["tok-primary-1234", "tok-backup-5678"],
            required_budget_usd=30.0,
        )
    finally:
        backend_app.apify_request = original_apify_request

    assert budget_pick["selected_token"] == "tok-primary-1234", budget_pick
    assert budget_miss["selected_token"] == "", budget_miss
    assert budget_miss["error_code"] == "APIFY_BUDGET_INSUFFICIENT", budget_miss

    post_calls = []

    def fake_apify_start(method, url, *, token, params=None, json_payload=None):
        assert method == "POST", method
        post_calls.append({"url": url, "token": token, "input": json_payload})
        return DummyApifyResponse(
            {
                "data": {
                    "id": "run-123",
                    "defaultDatasetId": "dataset-123",
                    "status": "RUNNING",
                }
            },
            status_code=201,
        )

    first_start = None
    second_start = None
    try:
        backend_app.apify_request = fake_apify_start
        backend_app.save_apify_run_guards({})
        first_start = backend_app.start_apify_run_guarded(
            "clockworks/tiktok-profile-scraper",
            {"profiles": ["creatorbeta"], "resultsPerPage": 20},
            "tok-primary-1234",
        )
        second_start = backend_app.start_apify_run_guarded(
            "clockworks/tiktok-profile-scraper",
            {"profiles": ["creatorbeta"], "resultsPerPage": 20},
            "tok-primary-1234",
        )
    finally:
        backend_app.apify_request = original_apify_request
        if first_start:
            backend_app.clear_apify_run_guard(first_start["guard_key"])

    assert len(post_calls) == 1, post_calls
    assert first_start["reused_guard"] is False, first_start
    assert second_start["reused_guard"] is True, second_start

    print("Runtime validation checks passed.")


if __name__ == "__main__":
    main()
