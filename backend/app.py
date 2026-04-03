from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import threading
import time
import uuid
import base64
import hashlib
import random
import weakref
from contextlib import contextmanager
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, as_completed, wait
from concurrent.futures.thread import _threads_queues, _worker
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import pandas as pd
import requests
from flask import Flask, jsonify, render_template, request, send_file
from backend import creator_cache
from harness.config import (
    load_env_file_snapshot,
    resolve_operator_launch_config,
    resolve_operator_task_candidates_config,
)
from harness.paths import resolve_operator_run_paths
from harness.spec import build_operator_task_spec, write_task_spec
from openpyxl import Workbook
from werkzeug.utils import secure_filename

try:
    from flask_cors import CORS
except Exception:  # pragma: no cover - optional dependency for local runtime only
    CORS = None

from backend import rules as rules_module
from backend import screening
from backend.final_export_merge import build_all_platforms_final_review_artifacts, extract_task_owner_context
from feishu_screening_bridge.feishu_api import DEFAULT_FEISHU_BASE_URL, FeishuOpenClient
from feishu_screening_bridge.task_upload_sync import inspect_task_upload_assignments


BASE_DIR = Path(__file__).resolve().parents[1]
DOTENV_LOCAL_PATH = BASE_DIR / ".env.local"


def parse_dotenv_file(path):
    parsed = {}
    if not path.exists():
        return parsed
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            parsed[key] = value
    return parsed


ENV_KEYS_BEFORE_DOTENV_LOCAL = set(os.environ.keys())
DOTENV_LOCAL_VALUES = parse_dotenv_file(DOTENV_LOCAL_PATH)
DOTENV_LOCAL_LOADED_KEYS = set()


def load_dotenv_local():
    if not DOTENV_LOCAL_PATH.exists():
        return
    for key, value in DOTENV_LOCAL_VALUES.items():
        if key and key not in os.environ:
            os.environ[key] = value
            DOTENV_LOCAL_LOADED_KEYS.add(key)


load_dotenv_local()


class DaemonThreadPoolExecutor(ThreadPoolExecutor):
    """Allow timed-out worker calls to stop blocking process exit."""

    def _adjust_thread_count(self):
        if self._idle_semaphore.acquire(timeout=0):
            return

        def weakref_cb(_, q=self._work_queue):
            q.put(None)

        num_threads = len(self._threads)
        if num_threads >= self._max_workers:
            return
        thread_name = "%s_%d" % (self._thread_name_prefix or self, num_threads)
        worker_thread = threading.Thread(
            name=thread_name,
            target=_worker,
            args=(weakref.ref(self, weakref_cb), self._work_queue, self._initializer, self._initargs),
        )
        worker_thread.daemon = True
        worker_thread.start()
        self._threads.add(worker_thread)
        _threads_queues[worker_thread] = self._work_queue


def parse_env_flag(name, default=False):
    raw_value = os.getenv(name)
    if raw_value is None:
        return bool(default)
    normalized = str(raw_value).strip().lower()
    if not normalized:
        return bool(default)
    return normalized not in {"0", "false", "no", "off"}


DATA_DIR = os.environ.get("SCREENING_DATA_DIR", str(BASE_DIR / "data"))
CONFIG_DIR = str(BASE_DIR / "config")
TEMP_DIR = str(BASE_DIR / "temp")
UPLOAD_FOLDER = str(Path(DATA_DIR) / "uploads")
ACTIVE_RULESPEC_PATH = str(Path(CONFIG_DIR) / "active_rulespec.json")
ACTIVE_VISUAL_PROMPTS_PATH = str(Path(CONFIG_DIR) / "active_visual_prompts.json")
FIELD_MATCH_REPORT_PATH = str(Path(CONFIG_DIR) / "field_match_report.json")
MISSING_CAPABILITIES_PATH = str(Path(CONFIG_DIR) / "missing_capabilities.json")
REVIEW_NOTES_PATH = str(Path(CONFIG_DIR) / "review_notes.md")
APIFY_API_BASE = "https://api.apify.com/v2"
APIFY_POLL_INTERVAL_SECONDS = int(os.getenv("APIFY_POLL_INTERVAL_SECONDS", "5"))
APIFY_REQUEST_TIMEOUT = int(os.getenv("APIFY_REQUEST_TIMEOUT", "60"))
APIFY_TRANSPORT_MAX_RETRIES = max(1, int(os.getenv("APIFY_TRANSPORT_MAX_RETRIES", "3")))
APIFY_TRANSPORT_RETRY_BACKOFF_SECONDS = max(
    0.5,
    float(os.getenv("APIFY_TRANSPORT_RETRY_BACKOFF_SECONDS", "1.5")),
)
APIFY_TOKEN_POOL_STATE_FILE = str(Path(DATA_DIR) / "apify_token_pool_state.json")
APIFY_BALANCE_CACHE_FILE = str(Path(DATA_DIR) / "apify_balance_cache.json")
APIFY_RUN_GUARDS_FILE = str(Path(DATA_DIR) / "apify_run_guards.json")
APIFY_BALANCE_REFRESH_INTERVAL_SECONDS = max(
    60,
    int(os.getenv("APIFY_BALANCE_REFRESH_INTERVAL_SECONDS", "300")),
)
APIFY_BALANCE_POLLER_ENABLED = parse_env_flag("APIFY_BALANCE_POLLER_ENABLED", default=True)
APIFY_GUARD_TTL_SECONDS = max(60, int(os.getenv("APIFY_GUARD_TTL_SECONDS", "1800")))
APIFY_BUDGET_SAFETY_MULTIPLIER = max(
    1.0,
    float(os.getenv("APIFY_BUDGET_SAFETY_MULTIPLIER", "1.1")),
)
APIFY_BUDGET_BUFFER_USD = max(
    0.0,
    float(os.getenv("APIFY_BUDGET_BUFFER_USD", "0.1")),
)
UPSTREAM_HTTP_STATUS_PUBLIC_LABELS = {
    400: "请求格式无效",
    401: "认证失败",
    403: "拒绝访问",
    404: "接口不存在",
    408: "请求超时",
    413: "请求体过大",
    415: "请求格式不受支持",
    422: "请求内容无效",
    429: "请求过于频繁",
    500: "上游内部错误",
    502: "上游网关错误",
    503: "上游服务暂不可用",
    504: "上游超时",
}
BACKEND_BIND_HOST = os.getenv("BACKEND_BIND_HOST", "127.0.0.1")
BACKEND_PORT = int(os.getenv("BACKEND_PORT", "5001"))
BACKEND_DEBUG = parse_env_flag("BACKEND_DEBUG", default=False)
BACKEND_USE_RELOADER = parse_env_flag("BACKEND_USE_RELOADER", default=BACKEND_DEBUG)
BACKEND_ALLOWED_ORIGINS = [
    item.strip()
    for item in (os.getenv("BACKEND_ALLOWED_ORIGINS") or "http://127.0.0.1:5173,http://localhost:5173").split(",")
    if item.strip()
]
JOB_TERMINAL_STATUSES = {"completed", "failed", "cancelled"}
PLATFORM_ACTORS = {
    "tiktok": "clockworks/tiktok-profile-scraper",
    "instagram": "apify/instagram-profile-scraper",
    "youtube": "streamers/youtube-scraper",
}
PLATFORM_BATCH_SIZES = {
    "tiktok": int(os.getenv("TIKTOK_BATCH_SIZE", "20")),
    "instagram": int(os.getenv("INSTAGRAM_BATCH_SIZE", "50")),
    "youtube": int(os.getenv("YOUTUBE_BATCH_SIZE", "5")),
}


def resolve_missing_retry_batch_size(platform):
    default_batch_size = max(1, int(PLATFORM_BATCH_SIZES.get(platform, 20) or 20))
    env_key = f"{str(platform or '').strip().upper()}_MISSING_RETRY_BATCH_SIZE"
    raw_value = str(os.getenv(env_key, "") or "").strip()
    if not raw_value:
        return min(default_batch_size, 10)
    try:
        parsed_value = int(raw_value)
    except Exception:
        return min(default_batch_size, 10)
    return max(1, min(default_batch_size, parsed_value))
DEFAULT_SCRAPE_MISSING_RETRY_ATTEMPTS = max(0, int(os.getenv("SCRAPE_MISSING_RETRY_ATTEMPTS", "1")))
PLATFORM_ESTIMATED_COST_PER_IDENTIFIER_USD = {
    "tiktok": float(os.getenv("APIFY_TIKTOK_COST_PER_PROFILE_USD", "0.0")),
    "instagram": float(os.getenv("APIFY_INSTAGRAM_COST_PER_PROFILE_USD", "0.0026")),
    "youtube": float(os.getenv("APIFY_YOUTUBE_COST_PER_PROFILE_USD", "0.0")),
}
PLATFORM_ESTIMATED_COST_PER_RESULT_USD = {
    "tiktok": float(os.getenv("APIFY_TIKTOK_COST_PER_RESULT_USD", "0.004")),
    "instagram": float(os.getenv("APIFY_INSTAGRAM_COST_PER_RESULT_USD", "0.0")),
    "youtube": float(os.getenv("APIFY_YOUTUBE_COST_PER_RESULT_USD", "0.004")),
}
INSTAGRAM_ABOUT_ADDON_COST_PER_PROFILE_USD = float(
    os.getenv("APIFY_INSTAGRAM_ABOUT_ADDON_COST_PER_PROFILE_USD", "0.007")
)
TRANSIENT_STATUS_CODES = {429, 500, 502, 503, 504}
UPLOAD_PLATFORM_ALIASES = {
    "tiktok": "tiktok",
    "tik_tok": "tiktok",
    "instagram": "instagram",
    "ig": "instagram",
    "youtube": "youtube",
    "yt": "youtube",
}
UPLOAD_PLATFORM_RESPONSE_LABELS = {
    "tiktok": "TikTok",
    "instagram": "Instagram",
    "youtube": "YouTube",
}
UPLOAD_METADATA_FIELD_ALIASES = {
    "nickname": "nickname",
    "description": "description",
    "region": "region",
    "language": "language",
    "followers": "followers",
    "avgviews": "avg_views",
    "avglikes": "avg_likes",
    "avgcomments": "avg_comments",
    "avgcollects": "avg_collects",
    "tags": "tags",
    "email": "email",
    "url": "url",
    "tiktokurl": "tiktok_url",
    "instagramurl": "instagram_url",
    "youtubeurl": "youtube_url",
    "platformattemptorder": "platform_attempt_order",
    "mailthreadkey": "mail_thread_key",
    "mailresolutionstage": "mail_resolution_stage",
    "mailresolutionconfidence": "mail_resolution_confidence",
    "mailapifygate": "mail_apify_gate",
    "mailevidence": "mail_evidence",
    "mailrawpath": "mail_raw_path",
    "latestexternalfrom": "latest_external_from",
    "latestexternalsentat": "latest_external_sent_at",
    "subject": "subject",
}
UPLOAD_METADATA_EXPORT_FIELDS = (
    ("upload_nickname", "nickname"),
    ("upload_handle", "handle"),
    ("upload_region", "region"),
    ("upload_language", "language"),
    ("upload_followers", "followers"),
    ("upload_avg_views", "avg_views"),
    ("upload_avg_likes", "avg_likes"),
    ("upload_avg_comments", "avg_comments"),
    ("upload_avg_collects", "avg_collects"),
)
VISION_REQUEST_TIMEOUT = int(os.getenv("VISION_REQUEST_TIMEOUT", "60"))
DEFAULT_VISION_MODEL = "gpt-5.4-mini"
VISION_MODEL = os.getenv("VISION_MODEL", DEFAULT_VISION_MODEL)
DEFAULT_QIANDAO_VISION_MODEL = "gemini-2.5-pro-preview-p"
DEFAULT_QIANDAO_FALLBACK_VISION_MODEL = "gemini-3-flash-preview-S"
QIANDAO_25P_VISION_MODEL = "gemini-2.5-pro-preview-p"
DEFAULT_QIANDAO_25P_VISUAL_REVIEW_MAX_WORKERS = 2
MAX_QIANDAO_25P_VISUAL_REVIEW_MAX_WORKERS = 3
DEFAULT_QIANDAO_MAX_TOKENS = 900
DEFAULT_QIANDAO_TEMPERATURE = 0.2
DEFAULT_MIMO_VISION_MODEL = "mimo-v2-omni"
DEFAULT_MIMO_MAX_COMPLETION_TOKENS = 2048
DEFAULT_QWEN_VISION_MODEL = "qwen-vl-max"
DEFAULT_REELX_BASE_URL = "https://llmxapi.com/v1beta"
DEFAULT_REELX_BASE_URL_FALLBACKS = (
    "https://reelxai.com/v1beta",
    "https://hk.llmxapi.com/v1beta",
    "https://hk.reelxai.com/v1beta",
)
VISION_API_STYLE_RESPONSES = "responses"
VISION_API_STYLE_CHAT_COMPLETIONS = "chat_completions"
VISION_API_STYLE_GENERATE_CONTENT = "generate_content"
VISUAL_REVIEW_REQUEST_COVER_LIMIT = max(1, int(os.getenv("VISUAL_REVIEW_REQUEST_COVER_LIMIT", "9")))
INLINE_TIKTOK_VISUAL_REVIEW_REQUEST_COVER_LIMIT = max(
    1,
    int(os.getenv("INLINE_TIKTOK_VISUAL_REVIEW_REQUEST_COVER_LIMIT", "5")),
)
VISUAL_REVIEW_CANDIDATE_COVER_LIMIT_FLOOR = max(
    VISUAL_REVIEW_REQUEST_COVER_LIMIT,
    int(os.getenv("VISUAL_REVIEW_CANDIDATE_COVER_LIMIT_FLOOR", "12")),
)
DEFAULT_VISUAL_REVIEW_MAX_WORKERS = max(1, int(os.getenv("VISUAL_REVIEW_MAX_WORKERS", "6")))
VISUAL_REVIEW_MAX_RETRIES = max(1, int(os.getenv("VISUAL_REVIEW_MAX_RETRIES", "3")))
DEFAULT_VISUAL_REVIEW_ITEM_TIMEOUT_SECONDS = max(
    90.0,
    float(os.getenv("VISUAL_REVIEW_ITEM_TIMEOUT_SECONDS", str(VISION_REQUEST_TIMEOUT * 2))),
)
VISUAL_REVIEW_RETRY_BASE_DELAY_SECONDS = max(
    0.5,
    float(os.getenv("VISUAL_REVIEW_RETRY_BASE_DELAY_SECONDS", "2")),
)
VISUAL_REVIEW_RETRY_MAX_DELAY_SECONDS = max(
    VISUAL_REVIEW_RETRY_BASE_DELAY_SECONDS,
    float(os.getenv("VISUAL_REVIEW_RETRY_MAX_DELAY_SECONDS", "12")),
)
VISUAL_IMAGE_DOWNLOAD_MAX_RETRIES = max(1, int(os.getenv("VISUAL_IMAGE_DOWNLOAD_MAX_RETRIES", "3")))
VISUAL_IMAGE_CACHE_ENABLED = parse_env_flag("VISUAL_IMAGE_CACHE_ENABLED", default=True)
VISUAL_IMAGE_CACHE_DIR = str(os.getenv("VISUAL_IMAGE_CACHE_DIR", "") or "").strip()
VISUAL_REVIEW_RETRYABLE_STATUS_CODES = {408, 429, 500, 502, 503, 504, 522}
DEFAULT_LOCAL_OPENAI_MAX_INFLIGHT_REQUESTS = 8
VISION_PROVIDER_CONFIGS = (
    {
        "name": "openai",
        "base_url_env_key": "OPENAI_BASE_URL",
        "default_base_url": "https://api.openai.com/v1",
        "env_key": "OPENAI_API_KEY",
        "api_style": VISION_API_STYLE_RESPONSES,
        "model_env_key": "OPENAI_VISION_MODEL",
        "max_inflight_env_key": "OPENAI_MAX_INFLIGHT_REQUESTS",
    },
    {
        "name": "reelx",
        "base_url_env_key": "VISION_REELX_BASE_URL",
        "default_base_url": DEFAULT_REELX_BASE_URL,
        "base_url_fallbacks_env_key": "VISION_REELX_BASE_URL_FALLBACKS",
        "default_base_url_fallbacks": DEFAULT_REELX_BASE_URL_FALLBACKS,
        "env_key": "VISION_REELX_API_KEY",
        "api_style": VISION_API_STYLE_GENERATE_CONTENT,
        "model_env_key": "VISION_REELX_MODEL",
        "default_model": DEFAULT_QWEN_VISION_MODEL,
        "fallback_model_env_key": "VISION_REELX_FALLBACK_MODEL",
        "default_fallback_model": "gemini-3-flash-preview",
    },
    {
        "name": "quan2go",
        "base_url_env_key": "VISION_QUAN2GO_BASE_URL",
        "default_base_url": "https://capi.quan2go.com/openai",
        "env_key": "VISION_QUAN2GO_API_KEY",
        "api_style": VISION_API_STYLE_CHAT_COMPLETIONS,
        "model_env_key": "VISION_QUAN2GO_MODEL",
    },
    {
        "name": "qiandao",
        "base_url_env_key": "VISION_QIANDAO_BASE_URL",
        "default_base_url": "https://api2.qiandao.mom/v1",
        "env_key": "VISION_QIANDAO_API_KEY",
        "api_style": VISION_API_STYLE_CHAT_COMPLETIONS,
        "model_env_key": "VISION_QIANDAO_MODEL",
        "default_model": DEFAULT_QIANDAO_VISION_MODEL,
        "fallback_model_env_key": "VISION_QIANDAO_FALLBACK_MODEL",
        "default_fallback_model": DEFAULT_QIANDAO_FALLBACK_VISION_MODEL,
        "max_tokens_env_key": "VISION_QIANDAO_MAX_TOKENS",
        "default_max_tokens": DEFAULT_QIANDAO_MAX_TOKENS,
        "temperature_env_key": "VISION_QIANDAO_TEMPERATURE",
        "default_temperature": DEFAULT_QIANDAO_TEMPERATURE,
    },
    {
        "name": "mimo",
        "base_url_env_key": "VISION_MIMO_BASE_URL",
        "default_base_url": "https://api.xiaomimimo.com/v1",
        "env_key": "VISION_MIMO_API_KEY",
        "api_style": VISION_API_STYLE_CHAT_COMPLETIONS,
        "model_env_key": "VISION_MIMO_MODEL",
        "default_model": DEFAULT_MIMO_VISION_MODEL,
        "auth_header_name": "api-key",
        "auth_header_value_prefix": "",
        "default_max_completion_tokens": DEFAULT_MIMO_MAX_COMPLETION_TOKENS,
        "max_completion_tokens_env_key": "VISION_MIMO_MAX_COMPLETION_TOKENS",
    },
    {
        "name": "lemonapi",
        "base_url_env_key": "VISION_LEMONAPI_BASE_URL",
        "default_base_url": "https://new.lemonapi.site/v1",
        "env_key": "VISION_LEMONAPI_API_KEY",
        "api_style": VISION_API_STYLE_CHAT_COMPLETIONS,
        "model_env_key": "VISION_LEMONAPI_MODEL",
    },
)
VISUAL_REVIEW_ROUTING_TIERED = "tiered"
VISUAL_REVIEW_ROUTING_PROBE_RANKED = "probe_ranked"
DEFAULT_VISUAL_REVIEW_ROUTING_PRIMARY_PROVIDER = "reelx"
DEFAULT_VISUAL_REVIEW_ROUTING_PRIMARY_MODEL = "gemini-3-flash-preview"
DEFAULT_VISUAL_REVIEW_ROUTING_PRIMARY_TIMEOUT_SECONDS = 20
DEFAULT_VISUAL_REVIEW_ROUTING_BACKUP_PROVIDER = "reelx"
DEFAULT_VISUAL_REVIEW_ROUTING_BACKUP_MODEL = "gemini-3.1-pro-preview"
DEFAULT_VISUAL_REVIEW_ROUTING_BACKUP_TIMEOUT_SECONDS = 25
DEFAULT_VISUAL_REVIEW_ROUTING_JUDGE_PROVIDER = "openai"
DEFAULT_VISUAL_REVIEW_ROUTING_JUDGE_MODEL = "gpt-5.4-mini"
DEFAULT_VISUAL_REVIEW_ROUTING_JUDGE_TIMEOUT_SECONDS = 30
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PROVIDER = "openai"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_MODEL = "gpt-5.4-mini"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_TIMEOUT_SECONDS = 60
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_PROVIDER = "reelx"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_MODEL = DEFAULT_QWEN_VISION_MODEL
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_TIMEOUT_SECONDS = 45
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_PROVIDER = "reelx"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_MODEL = "gemini-3-pro-preview"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_TIMEOUT_SECONDS = 35
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_PROVIDER = "reelx"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_MODEL = "gemini-3-flash-preview"
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_TIMEOUT_SECONDS = 30
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_DISABLE_AFTER_FAILURES = 2
DEFAULT_VISUAL_REVIEW_PROBE_RANKED_RETRY_ATTEMPTS = max(
    0,
    int(os.getenv("VISION_VISUAL_REVIEW_PROBE_RANKED_RETRY_ATTEMPTS", "1")),
)
VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED = "preferred"
VISUAL_REVIEW_PROBE_RANKED_GROUP_FALLBACK = "fallback"
VISUAL_REVIEW_PROBE_RANKED_SELECTED_STAGE_PREFERRED_POOL = "preferred_pool"
DEFAULT_VISUAL_REVIEW_ROUTING_HIGH_VALUE_FOLLOWER_THRESHOLD = 0
MINIMAL_VISUAL_REVIEW_PROBE_IMAGE_DATA_URL = (
    "data:image/png;base64,"
    "iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAIAAACQkWg2AAAAJUlEQVR4nGP47+AARw4N/+EIlzjDINRAjCJk8cGoYRAG60iMBwCgRZ8QleOc0QAAAABJRU5ErkJggg=="
)
VISUAL_REVIEW_BORDERLINE_MARKERS = (
    "可能",
    "疑似",
    "轻微",
    "不确定",
    "边界",
    "看不清",
    "较弱",
    "倾向",
)
VISION_PROMPT = """你是达人筛号流程中的视觉复核员。输入是同一位博主最近若干条内容的封面图，请综合全部图片一起判断。

只根据图片本身做初步判断，不要臆测看不到的信息。重点排查：
1. 是否明显过度性感、暴露或带强擦边倾向。
2. 是否存在明显低价平台/竞品合作痕迹，如 Temu、Shein、AliExpress、Wish、TikTok Shop。
3. 是否长期画面杂乱、昏暗、模糊、质感差。
4. 是否高度母婴/晒娃导向，主体大多是婴儿、儿童或孕期内容。
5. 是否整体过度商业化、广告摆拍感很重。
6. 是否存在大面积明显纹身等高风险视觉信号。

如果明显命中以上任一高风险情况，输出 Reject；否则输出 Pass。

请只返回 JSON，不要加 markdown，不要加额外说明，格式固定为：
{"decision":"Pass 或 Reject","reason":"一句中文原因","signals":["最多 3 个简短中文信号"]}"""
MIMO_VISION_PROMPT = """你是达人筛号视觉复核员。输入是同一位博主最近若干条内容的封面图，请综合全部图片一起判断。

只根据图片本身判断，不要臆测看不到的信息，不要逐图解释，不要输出推理过程。
命中以下任一项输出 Reject，否则输出 Pass：
1. 明显过度性感、暴露或擦边。
2. 明显低价平台/竞品合作痕迹，如 Temu、Shein、AliExpress、Wish、TikTok Shop。
3. 长期画面杂乱、昏暗、模糊、质感差。
4. 高度母婴/晒娃导向。
5. 整体过度商业化、广告摆拍感重。
6. 大面积明显纹身。

请只返回一行 JSON，不要 markdown，不要额外说明，格式固定为：
{"decision":"Pass 或 Reject","reason":"一句中文原因","signals":["最多 3 个简短中文信号"]}"""
QWEN_VISION_PROMPT_V2 = """你是保守但专业的达人筛号视觉复核员。输入是同一位博主最近若干条内容的封面图，请综合全部图片一起判断。

只能根据图片里直接可见的视觉证据判断。不要为了写得完整而乱猜；不要把偶发元素放大成长期标签；不要根据账号名、文案外信息或常识补全。

重点排查：
1. 是否明显过度性感、暴露或带强擦边倾向。
2. 是否存在明显低价平台/竞品合作痕迹，如 Temu、Shein、AliExpress、Wish、TikTok Shop。
3. 是否长期画面杂乱、昏暗、模糊、质感差。
4. 是否高度母婴/晒娃导向，主体大多是婴儿、儿童或孕期内容。
5. 是否整体过度商业化、广告摆拍感很重。
6. 是否存在大面积明显纹身等高风险视觉信号。

规则：
1. 只有当多张图都出现强且直接的视觉证据时，才输出 Reject。
2. 如果只是少量、模糊、边缘信号，不要硬判高风险。
3. `reason` 只写一句克制中文，先写最稳定的主因，不要写空泛套话。
4. `signals` 最多写 3 个简短中文点，只保留最直接的视觉证据。
5. 只返回 JSON，不要 markdown，不要解释。

格式固定为：
{"decision":"Pass 或 Reject","reason":"一句中文原因","signals":["最多 3 个简短中文信号"]}"""
POSITIONING_CARD_PROMPT = """你是达人筛号流程中的定位卡分析员。输入是同一位已通过视觉复核的博主最近若干条内容封面图，请综合全部图片一起判断其内容定位与品牌适配度。

只根据图片本身做初步判断，不要假设看不到的信息，不要输出推理过程。

请输出：
1. `positioning_labels`：1 到 4 个中文定位标签。
2. `fit_recommendation`：只能是 `High Fit`、`Medium Fit`、`Low Fit`、`Unclear` 之一。
3. `fit_summary`：一句中文结论，说明为什么适配或不适配。
4. `evidence_signals`：最多 4 个简短中文信号，只保留画面里能直接看到的证据。

请只返回 JSON，不要 markdown，不要额外说明，格式固定为：
{"positioning_labels":["标签1","标签2"],"fit_recommendation":"High Fit","fit_summary":"一句中文结论","evidence_signals":["信号1","信号2"]}"""

JOBS = {}
JOBS_LOCK = threading.Lock()
APIFY_TOKEN_LOCK = threading.Lock()
APIFY_BALANCE_CACHE_LOCK = threading.Lock()
APIFY_BALANCE_REFRESH_LOCK = threading.Lock()
APIFY_BALANCE_POLLER_LOCK = threading.Lock()
APIFY_RUN_GUARDS_LOCK = threading.Lock()
APIFY_BALANCE_POLLER_THREAD = None
APIFY_BALANCE_POLLER_STOP_EVENT = threading.Event()
OPERATOR_RUNS_LOCK = threading.Lock()
OPERATOR_RUNS: dict[str, dict[str, Any]] = {}
OPERATOR_RUN_PROCESSES: dict[str, subprocess.Popen[Any]] = {}
OPERATOR_RUN_LOG_HANDLES: dict[str, Any] = {}
VISION_PROVIDER_REQUEST_GATES_LOCK = threading.Lock()
VISION_PROVIDER_REQUEST_GATES: dict[tuple[str, str, int], threading.BoundedSemaphore] = {}
OPERATOR_RUNS_ROOT = BASE_DIR / "temp" / "operator_runs"
OPERATOR_RUN_TERMINAL_STATUSES = {
    "completed",
    "completed_with_quality_warnings",
    "completed_with_partial_scrape",
    "failed",
    "cancelled",
}

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["JSON_AS_ASCII"] = False
if CORS:
    CORS(app, resources={r"/api/*": {"origins": BACKEND_ALLOWED_ORIGINS}})


def ensure_runtime_dirs():
    Path(DATA_DIR).mkdir(parents=True, exist_ok=True)
    Path(CONFIG_DIR).mkdir(parents=True, exist_ok=True)
    Path(TEMP_DIR).mkdir(parents=True, exist_ok=True)
    Path(UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
    OPERATOR_RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    for platform in PLATFORM_ACTORS:
        Path(get_platform_dir(platform)).mkdir(parents=True, exist_ok=True)


def iso_now():
    return datetime.utcnow().isoformat() + "Z"


def safe_positive_float(value, default=0.0):
    try:
        numeric = float(value)
    except Exception:
        return float(default)
    if numeric < 0:
        return float(default)
    return numeric


def write_json_file(path, payload):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    return path


def load_json_payload(path, default=None):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as handle:
            raw_text = handle.read().strip()
        if not raw_text:
            return default
        return json.loads(raw_text)
    except Exception:
        return default


def sanitize_json_compatible(value):
    if isinstance(value, dict):
        return {str(key): sanitize_json_compatible(item) for key, item in value.items()}
    if isinstance(value, list):
        return [sanitize_json_compatible(item) for item in value]
    if isinstance(value, tuple):
        return [sanitize_json_compatible(item) for item in value]
    if pd.isna(value):
        return None
    if isinstance(value, (int, float, str, bool)) or value is None:
        return value
    return str(value)


def _path_is_within_root(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except Exception:
        return False


def _resolve_operator_env_snapshot(env_file: str = ""):
    raw = str(env_file or ".env").strip() or ".env"
    candidate = Path(raw).expanduser()
    if not candidate.is_absolute():
        candidate = (BASE_DIR / candidate).resolve()
    return load_env_file_snapshot(
        str(candidate),
        default_env_file=str((BASE_DIR / ".env").resolve()),
    )


def _resolve_operator_env_file(env_file: str = "") -> Path:
    return _resolve_operator_env_snapshot(env_file).path


def _build_operator_resolved_config_sources(
    *,
    env_file: str = "",
    task_upload_url: str = "",
    employee_info_url: str = "",
    matching_strategy: str = "",
    brand_keyword: str = "",
    platforms: list[str] | None = None,
    vision_provider: str = "",
    max_identifiers_per_platform: int = 0,
    mail_limit: int = 0,
    sent_since: str = "",
    reuse_existing: bool = True,
    probe_vision_provider_only: bool = False,
    skip_scrape: bool = False,
    skip_visual: bool = False,
    skip_positioning_card_analysis: bool = False,
) -> tuple[dict[str, Any], Any]:
    return resolve_operator_launch_config(
        env_file=env_file,
        task_upload_url=task_upload_url,
        employee_info_url=employee_info_url,
        matching_strategy=matching_strategy,
        brand_keyword=brand_keyword,
        platforms=platforms,
        vision_provider=vision_provider,
        max_identifiers_per_platform=max_identifiers_per_platform,
        mail_limit=mail_limit,
        sent_since=sent_since,
        reuse_existing=reuse_existing,
        probe_vision_provider_only=probe_vision_provider_only,
        skip_scrape=skip_scrape,
        skip_visual=skip_visual,
        skip_positioning_card_analysis=skip_positioning_card_analysis,
    )


def _build_operator_bootstrap_summary(
    *,
    started_at: str,
    task_name: str,
    operator_paths,
    env_snapshot,
    env_file_raw: str,
    resolved_config_sources: dict[str, Any],
) -> dict[str, Any]:
    return {
        "started_at": str(started_at or ""),
        "finished_at": "",
        "status": "running",
        "task_name": str(task_name or "").strip(),
        "run_id": operator_paths.run_id,
        "run_root": str(operator_paths.run_root),
        "env_file_raw": str(env_file_raw or ""),
        "env_file": str(env_snapshot.path),
        "output_root": str(operator_paths.output_root),
        "summary_json": str(operator_paths.summary_json),
        "task_spec_json": str(operator_paths.task_spec_json),
        "resolved_config_sources": resolved_config_sources,
        "resolved_inputs": {
            "env_file": {
                "path": str(env_snapshot.path),
                "exists": env_snapshot.exists,
                "source": env_snapshot.source,
            }
        },
        "resolved_paths": {},
        "artifacts": {},
    }


def _operator_file_ref(path_value: str) -> dict[str, Any]:
    raw_path = str(path_value or "").strip()
    if not raw_path:
        return {
            "path": "",
            "exists": False,
            "download_url": "",
        }
    resolved = Path(raw_path).expanduser().resolve()
    exists = resolved.exists()
    download_url = ""
    if exists and _path_is_within_root(resolved, BASE_DIR):
        download_url = f"/api/operator/file?path={quote(str(resolved))}"
    return {
        "path": str(resolved),
        "exists": exists,
        "download_url": download_url,
    }


def _normalize_operator_platforms(values: Any) -> list[str]:
    if isinstance(values, str):
        candidates = [item.strip() for item in values.split(",")]
    else:
        candidates = [str(item or "").strip() for item in (values or [])]
    normalized: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        value = UPLOAD_PLATFORM_ALIASES.get(candidate.casefold(), "")
        if value and value not in seen:
            seen.add(value)
            normalized.append(value)
    return normalized


def _boolish(value: Any, default: bool = False) -> bool:
    if value is None:
        return bool(default)
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if not normalized:
        return bool(default)
    return normalized not in {"0", "false", "no", "off"}


def _build_operator_error_payload(
    *,
    error_code: str,
    message: str,
    remediation: str,
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "success": False,
        "error_code": error_code,
        "error": message,
        "remediation": remediation,
        "details": details or {},
    }


def _classify_operator_config_error(exc: Exception) -> tuple[dict[str, Any], int]:
    message = str(exc) or exc.__class__.__name__
    error_code = "OPERATOR_CONFIG_ERROR"
    remediation = "检查 operator 页面里的 env / Feishu 配置后重试。"
    if "缺少 FEISHU_APP_ID" in message:
        error_code = "FEISHU_APP_ID_MISSING"
        remediation = "在 `.env` 或请求参数里填写 FEISHU_APP_ID。"
    elif "缺少 FEISHU_APP_SECRET" in message:
        error_code = "FEISHU_APP_SECRET_MISSING"
        remediation = "在 `.env` 或请求参数里填写 FEISHU_APP_SECRET。"
    elif "缺少 TASK_UPLOAD_URL" in message:
        error_code = "TASK_UPLOAD_URL_MISSING"
        remediation = "在 `.env` 或请求参数里填写 TASK_UPLOAD_URL。"
    elif "缺少 EMPLOYEE_INFO_URL" in message:
        error_code = "EMPLOYEE_INFO_URL_MISSING"
        remediation = "在 `.env` 或请求参数里填写 EMPLOYEE_INFO_URL。"
    elif "task_name" in message:
        error_code = "TASK_NAME_MISSING"
        remediation = "先从任务列表里选择任务，再启动 run。"
    status_code = 400
    return (
        _build_operator_error_payload(
            error_code=error_code,
            message=message,
            remediation=remediation,
            details={"exception_type": exc.__class__.__name__},
        ),
        status_code,
    )


def load_operator_task_candidates(
    *,
    env_file: str = "",
    task_upload_url: str = "",
    employee_info_url: str = "",
) -> dict[str, Any]:
    _, resolved_config = resolve_operator_task_candidates_config(
        env_file=env_file,
        task_upload_url=task_upload_url,
        employee_info_url=employee_info_url,
    )
    resolved_env_file = resolved_config["env_snapshot"].path
    app_id = resolved_config["feishu_app_id"].value
    if not resolved_config["feishu_app_id"].present:
        raise ValueError("缺少 FEISHU_APP_ID，请在本地 .env 或参数里填写。")
    app_secret = resolved_config["feishu_app_secret"].value
    if not resolved_config["feishu_app_secret"].present:
        raise ValueError("缺少 FEISHU_APP_SECRET，请在本地 .env 或参数里填写。")
    resolved_task_upload_url = resolved_config["task_upload_url"].value
    if not resolved_config["task_upload_url"].present:
        raise ValueError("缺少 TASK_UPLOAD_URL，请在本地 .env 或参数里填写。")
    resolved_employee_info_url = resolved_config["employee_info_url"].value
    if not resolved_config["employee_info_url"].present:
        raise ValueError("缺少 EMPLOYEE_INFO_URL，请在本地 .env 或参数里填写。")
    timeout_seconds = float(resolved_config["timeout_seconds"].value or "30")
    client = FeishuOpenClient(
        app_id=app_id,
        app_secret=app_secret,
        base_url=resolved_config["feishu_base_url"].value or DEFAULT_FEISHU_BASE_URL,
        timeout_seconds=timeout_seconds,
    )
    inspection = inspect_task_upload_assignments(
        client=client,
        task_upload_url=resolved_task_upload_url,
        employee_info_url=resolved_employee_info_url,
        download_dir=BASE_DIR / "downloads" / "task_upload_attachments",
        download_templates=False,
        parse_templates=False,
    )
    tasks = []
    for item in inspection.get("items") or []:
        tasks.append({
            "task_name": str(item.get("taskName") or "").strip(),
            "record_id": str(item.get("recordId") or "").strip(),
            "responsible_name": str(item.get("responsibleName") or "").strip(),
            "owner_name": str(item.get("ownerName") or "").strip(),
            "owner_email": str(item.get("ownerEmail") or "").strip(),
            "employee_email": str(item.get("employeeEmail") or "").strip(),
            "employee_matched": bool(item.get("employeeMatched")),
            "matched_by": str(item.get("matchedBy") or "").strip(),
            "linked_bitable_url": str(item.get("linkedBitableUrl") or "").strip(),
            "template_file_name": str(item.get("templateFileName") or "").strip(),
            "sending_list_file_name": str(item.get("sendingListFileName") or "").strip(),
        })
    tasks.sort(key=lambda entry: (str(entry.get("task_name") or "").casefold(), str(entry.get("record_id") or "")))
    return {
        "success": True,
        "env_file": str(resolved_env_file),
        "task_upload_url": resolved_task_upload_url,
        "employee_info_url": resolved_employee_info_url,
        "task_table_name": inspection.get("taskTableName") or "",
        "employee_table_name": inspection.get("employeeTableName") or "",
        "record_count": int(inspection.get("recordCount") or 0),
        "matched_count": int(inspection.get("matchedCount") or 0),
        "tasks": tasks,
    }


def _close_operator_log_handle(run_id: str) -> None:
    handle = OPERATOR_RUN_LOG_HANDLES.pop(run_id, None)
    if handle is None:
        return
    try:
        handle.close()
    except Exception:
        pass


def _collect_operator_summary_stage(summary: dict[str, Any]) -> str:
    summary_status = str(summary.get("status") or "").strip()
    if summary_status and summary_status != "running":
        return summary_status
    steps = summary.get("steps") or {}
    downstream_status = str((steps.get("downstream") or {}).get("status") or "").strip()
    if downstream_status:
        return f"downstream:{downstream_status}"
    upstream_status = str((steps.get("upstream") or {}).get("status") or "").strip()
    if upstream_status:
        return f"upstream:{upstream_status}"
    resolved_paths = summary.get("resolved_paths") or {}
    for key, label in (
        ("downstream_summary_json", "downstream"),
        ("upstream_summary_json", "upstream"),
    ):
        nested_path = str(resolved_paths.get(key) or "").strip()
        if not nested_path:
            continue
        nested_summary = load_json_payload(nested_path, default={}) or {}
        nested_status = str(nested_summary.get("status") or "").strip()
        if nested_status:
            return f"{label}:{nested_status}"
    return "running"


def _load_persisted_operator_run(run_id: str) -> dict[str, Any] | None:
    normalized_run_id = str(run_id or "").strip()
    if not normalized_run_id:
        return None
    output_root = (OPERATOR_RUNS_ROOT / normalized_run_id).resolve()
    if not output_root.exists() or not _path_is_within_root(output_root, OPERATOR_RUNS_ROOT):
        return None
    summary_path = output_root / "summary.json"
    if not summary_path.exists():
        return None
    summary = load_json_payload(summary_path, default={})
    if not isinstance(summary, dict) or not summary:
        return None
    log_path = output_root / "operator_run.log"
    status = str(summary.get("status") or "completed").strip() or "completed"
    return {
        "id": normalized_run_id,
        "task_name": str(summary.get("task_name") or "").strip() or normalized_run_id,
        "status": status,
        "stage": _collect_operator_summary_stage(summary),
        "env_file": str(summary.get("env_file") or "").strip(),
        "env_file_raw": str(summary.get("env_file_raw") or "").strip(),
        "run_root": str(summary.get("run_root") or output_root),
        "output_root": str(output_root),
        "task_spec_json": str(summary.get("task_spec_json") or (output_root / "task_spec.json")),
        "summary_path": str(summary_path),
        "log_path": str(log_path),
        "pid": None,
        "created_at": str(summary.get("started_at") or "").strip(),
        "updated_at": str(summary.get("finished_at") or summary.get("started_at") or "").strip(),
        "finished_at": str(summary.get("finished_at") or "").strip(),
        "return_code": 0 if status in OPERATOR_RUN_TERMINAL_STATUSES else None,
        "error": str(summary.get("error") or "").strip(),
        "requested_options": {},
        "summary": summary,
    }


def _build_operator_all_platforms_final_review(summary: dict[str, Any]) -> dict[str, Any]:
    artifacts = summary.get("artifacts") or {}
    existing_output = str(artifacts.get("all_platforms_final_review") or "").strip()
    if existing_output:
        existing_ref = _operator_file_ref(existing_output)
        if existing_ref.get("exists"):
            return existing_ref
    output_root_value = str(summary.get("output_root") or "").strip()
    if not output_root_value:
        return _operator_file_ref("")
    output_root = Path(output_root_value).expanduser().resolve()
    final_exports = artifacts.get("final_exports") or {}
    if not final_exports:
        return _operator_file_ref("")
    upstream_summary_json = str(artifacts.get("upstream_summary_json") or "").strip()
    task_owner_context: dict[str, Any] = {}
    if upstream_summary_json:
        upstream_path = Path(upstream_summary_json).expanduser().resolve()
        if upstream_path.exists():
            try:
                task_owner_context = extract_task_owner_context(
                    json.loads(upstream_path.read_text(encoding="utf-8"))
                )
            except Exception:
                task_owner_context = {}
    output_dir = output_root / "operator_exports"
    output_path = output_dir / "all_platforms_final_review.xlsx"
    payload_path = output_dir / "all_platforms_final_review_payload.json"
    try:
        build_all_platforms_final_review_artifacts(
            output_path=output_path,
            payload_json_path=payload_path,
            final_exports=final_exports,
            keep_workbook=str(artifacts.get("keep_workbook") or ""),
            task_owner=task_owner_context,
        )
    except Exception:
        return _operator_file_ref("")
    return _operator_file_ref(str(output_path))


def _build_operator_summary_view(summary: dict[str, Any]) -> dict[str, Any]:
    artifacts = summary.get("artifacts") or {}
    final_exports: dict[str, dict[str, Any]] = {}
    for platform, export_map in (artifacts.get("final_exports") or {}).items():
        if not isinstance(export_map, dict):
            continue
        final_exports[str(platform)] = {
            str(name): _operator_file_ref(str(path_value or ""))
            for name, path_value in export_map.items()
            if str(path_value or "").strip()
        }
    all_platforms_final_review = _build_operator_all_platforms_final_review(summary)
    return {
        "summary_json": _operator_file_ref(str(summary.get("summary_json") or "")),
        "upstream_summary_json": _operator_file_ref(
            str((summary.get("resolved_paths") or {}).get("upstream_summary_json") or "")
        ),
        "downstream_summary_json": _operator_file_ref(
            str((summary.get("resolved_paths") or {}).get("downstream_summary_json") or "")
        ),
        "keep_workbook": _operator_file_ref(str(artifacts.get("keep_workbook") or "")),
        "template_workbook": _operator_file_ref(str(artifacts.get("template_workbook") or "")),
        "all_platforms_final_review": all_platforms_final_review,
        "final_exports": final_exports,
    }


def _serialize_operator_run(run: dict[str, Any]) -> dict[str, Any]:
    payload = {key: value for key, value in run.items() if key != "summary"}
    summary = run.get("summary")
    if isinstance(summary, dict):
        payload["summary"] = sanitize_json_compatible(summary)
        payload["artifacts"] = _build_operator_summary_view(summary)
    else:
        payload["summary"] = None
        payload["artifacts"] = {
            "summary_json": _operator_file_ref(str(run.get("summary_path") or "")),
            "upstream_summary_json": _operator_file_ref(""),
            "downstream_summary_json": _operator_file_ref(""),
            "keep_workbook": _operator_file_ref(""),
            "template_workbook": _operator_file_ref(""),
            "all_platforms_final_review": _operator_file_ref(""),
            "final_exports": {},
        }
    return sanitize_json_compatible(payload)


def refresh_operator_run(run_id: str) -> dict[str, Any] | None:
    with OPERATOR_RUNS_LOCK:
        run = OPERATOR_RUNS.get(run_id)
    if not run:
        run = _load_persisted_operator_run(run_id)
        if run is None:
            return None
        with OPERATOR_RUNS_LOCK:
            OPERATOR_RUNS[run_id] = run
    run = dict(run)
    process = OPERATOR_RUN_PROCESSES.get(run_id)
    summary_path = Path(str(run.get("summary_path") or "")).expanduser()
    summary = load_json_payload(summary_path, default={}) if str(summary_path) else {}
    summary = summary if isinstance(summary, dict) else {}
    return_code = process.poll() if process is not None else run.get("return_code")
    status = str(run.get("status") or "queued").strip() or "queued"
    stage = str(run.get("stage") or "queued").strip() or "queued"
    error = str(run.get("error") or "").strip()
    finished_at = str(run.get("finished_at") or "").strip()
    if summary:
        summary_status = str(summary.get("status") or "").strip()
        if summary_status:
            status = summary_status
        stage = _collect_operator_summary_stage(summary)
        error = str(summary.get("error") or error).strip()
        finished_at = str(summary.get("finished_at") or finished_at).strip()
    elif process is not None and return_code is None:
        status = "running"
        stage = "running"
    if process is not None and return_code is not None and status == "running":
        status = "failed" if return_code else "completed"
        stage = status
        if not error and return_code:
            error = f"operator subprocess exited with code {return_code}"
        if not finished_at:
            finished_at = iso_now()
    if status in OPERATOR_RUN_TERMINAL_STATUSES or return_code is not None:
        _close_operator_log_handle(run_id)
    updated_run = dict(run)
    updated_run.update({
        "status": status,
        "stage": stage,
        "summary": summary or None,
        "return_code": return_code,
        "error": error,
        "finished_at": finished_at,
        "updated_at": iso_now(),
    })
    with OPERATOR_RUNS_LOCK:
        OPERATOR_RUNS[run_id] = updated_run
    return updated_run


def list_operator_runs() -> list[dict[str, Any]]:
    with OPERATOR_RUNS_LOCK:
        run_ids = list(OPERATOR_RUNS.keys())
    if OPERATOR_RUNS_ROOT.exists():
        for candidate in OPERATOR_RUNS_ROOT.iterdir():
            if not candidate.is_dir():
                continue
            if not (candidate / "summary.json").exists():
                continue
            run_ids.append(candidate.name)
    run_ids = list(dict.fromkeys(run_ids))
    runs = [refresh_operator_run(run_id) for run_id in run_ids]
    cleaned = [run for run in runs if run]
    cleaned.sort(key=lambda item: str(item.get("created_at") or ""), reverse=True)
    return [_serialize_operator_run(run) for run in cleaned]


def launch_operator_run(
    *,
    task_name: str,
    env_file: str = "",
    task_upload_url: str = "",
    employee_info_url: str = "",
    matching_strategy: str = "brand-keyword-fast-path",
    brand_keyword: str = "",
    brand_match_include_from: bool = False,
    platforms: list[str] | None = None,
    vision_provider: str = "",
    max_identifiers_per_platform: int = 0,
    mail_limit: int = 0,
    sent_since: str = "",
    reuse_existing: bool = True,
    probe_vision_provider_only: bool = False,
    skip_scrape: bool = False,
    skip_visual: bool = False,
    skip_positioning_card_analysis: bool = False,
) -> dict[str, Any]:
    normalized_task_name = str(task_name or "").strip()
    if not normalized_task_name:
        raise ValueError("task_name 不能为空。")
    ensure_runtime_dirs()
    OPERATOR_RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    resolved_config_sources, resolved_config = _build_operator_resolved_config_sources(
        env_file=env_file,
        task_upload_url=task_upload_url,
        employee_info_url=employee_info_url,
        matching_strategy=matching_strategy,
        brand_keyword=brand_keyword,
        platforms=_normalize_operator_platforms(platforms),
        vision_provider=vision_provider,
        max_identifiers_per_platform=max_identifiers_per_platform,
        mail_limit=mail_limit,
        sent_since=sent_since,
        reuse_existing=reuse_existing,
        probe_vision_provider_only=probe_vision_provider_only,
        skip_scrape=skip_scrape,
        skip_visual=skip_visual,
        skip_positioning_card_analysis=skip_positioning_card_analysis,
    )
    env_snapshot = resolved_config["env_snapshot"]
    operator_paths = resolve_operator_run_paths(
        task_name=normalized_task_name,
        runs_root=OPERATOR_RUNS_ROOT,
    )
    output_root = operator_paths.output_root
    output_root.mkdir(parents=True, exist_ok=True)
    summary_path = operator_paths.summary_json
    log_path = operator_paths.log_path
    started_at = iso_now()
    task_spec = build_operator_task_spec(
        generated_at=started_at,
        operator_paths=operator_paths,
        env_snapshot=env_snapshot,
        env_file_raw=str(env_file or ""),
        resolved_config_sources=resolved_config_sources,
        task_name=normalized_task_name,
        task_upload_url=resolved_config["task_upload_url"].value,
        employee_info_url=resolved_config["employee_info_url"].value,
        matching_strategy=str(matching_strategy or "brand-keyword-fast-path").strip() or "brand-keyword-fast-path",
        brand_keyword=str(brand_keyword or "").strip(),
        brand_match_include_from=bool(brand_match_include_from),
        requested_platforms=_normalize_operator_platforms(platforms),
        vision_provider=str(vision_provider or "").strip(),
        max_identifiers_per_platform=int(max_identifiers_per_platform or 0),
        mail_limit=int(mail_limit or 0),
        sent_since=str(sent_since or "").strip(),
        reuse_existing=bool(reuse_existing),
        probe_vision_provider_only=bool(probe_vision_provider_only),
        skip_scrape=bool(skip_scrape),
        skip_visual=bool(skip_visual),
        skip_positioning_card_analysis=bool(skip_positioning_card_analysis),
    )
    bootstrap_summary = _build_operator_bootstrap_summary(
        started_at=started_at,
        task_name=normalized_task_name,
        operator_paths=operator_paths,
        env_snapshot=env_snapshot,
        env_file_raw=str(env_file or ""),
        resolved_config_sources=resolved_config_sources,
    )
    write_task_spec(operator_paths.task_spec_json, task_spec)
    write_json_file(summary_path, bootstrap_summary)
    command = [
        sys.executable,
        str((BASE_DIR / "scripts" / "run_task_upload_to_final_export_pipeline.py").resolve()),
        "--task-name",
        normalized_task_name,
        "--env-file",
        str(env_snapshot.path),
        "--output-root",
        str(output_root),
        "--summary-json",
        str(summary_path),
        "--matching-strategy",
        str(matching_strategy or "brand-keyword-fast-path").strip() or "brand-keyword-fast-path",
    ]
    normalized_brand_keyword = str(brand_keyword or "").strip()
    if normalized_brand_keyword:
        command.extend(["--brand-keyword", normalized_brand_keyword])
    if brand_match_include_from:
        command.append("--brand-match-include-from")
    if task_upload_url:
        command.extend(["--task-upload-url", str(task_upload_url).strip()])
    if employee_info_url:
        command.extend(["--employee-info-url", str(employee_info_url).strip()])
    for platform in _normalize_operator_platforms(platforms):
        command.extend(["--platform", platform])
    if str(vision_provider or "").strip():
        command.extend(["--vision-provider", str(vision_provider).strip()])
    if int(max_identifiers_per_platform or 0) > 0:
        command.extend(["--max-identifiers-per-platform", str(int(max_identifiers_per_platform))])
    if int(mail_limit or 0) > 0:
        command.extend(["--mail-limit", str(int(mail_limit))])
    if str(sent_since or "").strip():
        command.extend(["--sent-since", str(sent_since).strip()])
    if not reuse_existing:
        command.append("--no-reuse-existing")
    if probe_vision_provider_only:
        command.append("--probe-vision-provider-only")
    if skip_scrape:
        command.append("--skip-scrape")
    if skip_visual:
        command.append("--skip-visual")
    if skip_positioning_card_analysis:
        command.append("--skip-positioning-card-analysis")
    log_handle = open(log_path, "a", encoding="utf-8")
    process = subprocess.Popen(
        command,
        cwd=str(BASE_DIR),
        stdout=log_handle,
        stderr=subprocess.STDOUT,
        env={**os.environ, "PYTHONUNBUFFERED": "1"},
        text=True,
    )
    run = {
        "id": operator_paths.run_id,
        "task_name": normalized_task_name,
        "status": "running",
        "stage": "starting",
        "env_file_raw": str(env_file or ""),
        "env_file": str(env_snapshot.path),
        "run_root": str(operator_paths.run_root),
        "output_root": str(output_root),
        "task_spec_json": str(operator_paths.task_spec_json),
        "summary_path": str(summary_path),
        "log_path": str(log_path),
        "command": command,
        "pid": process.pid,
        "created_at": iso_now(),
        "updated_at": iso_now(),
        "finished_at": "",
        "return_code": None,
        "error": "",
        "summary": bootstrap_summary,
        "requested_options": {
            "task_upload_url": str(task_upload_url or "").strip(),
            "employee_info_url": str(employee_info_url or "").strip(),
            "matching_strategy": str(matching_strategy or "").strip() or "brand-keyword-fast-path",
            "brand_keyword": normalized_brand_keyword,
            "brand_match_include_from": bool(brand_match_include_from),
            "platforms": _normalize_operator_platforms(platforms),
            "vision_provider": str(vision_provider or "").strip(),
            "max_identifiers_per_platform": int(max_identifiers_per_platform or 0),
            "mail_limit": int(mail_limit or 0),
            "sent_since": str(sent_since or "").strip(),
            "reuse_existing": bool(reuse_existing),
            "probe_vision_provider_only": bool(probe_vision_provider_only),
            "skip_scrape": bool(skip_scrape),
            "skip_visual": bool(skip_visual),
            "skip_positioning_card_analysis": bool(skip_positioning_card_analysis),
        },
    }
    with OPERATOR_RUNS_LOCK:
        OPERATOR_RUNS[operator_paths.run_id] = run
        OPERATOR_RUN_PROCESSES[operator_paths.run_id] = process
        OPERATOR_RUN_LOG_HANDLES[operator_paths.run_id] = log_handle
    refreshed = refresh_operator_run(operator_paths.run_id)
    return _serialize_operator_run(refreshed or run)


def get_platform_dir(platform):
    return str(Path(DATA_DIR) / platform)


def get_upload_metadata_path(platform):
    return str(Path(get_platform_dir(platform)) / f"{platform}_upload_metadata.json")


def get_raw_data_path(platform):
    return str(Path(get_platform_dir(platform)) / f"{platform}_data.json")


def get_profile_reviews_path(platform):
    return str(Path(get_platform_dir(platform)) / f"{platform}_profile_reviews.json")


def get_visual_results_path(platform):
    return str(Path(get_platform_dir(platform)) / f"{platform}_visual_results.json")


def get_positioning_card_results_path(platform):
    return str(Path(get_platform_dir(platform)) / f"{platform}_positioning_card_results.json")


def get_visual_image_cache_dir(platform):
    if VISUAL_IMAGE_CACHE_DIR:
        return str(Path(VISUAL_IMAGE_CACHE_DIR) / platform)
    return str(Path(get_platform_dir(platform)) / "covers")


def load_upload_metadata(platform):
    return load_json_payload(get_upload_metadata_path(platform), default={}) or {}


def save_upload_metadata(platform, metadata_map, replace=False):
    current = {} if replace else load_upload_metadata(platform)
    merged = dict(current)
    merged.update(metadata_map or {})
    write_json_file(get_upload_metadata_path(platform), merged)
    return merged


def load_profile_reviews(platform):
    return load_json_payload(get_profile_reviews_path(platform), default=[]) or []


def save_profile_reviews(platform, profile_reviews):
    return write_json_file(get_profile_reviews_path(platform), profile_reviews or [])


def load_visual_results(platform):
    return load_json_payload(get_visual_results_path(platform), default={}) or {}


def save_visual_results(platform, visual_results):
    return write_json_file(get_visual_results_path(platform), visual_results or {})


def load_positioning_card_results(platform):
    return load_json_payload(get_positioning_card_results_path(platform), default={}) or {}


def save_positioning_card_results(platform, positioning_results):
    return write_json_file(get_positioning_card_results_path(platform), positioning_results or {})


def load_apify_token_state():
    return load_json_payload(APIFY_TOKEN_POOL_STATE_FILE, default={}) or {}


def save_apify_token_state(state):
    write_json_file(APIFY_TOKEN_POOL_STATE_FILE, state or {})
    return state or {}


def load_apify_balance_cache():
    return load_json_payload(APIFY_BALANCE_CACHE_FILE, default={}) or {}


def save_apify_balance_cache(cache_payload):
    with APIFY_BALANCE_CACHE_LOCK:
        write_json_file(APIFY_BALANCE_CACHE_FILE, cache_payload or {})
    return cache_payload or {}


def get_apify_token():
    pool = get_apify_token_pool()
    return pool[0] if pool else ""


def mask_apify_token(token):
    cleaned = str(token or "").strip()
    if not cleaned:
        return ""
    if len(cleaned) <= 8:
        return cleaned[:2] + "***"
    return f"{cleaned[:4]}...{cleaned[-4:]}"


def sanitize_apify_error_text(text, tokens=None):
    sanitized = str(text or "")
    for token in tokens or []:
        cleaned = str(token or "").strip()
        if not cleaned:
            continue
        sanitized = sanitized.replace(cleaned, mask_apify_token(cleaned))
    sanitized = re.sub(r"""(token=)([^&\s)"']+)""", r"\1***", sanitized)
    return sanitized


def redact_url_like_text(text):
    return re.sub(r"https?://[^\s)\"'>]+", "[redacted-url]", str(text or ""))


def split_apify_token_list(value):
    return [
        item.strip()
        for item in re.split(r"[\s,]+", str(value or "").strip())
        if item.strip()
    ]


def get_apify_token_pool():
    seen = set()
    pool = []
    for source in (
        os.getenv("APIFY_TOKEN"),
        os.getenv("APIFY_API_TOKEN"),
        os.getenv("APIFY_BACKUP_TOKENS"),
        os.getenv("APIFY_FREE_TOKENS"),
    ):
        for token in split_apify_token_list(source):
            if token in seen:
                continue
            seen.add(token)
            pool.append(token)
    auth_payload = load_json_payload(str(Path.home() / ".apify" / "auth.json"), default={}) or {}
    auth_token = str(auth_payload.get("token") or "").strip()
    if auth_token and auth_token not in seen:
        pool.append(auth_token)
    return pool


def load_apify_run_guards():
    return load_json_payload(APIFY_RUN_GUARDS_FILE, default={}) or {}


def save_apify_run_guards(guards):
    with APIFY_RUN_GUARDS_LOCK:
        write_json_file(APIFY_RUN_GUARDS_FILE, guards or {})
    return guards or {}


def purge_expired_apify_run_guards(guards):
    now_ts = time.time()
    cleaned = {}
    for key, record in (guards or {}).items():
        try:
            expires_at_ts = float((record or {}).get("expires_at_ts") or 0)
        except Exception:
            expires_at_ts = 0
        if expires_at_ts and expires_at_ts <= now_ts:
            continue
        cleaned[key] = record
    return cleaned


def get_apify_run_guard(guard_key):
    with APIFY_RUN_GUARDS_LOCK:
        guards = purge_expired_apify_run_guards(load_apify_run_guards())
        write_json_file(APIFY_RUN_GUARDS_FILE, guards)
    return guards.get(guard_key)


def remember_apify_run_guard(guard_key, record):
    with APIFY_RUN_GUARDS_LOCK:
        guards = purge_expired_apify_run_guards(load_apify_run_guards())
        guards[guard_key] = record
        write_json_file(APIFY_RUN_GUARDS_FILE, guards)
    return record


def clear_apify_run_guard(guard_key):
    with APIFY_RUN_GUARDS_LOCK:
        guards = purge_expired_apify_run_guards(load_apify_run_guards())
        if guard_key in guards:
            guards.pop(guard_key, None)
            write_json_file(APIFY_RUN_GUARDS_FILE, guards)


def build_apify_guard_key(actor_id, input_data):
    canonical_payload = json.dumps(
        {
            "actor_id": str(actor_id or "").strip(),
            "input_data": input_data or {},
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest()


def build_apify_run_guard_record(actor_id, input_data, token, **extra):
    return {
        "actor_id": str(actor_id or "").strip(),
        "input_data": input_data or {},
        "token_masked": mask_apify_token(token),
        "created_at": iso_now(),
        "expires_at_ts": time.time() + APIFY_GUARD_TTL_SECONDS,
        **extra,
    }


def get_requested_results_per_identifier(platform, limit):
    normalized_platform = str(platform or "").strip().lower()
    try:
        resolved_limit = max(1, int(limit or 1))
    except Exception:
        resolved_limit = 1
    if normalized_platform in {"tiktok", "youtube"}:
        return resolved_limit
    return 1


def estimate_identifier_cost_usd(platform, limit):
    normalized_platform = str(platform or "").strip().lower()
    per_identifier_cost = safe_positive_float(
        PLATFORM_ESTIMATED_COST_PER_IDENTIFIER_USD.get(normalized_platform),
        0.0,
    )
    per_result_cost = safe_positive_float(
        PLATFORM_ESTIMATED_COST_PER_RESULT_USD.get(normalized_platform),
        0.0,
    )
    requested_results = get_requested_results_per_identifier(normalized_platform, limit)
    if per_result_cost > 0 and requested_results > 0:
        per_identifier_cost = max(per_identifier_cost, per_result_cost * requested_results)
    return round(per_identifier_cost, 6)


def estimate_apify_batch_cost_usd(platform, batch, payload):
    total_cost = estimate_identifier_cost_usd(platform, payload.get("limit", 1)) * len(batch or [])
    if str(platform or "").strip().lower() == "instagram" and bool(payload.get("includeAbout", True)):
        total_cost += INSTAGRAM_ABOUT_ADDON_COST_PER_PROFILE_USD * len(batch or [])
    return round(total_cost, 6)


def apply_apify_budget_guard_band(estimated_cost_usd):
    guarded_cost = safe_positive_float(estimated_cost_usd, 0.0) * APIFY_BUDGET_SAFETY_MULTIPLIER
    guarded_cost += APIFY_BUDGET_BUFFER_USD
    return round(guarded_cost, 6)


def acquire_apify_token_candidates():
    pool = get_apify_token_pool()
    if not pool:
        return []
    with APIFY_TOKEN_LOCK:
        state = load_apify_token_state()
        next_index = int(state.get("next_index") or 0)
        slot = next_index % len(pool)
        state.update({
            "next_index": next_index + 1,
            "last_slot": slot,
            "token_count": len(pool),
            "updated_at": iso_now(),
        })
        save_apify_token_state(state)
    return pool[slot:] + pool[:slot]


def remember_apify_budget_snapshot(snapshot):
    token = str((snapshot or {}).get("token") or "").strip()
    if not token:
        return
    with APIFY_TOKEN_LOCK:
        state = load_apify_token_state()
        tokens = state.setdefault("tokens", {})
        token_state = tokens.get(token, {})
        token_state.update({
            "masked": snapshot.get("masked") or mask_apify_token(token),
            "max_monthly_usage_usd": snapshot.get("max_monthly_usage_usd"),
            "monthly_usage_usd": snapshot.get("monthly_usage_usd"),
            "remaining_monthly_usage_usd": snapshot.get("remaining_monthly_usage_usd"),
            "monthly_usage_cycle_start_at": snapshot.get("monthly_usage_cycle_start_at"),
            "monthly_usage_cycle_end_at": snapshot.get("monthly_usage_cycle_end_at"),
            "budget_checked_at": snapshot.get("checked_at") or iso_now(),
        })
        tokens[token] = token_state
        save_apify_token_state(state)


def fetch_apify_budget_snapshot(token):
    cleaned_token = str(token or "").strip()
    if not cleaned_token:
        raise RuntimeError("缺少 Apify token，无法查询月额度。")

    response = apify_request(
        "GET",
        f"{APIFY_API_BASE}/users/me/limits",
        token=cleaned_token,
    )
    if response.status_code != 200:
        raise RuntimeError(f"查询 Apify 月额度失败：{extract_apify_response_error(response)}")

    payload = (response.json() or {}).get("data") or {}
    limits = payload.get("limits") or {}
    current = payload.get("current") or {}
    cycle = payload.get("monthlyUsageCycle") or {}
    max_monthly_usage_usd = round(safe_positive_float(limits.get("maxMonthlyUsageUsd"), 0.0), 6)
    monthly_usage_usd = round(safe_positive_float(current.get("monthlyUsageUsd"), 0.0), 6)
    snapshot = {
        "token": cleaned_token,
        "masked": mask_apify_token(cleaned_token),
        "max_monthly_usage_usd": max_monthly_usage_usd,
        "monthly_usage_usd": monthly_usage_usd,
        "remaining_monthly_usage_usd": round(max(0.0, max_monthly_usage_usd - monthly_usage_usd), 6),
        "monthly_usage_cycle_start_at": (
            cycle.get("startedAt")
            or cycle.get("startAt")
            or cycle.get("startDate")
            or current.get("monthlyUsageCycleStartedAt")
        ),
        "monthly_usage_cycle_end_at": (
            cycle.get("endsAt")
            or cycle.get("endAt")
            or cycle.get("endDate")
            or current.get("monthlyUsageCycleEndsAt")
        ),
        "checked_at": iso_now(),
    }
    remember_apify_budget_snapshot(snapshot)
    return snapshot


def collect_apify_budget_snapshots_for_tokens(tokens):
    snapshots = []
    errors = []
    for token in tokens or []:
        try:
            snapshots.append(fetch_apify_budget_snapshot(token))
        except Exception as exc:
            errors.append({
                "token_masked": mask_apify_token(token),
                "error": sanitize_apify_error_text(str(exc), tokens=tokens),
            })
    return snapshots, errors


def collect_apify_budget_snapshots():
    token_pool = get_apify_token_pool()
    return collect_apify_budget_snapshots_for_tokens(token_pool)


def select_apify_token_for_batch(token_candidates, required_budget_usd):
    checked_snapshots = []
    query_errors = []
    insufficient_snapshots = []
    normalized_required_budget = round(safe_positive_float(required_budget_usd, 0.0), 6)

    for token in token_candidates or []:
        try:
            snapshot = fetch_apify_budget_snapshot(token)
        except Exception as exc:
            query_errors.append({
                "token_masked": mask_apify_token(token),
                "error": sanitize_apify_error_text(str(exc), tokens=token_candidates),
            })
            continue

        checked_snapshots.append(snapshot)
        remaining = safe_positive_float(snapshot.get("remaining_monthly_usage_usd"), 0.0)
        if remaining >= normalized_required_budget:
            return {
                "selected_token": token,
                "selected_snapshot": snapshot,
                "required_budget_usd": normalized_required_budget,
                "checked_snapshots": checked_snapshots,
                "insufficient_snapshots": insufficient_snapshots,
                "query_errors": query_errors,
            }

        insufficient_snapshots.append({
            "token_masked": snapshot.get("masked"),
            "remaining_monthly_usage_usd": snapshot.get("remaining_monthly_usage_usd"),
        })

    best_remaining = max(
        (
            safe_positive_float(item.get("remaining_monthly_usage_usd"), 0.0)
            for item in checked_snapshots
        ),
        default=0.0,
    )
    if checked_snapshots:
        error = (
            f"Apify 预算不足：当前批次需要约 {normalized_required_budget:.6f} USD，"
            f"但可用 token 的最高剩余额度只有 {best_remaining:.6f} USD。"
        )
        error_code = "APIFY_BUDGET_INSUFFICIENT"
    else:
        joined_errors = "；".join(item["error"] for item in query_errors if item.get("error"))
        error = f"Apify 预算查询失败{f'：{joined_errors}' if joined_errors else ''}"
        error_code = "APIFY_BUDGET_CHECK_FAILED"

    return {
        "selected_token": "",
        "selected_snapshot": None,
        "required_budget_usd": normalized_required_budget,
        "checked_snapshots": checked_snapshots,
        "insufficient_snapshots": insufficient_snapshots,
        "query_errors": query_errors,
        "error": error,
        "error_code": error_code,
    }


def summarize_apify_budget_snapshots(snapshots):
    return {
        "max_monthly_usage_usd_total": round(
            sum(safe_positive_float(item.get("max_monthly_usage_usd"), 0.0) for item in snapshots),
            6,
        ),
        "monthly_usage_usd_total": round(
            sum(safe_positive_float(item.get("monthly_usage_usd"), 0.0) for item in snapshots),
            6,
        ),
        "remaining_monthly_usage_usd_total": round(
            sum(safe_positive_float(item.get("remaining_monthly_usage_usd"), 0.0) for item in snapshots),
            6,
        ),
    }


def iso_to_epoch_seconds(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        normalized = text.replace("Z", "+00:00")
        return datetime.fromisoformat(normalized).timestamp()
    except Exception:
        return None


def enrich_apify_balance_payload(payload, status_code, *, data_source, cache_record=None):
    normalized_payload = dict(payload or {})
    normalized_payload["status_code"] = int(status_code or 200)
    normalized_payload["data_source"] = data_source
    normalized_payload["refresh_interval_seconds"] = APIFY_BALANCE_REFRESH_INTERVAL_SECONDS

    record = cache_record or {}
    fetched_at = str(record.get("fetched_at") or normalized_payload.get("checked_at") or "").strip()
    normalized_payload["cache_fetched_at"] = fetched_at or None
    cache_age_seconds = None
    fetched_ts = iso_to_epoch_seconds(fetched_at)
    if fetched_ts is not None:
        cache_age_seconds = max(0, int(time.time() - fetched_ts))
    normalized_payload["cache_age_seconds"] = cache_age_seconds
    normalized_payload["cache_is_stale"] = (
        cache_age_seconds is not None and cache_age_seconds > APIFY_BALANCE_REFRESH_INTERVAL_SECONDS * 2
    )
    normalized_payload["last_refresh_attempt_at"] = record.get("last_refresh_attempt_at") or normalized_payload.get("checked_at")
    normalized_payload["background_refresh_error"] = record.get("last_refresh_error")
    normalized_payload["background_refresh_error_code"] = record.get("last_refresh_error_code")
    return normalized_payload


def build_live_apify_balance_payload():
    ensure_runtime_dirs()
    token_pool = get_apify_token_pool()
    if not token_pool:
        return {
            "success": False,
            "error_code": "MISSING_APIFY_CONFIG",
            "error": "缺少 Apify 配置：请设置 APIFY_TOKEN 或 APIFY_API_TOKEN。",
        }, 400

    snapshots, errors = collect_apify_budget_snapshots()
    if not snapshots:
        joined_errors = "；".join(item["error"] for item in errors if item.get("error"))
        return {
            "success": False,
            "error_code": "APIFY_BALANCE_QUERY_FAILED",
            "error": f"查询 Apify 余额失败{f'：{joined_errors}' if joined_errors else ''}",
            "token_pool_size": len(token_pool),
            "checked_token_count": 0,
            "failed_token_count": len(errors),
            "checked_at": iso_now(),
            "errors": errors,
            "tokens": [],
            "summary": {},
        }, 502

    return {
        "success": True,
        "token_pool_size": len(token_pool),
        "checked_token_count": len(snapshots),
        "failed_token_count": len(errors),
        "checked_at": iso_now(),
        "summary": summarize_apify_budget_snapshots(snapshots),
        "tokens": [
            {
                "token_masked": item.get("masked"),
                "max_monthly_usage_usd": item.get("max_monthly_usage_usd"),
                "monthly_usage_usd": item.get("monthly_usage_usd"),
                "remaining_monthly_usage_usd": item.get("remaining_monthly_usage_usd"),
                "monthly_usage_cycle_start_at": item.get("monthly_usage_cycle_start_at"),
                "monthly_usage_cycle_end_at": item.get("monthly_usage_cycle_end_at"),
                "checked_at": item.get("checked_at"),
            }
            for item in snapshots
        ],
        "errors": errors,
    }, 200


def read_cached_apify_balance_payload():
    cache_record = load_apify_balance_cache()
    payload = cache_record.get("payload")
    if not isinstance(payload, dict):
        return None, None
    status_code = int(cache_record.get("status_code") or 200)
    return enrich_apify_balance_payload(
        payload,
        status_code,
        data_source="cache",
        cache_record=cache_record,
    ), status_code


def refresh_apify_balance_cache():
    with APIFY_BALANCE_REFRESH_LOCK:
        previous_cache = load_apify_balance_cache()
        live_payload, status_code = build_live_apify_balance_payload()
        now = iso_now()
        cache_record = {
            "refresh_interval_seconds": APIFY_BALANCE_REFRESH_INTERVAL_SECONDS,
            "last_refresh_attempt_at": now,
        }

        previous_payload = previous_cache.get("payload")
        if status_code < 400 and live_payload.get("success") is True:
            cache_record["payload"] = live_payload
            cache_record["status_code"] = status_code
            cache_record["fetched_at"] = live_payload.get("checked_at") or now
            cache_record["last_refresh_error"] = None
            cache_record["last_refresh_error_code"] = None
        else:
            cache_record["last_refresh_error"] = live_payload.get("error")
            cache_record["last_refresh_error_code"] = live_payload.get("error_code")
            if isinstance(previous_payload, dict):
                cache_record["payload"] = previous_payload
                cache_record["status_code"] = int(previous_cache.get("status_code") or 200)
                cache_record["fetched_at"] = previous_cache.get("fetched_at") or previous_payload.get("checked_at") or now
            else:
                cache_record["payload"] = live_payload
                cache_record["status_code"] = status_code
                cache_record["fetched_at"] = now

        save_apify_balance_cache(cache_record)
        return enrich_apify_balance_payload(
            cache_record.get("payload"),
            cache_record.get("status_code"),
            data_source="live",
            cache_record=cache_record,
        ), int(cache_record.get("status_code") or status_code or 200)


def build_apify_balance_payload(force_refresh=False):
    if force_refresh:
        return refresh_apify_balance_cache()

    cached_payload, cached_status_code = read_cached_apify_balance_payload()
    if cached_payload is not None:
        return cached_payload, cached_status_code
    return refresh_apify_balance_cache()


def apify_balance_poller_loop():
    while not APIFY_BALANCE_POLLER_STOP_EVENT.is_set():
        try:
            refresh_apify_balance_cache()
        except Exception as exc:
            error_payload = {
                "success": False,
                "error_code": "APIFY_BALANCE_POLLER_FAILED",
                "error": sanitize_apify_error_text(str(exc), tokens=get_apify_token_pool()),
                "checked_at": iso_now(),
                "errors": [],
                "tokens": [],
                "summary": {},
            }
            cache_record = load_apify_balance_cache()
            cache_record["last_refresh_attempt_at"] = iso_now()
            cache_record["last_refresh_error"] = error_payload["error"]
            cache_record["last_refresh_error_code"] = error_payload["error_code"]
            if not isinstance(cache_record.get("payload"), dict):
                cache_record["payload"] = error_payload
                cache_record["status_code"] = 502
                cache_record["fetched_at"] = error_payload["checked_at"]
            save_apify_balance_cache(cache_record)
        if APIFY_BALANCE_POLLER_STOP_EVENT.wait(APIFY_BALANCE_REFRESH_INTERVAL_SECONDS):
            return


def start_apify_balance_poller():
    global APIFY_BALANCE_POLLER_THREAD
    if not APIFY_BALANCE_POLLER_ENABLED:
        return None
    with APIFY_BALANCE_POLLER_LOCK:
        if APIFY_BALANCE_POLLER_THREAD and APIFY_BALANCE_POLLER_THREAD.is_alive():
            return APIFY_BALANCE_POLLER_THREAD
        APIFY_BALANCE_POLLER_STOP_EVENT.clear()
        APIFY_BALANCE_POLLER_THREAD = threading.Thread(
            target=apify_balance_poller_loop,
            daemon=True,
            name="apify-balance-poller",
        )
        APIFY_BALANCE_POLLER_THREAD.start()
        return APIFY_BALANCE_POLLER_THREAD


def normalize_vision_provider_name(provider_name):
    return str(provider_name or "").strip().lower()


def resolve_vision_provider_request(provider_name=None):
    explicit_provider = normalize_vision_provider_name(provider_name)
    if explicit_provider:
        return explicit_provider, "explicit"
    env_provider = normalize_vision_provider_name(os.getenv("VISION_PROVIDER_PREFERENCE", ""))
    if env_provider:
        return env_provider, "env"
    return "", "default"


def resolve_env_source(env_key):
    cleaned_key = str(env_key or "").strip()
    if not cleaned_key:
        return "static"
    value = str(os.getenv(cleaned_key, "") or "").strip()
    if not value:
        return "missing"
    if cleaned_key in DOTENV_LOCAL_LOADED_KEYS:
        return "env.local"
    return "process_env"


def build_env_resolution_details(env_key):
    cleaned_key = str(env_key or "").strip()
    value = str(os.getenv(cleaned_key, "") or "").strip()
    return {
        "env_key": cleaned_key,
        "source": resolve_env_source(cleaned_key),
        "present": bool(value),
        "dotenv_local_path": str(DOTENV_LOCAL_PATH),
        "dotenv_local_exists": DOTENV_LOCAL_PATH.exists(),
        "dotenv_local_has_key": cleaned_key in DOTENV_LOCAL_VALUES,
        "dotenv_local_loaded": cleaned_key in DOTENV_LOCAL_LOADED_KEYS,
        "process_env_present_at_boot": cleaned_key in ENV_KEYS_BEFORE_DOTENV_LOCAL,
    }


def resolve_vision_provider_api_key(provider):
    env_key = str((provider or {}).get("env_key") or "").strip()
    return str(os.getenv(env_key, "") or "").strip()


def resolve_vision_provider_base_url(provider):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    base_url_env_key = str((provider or {}).get("base_url_env_key") or "").strip()
    default_base_url = str((provider or {}).get("default_base_url") or "").strip()
    base_url = str(os.getenv(base_url_env_key, default_base_url) or "").strip()
    normalized = base_url.rstrip("/")
    if provider_name == "quan2go":
        if normalized.endswith("/openai"):
            return f"{normalized[:-len('/openai')]}/v1"
        if normalized.endswith("/openai/v1"):
            return f"{normalized[:-len('/openai/v1')]}/v1"
    return normalized


def resolve_vision_provider_base_urls(provider):
    primary_base_url = resolve_vision_provider_base_url(provider)
    candidates = []

    def _append(value):
        normalized = str(value or "").strip().rstrip("/")
        if not normalized or normalized in candidates:
            return
        candidates.append(normalized)

    _append(primary_base_url)
    fallback_env_key = str((provider or {}).get("base_url_fallbacks_env_key") or "").strip()
    fallback_values = []
    if fallback_env_key:
        raw_fallbacks = str(os.getenv(fallback_env_key, "") or "").strip()
        if raw_fallbacks:
            fallback_values = [item.strip() for item in raw_fallbacks.split(",")]
    if not fallback_values:
        fallback_values = list((provider or {}).get("default_base_url_fallbacks") or [])
    for value in fallback_values:
        _append(value)
    return candidates


def resolve_vision_provider_model(provider):
    model_override = str((provider or {}).get("model_override") or "").strip()
    if model_override:
        return model_override
    model_env_key = str((provider or {}).get("model_env_key") or "").strip()
    provider_model = str(os.getenv(model_env_key, "") or "").strip()
    if provider_model:
        return provider_model
    default_model = str((provider or {}).get("default_model") or "").strip()
    if default_model:
        return default_model
    return str(os.getenv("VISION_MODEL", DEFAULT_VISION_MODEL) or "").strip() or DEFAULT_VISION_MODEL


def resolve_vision_provider_fallback_model(provider):
    fallback_override = str((provider or {}).get("fallback_model_override") or "").strip()
    if fallback_override:
        return fallback_override
    env_key = str((provider or {}).get("fallback_model_env_key") or "").strip()
    env_value = str(os.getenv(env_key, "") or "").strip() if env_key else ""
    if env_value:
        return env_value
    return str((provider or {}).get("default_fallback_model") or "").strip()


def resolve_vision_provider_model_candidates(provider):
    candidates = [resolve_vision_provider_model(provider), resolve_vision_provider_fallback_model(provider)]
    normalized = []
    seen = set()
    for candidate in candidates:
        cleaned = str(candidate or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        normalized.append(cleaned)
    return normalized


def resolve_vision_provider_max_completion_tokens(provider):
    env_key = str((provider or {}).get("max_completion_tokens_env_key") or "").strip()
    raw_value = str(os.getenv(env_key, "") or "").strip() if env_key else ""
    if raw_value:
        try:
            return max(1, int(raw_value))
        except Exception:
            pass
    default_value = (provider or {}).get("default_max_completion_tokens")
    if default_value in (None, ""):
        return None
    try:
        return max(1, int(default_value))
    except Exception:
        return None


def resolve_vision_provider_max_tokens(provider):
    env_key = str((provider or {}).get("max_tokens_env_key") or "").strip()
    raw_value = str(os.getenv(env_key, "") or "").strip() if env_key else ""
    if raw_value:
        try:
            return max(1, int(raw_value))
        except Exception:
            pass
    default_value = (provider or {}).get("default_max_tokens")
    if default_value in (None, ""):
        return None
    try:
        return max(1, int(default_value))
    except Exception:
        return None


def resolve_vision_provider_temperature(provider):
    env_key = str((provider or {}).get("temperature_env_key") or "").strip()
    raw_value = str(os.getenv(env_key, "") or "").strip() if env_key else ""
    if raw_value:
        try:
            return float(raw_value)
        except Exception:
            pass
    default_value = (provider or {}).get("default_temperature")
    if default_value in (None, ""):
        return None
    try:
        return float(default_value)
    except Exception:
        return None


def resolve_vision_provider_request_timeout(provider):
    value = (provider or {}).get("request_timeout_seconds")
    if value not in (None, ""):
        try:
            return max(1, int(value))
        except Exception:
            pass
    return VISION_REQUEST_TIMEOUT


def build_vision_provider_headers(provider):
    header_name = str((provider or {}).get("auth_header_name") or "Authorization").strip() or "Authorization"
    value_prefix = str((provider or {}).get("auth_header_value_prefix") or "").strip()
    if not value_prefix and header_name.lower() == "authorization":
        value_prefix = "Bearer"
    header_value = f"{value_prefix} {provider['api_key']}".strip() if value_prefix else str(provider["api_key"])
    return {
        header_name: header_value,
        "Content-Type": "application/json",
    }


def build_vision_provider_chat_messages(provider, content):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    if provider_name == "mimo" and isinstance(content, str):
        return [{"role": "user", "content": [{"type": "text", "text": content}]}]
    return [{"role": "user", "content": content}]


def build_vision_provider_chat_body(provider, messages, model_override=""):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    body = {
        "model": str(model_override or resolve_vision_provider_model(provider)).strip(),
        "messages": messages,
    }
    if provider_name == "qiandao":
        max_tokens = resolve_vision_provider_max_tokens(provider)
        if max_tokens:
            body["max_tokens"] = max_tokens
        temperature = resolve_vision_provider_temperature(provider)
        if temperature is not None:
            body["temperature"] = temperature
        body["stream"] = False
        return body
    max_completion_tokens = resolve_vision_provider_max_completion_tokens(provider)
    if max_completion_tokens:
        body["max_completion_tokens"] = max_completion_tokens
    return body


def build_vision_provider_generate_content_body(parts):
    return {
        "contents": [{"role": "user", "parts": parts}],
        "generationConfig": {
            "temperature": 0,
            "maxOutputTokens": 400,
            "responseMimeType": "application/json",
        },
    }


def is_qwen_visual_model(provider_name, model_name=""):
    normalized_provider = normalize_vision_provider_name(provider_name)
    normalized_model = str(model_name or "").strip().lower()
    return "qwen" in normalized_provider or "qwen" in normalized_model


def extract_visual_prompt_text(prompt_candidate):
    if isinstance(prompt_candidate, str):
        return prompt_candidate.strip()
    if isinstance(prompt_candidate, dict):
        for key in ("prompt", "text", "content"):
            value = str(prompt_candidate.get(key) or "").strip()
            if value:
                return value
    return ""


def normalize_visual_prompt_lookup_key(value):
    return str(value or "").strip().lower()


def resolve_visual_prompt_mapping_prompt(mapping, lookup_key):
    normalized_lookup = normalize_visual_prompt_lookup_key(lookup_key)
    if not normalized_lookup or not isinstance(mapping, dict):
        return ""
    direct = extract_visual_prompt_text(mapping.get(normalized_lookup))
    if direct:
        return direct
    for candidate_key, candidate_value in mapping.items():
        if normalize_visual_prompt_lookup_key(candidate_key) == normalized_lookup:
            return extract_visual_prompt_text(candidate_value)
    return ""


def resolve_active_visual_prompt_bundle(active_visual_prompts, platform):
    if not isinstance(active_visual_prompts, dict):
        return {}
    normalized_platform = normalize_visual_prompt_lookup_key(platform)
    if not normalized_platform:
        return {}
    for container_key in ("platform_prompts", "platforms"):
        container = active_visual_prompts.get(container_key)
        if isinstance(container, dict):
            for candidate_key, candidate_value in container.items():
                if normalize_visual_prompt_lookup_key(candidate_key) == normalized_platform and isinstance(candidate_value, dict):
                    return candidate_value
    for candidate_key, candidate_value in active_visual_prompts.items():
        if normalize_visual_prompt_lookup_key(candidate_key) == normalized_platform and isinstance(candidate_value, dict):
            return candidate_value
    return {}


def load_active_visual_prompts():
    payload = load_json_payload(ACTIVE_VISUAL_PROMPTS_PATH, default={})
    return payload if isinstance(payload, dict) else {}


def resolve_generic_visual_review_prompt(provider_name, model_name=""):
    normalized_provider = normalize_vision_provider_name(provider_name)
    if normalized_provider == "mimo":
        return MIMO_VISION_PROMPT
    if is_qwen_visual_model(normalized_provider, model_name):
        return QWEN_VISION_PROMPT_V2
    return VISION_PROMPT


VISUAL_CONTRACT_SENSITIVE_TERMS = (
    "黑人",
    "白人",
    "黄种人",
    "亚裔",
    "拉丁裔",
    "中老年",
    "老年",
    "老人",
    "种族",
    "民族",
    "年龄",
    "race",
    "ethnic",
    "black",
    "white",
)


def normalize_visual_contract_prompt_text(value):
    return str(value or "").strip()


def render_visual_contract_manual_item(item):
    if not isinstance(item, dict):
        return normalize_visual_contract_prompt_text(item)
    label = normalize_visual_contract_prompt_text(item.get("label"))
    value = normalize_visual_contract_prompt_text(item.get("value"))
    note = normalize_visual_contract_prompt_text(item.get("note"))
    if label and value and value != label:
        return f"{label}：{value}"
    if label and note and note != label:
        return f"{label}：{note}"
    return label or value or note


def render_visual_contract_compliance_note(item):
    if not isinstance(item, dict):
        return normalize_visual_contract_prompt_text(item)
    label = normalize_visual_contract_prompt_text(item.get("label"))
    value = normalize_visual_contract_prompt_text(item.get("value"))
    note = normalize_visual_contract_prompt_text(item.get("note"))
    if label and value and value != label:
        return f"{label}：{value}"
    if value:
        return value
    if label and note and note != label:
        return f"{label}：{note}"
    return label or note


def visual_contract_has_protected_attribute_notice(compliance_notes):
    for item in compliance_notes or []:
        if not isinstance(item, dict):
            continue
        combined = " ".join(
            normalize_visual_contract_prompt_text(item.get(field))
            for field in ("key", "label", "value", "note", "policy")
        ).lower()
        if "protected_attribute" in combined or "受保护属性" in combined:
            return True
    return False


def visual_contract_is_sensitive_compliance_item(item):
    if not isinstance(item, dict):
        combined = normalize_visual_contract_prompt_text(item).lower()
    else:
        combined = " ".join(
            normalize_visual_contract_prompt_text(item.get(field))
            for field in ("key", "label", "value", "note", "policy")
        ).lower()
    return any(term.lower() in combined for term in VISUAL_CONTRACT_SENSITIVE_TERMS)


def build_rulespec_visual_contract_prompt(platform, visual_contract):
    if not isinstance(visual_contract, dict):
        return ""

    goal = str(visual_contract.get("goal") or "").strip()
    feature_labels = [
        str(item or "").strip()
        for item in (visual_contract.get("positive_feature_labels") or [])
        if str(item or "").strip()
    ]
    exclusion_summaries = [
        str(item or "").strip()
        for item in (visual_contract.get("exclusion_summaries") or [])
        if str(item or "").strip()
    ]
    manual_review_items = [
        item
        for item in (
            render_visual_contract_manual_item(raw_item)
            for raw_item in (visual_contract.get("manual_review_items") or [])
        )
        if item
    ]
    compliance_notes_raw = list(visual_contract.get("compliance_notes") or [])
    compliance_note_pairs = [
        (raw_item, render_visual_contract_compliance_note(raw_item))
        for raw_item in compliance_notes_raw
        if render_visual_contract_compliance_note(raw_item)
    ]
    include_protected_notice = visual_contract_has_protected_attribute_notice(compliance_notes_raw)
    cover_count = screening.coerce_positive_int(visual_contract.get("cover_count")) or VISUAL_REVIEW_REQUEST_COVER_LIMIT
    min_hit_features = screening.coerce_positive_int(visual_contract.get("min_hit_features")) or 1
    if not goal and not feature_labels and not exclusion_summaries:
        return ""

    lines = [
        f"你是 {UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 达人初筛流程中的视觉复核员。输入图片是一位博主最近最多 {cover_count} 张封面，请综合全部输入图片一起判断。",
        "",
        "只根据图片画面做初步判断，不要假设看不到的内容。",
    ]
    if goal:
        lines.extend(["", f"审核目标：{goal}"])
    if feature_labels:
        lines.extend(["", f"优先确认是否命中以下至少 {min_hit_features} 类视觉特征："])
        for label in feature_labels:
            lines.append(f"- {label}")
        lines.append(f'如果以上特征全部未命中，输出 Reject，reason 写"未命中{"/".join(feature_labels)}"。')
    if exclusion_summaries:
        lines.extend(["", "同时排除以下视觉风险："])
        for label in exclusion_summaries:
            lines.append(f"- {label}")
    if manual_review_items:
        lines.extend(
            [
                "",
                "人工复核提醒：以下情况不要直接当作自动通过或自动拒绝条件；若明显出现，请在 reason 或 signals 里点明，供人工复核：",
            ]
        )
        for item in manual_review_items:
            lines.append(f"- {item}")
    if include_protected_notice or compliance_note_pairs:
        lines.extend(["", "合规提醒："])
        if include_protected_notice:
            lines.append("- 不要根据年龄、种族、民族、肤色、宗教等受保护属性做判断。")
        for raw_item, rendered_note in compliance_note_pairs:
            if include_protected_notice and "受保护属性" in rendered_note:
                continue
            if include_protected_notice and visual_contract_is_sensitive_compliance_item(raw_item):
                continue
            lines.append(f"- {rendered_note}")
    lines.extend(
        [
            "",
            "如果通过以上检查，输出 Pass。",
            "",
            "请只返回 JSON，不要 markdown，不要额外说明，格式固定为：",
            '{"decision":"Pass 或 Reject","reason":"一句中文原因","signals":["最多 3 个简短中文信号"]}',
        ]
    )
    return "\n".join(lines)


def resolve_visual_review_prompt_selection(
    provider_name,
    platform,
    model_name="",
    active_visual_prompts=None,
    active_rulespec=None,
):
    normalized_provider = normalize_vision_provider_name(provider_name)
    normalized_model = normalize_visual_prompt_lookup_key(model_name)
    resolved_active_rulespec = load_active_rulespec() if active_rulespec is None else active_rulespec
    runtime_rules = screening.get_runtime_rules(resolved_active_rulespec, platform)
    visual_contract = dict(runtime_rules.get("visual_runtime_contract") or {})
    visual_contract_source = str(runtime_rules.get("visual_contract_source") or "").strip()
    resolved_cover_limit = screening.resolve_visual_review_request_cover_limit(runtime_rules)
    prompt_bundle = resolve_active_visual_prompt_bundle(
        load_active_visual_prompts() if active_visual_prompts is None else active_visual_prompts,
        platform,
    )
    base_payload = {
        "platform": normalize_visual_prompt_lookup_key(platform),
        "provider": normalized_provider,
        "model": normalized_model,
        "active_visual_prompts_path": ACTIVE_VISUAL_PROMPTS_PATH,
        "visual_contract_source": visual_contract_source,
        "visual_runtime_contract": visual_contract,
        "resolved_cover_limit": resolved_cover_limit,
    }
    provider_prompt = resolve_visual_prompt_mapping_prompt(prompt_bundle.get("provider_prompts"), normalized_provider)
    if provider_prompt:
        return {
            **base_payload,
            "prompt": provider_prompt,
            "source": "provider_prompts",
        }
    model_prompt = resolve_visual_prompt_mapping_prompt(prompt_bundle.get("model_prompts"), normalized_model)
    if model_prompt:
        return {
            **base_payload,
            "prompt": model_prompt,
            "source": "model_prompts",
        }
    platform_prompt = extract_visual_prompt_text(prompt_bundle.get("prompt"))
    if platform_prompt:
        return {
            **base_payload,
            "prompt": platform_prompt,
            "source": "platform_prompt",
        }
    contract_prompt = build_rulespec_visual_contract_prompt(platform, visual_contract)
    if contract_prompt:
        return {
            **base_payload,
            "prompt": contract_prompt,
            "source": "rulespec_visual_contract",
        }
    return {
        **base_payload,
        "prompt": resolve_generic_visual_review_prompt(normalized_provider, normalized_model),
        "source": "generic_fallback",
    }


def build_visual_review_prompt(provider_name, platform, username, model_name=""):
    selection = resolve_visual_review_prompt_selection(provider_name, platform, model_name=model_name)
    return (
        f"平台：{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)}\n"
        f"达人：{username or 'unknown'}\n"
        f"{selection['prompt']}"
    )


def build_generic_positioning_card_prompt(platform, visual_contract):
    goal = str((visual_contract or {}).get("goal") or "").strip()
    feature_labels = [
        str(item or "").strip()
        for item in ((visual_contract or {}).get("positive_feature_labels") or [])
        if str(item or "").strip()
    ]
    exclusion_summaries = [
        str(item or "").strip()
        for item in ((visual_contract or {}).get("exclusion_summaries") or [])
        if str(item or "").strip()
    ]
    cover_count = screening.coerce_positive_int((visual_contract or {}).get("cover_count")) or VISUAL_REVIEW_REQUEST_COVER_LIMIT
    lines = [
        f"你是 {UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 达人初筛流程中的定位卡分析员。输入图片是一位已通过视觉复核的博主最近最多 {cover_count} 张封面，请综合全部输入图片判断她/他的内容定位与品牌适配度。",
        "",
        "只根据图片画面做定位判断，不要假设文案、口播或评论区里出现了什么。",
    ]
    if goal:
        lines.extend(["", f"品牌审核目标：{goal}"])
    if feature_labels:
        lines.extend(["", "优先识别以下品牌相关视觉特征或内容方向："])
        for label in feature_labels:
            lines.append(f"- {label}")
    if exclusion_summaries:
        lines.extend(["", "以下风险或不适配信号若明显出现，应下调 fit_recommendation："])
        for label in exclusion_summaries:
            lines.append(f"- {label}")
    lines.extend(["", POSITIONING_CARD_PROMPT])
    return "\n".join(lines)


def resolve_positioning_card_prompt_selection(provider_name, platform, model_name="", active_rulespec=None):
    normalized_provider = normalize_vision_provider_name(provider_name)
    normalized_model = normalize_visual_prompt_lookup_key(model_name)
    resolved_active_rulespec = load_active_rulespec() if active_rulespec is None else active_rulespec
    runtime_rules = screening.get_runtime_rules(resolved_active_rulespec, platform)
    visual_contract = dict(runtime_rules.get("visual_runtime_contract") or {})
    visual_contract_source = str(runtime_rules.get("visual_contract_source") or "").strip()
    resolved_cover_limit = screening.resolve_visual_review_request_cover_limit(runtime_rules)
    return {
        "platform": normalize_visual_prompt_lookup_key(platform),
        "provider": normalized_provider,
        "model": normalized_model,
        "visual_contract_source": visual_contract_source,
        "visual_runtime_contract": visual_contract,
        "resolved_cover_limit": resolved_cover_limit,
        "prompt": build_generic_positioning_card_prompt(platform, visual_contract),
        "source": "generic_brand_fit",
    }


def build_positioning_card_prompt(provider_name, platform, username, model_name=""):
    selection = resolve_positioning_card_prompt_selection(provider_name, platform, model_name=model_name)
    return (
        f"平台：{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)}\n"
        f"达人：{username or 'unknown'}\n"
        f"{selection['prompt']}"
    )


def build_vision_provider_snapshot(provider):
    normalized_name = normalize_vision_provider_name((provider or {}).get("name"))
    api_key = resolve_vision_provider_api_key(provider)
    base_url = resolve_vision_provider_base_url(provider)
    base_url_candidates = resolve_vision_provider_base_urls(provider)
    api_style = str((provider or {}).get("api_style") or VISION_API_STYLE_RESPONSES).strip().lower()
    model_env_key = str((provider or {}).get("model_env_key") or "").strip()
    provider_model = str(os.getenv(model_env_key, "") or "").strip()
    model = resolve_vision_provider_model(provider)
    issues = []
    parsed_base_url = urlparse(base_url) if base_url else None
    if not api_key:
        issues.append("missing_api_key")
    if not base_url:
        issues.append("missing_base_url")
    elif parsed_base_url.scheme not in {"http", "https"} or not parsed_base_url.netloc:
        issues.append("invalid_base_url")
    if api_style not in {VISION_API_STYLE_RESPONSES, VISION_API_STYLE_CHAT_COMPLETIONS, VISION_API_STYLE_GENERATE_CONTENT}:
        issues.append("unsupported_api_style")
    if not model:
        issues.append("missing_model")

    runnable = not issues
    status = "runnable" if runnable else ("configured_with_issues" if api_key else "missing_config")
    model_source_key = model_env_key if provider_model else "VISION_MODEL"
    model_source = resolve_env_source(model_source_key)
    if model_source == "missing":
        model_source = "default"
    base_url_env_key = str((provider or {}).get("base_url_env_key") or "").strip()
    base_url_source = resolve_env_source(base_url_env_key)
    if base_url_source == "missing" and base_url == str((provider or {}).get("default_base_url") or "").strip():
        base_url_source = "default"
    return {
        "name": normalized_name,
        "status": status,
        "runnable": runnable,
        "env_key": str((provider or {}).get("env_key") or "").strip(),
        "api_key_present": bool(api_key),
        "api_key_masked": mask_apify_token(api_key),
        "api_key_source": resolve_env_source((provider or {}).get("env_key")),
        "api_key_resolution": build_env_resolution_details((provider or {}).get("env_key")),
        "base_url": base_url,
        "base_url_candidates": base_url_candidates,
        "base_url_env_key": base_url_env_key,
        "base_url_source": base_url_source,
        "base_url_resolution": build_env_resolution_details(base_url_env_key) if base_url_env_key else {},
        "model": model,
        "model_env_key": model_source_key,
        "model_source": model_source,
        "model_uses_global_fallback": not bool(provider_model),
        "model_resolution": build_env_resolution_details(model_source_key),
        "api_style": api_style,
        "issues": issues,
    }


def build_vision_provider_snapshots():
    return [build_vision_provider_snapshot(provider) for provider in VISION_PROVIDER_CONFIGS]


def build_vision_preflight(provider_name=None):
    providers = build_vision_provider_snapshots()
    configured_provider_names = [item["name"] for item in providers if item.get("api_key_present")]
    runnable_provider_names = [item["name"] for item in providers if item.get("runnable")]
    requested_provider, provider_request_source = resolve_vision_provider_request(provider_name)
    requested_provider_declared = bool(requested_provider and requested_provider in {item["name"] for item in providers})
    requested_provider_runnable = bool(requested_provider and requested_provider in runnable_provider_names)
    selected_provider = ""
    if requested_provider_runnable:
        selected_provider = requested_provider
    elif runnable_provider_names:
        selected_provider = runnable_provider_names[0]

    if requested_provider and not requested_provider_declared:
        status = "degraded" if runnable_provider_names else "unconfigured"
        error_code = "UNKNOWN_VISION_PROVIDER"
        message = f"指定视觉 provider 不存在：{requested_provider}"
    elif requested_provider and not requested_provider_runnable:
        status = "degraded" if configured_provider_names else "unconfigured"
        error_code = "VISION_PROVIDER_NOT_RUNNABLE"
        message = f"指定视觉 provider 当前不可运行：{requested_provider}"
    elif runnable_provider_names:
        status = "configured"
        error_code = ""
        message = f"视觉模型已就绪：{selected_provider or ', '.join(runnable_provider_names)}"
    elif configured_provider_names:
        status = "degraded"
        error_code = "VISION_PROVIDER_PREFLIGHT_FAILED"
        message = "视觉模型预检未通过：已检测到 provider key，但当前没有可运行 provider。请检查 base_url、api_style 和 model。"
    else:
        status = "unconfigured"
        error_code = "MISSING_VISION_CONFIG"
        message = "缺少视觉模型配置：请设置 OPENAI_API_KEY、VISION_REELX_API_KEY、VISION_MIMO_API_KEY、VISION_QIANDAO_API_KEY、VISION_QUAN2GO_API_KEY 或 VISION_LEMONAPI_API_KEY。"
    return {
        "status": status,
        "error_code": error_code,
        "message": message,
        "provider_names": [item["name"] for item in providers],
        "configured_provider_names": configured_provider_names,
        "runnable_provider_names": runnable_provider_names,
        "preferred_provider": selected_provider,
        "requested_provider": requested_provider,
        "requested_provider_source": provider_request_source,
        "requested_provider_declared": requested_provider_declared,
        "requested_provider_runnable": requested_provider_runnable,
        "providers": providers,
        "backend_env_bootstrap": {
            "dotenv_local_path": str(DOTENV_LOCAL_PATH),
            "dotenv_local_exists": DOTENV_LOCAL_PATH.exists(),
            "dotenv_local_declared_keys": sorted(DOTENV_LOCAL_VALUES.keys()),
            "dotenv_local_loaded_keys": sorted(DOTENV_LOCAL_LOADED_KEYS),
        },
    }


def build_vision_preflight_error_payload(provider_name=None):
    preflight = build_vision_preflight(provider_name)
    return {
        "success": False,
        "error_code": preflight.get("error_code") or "VISION_PROVIDER_PREFLIGHT_FAILED",
        "error": preflight.get("message") or "视觉模型预检未通过",
        "vision_preflight": preflight,
    }


def get_available_vision_providers(provider_name=None):
    snapshots = {item["name"]: item for item in build_vision_provider_snapshots()}
    requested_provider, _provider_request_source = resolve_vision_provider_request(provider_name)
    providers = []
    for provider in VISION_PROVIDER_CONFIGS:
        normalized_name = normalize_vision_provider_name(provider.get("name"))
        if requested_provider and normalized_name != requested_provider:
            continue
        snapshot = snapshots.get(normalized_name) or {}
        if not snapshot.get("runnable"):
            continue
        providers.append({
            **provider,
            "name": normalized_name,
            "api_key": resolve_vision_provider_api_key(provider),
            "base_url": snapshot.get("base_url") or resolve_vision_provider_base_url(provider),
            "base_url_candidates": snapshot.get("base_url_candidates") or resolve_vision_provider_base_urls(provider),
            "model": snapshot.get("model") or resolve_vision_provider_model(provider),
            "api_style": snapshot.get("api_style") or str(provider.get("api_style") or VISION_API_STYLE_RESPONSES).strip().lower(),
        })
    return providers


def get_available_vision_provider_names(provider_name=None):
    return [provider["name"] for provider in get_available_vision_providers(provider_name)]


def normalize_visual_review_routing_strategy(value):
    normalized = str(value or "").strip().lower()
    if normalized in {"tiered", "smart", "auto"}:
        return VISUAL_REVIEW_ROUTING_TIERED
    if normalized in {"probe_ranked", "ranked_probe", "probe-ranked", "race", "probe_race"}:
        return VISUAL_REVIEW_ROUTING_PROBE_RANKED
    return ""


def resolve_visual_review_routing_strategy(payload=None):
    payload_value = normalize_visual_review_routing_strategy((payload or {}).get("routing_strategy"))
    if payload_value:
        return payload_value
    return normalize_visual_review_routing_strategy(os.getenv("VISION_VISUAL_REVIEW_ROUTING_STRATEGY", ""))


def _resolve_visual_review_routing_timeout(env_key, default_value):
    raw_value = str(os.getenv(env_key, "") or "").strip()
    if raw_value:
        try:
            return max(1, int(raw_value))
        except Exception:
            pass
    return int(default_value)


def _resolve_visual_review_routing_follower_threshold():
    raw_value = str(os.getenv("VISION_VISUAL_REVIEW_HIGH_VALUE_FOLLOWER_THRESHOLD", "") or "").strip()
    if raw_value:
        try:
            return max(0, int(raw_value))
        except Exception:
            pass
    return int(DEFAULT_VISUAL_REVIEW_ROUTING_HIGH_VALUE_FOLLOWER_THRESHOLD)


def build_visual_review_routing_plan():
    return [
        {
            "stage": "primary",
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PRIMARY_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_ROUTING_PRIMARY_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PRIMARY_MODEL",
                    DEFAULT_VISUAL_REVIEW_ROUTING_PRIMARY_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_PRIMARY_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_ROUTING_PRIMARY_TIMEOUT_SECONDS,
            ),
        },
        {
            "stage": "backup",
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_BACKUP_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_ROUTING_BACKUP_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_BACKUP_MODEL",
                    DEFAULT_VISUAL_REVIEW_ROUTING_BACKUP_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_BACKUP_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_ROUTING_BACKUP_TIMEOUT_SECONDS,
            ),
        },
        {
            "stage": "judge",
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_JUDGE_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_ROUTING_JUDGE_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_JUDGE_MODEL",
                    DEFAULT_VISUAL_REVIEW_ROUTING_JUDGE_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_JUDGE_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_ROUTING_JUDGE_TIMEOUT_SECONDS,
            ),
        },
    ]


def build_visual_review_probe_ranked_plan():
    return [
        {
            "stage": "preferred",
            "group": VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED,
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_MODEL",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_TIMEOUT_SECONDS,
            ),
        },
        {
            "stage": "preferred_parallel",
            "group": VISUAL_REVIEW_PROBE_RANKED_GROUP_FALLBACK,
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_MODEL",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_PROBE_RANKED_PREFERRED_PARALLEL_TIMEOUT_SECONDS,
            ),
        },
        {
            "stage": "secondary",
            "group": VISUAL_REVIEW_PROBE_RANKED_GROUP_FALLBACK,
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_MODEL",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_PROBE_RANKED_SECONDARY_TIMEOUT_SECONDS,
            ),
        },
        {
            "stage": "tertiary",
            "group": VISUAL_REVIEW_PROBE_RANKED_GROUP_FALLBACK,
            "provider": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_PROVIDER",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_PROVIDER,
                )
                or ""
            ).strip().lower(),
            "model": str(
                os.getenv(
                    "VISION_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_MODEL",
                    DEFAULT_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_MODEL,
                )
                or ""
            ).strip(),
            "timeout_seconds": _resolve_visual_review_routing_timeout(
                "VISION_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_TIMEOUT_SECONDS",
                DEFAULT_VISUAL_REVIEW_PROBE_RANKED_TERTIARY_TIMEOUT_SECONDS,
            ),
        },
    ]


def clone_vision_provider_with_overrides(provider, *, model="", timeout_seconds=None):
    cloned = dict(provider or {})
    if model:
        cloned["model_override"] = str(model).strip()
        cloned["default_model"] = str(model).strip()
        cloned["fallback_model_override"] = ""
        cloned["fallback_model_env_key"] = ""
    cloned["default_fallback_model"] = ""
    if timeout_seconds not in (None, ""):
        try:
            cloned["request_timeout_seconds"] = max(1, int(timeout_seconds))
        except Exception:
            pass
    return cloned


def get_runnable_vision_provider(provider_name, *, model="", timeout_seconds=None):
    providers = get_available_vision_providers(provider_name)
    if not providers:
        return None
    return clone_vision_provider_with_overrides(
        providers[0],
        model=model,
        timeout_seconds=timeout_seconds,
    )


def normalize_visual_review_signals(value):
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value if str(item or "").strip()]


def is_visual_review_result_structurally_valid(result):
    if not isinstance(result, dict):
        return False
    decision = str(result.get("decision") or "").strip()
    reason = str(result.get("reason") or "").strip()
    signals = normalize_visual_review_signals(result.get("signals"))
    return decision in {"Pass", "Reject"} and bool(reason) and bool(signals)


def is_visual_review_result_borderline(result):
    if not isinstance(result, dict):
        return False
    text = " ".join(
        [
            str(result.get("decision") or "").strip(),
            str(result.get("reason") or "").strip(),
            *normalize_visual_review_signals(result.get("signals")),
        ]
    )
    return any(marker in text for marker in VISUAL_REVIEW_BORDERLINE_MARKERS)


def parse_visual_review_followers(value):
    text = str(value or "").strip()
    if not text:
        return 0
    normalized = text.replace(",", "").replace("_", "").strip().lower()
    multiplier = 1
    if normalized.endswith("k"):
        multiplier = 1000
        normalized = normalized[:-1]
    elif normalized.endswith("m"):
        multiplier = 1000000
        normalized = normalized[:-1]
    try:
        return int(float(normalized) * multiplier)
    except Exception:
        return 0


def is_high_value_visual_review_item(review_item):
    threshold = _resolve_visual_review_routing_follower_threshold()
    if threshold <= 0:
        return False
    metadata = dict((review_item or {}).get("upload_metadata") or {})
    candidates = [
        metadata.get("followers"),
        (review_item or {}).get("followers"),
        ((review_item or {}).get("stats") or {}).get("followers"),
    ]
    follower_count = max(parse_visual_review_followers(candidate) for candidate in candidates)
    return follower_count >= threshold


def collect_visual_review_escalation_reasons(result, review_item=None):
    reasons = []
    if not isinstance(result, dict):
        return ["missing_result"]
    decision = str(result.get("decision") or "").strip()
    reason = str(result.get("reason") or "").strip()
    signals = normalize_visual_review_signals(result.get("signals"))
    if decision not in {"Pass", "Reject"}:
        reasons.append("invalid_decision")
    if not reason:
        reasons.append("missing_reason")
    if not signals:
        reasons.append("missing_signals")
    if is_visual_review_result_borderline(result):
        reasons.append("borderline_output")
    if is_high_value_visual_review_item(review_item):
        reasons.append("high_value_account")
    return reasons


def build_visual_review_trace_entry(
    stage,
    provider,
    *,
    ok,
    model="",
    configured_model="",
    requested_model="",
    response_model="",
    effective_model="",
    decision="",
    escalation_reasons=None,
    error="",
    group="",
    retryable=None,
):
    configured_model = str(configured_model or "").strip()
    requested_model = str(requested_model or "").strip()
    response_model = str(response_model or "").strip()
    effective_model = (
        str(effective_model or "").strip()
        or str(model or "").strip()
        or response_model
        or requested_model
        or configured_model
        or str(resolve_vision_provider_model(provider)).strip()
    )
    payload = {
        "stage": str(stage or "").strip(),
        "provider": normalize_vision_provider_name((provider or {}).get("name")),
        "model": effective_model,
        "ok": bool(ok),
    }
    if configured_model:
        payload["configured_model"] = configured_model
    if requested_model:
        payload["requested_model"] = requested_model
    if response_model:
        payload["response_model"] = response_model
    if effective_model:
        payload["effective_model"] = effective_model
    if group:
        payload["group"] = str(group).strip()
    if decision:
        payload["decision"] = str(decision).strip()
    if escalation_reasons:
        payload["escalation_reasons"] = list(escalation_reasons)
    if error:
        payload["error"] = str(error).strip()
    if retryable is not None:
        payload["retryable"] = bool(retryable)
    return payload


def extract_vision_response_model(payload):
    if not isinstance(payload, dict):
        return ""
    return str(payload.get("model") or payload.get("modelVersion") or "").strip()


def _resolve_probe_ranked_disable_after_failures():
    raw_value = str(os.getenv("VISION_VISUAL_REVIEW_PROBE_RANKED_DISABLE_AFTER_FAILURES", "") or "").strip()
    if raw_value:
        try:
            return max(1, int(raw_value))
        except Exception:
            pass
    return int(DEFAULT_VISUAL_REVIEW_PROBE_RANKED_DISABLE_AFTER_FAILURES)


def _resolve_probe_ranked_retry_attempts():
    raw_value = str(os.getenv("VISION_VISUAL_REVIEW_PROBE_RANKED_RETRY_ATTEMPTS", "") or "").strip()
    if raw_value:
        try:
            return max(0, int(raw_value))
        except Exception:
            pass
    return int(DEFAULT_VISUAL_REVIEW_PROBE_RANKED_RETRY_ATTEMPTS)


def resolve_probe_ranked_channel_race(ranked_context):
    if isinstance(ranked_context, dict) and isinstance(ranked_context.get("channel_race"), dict):
        return dict(ranked_context.get("channel_race") or {})
    if isinstance(ranked_context, dict):
        return dict(ranked_context)
    return {}


def build_probe_ranked_runtime_context(channel_race):
    return {
        "channel_race": sanitize_json_compatible(channel_race or {}),
        "_lock": threading.Lock(),
        "_disabled_stages": set(),
        "_failure_counts": {},
        "_disable_after_failures": _resolve_probe_ranked_disable_after_failures(),
    }


def snapshot_probe_ranked_channel_race(ranked_context):
    payload = resolve_probe_ranked_channel_race(ranked_context)
    if not (isinstance(ranked_context, dict) and isinstance(ranked_context.get("channel_race"), dict)):
        return payload
    lock = ranked_context.get("_lock")
    if lock is None:
        payload["runtime_disabled_stages"] = sorted(str(item) for item in ranked_context.get("_disabled_stages", set()) if str(item))
        payload["runtime_failure_counts"] = {
            str(key): int(value)
            for key, value in dict(ranked_context.get("_failure_counts") or {}).items()
            if str(key)
        }
        return payload
    with lock:
        payload["runtime_disabled_stages"] = sorted(
            str(item)
            for item in set(ranked_context.get("_disabled_stages") or set())
            if str(item)
        )
        payload["runtime_failure_counts"] = {
            str(key): int(value)
            for key, value in dict(ranked_context.get("_failure_counts") or {}).items()
            if str(key)
        }
    return payload


def record_probe_ranked_candidate_outcome(ranked_context, candidate, ok):
    if not (isinstance(ranked_context, dict) and isinstance(ranked_context.get("channel_race"), dict)):
        return
    stage_name = str((candidate or {}).get("stage") or "").strip()
    if not stage_name:
        return
    lock = ranked_context.get("_lock")
    if lock is None:
        return
    with lock:
        failure_counts = ranked_context.setdefault("_failure_counts", {})
        disabled_stages = ranked_context.setdefault("_disabled_stages", set())
        if ok:
            failure_counts[stage_name] = 0
            disabled_stages.discard(stage_name)
            return
        next_failure_count = int(failure_counts.get(stage_name) or 0) + 1
        failure_counts[stage_name] = next_failure_count
        threshold = max(1, int(ranked_context.get("_disable_after_failures") or _resolve_probe_ranked_disable_after_failures()))
        if next_failure_count >= threshold:
            disabled_stages.add(stage_name)


def build_probe_ranked_candidate_order(platform, review_item, ranked_context):
    race = resolve_probe_ranked_channel_race(ranked_context)
    successful_candidates = [
        dict(item)
        for item in (race.get("candidates") or [])
        if isinstance(item, dict) and item.get("ok")
    ]
    if not successful_candidates:
        return []

    disabled_stages = set()
    if isinstance(ranked_context, dict) and isinstance(ranked_context.get("channel_race"), dict):
        lock = ranked_context.get("_lock")
        if lock is not None:
            with lock:
                disabled_stages = {
                    str(item)
                    for item in set(ranked_context.get("_disabled_stages") or set())
                    if str(item)
                }
    filtered_candidates = [
        item
        for item in successful_candidates
        if str(item.get("stage") or "").strip() not in disabled_stages
    ]
    if not filtered_candidates:
        filtered_candidates = successful_candidates

    preferred_candidates = [
        item for item in filtered_candidates
        if str(item.get("group") or "").strip() == VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED
    ]
    fallback_candidates = [
        item for item in filtered_candidates
        if str(item.get("group") or "").strip() != VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED
    ]
    if preferred_candidates:
        identifier = (
            screening.resolve_profile_review_identifier(platform, review_item)
            or screening.normalize_identifier((review_item or {}).get("username"))
            or str((review_item or {}).get("username") or "")
        )
        if identifier:
            shard_seed = int(hashlib.sha256(identifier.encode("utf-8")).hexdigest()[:8], 16)
            start_index = shard_seed % len(preferred_candidates)
            preferred_candidates = preferred_candidates[start_index:] + preferred_candidates[:start_index]
    return preferred_candidates + fallback_candidates


def probe_vision_provider_with_image(provider, platform="instagram", cover_urls=None):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    probe_cover_urls = dedupe_non_empty_strings(cover_urls or []) or [MINIMAL_VISUAL_REVIEW_PROBE_IMAGE_DATA_URL]
    result = call_vision_provider(
        provider,
        platform,
        "vision-probe",
        probe_cover_urls,
    )
    return {
        "success": True,
        "provider": provider_name,
        "api_style": str((provider or {}).get("api_style") or "").strip().lower(),
        "base_url": str(result.get("base_url") or (provider or {}).get("base_url") or "").rstrip("/"),
        "model": str(result.get("effective_model") or result.get("model") or resolve_vision_provider_model(provider)).strip(),
        "configured_model": str(result.get("configured_model") or resolve_vision_provider_model(provider)).strip(),
        "requested_model": str(result.get("requested_model") or "").strip(),
        "response_model": str(result.get("response_model") or "").strip(),
        "effective_model": str(result.get("effective_model") or result.get("model") or resolve_vision_provider_model(provider)).strip(),
        "checked_at": iso_now(),
        "decision": str(result.get("decision") or "").strip(),
        "reason": str(result.get("reason") or "").strip(),
        "signals": normalize_visual_review_signals(result.get("signals")),
        "response_excerpt": str(result.get("reason") or "").strip()[:160],
    }


def probe_visual_review_ranked_candidate(stage, platform="instagram", cover_urls=None):
    provider = get_runnable_vision_provider(
        stage.get("provider"),
        model=stage.get("model"),
        timeout_seconds=stage.get("timeout_seconds"),
    )
    configured_model = str((stage or {}).get("model") or "").strip()
    if provider:
        configured_model = str(resolve_vision_provider_model(provider) or configured_model).strip()
    result = {
        "stage": str(stage.get("stage") or "").strip(),
        "group": str(stage.get("group") or "").strip(),
        "provider": normalize_vision_provider_name(stage.get("provider")),
        "model": configured_model,
        "configured_model": configured_model,
        "requested_model": configured_model,
        "response_model": "",
        "effective_model": configured_model,
        "timeout_seconds": int(stage.get("timeout_seconds") or 0),
        "ok": False,
        "retryable": False,
        "attempt_count": 1,
    }
    if not provider:
        result["error"] = "provider_unavailable"
        return result
    try:
        probe_result = probe_vision_provider_with_image(provider, platform=platform, cover_urls=cover_urls)
    except Exception as exc:
        result["error"] = str(exc)
        result["retryable"] = is_retryable_visual_exception(exc)
        return result
    result.update({
        "ok": True,
        "provider": probe_result.get("provider") or result["provider"],
        "model": probe_result.get("model") or result["model"],
        "configured_model": probe_result.get("configured_model") or result["configured_model"],
        "requested_model": probe_result.get("requested_model") or result["requested_model"],
        "response_model": probe_result.get("response_model") or "",
        "effective_model": probe_result.get("effective_model") or probe_result.get("model") or result["effective_model"],
        "decision": probe_result.get("decision"),
        "reason": probe_result.get("reason"),
        "signals": probe_result.get("signals") or [],
        "checked_at": probe_result.get("checked_at"),
        "response_excerpt": probe_result.get("response_excerpt"),
    })
    return result


def _run_probe_ranked_visual_provider_probe_pass(stages, platform="instagram", cover_urls=None):
    results_by_stage = {}
    with ThreadPoolExecutor(max_workers=max(1, len(stages)), thread_name_prefix="vision-probe-race") as executor:
        future_map = {
            executor.submit(
                probe_visual_review_ranked_candidate,
                stage,
                platform,
                cover_urls,
            ): str(stage.get("stage") or "").strip()
            for stage in stages
        }
        for future in as_completed(future_map):
            stage_name = future_map[future]
            try:
                candidate_result = future.result()
            except Exception as exc:
                candidate_result = {
                    "stage": stage_name,
                    "provider": "",
                    "model": "",
                    "timeout_seconds": 0,
                    "ok": False,
                    "retryable": is_retryable_visual_exception(exc),
                    "attempt_count": 1,
                    "error": str(exc),
                }
            results_by_stage[stage_name] = candidate_result
    return results_by_stage


def run_probe_ranked_visual_provider_race(platform="instagram", cover_urls=None):
    stages = build_visual_review_probe_ranked_plan()
    results_by_stage = _run_probe_ranked_visual_provider_probe_pass(stages, platform=platform, cover_urls=cover_urls)
    retry_history = []

    # Retry transient probe failures once or twice so brief provider blips do not abort the whole visual chain.
    for retry_attempt in range(1, _resolve_probe_ranked_retry_attempts() + 1):
        preferred_success_exists = any(
            bool((item or {}).get("ok"))
            and str((item or {}).get("group") or "").strip() == VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED
            for item in results_by_stage.values()
        )
        any_success_exists = any(bool((item or {}).get("ok")) for item in results_by_stage.values())
        if preferred_success_exists:
            break
        if any_success_exists:
            retry_stages = [
                stage
                for stage in stages
                if str(stage.get("group") or "").strip() == VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED
                and bool((results_by_stage.get(str(stage.get("stage") or "").strip()) or {}).get("retryable"))
            ]
            if not retry_stages:
                break
        else:
            retry_stages = [
                stage
                for stage in stages
                if bool((results_by_stage.get(str(stage.get("stage") or "").strip()) or {}).get("retryable"))
            ]
        if not retry_stages:
            break
        delay_seconds = compute_visual_retry_delay_seconds(retry_attempt)
        if delay_seconds > 0:
            time.sleep(delay_seconds)
        retried_results = _run_probe_ranked_visual_provider_probe_pass(
            retry_stages,
            platform=platform,
            cover_urls=cover_urls,
        )
        retry_snapshot = {
            "attempt": retry_attempt,
            "delay_seconds": delay_seconds,
            "stage_names": [str(stage.get("stage") or "").strip() for stage in retry_stages if str(stage.get("stage") or "").strip()],
            "results": [],
        }
        for stage in retry_stages:
            stage_name = str(stage.get("stage") or "").strip()
            previous_result = dict(results_by_stage.get(stage_name) or {})
            retried_result = dict(retried_results.get(stage_name) or previous_result)
            retried_result["attempt_count"] = max(1, int(previous_result.get("attempt_count") or 1)) + 1
            retried_result["retried"] = True
            results_by_stage[stage_name] = retried_result
            retry_snapshot["results"].append(
                {
                    "stage": stage_name,
                    "provider": str(retried_result.get("provider") or "").strip(),
                    "model": str(retried_result.get("model") or "").strip(),
                    "ok": bool(retried_result.get("ok")),
                    "retryable": bool(retried_result.get("retryable")),
                    "error": str(retried_result.get("error") or "").strip(),
                    "attempt_count": int(retried_result.get("attempt_count") or 1),
                }
            )
        retry_history.append(retry_snapshot)

    ordered_candidates = []
    selected_candidate = None
    active_preferred_candidates = []
    for stage in stages:
        stage_name = str(stage.get("stage") or "").strip()
        candidate_result = results_by_stage.get(stage_name) or {
            "stage": stage_name,
            "group": str(stage.get("group") or "").strip(),
            "provider": normalize_vision_provider_name(stage.get("provider")),
            "model": str(stage.get("model") or "").strip(),
            "timeout_seconds": int(stage.get("timeout_seconds") or 0),
            "ok": False,
            "error": "probe_missing",
        }
        if candidate_result.get("ok") and str(candidate_result.get("group") or "").strip() == VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED:
            active_preferred_candidates.append(candidate_result)
        if selected_candidate is None and candidate_result.get("ok"):
            candidate_result["selected"] = True
            selected_candidate = candidate_result
        else:
            candidate_result["selected"] = False
        ordered_candidates.append(candidate_result)

    successful_candidates = [item for item in ordered_candidates if item.get("ok")]
    selected_stage = str((selected_candidate or {}).get("stage") or "").strip()
    if len(active_preferred_candidates) > 1:
        selected_stage = VISUAL_REVIEW_PROBE_RANKED_SELECTED_STAGE_PREFERRED_POOL
    selected_provider = str((selected_candidate or {}).get("provider") or "").strip()
    selected_model = str((selected_candidate or {}).get("model") or "").strip()
    if active_preferred_candidates:
        selected_provider = str(active_preferred_candidates[0].get("provider") or "").strip() or selected_provider
        selected_model = str(active_preferred_candidates[0].get("model") or "").strip() or selected_model
    return {
        "strategy": VISUAL_REVIEW_ROUTING_PROBE_RANKED,
        "checked_at": iso_now(),
        "success": bool(selected_candidate),
        "selected_stage": selected_stage,
        "selected_provider": selected_provider,
        "selected_model": selected_model,
        "active_preferred_candidates": [
            {
                "stage": str(item.get("stage") or "").strip(),
                "provider": str(item.get("provider") or "").strip(),
                "model": str(item.get("model") or "").strip(),
                "configured_model": str(item.get("configured_model") or "").strip(),
                "requested_model": str(item.get("requested_model") or "").strip(),
                "response_model": str(item.get("response_model") or "").strip(),
                "effective_model": str(item.get("effective_model") or item.get("model") or "").strip(),
            }
            for item in active_preferred_candidates
        ],
        "dual_active_enabled": len(active_preferred_candidates) > 1,
        "fallback_candidates": [
            {
                "stage": str(item.get("stage") or "").strip(),
                "provider": str(item.get("provider") or "").strip(),
                "model": str(item.get("model") or "").strip(),
                "configured_model": str(item.get("configured_model") or "").strip(),
                "requested_model": str(item.get("requested_model") or "").strip(),
                "response_model": str(item.get("response_model") or "").strip(),
                "effective_model": str(item.get("effective_model") or item.get("model") or "").strip(),
            }
            for item in successful_candidates
            if str(item.get("group") or "").strip() != VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED
        ],
        "retry_history": retry_history,
        "candidates": ordered_candidates,
    }


def run_tiered_visual_review(platform, review_item, cover_urls=None):
    trace = []
    last_error = None
    cover_urls = dedupe_non_empty_strings(cover_urls or build_visual_review_candidate_cover_urls(platform, review_item))
    if not cover_urls:
        raise ValueError("没有可送审的封面 URL")

    stages = build_visual_review_routing_plan()
    for stage in stages:
        provider = get_runnable_vision_provider(
            stage.get("provider"),
            model=stage.get("model"),
            timeout_seconds=stage.get("timeout_seconds"),
        )
        if not provider:
            trace.append(
                build_visual_review_trace_entry(
                    stage.get("stage"),
                    {"name": stage.get("provider")},
                    ok=False,
                    model=stage.get("model"),
                    error="provider_unavailable",
                )
            )
            continue
        try:
            result = call_vision_provider(
                provider,
                platform,
                screening.resolve_profile_review_identifier(platform, review_item) or review_item.get("username") or "",
                cover_urls,
            )
        except Exception as exc:
            last_error = exc
            trace.append(
                build_visual_review_trace_entry(
                    stage.get("stage"),
                    provider,
                    ok=False,
                    model=stage.get("model"),
                    error=str(exc),
                )
            )
            continue

        escalation_reasons = []
        if stage.get("stage") != "judge":
            escalation_reasons = collect_visual_review_escalation_reasons(result, review_item)
        trace.append(
            build_visual_review_trace_entry(
                stage.get("stage"),
                provider,
                ok=True,
                model=result.get("model") or stage.get("model"),
                decision=result.get("decision"),
                escalation_reasons=escalation_reasons,
            )
        )
        if escalation_reasons:
            last_error = RuntimeError(f"{stage.get('stage')} escalated: {', '.join(escalation_reasons)}")
            continue

        result["provider"] = normalize_vision_provider_name(provider.get("name"))
        result["route"] = str(stage.get("stage") or "").strip()
        result["routing_strategy"] = VISUAL_REVIEW_ROUTING_TIERED
        result["trace"] = trace
        result["judge_used"] = any(item.get("stage") == "judge" and item.get("ok") for item in trace)
        result["escalation_reasons"] = []
        return result

    if last_error is not None:
        raise RuntimeError(str(last_error))
    raise RuntimeError("tiered visual routing did not produce a result")


def run_probe_ranked_visual_review(platform, review_item, ranked_race, cover_urls=None):
    trace = []
    last_error = None
    cover_urls = dedupe_non_empty_strings(cover_urls or build_visual_review_candidate_cover_urls(platform, review_item))
    if not cover_urls:
        raise ValueError("没有可送审的封面 URL")

    ordered_candidates = build_probe_ranked_candidate_order(platform, review_item, ranked_race)
    if not ordered_candidates:
        raise RuntimeError("probe_ranked visual routing did not produce any runnable candidate")

    def attempt_candidates(candidates):
        nonlocal last_error
        attempts = []
        for candidate in candidates:
            provider = get_runnable_vision_provider(
                candidate.get("provider"),
                model=candidate.get("model"),
                timeout_seconds=candidate.get("timeout_seconds"),
            )
            configured_model = str((candidate or {}).get("configured_model") or (candidate or {}).get("model") or "").strip()
            requested_model = str((candidate or {}).get("requested_model") or configured_model).strip()
            if provider:
                configured_model = str(resolve_vision_provider_model(provider) or configured_model).strip()
                requested_model = str((candidate or {}).get("requested_model") or (candidate or {}).get("model") or configured_model).strip()
            if not provider:
                error = RuntimeError("provider_unavailable")
                retryable = False
                trace.append(
                    build_visual_review_trace_entry(
                        candidate.get("stage"),
                        {"name": candidate.get("provider")},
                        ok=False,
                        model=candidate.get("model"),
                        configured_model=configured_model,
                        requested_model=requested_model,
                        error=str(error),
                        group=candidate.get("group"),
                        retryable=retryable,
                    )
                )
                record_probe_ranked_candidate_outcome(ranked_race, candidate, ok=False)
                last_error = error
                attempts.append({"candidate": candidate, "retryable": retryable})
                continue
            try:
                result = call_vision_provider(
                    provider,
                    platform,
                    screening.resolve_profile_review_identifier(platform, review_item) or review_item.get("username") or "",
                    cover_urls,
                )
            except Exception as exc:
                retryable = is_retryable_visual_exception(exc)
                last_error = exc
                record_probe_ranked_candidate_outcome(ranked_race, candidate, ok=False)
                trace.append(
                    build_visual_review_trace_entry(
                        candidate.get("stage"),
                        provider,
                        ok=False,
                        model=candidate.get("model"),
                        configured_model=configured_model,
                        requested_model=requested_model,
                        error=str(exc),
                        group=candidate.get("group"),
                        retryable=retryable,
                    )
                )
                attempts.append({"candidate": candidate, "retryable": retryable})
                continue
            record_probe_ranked_candidate_outcome(ranked_race, candidate, ok=True)
            trace.append(
                build_visual_review_trace_entry(
                    candidate.get("stage"),
                    provider,
                    ok=True,
                    model=result.get("model") or candidate.get("model"),
                    configured_model=result.get("configured_model") or configured_model,
                    requested_model=result.get("requested_model") or requested_model,
                    response_model=result.get("response_model"),
                    effective_model=result.get("effective_model") or result.get("model"),
                    decision=result.get("decision"),
                    group=candidate.get("group"),
                )
            )
            result["provider"] = normalize_vision_provider_name(provider.get("name"))
            result["route"] = str(candidate.get("stage") or "").strip()
            result["routing_strategy"] = VISUAL_REVIEW_ROUTING_PROBE_RANKED
            result["trace"] = trace
            result["judge_used"] = False
            result["escalation_reasons"] = []
            result["channel_race"] = snapshot_probe_ranked_channel_race(ranked_race)
            return result, attempts
        return None, attempts

    result, first_pass_attempts = attempt_candidates(ordered_candidates)
    if result is not None:
        return result

    preferred_retry_candidates = [
        item["candidate"]
        for item in first_pass_attempts
        if item.get("retryable")
        and str((item.get("candidate") or {}).get("group") or "").strip() == VISUAL_REVIEW_PROBE_RANKED_GROUP_PREFERRED
    ]
    if preferred_retry_candidates and first_pass_attempts and all(item.get("retryable") for item in first_pass_attempts):
        result, _ = attempt_candidates(preferred_retry_candidates)
        if result is not None:
            return result

    if last_error is not None:
        raise RuntimeError(str(last_error))
    raise RuntimeError("probe_ranked visual routing did not produce a result")


def strip_code_fences(text):
    stripped = str(text or "").strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```[a-zA-Z0-9_]*\n?", "", stripped)
        stripped = re.sub(r"\n?```$", "", stripped)
    return stripped.strip()


def _contains_cjk_text(value):
    return any("\u4e00" <= char <= "\u9fff" for char in str(value or ""))


def repair_mojibake_text(value):
    text = str(value or "")
    if not text or _contains_cjk_text(text):
        return text
    try:
        repaired = text.encode("latin-1").decode("utf-8")
    except Exception:
        return text
    if _contains_cjk_text(repaired):
        return repaired
    return text


def normalize_vision_payload_text(payload):
    if isinstance(payload, str):
        return repair_mojibake_text(payload)
    if isinstance(payload, list):
        return [normalize_vision_payload_text(item) for item in payload]
    if isinstance(payload, dict):
        return {key: normalize_vision_payload_text(value) for key, value in payload.items()}
    return payload


def parse_json_contract_payload(raw_text):
    cleaned = repair_mojibake_text(strip_code_fences(raw_text))
    payload = None
    try:
        payload = json.loads(cleaned)
    except json.JSONDecodeError:
        start_idx = cleaned.find("{")
        end_idx = cleaned.rfind("}")
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            snippet = cleaned[start_idx:end_idx + 1]
            try:
                payload = json.loads(snippet)
            except json.JSONDecodeError:
                payload = None
    return cleaned, payload


def parse_visual_review_payload(raw_text):
    return parse_json_contract_payload(raw_text)


def is_valid_visual_review_payload(payload):
    if not isinstance(payload, dict):
        return False
    decision = str(payload.get("decision") or "").strip().title()
    if decision not in {"Pass", "Reject"}:
        return False
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        return False
    signals = payload.get("signals")
    if signals is None:
        return True
    return isinstance(signals, list)


def parse_visual_review_result(raw_text):
    cleaned, payload = parse_visual_review_payload(raw_text)

    if isinstance(payload, dict):
        decision = str(payload.get("decision") or "").strip().title()
        if decision not in {"Pass", "Reject"}:
            decision = "Reject" if "reject" in cleaned.lower() else "Pass"
        signals = payload.get("signals")
        if not isinstance(signals, list):
            signals = []
        return {
            "decision": decision,
            "reason": str(payload.get("reason") or cleaned).strip() or "模型未返回原因",
            "signals": [str(item).strip() for item in signals if str(item or "").strip()][:3],
        }

    return {
        "decision": "Reject" if "reject" in cleaned.lower() else "Pass",
        "reason": cleaned or "模型未返回可解析内容",
        "signals": [],
    }


def is_valid_positioning_card_payload(payload):
    if not isinstance(payload, dict):
        return False
    fit_summary = str(payload.get("fit_summary") or "").strip()
    if not fit_summary:
        return False
    labels = payload.get("positioning_labels")
    if labels is not None and not isinstance(labels, list):
        return False
    signals = payload.get("evidence_signals")
    if signals is not None and not isinstance(signals, list):
        return False
    return True


def parse_positioning_card_result(raw_text):
    cleaned, payload = parse_json_contract_payload(raw_text)
    labels = []
    signals = []
    fit_recommendation = "Unclear"
    fit_summary = cleaned or "模型未返回可解析内容"

    if isinstance(payload, dict):
        raw_labels = payload.get("positioning_labels")
        if isinstance(raw_labels, list):
            labels = [str(item).strip() for item in raw_labels if str(item or "").strip()][:4]
        raw_signals = payload.get("evidence_signals")
        if isinstance(raw_signals, list):
            signals = [str(item).strip() for item in raw_signals if str(item or "").strip()][:4]
        fit_recommendation = screening.normalize_fit_recommendation(payload.get("fit_recommendation"))
        fit_summary = str(payload.get("fit_summary") or cleaned).strip() or "模型未返回结论"

    return {
        "positioning_labels": labels,
        "fit_recommendation": fit_recommendation,
        "fit_summary": fit_summary,
        "evidence_signals": signals,
    }


def extract_vision_provider_text_error(raw_text):
    cleaned = repair_mojibake_text(strip_code_fences(raw_text)).strip()
    if not cleaned:
        return None
    lowered = cleaned.lower()
    status_code = None
    status_match = re.search(r"(?:status|http)\s*[:=]?\s*(\d{3})", lowered)
    if status_match:
        try:
            status_code = int(status_match.group(1))
        except Exception:
            status_code = None
    error_markers = (
        "internal server error",
        "service unavailable",
        "bad gateway",
        "gateway timeout",
        "upstream connect error",
        "upstream request timeout",
        "server error",
        "status=500",
        "status 500",
        "http 500",
        "status=502",
        "status 502",
        "http 502",
        "status=503",
        "status 503",
        "http 503",
        "status=504",
        "status 504",
        "http 504",
        "status=522",
        "status 522",
        "http 522",
        "not supported",
        "unsupported",
        "model not found",
        "unknown model",
        "invalid model",
        "not available for model",
    )
    if (status_code is not None and status_code >= 400) or any(marker in lowered for marker in error_markers):
        return {
            "message": cleaned,
            "status_code": status_code,
            "retryable": (status_code in VISUAL_REVIEW_RETRYABLE_STATUS_CODES) if status_code is not None else True,
        }
    return None


def extract_vision_response_text(payload):
    if isinstance(payload, str):
        return payload.strip()
    if not isinstance(payload, dict):
        return str(payload)

    output_text = payload.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text

    output = payload.get("output")
    if isinstance(output, list):
        text_parts = []
        for item in output:
            for content_item in (item or {}).get("content") or []:
                text_value = content_item.get("text")
                if isinstance(text_value, str) and text_value.strip():
                    text_parts.append(text_value)
            if text_parts:
                return "\n".join(text_parts)

    choices = payload.get("choices")
    if isinstance(choices, list):
        text_parts = []
        for choice in choices:
            message = (choice or {}).get("message") or {}
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                text_parts.append(content)
                continue
            if isinstance(content, list):
                for content_item in content:
                    text_value = (content_item or {}).get("text")
                    if isinstance(text_value, str) and text_value.strip():
                        text_parts.append(text_value)
            if text_parts:
                return "\n".join(text_parts)
        if text_parts:
            return "\n".join(text_parts)

    candidates = payload.get("candidates")
    if isinstance(candidates, list):
        text_parts = []
        for candidate in candidates:
            content = (candidate or {}).get("content") or {}
            parts = content.get("parts")
            if not isinstance(parts, list):
                continue
            for part in parts:
                text_value = (part or {}).get("text")
                if isinstance(text_value, str) and text_value.strip():
                    text_parts.append(text_value)
            if text_parts:
                return "\n".join(text_parts)

    return json.dumps(payload, ensure_ascii=False)


def parse_streaming_chat_completion_payload(raw_text):
    cleaned = repair_mojibake_text(str(raw_text or "").strip())
    if not cleaned:
        return {}

    content_parts = []
    model = ""
    finish_reason = ""
    for raw_line in cleaned.splitlines():
        line = raw_line.strip()
        if not line.startswith("data:"):
            continue
        data = line[len("data:") :].strip()
        if not data or data == "[DONE]":
            continue
        try:
            chunk = json.loads(data)
        except Exception:
            continue
        if not model:
            model = str(chunk.get("model") or "").strip()
        choices = chunk.get("choices")
        if not isinstance(choices, list):
            continue
        for choice in choices:
            delta = (choice or {}).get("delta") or {}
            text = delta.get("content")
            if isinstance(text, str) and text:
                content_parts.append(text)
            if not finish_reason:
                finish_reason = str((choice or {}).get("finish_reason") or "").strip()

    if not content_parts and not finish_reason:
        return {}

    payload = {
        "choices": [
            {
                "message": {
                    "role": "assistant",
                    "content": "".join(content_parts),
                },
                "finish_reason": finish_reason or None,
            }
        ]
    }
    if model:
        payload["model"] = model
    return normalize_vision_payload_text(payload)


def vision_payload_has_text_content(payload):
    if isinstance(payload, str):
        return bool(payload.strip())
    if not isinstance(payload, dict):
        return False

    output_text = payload.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return True

    output = payload.get("output")
    if isinstance(output, list):
        for item in output:
            for content_item in (item or {}).get("content") or []:
                text_value = content_item.get("text")
                if isinstance(text_value, str) and text_value.strip():
                    return True

    choices = payload.get("choices")
    if isinstance(choices, list):
        for choice in choices:
            message = (choice or {}).get("message") or {}
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                return True
            if isinstance(content, list):
                for content_item in content:
                    text_value = (content_item or {}).get("text")
                    if isinstance(text_value, str) and text_value.strip():
                        return True
    candidates = payload.get("candidates")
    if isinstance(candidates, list):
        for candidate in candidates:
            content = (candidate or {}).get("content") or {}
            parts = content.get("parts")
            if not isinstance(parts, list):
                continue
            for part in parts:
                text_value = (part or {}).get("text")
                if isinstance(text_value, str) and text_value.strip():
                    return True
    return False


def should_retry_vision_payload(provider_name, payload):
    if normalize_vision_provider_name(provider_name) != "mimo" or not isinstance(payload, dict):
        return False
    if vision_payload_has_text_content(payload):
        return False
    choices = payload.get("choices")
    if not isinstance(choices, list):
        return False
    for choice in choices:
        if str((choice or {}).get("finish_reason") or "").strip().lower() == "length":
            return True
    return False


def extract_vision_usage(payload):
    if not isinstance(payload, dict):
        return {}
    usage = payload.get("usage")
    if not isinstance(usage, dict):
        return {}

    normalized = {}
    for key in ("prompt_tokens", "completion_tokens", "total_tokens"):
        value = usage.get(key)
        try:
            normalized[key] = int(value)
        except Exception:
            continue

    completion_details = usage.get("completion_tokens_details")
    if isinstance(completion_details, dict):
        try:
            normalized["reasoning_tokens"] = int(completion_details.get("reasoning_tokens"))
        except Exception:
            pass

    prompt_details = usage.get("prompt_tokens_details")
    if isinstance(prompt_details, dict):
        for source_key, target_key in (
            ("image_tokens", "image_tokens"),
            ("cached_tokens", "cached_tokens"),
        ):
            try:
                normalized[target_key] = int(prompt_details.get(source_key))
            except Exception:
                continue
    usage_metadata = payload.get("usageMetadata")
    if isinstance(usage_metadata, dict):
        for source_key, target_key in (
            ("promptTokenCount", "prompt_tokens"),
            ("candidatesTokenCount", "completion_tokens"),
            ("totalTokenCount", "total_tokens"),
        ):
            try:
                normalized[target_key] = int(usage_metadata.get(source_key))
            except Exception:
                continue
    return normalized


def decode_vision_response_text(response):
    content = getattr(response, "content", b"")
    if isinstance(content, bytes) and content:
        for encoding in ("utf-8", "utf-8-sig"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="replace")
    return str(getattr(response, "text", "") or "")


def parse_vision_provider_response_payload(response):
    try:
        payload = response.json()
        if isinstance(payload, (dict, list)):
            return normalize_vision_payload_text(payload)
        return payload
    except ValueError:
        payload = parse_streaming_chat_completion_payload(decode_vision_response_text(response))
        if payload:
            return payload
        return {}


def summarize_visual_usage(results):
    totals = {
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "total_tokens": 0,
        "reasoning_tokens": 0,
        "image_tokens": 0,
        "cached_tokens": 0,
    }
    found = False
    for item in (results or {}).values():
        if not isinstance(item, dict):
            continue
        usage = item.get("usage")
        if not isinstance(usage, dict):
            continue
        found = True
        for key in totals:
            try:
                totals[key] += int(usage.get(key) or 0)
            except Exception:
                continue
    return totals if found else {}


def guess_image_mime_type(response, fallback_url=""):
    content_type = str((response.headers or {}).get("Content-Type") or "").split(";", 1)[0].strip().lower()
    if content_type.startswith("image/"):
        return content_type
    fallback_url = str(fallback_url or "").lower()
    if fallback_url.endswith(".png"):
        return "image/png"
    if fallback_url.endswith(".webp"):
        return "image/webp"
    return "image/jpeg"


def guess_image_extension(mime_type, fallback_url=""):
    normalized_mime_type = str(mime_type or "").split(";", 1)[0].strip().lower()
    if normalized_mime_type in {"image/jpeg", "image/jpg"}:
        return ".jpg"
    if normalized_mime_type == "image/png":
        return ".png"
    if normalized_mime_type == "image/webp":
        return ".webp"
    if normalized_mime_type == "image/gif":
        return ".gif"
    if normalized_mime_type == "image/bmp":
        return ".bmp"
    if normalized_mime_type == "image/avif":
        return ".avif"
    fallback_path = urlparse(str(fallback_url or "")).path.lower()
    for extension in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".avif"):
        if fallback_path.endswith(extension):
            return ".jpg" if extension == ".jpeg" else extension
    return ".jpg"


def build_image_data_url(image_bytes, mime_type):
    encoded = base64.b64encode(image_bytes).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def build_visual_image_cache_key(image_url):
    normalized = str(image_url or "").strip()
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def load_cached_visual_image(platform, image_url):
    if not VISUAL_IMAGE_CACHE_ENABLED:
        return None
    cache_dir = Path(get_visual_image_cache_dir(platform))
    metadata_path = cache_dir / f"{build_visual_image_cache_key(image_url)}.json"
    metadata = load_json_payload(str(metadata_path), default={}) or {}
    filename = str(metadata.get("filename") or "").strip()
    mime_type = str(metadata.get("mime_type") or "").strip().lower()
    if not filename or not mime_type:
        return None
    image_path = cache_dir / filename
    if not image_path.exists() or not image_path.is_file():
        return None
    try:
        return {
            "bytes": image_path.read_bytes(),
            "mime_type": mime_type,
            "path": str(image_path),
        }
    except Exception:
        return None


def persist_visual_image_cache(platform, image_url, image_bytes, mime_type):
    if not VISUAL_IMAGE_CACHE_ENABLED or not image_bytes:
        return None
    cache_dir = Path(get_visual_image_cache_dir(platform))
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_key = build_visual_image_cache_key(image_url)
    image_extension = guess_image_extension(mime_type, fallback_url=image_url)
    image_filename = f"{cache_key}{image_extension}"
    image_path = cache_dir / image_filename
    metadata_path = cache_dir / f"{cache_key}.json"
    temp_image_path = cache_dir / f".{image_filename}.{uuid.uuid4().hex}.tmp"
    temp_metadata_path = cache_dir / f".{cache_key}.{uuid.uuid4().hex}.json.tmp"
    try:
        temp_image_path.write_bytes(image_bytes)
        os.replace(temp_image_path, image_path)
        metadata_payload = {
            "source_url": str(image_url or "").strip(),
            "filename": image_filename,
            "mime_type": str(mime_type or "").strip().lower() or "image/jpeg",
            "size_bytes": len(image_bytes),
            "cached_at": iso_now(),
        }
        temp_metadata_path.write_text(
            json.dumps(metadata_payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        os.replace(temp_metadata_path, metadata_path)
        return {
            "path": str(image_path),
            "mime_type": metadata_payload["mime_type"],
        }
    finally:
        for temp_path in (temp_image_path, temp_metadata_path):
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                pass


def build_data_url_from_remote_image(platform, image_url):
    cached_image = load_cached_visual_image(platform, image_url)
    if cached_image:
        return build_image_data_url(cached_image["bytes"], cached_image["mime_type"])
    last_error = None
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "image/*,*/*;q=0.8",
        "Referer": image_url,
    }
    for attempt_index in range(1, VISUAL_IMAGE_DOWNLOAD_MAX_RETRIES + 1):
        try:
            response = requests.get(
                image_url,
                timeout=VISION_REQUEST_TIMEOUT,
                headers=headers,
            )
            if response.status_code >= 400:
                retryable = response.status_code in VISUAL_REVIEW_RETRYABLE_STATUS_CODES
                if retryable and attempt_index < VISUAL_IMAGE_DOWNLOAD_MAX_RETRIES:
                    time.sleep(compute_visual_retry_delay_seconds(attempt_index))
                    continue
                raise RuntimeError(f"下载封面失败：HTTP {response.status_code}")
            mime_type = guess_image_mime_type(response, fallback_url=image_url)
            image_bytes = response.content
            persist_visual_image_cache(platform, image_url, image_bytes, mime_type)
            return build_image_data_url(image_bytes, mime_type)
        except requests.exceptions.RequestException as exc:
            last_error = exc
            if attempt_index >= VISUAL_IMAGE_DOWNLOAD_MAX_RETRIES:
                break
            time.sleep(compute_visual_retry_delay_seconds(attempt_index))
    if last_error is not None:
        raise last_error
    raise RuntimeError("下载封面失败")


def normalize_visual_image_source(platform, image_url, *, force_data_url=False):
    image_url = str(image_url or "").strip()
    if not image_url:
        return ""
    if image_url.startswith("data:"):
        return image_url
    if not force_data_url:
        return image_url
    return build_data_url_from_remote_image(platform, image_url)


def reduce_visual_review_cover_limit(selected_cover_count):
    try:
        current_limit = max(1, int(selected_cover_count or 0))
    except (TypeError, ValueError):
        return None
    if current_limit <= 1:
        return None
    if current_limit <= 3:
        return current_limit - 1
    return max(1, (current_limit + 1) // 2)


def resolve_inline_visual_retry_cover_limit(selected_cover_count):
    reduced_limit = reduce_visual_review_cover_limit(selected_cover_count)
    if reduced_limit is None:
        try:
            reduced_limit = max(1, int(selected_cover_count or 1))
        except (TypeError, ValueError):
            reduced_limit = 1
    return max(1, min(5, int(reduced_limit)))


def resolve_visual_transport_cover_limit(platform, api_style, request_cover_limit):
    try:
        resolved_limit = max(1, int(request_cover_limit or VISUAL_REVIEW_REQUEST_COVER_LIMIT))
    except (TypeError, ValueError):
        resolved_limit = VISUAL_REVIEW_REQUEST_COVER_LIMIT
    normalized_style = str(api_style or "").strip().lower()
    if normalized_style in {VISION_API_STYLE_RESPONSES, VISION_API_STYLE_CHAT_COMPLETIONS}:
        if str(platform or "").strip().lower() == "tiktok":
            return min(resolved_limit, INLINE_TIKTOK_VISUAL_REVIEW_REQUEST_COVER_LIMIT)
    return resolved_limit


def data_url_to_inline_part(image_source):
    cleaned = str(image_source or "").strip()
    match = re.match(r"^data:(image/[^;]+);base64,(.+)$", cleaned, re.IGNORECASE | re.DOTALL)
    if not match:
        raise ValueError("invalid data url image source")
    return {
        "inlineData": {
            "mimeType": str(match.group(1) or "image/jpeg").strip().lower(),
            "data": str(match.group(2) or "").strip(),
        }
    }


def dedupe_non_empty_strings(values):
    seen = set()
    normalized = []
    for value in values or []:
        cleaned = str(value or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        normalized.append(cleaned)
    return normalized


def extract_tiktok_raw_identifier(item):
    return screening.extract_platform_identifier(
        "tiktok",
        ((item.get("authorMeta") or {}).get("profileUrl")) or ((item.get("authorMeta") or {}).get("name")),
    )


def extract_instagram_raw_identifier(item):
    return (
        screening.extract_platform_identifier("instagram", item.get("url"))
        or screening.extract_platform_identifier("instagram", item.get("username"))
    )


def extract_youtube_raw_identifier(item):
    return (
        screening.extract_platform_identifier("youtube", item.get("inputChannelUrl"))
        or screening.extract_platform_identifier("youtube", item.get("input"))
        or screening.extract_platform_identifier("youtube", item.get("channelUsername"))
        or screening.extract_platform_identifier("youtube", item.get("channelUrl"))
        or screening.extract_platform_identifier("youtube", item.get("channelName"))
        or screening.extract_platform_identifier("youtube", ((item.get("aboutChannelInfo") or {}).get("channelUrl")))
        or screening.extract_platform_identifier("youtube", ((item.get("aboutChannelInfo") or {}).get("inputChannelUrl")))
        or screening.extract_platform_identifier("youtube", ((item.get("aboutChannelInfo") or {}).get("channelUsername")))
    )


def derive_visual_review_candidate_cover_urls_from_raw(platform, identifier, candidate_limit):
    identifier = screening.normalize_identifier(identifier)
    if not identifier:
        return []
    raw_items = load_json_payload(get_raw_data_path(platform), default=[]) or []
    if not raw_items:
        return []

    if platform == "tiktok":
        matched_items = [
            item
            for item in raw_items
            if extract_tiktok_raw_identifier(item) == identifier
        ]
        sorted_items = screening.sort_items_by_latest(matched_items, "createTimeISO")
        return screening.extract_tiktok_cover_urls(sorted_items, candidate_limit)

    if platform == "instagram":
        for item in raw_items:
            if extract_instagram_raw_identifier(item) != identifier:
                continue
            sorted_posts = screening.sort_items_by_latest(item.get("latestPosts") or [], "timestamp")
            return screening.extract_instagram_cover_urls(sorted_posts, candidate_limit)
        return []

    if platform == "youtube":
        matched_items = [
            item
            for item in raw_items
            if extract_youtube_raw_identifier(item) == identifier
        ]
        sorted_items = screening.sort_items_by_latest(matched_items, "date")
        return screening.extract_youtube_cover_urls(sorted_items, candidate_limit)

    return []


def build_visual_review_candidate_cover_urls(platform, review_item):
    existing_covers = dedupe_non_empty_strings((review_item or {}).get("covers") or [])
    identifier = screening.resolve_profile_review_identifier(platform, review_item)
    candidate_limit = max(
        VISUAL_REVIEW_REQUEST_COVER_LIMIT,
        VISUAL_REVIEW_CANDIDATE_COVER_LIMIT_FLOOR,
        len(existing_covers),
    )
    derived_covers = derive_visual_review_candidate_cover_urls_from_raw(platform, identifier, candidate_limit)
    return dedupe_non_empty_strings(existing_covers + derived_covers)[:candidate_limit]


def build_multimodal_prompt_input(
    platform,
    prompt_text,
    cover_urls,
    *,
    api_style=VISION_API_STYLE_RESPONSES,
    force_data_url_override=False,
    request_cover_limit=VISUAL_REVIEW_REQUEST_COVER_LIMIT,
    prompt_source="",
    prompt_selection=None,
):
    candidate_cover_urls = dedupe_non_empty_strings(cover_urls)
    normalized_image_sources = []
    skipped_cover_count = 0
    for cover_url in candidate_cover_urls:
        try:
            image_source = normalize_visual_image_source(
                platform,
                cover_url,
                force_data_url=(
                    bool(force_data_url_override)
                    or str(api_style or "").strip().lower()
                    in {
                        VISION_API_STYLE_GENERATE_CONTENT,
                        VISION_API_STYLE_RESPONSES,
                        VISION_API_STYLE_CHAT_COMPLETIONS,
                    }
                ),
            )
        except Exception:
            skipped_cover_count += 1
            continue
        if not image_source:
            skipped_cover_count += 1
            continue
        normalized_image_sources.append(image_source)
        if len(normalized_image_sources) >= resolve_visual_transport_cover_limit(
            platform,
            api_style,
            request_cover_limit,
        ):
            break
    if not normalized_image_sources:
        raise ValueError("没有可送审的有效图片")
    response_content = [{"type": "input_text", "text": prompt_text}]
    response_content.extend(
        {"type": "input_image", "image_url": image_source}
        for image_source in normalized_image_sources
    )
    chat_content = [{"type": "text", "text": prompt_text}]
    chat_content.extend(
        {"type": "image_url", "image_url": {"url": image_source}}
        for image_source in normalized_image_sources
    )
    payload = {
        "responses": [{"role": "user", "content": response_content}],
        "chat": [{"role": "user", "content": chat_content}],
        "selected_cover_count": len(normalized_image_sources),
        "candidate_cover_count": len(candidate_cover_urls),
        "skipped_cover_count": skipped_cover_count,
        "prompt_selection": dict(prompt_selection or {}),
        "prompt_source": str(prompt_source or "").strip(),
        "request_cover_limit": max(1, int(request_cover_limit or VISUAL_REVIEW_REQUEST_COVER_LIMIT)),
    }
    if str(api_style or "").strip().lower() == VISION_API_STYLE_GENERATE_CONTENT:
        generate_content_parts = [{"text": prompt_text}]
        generate_content_parts.extend(
            data_url_to_inline_part(image_source)
            for image_source in normalized_image_sources
        )
        payload["generate_content"] = build_vision_provider_generate_content_body(generate_content_parts)
    return payload


def build_visual_review_input(
    provider_name,
    platform,
    username,
    cover_urls,
    model_name="",
    api_style=VISION_API_STYLE_RESPONSES,
    force_data_url_override=False,
    request_cover_limit_override=None,
):
    prompt_selection = resolve_visual_review_prompt_selection(provider_name, platform, model_name=model_name)
    header = (
        f"平台：{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)}\n"
        f"达人：{username or 'unknown'}\n"
        f"{prompt_selection['prompt']}"
    )
    resolved_cover_limit = prompt_selection.get("resolved_cover_limit") or VISUAL_REVIEW_REQUEST_COVER_LIMIT
    if request_cover_limit_override not in (None, ""):
        try:
            resolved_cover_limit = max(1, int(request_cover_limit_override))
        except (TypeError, ValueError):
            pass
    return build_multimodal_prompt_input(
        platform,
        header,
        cover_urls,
        api_style=api_style,
        force_data_url_override=force_data_url_override,
        request_cover_limit=resolved_cover_limit,
        prompt_source=prompt_selection.get("source"),
        prompt_selection=prompt_selection,
    )


def build_positioning_card_input(
    provider_name,
    platform,
    username,
    cover_urls,
    model_name="",
    api_style=VISION_API_STYLE_RESPONSES,
    force_data_url_override=False,
    request_cover_limit_override=None,
):
    prompt_selection = resolve_positioning_card_prompt_selection(provider_name, platform, model_name=model_name)
    header = (
        f"平台：{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)}\n"
        f"达人：{username or 'unknown'}\n"
        f"{prompt_selection['prompt']}"
    )
    resolved_cover_limit = prompt_selection.get("resolved_cover_limit") or VISUAL_REVIEW_REQUEST_COVER_LIMIT
    if request_cover_limit_override not in (None, ""):
        try:
            resolved_cover_limit = max(1, int(request_cover_limit_override))
        except (TypeError, ValueError):
            pass
    return build_multimodal_prompt_input(
        platform,
        header,
        cover_urls,
        api_style=api_style,
        force_data_url_override=force_data_url_override,
        request_cover_limit=resolved_cover_limit,
        prompt_source=prompt_selection.get("source"),
        prompt_selection=prompt_selection,
    )


class VisionProviderError(RuntimeError):
    def __init__(self, provider_name, message, status_code=None, retryable=False):
        self.provider_name = normalize_vision_provider_name(provider_name)
        self.status_code = status_code
        self.retryable = bool(retryable)
        sanitized_message = normalize_upstream_error_text(message, max_length=240) or str(message or "").strip()
        super().__init__(f"{self.provider_name}: {sanitized_message}")


def is_retryable_visual_exception(exc):
    if isinstance(exc, VisionProviderError):
        return exc.retryable
    if isinstance(exc, requests.exceptions.Timeout):
        return True
    if isinstance(exc, requests.exceptions.ConnectionError):
        return True
    lowered = str(exc or "").strip().lower()
    if not lowered:
        return False
    status_match = re.search(r"(?:status|http)\s*[:=]?\s*(\d{3})", lowered)
    if status_match:
        try:
            if int(status_match.group(1)) in VISUAL_REVIEW_RETRYABLE_STATUS_CODES:
                return True
        except Exception:
            pass
    retryable_markers = (
        "timed out",
        "timeout",
        "ssl",
        "eof occurred in violation of protocol",
        "connection aborted",
        "bad gateway",
        "gateway timeout",
        "service unavailable",
        "internal server error",
    )
    return any(marker in lowered for marker in retryable_markers)


def compute_visual_retry_delay_seconds(attempt_index):
    capped_attempt = max(1, int(attempt_index or 1))
    delay = VISUAL_REVIEW_RETRY_BASE_DELAY_SECONDS * (2 ** (capped_attempt - 1))
    delay = min(delay, VISUAL_REVIEW_RETRY_MAX_DELAY_SECONDS)
    jitter = random.uniform(0, min(0.5, delay * 0.2))
    return round(delay + jitter, 3)


def should_retry_visual_provider_with_inline_images(api_style, error_message):
    normalized_style = str(api_style or "").strip().lower()
    if normalized_style not in {VISION_API_STYLE_RESPONSES, VISION_API_STYLE_CHAT_COMPLETIONS}:
        return False
    lowered = str(error_message or "").strip().lower()
    if not lowered:
        return False
    return (
        "error while downloading" in lowered
        or "upstream status code: 403" in lowered
        or "下载图片失败" in lowered
    )


def is_local_cliproxyapi_base_url(base_url):
    parsed = urlparse(str(base_url or "").strip())
    hostname = str(parsed.hostname or "").strip().lower()
    try:
        port = int(parsed.port) if parsed.port is not None else None
    except Exception:
        port = None
    return hostname in {"127.0.0.1", "localhost"} and port == 8317


def resolve_vision_provider_max_inflight_requests(provider, base_url=""):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    env_key = str((provider or {}).get("max_inflight_env_key") or "").strip()
    if env_key:
        raw_value = os.getenv(env_key, "")
        if str(raw_value or "").strip():
            try:
                return max(0, int(raw_value))
            except (TypeError, ValueError):
                return 0
    if provider_name == "openai" and is_local_cliproxyapi_base_url(base_url or (provider or {}).get("base_url")):
        return DEFAULT_LOCAL_OPENAI_MAX_INFLIGHT_REQUESTS
    return 0


def get_vision_provider_request_gate(provider_name, base_url, limit):
    normalized_provider = normalize_vision_provider_name(provider_name)
    normalized_base_url = str(base_url or "").strip().rstrip("/")
    normalized_limit = max(0, int(limit or 0))
    if normalized_limit <= 0:
        return None
    key = (normalized_provider, normalized_base_url, normalized_limit)
    with VISION_PROVIDER_REQUEST_GATES_LOCK:
        gate = VISION_PROVIDER_REQUEST_GATES.get(key)
        if gate is None:
            gate = threading.BoundedSemaphore(normalized_limit)
            VISION_PROVIDER_REQUEST_GATES[key] = gate
        return gate


@contextmanager
def acquire_vision_provider_request_slot(provider, base_url, request_timeout):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    limit = resolve_vision_provider_max_inflight_requests(provider, base_url=base_url)
    gate = get_vision_provider_request_gate(provider_name, base_url, limit)
    if gate is None:
        yield
        return
    acquired = gate.acquire(timeout=max(float(request_timeout or 0), 30.0))
    if not acquired:
        raise VisionProviderError(
            provider_name,
            f"{provider_name} 并发闸门等待超时（limit={limit}）",
            retryable=True,
        )
    try:
        yield
    finally:
        gate.release()


def is_qiandao_25p_visual_model(model_name):
    return str(model_name or "").strip() == QIANDAO_25P_VISION_MODEL


def should_use_qiandao_25p_visual_worker_profile(payload=None, requested_provider="", requested_model="", routing_strategy=""):
    normalized_routing_strategy = normalize_visual_review_routing_strategy(
        routing_strategy or (payload or {}).get("routing_strategy")
    )
    if normalized_routing_strategy == VISUAL_REVIEW_ROUTING_TIERED:
        for stage in build_visual_review_routing_plan():
            if is_qiandao_25p_visual_model(stage.get("model")):
                return True
        return False

    provider_name = normalize_vision_provider_name(requested_provider or (payload or {}).get("provider"))
    if not provider_name:
        provider_name, _ = resolve_vision_provider_request()
    if requested_model:
        return is_qiandao_25p_visual_model(requested_model)

    if provider_name == "reelx":
        reelx_provider = next(
            (item for item in VISION_PROVIDER_CONFIGS if normalize_vision_provider_name(item.get("name")) == "reelx"),
            None,
        )
        if not reelx_provider:
            return False
        return any(
            is_qiandao_25p_visual_model(candidate)
            for candidate in resolve_vision_provider_model_candidates(reelx_provider)
        )

    if provider_name != "qiandao":
        return False

    qiandao_provider = next(
        (item for item in VISION_PROVIDER_CONFIGS if normalize_vision_provider_name(item.get("name")) == "qiandao"),
        None,
    )
    if not qiandao_provider:
        return False
    return is_qiandao_25p_visual_model(resolve_vision_provider_model(qiandao_provider))


def resolve_visual_review_max_workers(payload, target_count, requested_provider="", requested_model="", routing_strategy=""):
    default_value = DEFAULT_VISUAL_REVIEW_MAX_WORKERS
    max_allowed = None
    if should_use_qiandao_25p_visual_worker_profile(
        payload,
        requested_provider=requested_provider,
        requested_model=requested_model,
        routing_strategy=routing_strategy,
    ):
        default_value = DEFAULT_QIANDAO_25P_VISUAL_REVIEW_MAX_WORKERS
        max_allowed = MAX_QIANDAO_25P_VISUAL_REVIEW_MAX_WORKERS

    requested = (payload or {}).get("max_workers")
    try:
        requested_value = int(requested) if requested not in (None, "") else default_value
    except Exception:
        requested_value = default_value
    if max_allowed is not None:
        requested_value = min(requested_value, max_allowed)
    return max(1, min(int(target_count or 1), requested_value))


def resolve_visual_review_item_timeout_seconds(payload=None):
    raw_value = None
    if isinstance(payload, dict):
        raw_value = payload.get("item_timeout_seconds")
    if raw_value in (None, ""):
        raw_value = os.getenv("VISUAL_REVIEW_ITEM_TIMEOUT_SECONDS", "")
    try:
        resolved = float(raw_value)
    except (TypeError, ValueError):
        resolved = DEFAULT_VISUAL_REVIEW_ITEM_TIMEOUT_SECONDS
    return max(1.0, resolved)


def call_vision_provider_with_json_contract(
    provider,
    *,
    platform,
    username,
    cover_urls,
    build_input_fn,
    validate_payload_fn,
    parse_result_fn,
    contract_label,
):
    provider_name = normalize_vision_provider_name(provider.get("name"))
    api_style = str(provider.get("api_style") or VISION_API_STYLE_RESPONSES).strip().lower()
    base_urls = [
        str(item or "").strip().rstrip("/")
        for item in (provider.get("base_url_candidates") or resolve_vision_provider_base_urls(provider))
        if str(item or "").strip()
    ]
    base_url = base_urls[0] if base_urls else str(provider.get("base_url") or "").rstrip("/")
    if not base_urls and base_url:
        base_urls = [base_url]
    if not base_url:
        raise VisionProviderError(provider_name, "base_url 未配置")
    configured_model = str(resolve_vision_provider_model(provider)).strip()

    headers = build_vision_provider_headers(provider)
    last_error = None
    request_timeout = resolve_vision_provider_request_timeout(provider)
    model_candidates = (
        resolve_vision_provider_model_candidates(provider)
        if api_style in {VISION_API_STYLE_CHAT_COMPLETIONS, VISION_API_STYLE_GENERATE_CONTENT}
        else [resolve_vision_provider_model(provider)]
    )
    for model_name in model_candidates:
        for candidate_index, candidate_base_url in enumerate(base_urls):
            request_cover_limit_override = None
            force_data_url_override = False
            input_payload = None
            is_last_candidate = (
                model_name == model_candidates[-1] and candidate_index == len(base_urls) - 1
            )
            while True:
                try:
                    input_payload = build_input_fn(
                        provider_name,
                        platform,
                        username,
                        cover_urls,
                        model_name=model_name or configured_model,
                        api_style=api_style,
                        force_data_url_override=force_data_url_override,
                        request_cover_limit_override=request_cover_limit_override,
                    )
                except requests.exceptions.RequestException as exc:
                    raise VisionProviderError(
                        provider_name,
                        sanitize_vision_provider_exception_message(exc),
                        retryable=True,
                    ) from exc
                if api_style == VISION_API_STYLE_CHAT_COMPLETIONS:
                    url = f"{candidate_base_url}/chat/completions"
                    body = build_vision_provider_chat_body(provider, input_payload["chat"], model_override=model_name)
                elif api_style == VISION_API_STYLE_GENERATE_CONTENT:
                    url = f"{candidate_base_url}/models/{quote(str(model_name or '').strip(), safe='')}:generateContent"
                    body = input_payload["generate_content"]
                else:
                    url = f"{candidate_base_url}/responses"
                    body = {
                        "model": model_name,
                        "input": input_payload["responses"],
                    }

                try:
                    with acquire_vision_provider_request_slot(provider, candidate_base_url, request_timeout):
                        response = requests.post(url, headers=headers, json=body, timeout=request_timeout)
                except requests.exceptions.RequestException as exc:
                    last_error = VisionProviderError(
                        provider_name,
                        sanitize_vision_provider_exception_message(exc),
                        retryable=True,
                    )
                    if not is_last_candidate:
                        break
                    raise last_error from exc
                if response.status_code >= 400:
                    error_message = build_vision_provider_http_error_message(response)
                    if should_retry_visual_provider_with_inline_images(api_style, error_message) and not force_data_url_override:
                        force_data_url_override = True
                        request_cover_limit_override = resolve_inline_visual_retry_cover_limit(
                            input_payload.get("selected_cover_count")
                        )
                        continue
                    if api_style == VISION_API_STYLE_GENERATE_CONTENT and response.status_code == 413:
                        reduced_cover_limit = reduce_visual_review_cover_limit(input_payload.get("selected_cover_count"))
                        if reduced_cover_limit and reduced_cover_limit != request_cover_limit_override:
                            request_cover_limit_override = reduced_cover_limit
                            continue
                    last_error = VisionProviderError(
                        provider_name,
                        error_message,
                        status_code=response.status_code,
                        retryable=response.status_code in VISUAL_REVIEW_RETRYABLE_STATUS_CODES,
                    )
                    if not is_last_candidate:
                        break
                    raise last_error
                payload = parse_vision_provider_response_payload(response)
                if should_retry_vision_payload(provider_name, payload):
                    last_error = VisionProviderError(
                        provider_name,
                        "模型输出被截断，未返回最终 JSON",
                        retryable=True,
                    )
                    if not is_last_candidate:
                        break
                    raise last_error
                raw_text = extract_vision_response_text(payload)
                text_error = extract_vision_provider_text_error(raw_text)
                if text_error:
                    last_error = VisionProviderError(
                        provider_name,
                        text_error["message"],
                        status_code=text_error.get("status_code"),
                        retryable=text_error.get("retryable", False),
                    )
                    if not is_last_candidate:
                        break
                    raise last_error
                _cleaned_contract_text, contract_payload = parse_json_contract_payload(raw_text)
                if not validate_payload_fn(contract_payload):
                    last_error = VisionProviderError(
                        provider_name,
                        f"模型未返回合法的{contract_label} JSON contract",
                        retryable=True,
                    )
                    if not is_last_candidate:
                        break
                    raise last_error
                response_model = extract_vision_response_model(payload)
                effective_model = str(response_model or model_name or configured_model).strip()
                parsed = parse_result_fn(raw_text)
                parsed["provider"] = provider_name
                parsed["model"] = effective_model
                parsed["configured_model"] = configured_model
                parsed["requested_model"] = str(model_name or "").strip()
                parsed["response_model"] = response_model
                parsed["effective_model"] = effective_model
                parsed["raw_text"] = raw_text
                parsed["usage"] = extract_vision_usage(payload)
                parsed["cover_count"] = input_payload.get("selected_cover_count")
                parsed["candidate_cover_count"] = input_payload.get("candidate_cover_count")
                parsed["skipped_cover_count"] = input_payload.get("skipped_cover_count")
                parsed["prompt_source"] = input_payload.get("prompt_source")
                parsed["prompt_selection"] = dict(input_payload.get("prompt_selection") or {})
                parsed["base_url"] = candidate_base_url
                return parsed
    if last_error is not None:
        raise last_error
    raise VisionProviderError(provider_name, "视觉模型调用失败")


def call_vision_provider(provider, platform, username, cover_urls):
    return call_vision_provider_with_json_contract(
        provider,
        platform=platform,
        username=username,
        cover_urls=cover_urls,
        build_input_fn=build_visual_review_input,
        validate_payload_fn=is_valid_visual_review_payload,
        parse_result_fn=parse_visual_review_result,
        contract_label="视觉复核",
    )


def call_positioning_card_provider(provider, platform, username, cover_urls):
    return call_vision_provider_with_json_contract(
        provider,
        platform=platform,
        username=username,
        cover_urls=cover_urls,
        build_input_fn=build_positioning_card_input,
        validate_payload_fn=is_valid_positioning_card_payload,
        parse_result_fn=parse_positioning_card_result,
        contract_label="定位卡分析",
    )


def build_vision_provider_probe_request(provider):
    provider_name = normalize_vision_provider_name((provider or {}).get("name"))
    api_style = str((provider or {}).get("api_style") or VISION_API_STYLE_RESPONSES).strip().lower()
    base_url = str((provider or {}).get("base_url") or "").rstrip("/")
    if not base_url:
        raise VisionProviderError(provider_name, "base_url 未配置")
    headers = build_vision_provider_headers(provider)
    prompt = "Reply with a short ok."
    if api_style == VISION_API_STYLE_CHAT_COMPLETIONS:
        return {
            "provider_name": provider_name,
            "api_style": api_style,
            "url": f"{base_url}/chat/completions",
            "headers": headers,
            "body": build_vision_provider_chat_body(
                provider,
                build_vision_provider_chat_messages(provider, prompt),
            ),
        }
    if api_style == VISION_API_STYLE_GENERATE_CONTENT:
        return {
            "provider_name": provider_name,
            "api_style": api_style,
            "url": f"{base_url}/models/{quote(str((provider or {}).get('model') or resolve_vision_provider_model(provider)).strip(), safe='')}:generateContent",
            "headers": headers,
            "body": build_vision_provider_generate_content_body([{"text": prompt}]),
        }
    return {
        "provider_name": provider_name,
        "api_style": api_style,
        "url": f"{base_url}/responses",
        "headers": headers,
        "body": {
            "model": str((provider or {}).get("model") or resolve_vision_provider_model(provider)),
            "input": [{"role": "user", "content": [{"type": "input_text", "text": prompt}]}],
        },
    }


def probe_vision_provider(provider):
    base_urls = [
        str(item or "").strip().rstrip("/")
        for item in ((provider or {}).get("base_url_candidates") or resolve_vision_provider_base_urls(provider))
        if str(item or "").strip()
    ]
    if not base_urls:
        fallback_base_url = str((provider or {}).get("base_url") or "").rstrip("/")
        if fallback_base_url:
            base_urls = [fallback_base_url]
    last_error = None
    for candidate_index, candidate_base_url in enumerate(base_urls):
        request_payload = build_vision_provider_probe_request({**provider, "base_url": candidate_base_url})
        try:
            probe_timeout = min(VISION_REQUEST_TIMEOUT, 20)
            with acquire_vision_provider_request_slot(provider, candidate_base_url, probe_timeout):
                response = requests.post(
                    request_payload["url"],
                    headers=request_payload["headers"],
                    json=request_payload["body"],
                    timeout=probe_timeout,
                )
        except requests.exceptions.RequestException as exc:
            last_error = VisionProviderError(
                request_payload["provider_name"],
                sanitize_vision_provider_exception_message(exc),
                retryable=True,
            )
            if candidate_index != len(base_urls) - 1:
                continue
            raise last_error from exc
        if response.status_code >= 400:
            last_error = VisionProviderError(
                request_payload["provider_name"],
                build_vision_provider_http_error_message(response),
                status_code=response.status_code,
                retryable=response.status_code in VISUAL_REVIEW_RETRYABLE_STATUS_CODES,
            )
            if candidate_index != len(base_urls) - 1:
                continue
            raise last_error
        payload = parse_vision_provider_response_payload(response)
        raw_text = extract_vision_response_text(payload) if isinstance(payload, dict) else ""
        text_error = extract_vision_provider_text_error(raw_text)
        if text_error:
            last_error = VisionProviderError(
                request_payload["provider_name"],
                text_error["message"],
                status_code=text_error.get("status_code"),
                retryable=text_error.get("retryable", False),
            )
            if candidate_index != len(base_urls) - 1:
                continue
            raise last_error
        return {
            "success": True,
            "provider": request_payload["provider_name"],
            "api_style": request_payload["api_style"],
            "base_url": candidate_base_url,
            "model": str((provider or {}).get("model") or resolve_vision_provider_model(provider)),
            "checked_at": iso_now(),
            "response_excerpt": str(raw_text or "").strip()[:160],
        }
    if last_error is not None:
        raise last_error
    raise VisionProviderError(str((provider or {}).get("name") or "").strip(), "视觉模型探测失败")


def evaluate_profile_visual_review(platform, review_item, requested_provider="", routing_strategy="", routing_context=None):
    identifier = screening.resolve_profile_review_identifier(platform, review_item)
    cover_urls = build_visual_review_candidate_cover_urls(platform, review_item)
    if not cover_urls:
        raise ValueError("没有可送审的封面 URL")

    normalized_routing_strategy = normalize_visual_review_routing_strategy(routing_strategy)
    if normalized_routing_strategy == VISUAL_REVIEW_ROUTING_TIERED:
        result = run_tiered_visual_review(platform, review_item, cover_urls=cover_urls)
        result["success"] = True
        result["reviewed_at"] = iso_now()
        result["cover_count"] = int(result.get("cover_count") or 0)
        result["candidate_cover_count"] = int(result.get("candidate_cover_count") or len(cover_urls))
        result["skipped_cover_count"] = int(result.get("skipped_cover_count") or 0)
        result["attempt_count"] = int(result.get("attempt_count") or 1)
        return result
    if normalized_routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED:
        result = run_probe_ranked_visual_review(
            platform,
            review_item,
            routing_context or {},
            cover_urls=cover_urls,
        )
        result["success"] = True
        result["reviewed_at"] = iso_now()
        result["cover_count"] = int(result.get("cover_count") or 0)
        result["candidate_cover_count"] = int(result.get("candidate_cover_count") or len(cover_urls))
        result["skipped_cover_count"] = int(result.get("skipped_cover_count") or 0)
        result["attempt_count"] = int(result.get("attempt_count") or 1)
        return result

    last_error = None
    for provider in get_available_vision_providers(requested_provider):
        attempt_count = 0
        while attempt_count < VISUAL_REVIEW_MAX_RETRIES:
            attempt_count += 1
            try:
                result = call_vision_provider(
                    provider,
                    platform,
                    identifier or review_item.get("username") or "",
                    cover_urls,
                )
                result["success"] = True
                result["reviewed_at"] = iso_now()
                result["cover_count"] = int(result.get("cover_count") or 0)
                result["candidate_cover_count"] = int(result.get("candidate_cover_count") or len(cover_urls))
                result["skipped_cover_count"] = int(result.get("skipped_cover_count") or 0)
                result["attempt_count"] = attempt_count
                return result
            except Exception as exc:
                last_error = exc
                should_retry = is_retryable_visual_exception(exc) and attempt_count < VISUAL_REVIEW_MAX_RETRIES
                if not should_retry:
                    break
                time.sleep(compute_visual_retry_delay_seconds(attempt_count))
                continue

    raise RuntimeError(str(last_error) if last_error else "缺少视觉模型配置")


def evaluate_profile_positioning_card_analysis(platform, review_item, requested_provider=""):
    identifier = screening.resolve_profile_review_identifier(platform, review_item)
    cover_urls = build_visual_review_candidate_cover_urls(platform, review_item)
    if not cover_urls:
        raise ValueError("没有可送审的封面 URL")

    last_error = None
    for provider in get_available_vision_providers(requested_provider):
        attempt_count = 0
        while attempt_count < VISUAL_REVIEW_MAX_RETRIES:
            attempt_count += 1
            try:
                result = call_positioning_card_provider(
                    provider,
                    platform,
                    identifier or review_item.get("username") or "",
                    cover_urls,
                )
                result["success"] = True
                result["reviewed_at"] = iso_now()
                result["cover_count"] = int(result.get("cover_count") or 0)
                result["candidate_cover_count"] = int(result.get("candidate_cover_count") or len(cover_urls))
                result["skipped_cover_count"] = int(result.get("skipped_cover_count") or 0)
                result["attempt_count"] = attempt_count
                return result
            except Exception as exc:
                last_error = exc
                should_retry = is_retryable_visual_exception(exc) and attempt_count < VISUAL_REVIEW_MAX_RETRIES
                if not should_retry:
                    break
                time.sleep(compute_visual_retry_delay_seconds(attempt_count))
                continue

    raise RuntimeError(str(last_error) if last_error else "缺少视觉模型配置")


def build_visual_review_partial_result(platform, results, targets, channel_race=None):
    target_identifiers = {
        screening.resolve_profile_review_identifier(platform, item)
        for item in (targets or [])
        if screening.resolve_profile_review_identifier(platform, item)
    }
    passed = 0
    rejected = 0
    failed = 0
    filtered_results = {}
    for key, item in (results or {}).items():
        if not isinstance(item, dict):
            continue
        identifier = screening.normalize_identifier(key or item.get("username"))
        if target_identifiers and identifier not in target_identifiers:
            continue
        filtered_results[identifier] = item
        if item.get("success") is False:
            failed += 1
            continue
        if str(item.get("decision") or "").strip() == "Reject":
            rejected += 1
        else:
            passed += 1
    payload = {
        "platform": platform,
        "target_total": len(targets),
        "reviewed_total": passed + rejected + failed,
        "summary": {
            "pass": passed,
            "reject": rejected,
            "error": failed,
        },
        "usage": summarize_visual_usage(filtered_results),
        "visual_results_path": get_visual_results_path(platform),
        "visual_results": filtered_results,
    }
    if channel_race:
        payload["channel_race"] = sanitize_json_compatible(channel_race)
    return payload


def build_positioning_card_partial_result(platform, results, targets):
    target_identifiers = {
        screening.resolve_profile_review_identifier(platform, item)
        for item in (targets or [])
        if screening.resolve_profile_review_identifier(platform, item)
    }
    completed = 0
    failed = 0
    recommendation_counts = {
        "High Fit": 0,
        "Medium Fit": 0,
        "Low Fit": 0,
        "Unclear": 0,
    }
    filtered_results = {}
    for key, item in (results or {}).items():
        if not isinstance(item, dict):
            continue
        identifier = screening.normalize_identifier(key or item.get("username"))
        if target_identifiers and identifier not in target_identifiers:
            continue
        filtered_results[identifier] = item
        if item.get("success") is False:
            failed += 1
            continue
        completed += 1
        fit_recommendation = screening.normalize_fit_recommendation(item.get("fit_recommendation"))
        recommendation_counts[fit_recommendation] = recommendation_counts.get(fit_recommendation, 0) + 1
    return {
        "platform": platform,
        "target_total": len(targets),
        "reviewed_total": completed + failed,
        "summary": {
            "completed": completed,
            "error": failed,
            "fit_recommendations": recommendation_counts,
        },
        "usage": summarize_visual_usage(filtered_results),
        "positioning_card_results_path": get_positioning_card_results_path(platform),
        "positioning_card_results": filtered_results,
    }


class ApifyStartError(RuntimeError):
    def __init__(self, status_code, message):
        self.status_code = int(status_code or 0)
        super().__init__(message)

    @property
    def retryable_with_next_token(self):
        return self.status_code in {401, 402, 403, 429} or self.status_code in TRANSIENT_STATUS_CODES

    @property
    def uncertain_submission(self):
        return self.status_code in TRANSIENT_STATUS_CODES


class ApifyRuntimeError(RuntimeError):
    def __init__(
        self,
        failure_stage,
        message,
        *,
        retryable=False,
        apify=None,
        partial_result=None,
    ):
        self.failure_stage = str(failure_stage or "failed").strip() or "failed"
        self.retryable = bool(retryable)
        self.apify = sanitize_json_compatible(apify or {})
        self.partial_result = sanitize_json_compatible(partial_result) if partial_result is not None else None
        super().__init__(message)


def normalize_upload_column_name(name):
    text = str(name or "").strip().lower()
    if not text:
        return ""
    return re.sub(r"[^a-z0-9]+", "", text)


def clean_upload_metadata_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if pd.isna(value):
        return ""
    return value


def normalize_upload_platform_value(value):
    normalized = normalize_upload_column_name(value)
    return UPLOAD_PLATFORM_ALIASES.get(normalized, "")


def infer_upload_platform_value(value):
    normalized = normalize_upload_column_name(value)
    if not normalized:
        return ""
    for alias, platform in UPLOAD_PLATFORM_ALIASES.items():
        if alias and alias in normalized:
            return platform
    return ""


def is_probable_upload_profile_url(value):
    text = str(clean_upload_metadata_value(value) or "").strip().lower()
    if not text:
        return False
    return any(
        marker in text
        for marker in (
            "http://",
            "https://",
            "www.instagram.com",
            "instagram.com/",
            "www.tiktok.com",
            "tiktok.com/",
            "www.youtube.com",
            "youtube.com/",
            "youtu.be/",
        )
    )


def is_empty_upload_row(row_dict):
    for key, value in (row_dict or {}).items():
        if str(key).startswith("__"):
            continue
        cleaned = clean_upload_metadata_value(value)
        if cleaned not in ("", None):
            return False
    return True


def format_upload_row_location(row_dict, fallback_row_num):
    sheet_name = str((row_dict or {}).get("__sheet_name") or "").strip()
    row_num = clean_upload_metadata_value((row_dict or {}).get("__sheet_row_num"))
    try:
        row_num = int(row_num)
    except Exception:
        row_num = fallback_row_num
    if sheet_name:
        return f"Sheet `{sheet_name}` 第 {row_num} 行"
    return f"第 {row_num} 行"


def build_upload_validation_error(message, details, error_code="UPLOAD_TEMPLATE_INVALID"):
    return jsonify({
        "success": False,
        "error": message,
        "error_code": error_code,
        "details": [str(item) for item in details if str(item).strip()],
    }), 400


def load_canonical_upload_workbook_frames(filepath):
    workbook = pd.read_excel(filepath, sheet_name=None)
    frames = []
    for sheet_name, df in (workbook or {}).items():
        if df is None or df.empty:
            continue
        prepared = df.copy()
        prepared["__sheet_name"] = str(sheet_name or "").strip() or "Sheet1"
        prepared["__sheet_row_num"] = [index + 2 for index in range(len(prepared))]
        if not any(
            clean_upload_metadata_value(value) not in ("", None)
            for column in prepared.columns
            if not str(column).startswith("__")
            for value in prepared[column].tolist()
        ):
            continue
        frames.append(prepared)
    return frames


def resolve_canonical_upload_columns(columns):
    normalized = {}
    for column in columns:
        normalized[normalize_upload_column_name(column)] = column
    resolved = {
        "platform": normalized.get("platform"),
        "handle": normalized.get("username") or normalized.get("handle"),
        "url": normalized.get("url"),
    }
    missing = []
    if not resolved["platform"]:
        missing.append("Platform")
    if not resolved["handle"]:
        missing.append("@username")
    return resolved, missing


def build_upload_metadata_record(row_dict, platform, canonical_url, source_filename):
    metadata = {
        "platform": platform,
        "url": canonical_url,
        "source_filename": source_filename,
    }
    for column_name, raw_value in (row_dict or {}).items():
        normalized_column = normalize_upload_column_name(column_name)
        mapped_field = UPLOAD_METADATA_FIELD_ALIASES.get(normalized_column)
        if not mapped_field:
            continue
        if mapped_field == "url" and metadata.get("url"):
            continue
        cleaned = clean_upload_metadata_value(raw_value)
        if cleaned in ("", None):
            continue
        metadata[mapped_field] = cleaned
    handle = screening.extract_platform_identifier(platform, canonical_url) or screening.extract_platform_identifier(platform, metadata.get("handle"))
    metadata["handle"] = handle
    return sanitize_json_compatible(metadata)


def parse_canonical_upload_workbook(df, source_filename):
    resolved_columns, missing = resolve_canonical_upload_columns(df.columns)
    if missing:
        details = [f"缺少必填列：{label}" for label in missing]
        details.append("固定模板至少需要 `Platform` 和 `@username` 两列。")
        return None, build_upload_validation_error(
            f"上传模板缺少必填列：{', '.join(missing)}",
            details,
        )

    grouped_data = {platform: [] for platform in PLATFORM_ACTORS}
    metadata_by_platform = {platform: {} for platform in PLATFORM_ACTORS}
    preview_rows = []
    invalid_rows = []
    processed = 0

    for index, (_, row_series) in enumerate(df.iterrows(), start=2):
        row_dict = row_series.to_dict()
        if is_empty_upload_row(row_dict):
            continue
        processed += 1
        row_location = format_upload_row_location(row_dict, index)
        platform = normalize_upload_platform_value(row_dict.get(resolved_columns["platform"]))
        if not platform:
            fallback_values = [
                row_dict.get(resolved_columns["url"]) if resolved_columns.get("url") else "",
                row_dict.get(resolved_columns["handle"]) if resolved_columns.get("handle") else "",
                row_dict.get("profile_dedupe_key"),
                row_dict.get("URL"),
                row_dict.get("@username"),
            ]
            for fallback_value in fallback_values:
                platform = infer_upload_platform_value(fallback_value)
                if platform:
                    break
        if not platform:
            invalid_rows.append(f"{row_location} Platform 无效，只支持 Instagram / TikTok / YouTube。")
            continue

        raw_handle = clean_upload_metadata_value(row_dict.get(resolved_columns["handle"]))
        raw_url = clean_upload_metadata_value(row_dict.get(resolved_columns["url"])) if resolved_columns.get("url") else ""
        raw_url_lower = str(raw_url or "").strip().lower()
        raw_url_is_search = platform == "tiktok" and "tiktok.com/search" in raw_url_lower
        raw_url_identifier = (
            screening.extract_platform_identifier(platform, raw_url)
            if raw_url and (is_probable_upload_profile_url(raw_url) or raw_url_is_search)
            else ""
        )
        identifier = (
            raw_url_identifier
            or screening.extract_platform_identifier(platform, raw_handle)
        )
        if not identifier:
            invalid_rows.append(f"{row_location} @username 为空或无法识别：`{raw_handle}`。")
            continue

        canonical_url = (
            raw_url
            if raw_url and is_probable_upload_profile_url(raw_url) and not raw_url_is_search and raw_url_identifier
            else screening.build_canonical_profile_url(platform, identifier)
        )
        metadata = build_upload_metadata_record(row_dict, platform, canonical_url, source_filename)
        metadata_by_platform[platform][identifier] = metadata
        grouped_data[platform].append(canonical_url)

        if len(preview_rows) < 5:
            preview_rows.append({
                "Platform": UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform),
                "@username": identifier,
                "URL": canonical_url,
                "nickname": metadata.get("nickname", ""),
                "Region": metadata.get("region", ""),
                "Language": metadata.get("language", ""),
                "Followers": metadata.get("followers", ""),
            })

    if processed == 0:
        return None, build_upload_validation_error(
            "上传表没有可用数据行",
            ["请确认表头下方至少有一行 `Platform` 和 `@username` 都已填写的账号数据。"],
        )
    if invalid_rows:
        return None, build_upload_validation_error(
            "上传模板存在无效数据，请修正后重试",
            invalid_rows[:20],
        )

    deduped_grouped = {platform: list(dict.fromkeys(values)) for platform, values in grouped_data.items()}
    stats = {UPLOAD_PLATFORM_RESPONSE_LABELS[platform]: len(values) for platform, values in deduped_grouped.items()}
    stats["Unknown"] = 0
    return {
        "grouped_data": deduped_grouped,
        "metadata_by_platform": metadata_by_platform,
        "preview": preview_rows,
        "stats": stats,
    }, None


def load_active_rulespec():
    payload = load_json_payload(ACTIVE_RULESPEC_PATH, default={})
    return payload if isinstance(payload, dict) else {}


def persist_active_rulespec(compiled):
    rulespec = compiled.get("rule_spec") or {}
    rules_module.write_json(ACTIVE_RULESPEC_PATH, rulespec)
    rules_module.write_json(FIELD_MATCH_REPORT_PATH, compiled.get("field_match_report") or {})
    rules_module.write_json(MISSING_CAPABILITIES_PATH, compiled.get("missing_capabilities") or {})
    Path(REVIEW_NOTES_PATH).write_text(compiled.get("review_notes_markdown") or "", encoding="utf-8")


def create_job(job_type, platform=None, message="任务已创建"):
    job = {
        "id": uuid.uuid4().hex,
        "type": job_type,
        "platform": platform,
        "status": "queued",
        "stage": "queued",
        "message": message,
        "created_at": iso_now(),
        "updated_at": iso_now(),
        "progress": {"done": 0, "total": None, "percent": 0, "determinate": False},
        "partial_result": None,
        "result": None,
        "error": None,
        "cancel_requested": False,
    }
    with JOBS_LOCK:
        JOBS[job["id"]] = job
    return job


def get_job(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        return dict(job) if job else None


def update_job(job_id, **updates):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return None
        job.update(updates)
        job["updated_at"] = iso_now()
        return dict(job)


def build_job_progress(done=None, total=None):
    resolved_done = 0 if done is None else done
    resolved_total = total
    determinate = isinstance(resolved_done, (int, float)) and isinstance(resolved_total, (int, float)) and resolved_total > 0
    percent = int((resolved_done / resolved_total) * 100) if determinate else 0
    return {
        "done": resolved_done,
        "total": resolved_total,
        "percent": max(0, min(100, percent)),
        "determinate": determinate,
    }


def start_background_job(job, worker):
    def runner():
        update_job(job["id"], status="running", stage="starting")

        def progress_callback(stage, message=None, done=None, total=None, partial_result=None, **extra):
            payload = {
                "stage": stage,
                "message": message or stage,
                "progress": build_job_progress(done=done, total=total),
            }
            if partial_result is not None:
                payload["partial_result"] = sanitize_json_compatible(partial_result)
            if extra:
                payload["metadata"] = sanitize_json_compatible(extra)
            update_job(job["id"], **payload)

        def cancel_check():
            with JOBS_LOCK:
                current = JOBS.get(job["id"]) or {}
                return bool(current.get("cancel_requested"))

        try:
            result = worker(progress_callback, cancel_check)
            if result and result.get("cancelled"):
                update_job(
                    job["id"],
                    status="cancelled",
                    stage="cancelled",
                    message=result.get("message") or "任务已取消",
                    result=sanitize_json_compatible(result),
                    progress=build_job_progress(done=1, total=1),
                )
                return
            if not result or not result.get("success"):
                failure_stage = str((result or {}).get("failure_stage") or (result or {}).get("stage") or "failed").strip() or "failed"
                updates = {
                    "status": "failed",
                    "stage": failure_stage,
                    "message": (result or {}).get("error") or "任务失败",
                    "error": (result or {}).get("error") or "任务失败",
                    "result": sanitize_json_compatible(result),
                    "progress": build_job_progress(done=1, total=1),
                }
                if isinstance(result, dict) and result.get("partial_result") is not None:
                    updates["partial_result"] = sanitize_json_compatible(result.get("partial_result"))
                update_job(job["id"], **updates)
                return
            update_job(
                job["id"],
                status="completed",
                stage="completed",
                message=result.get("message") or "任务完成",
                result=sanitize_json_compatible(result),
                partial_result=sanitize_json_compatible(result),
                progress=build_job_progress(done=1, total=1),
            )
        except Exception as exc:  # pragma: no cover - runtime guard
            failure_stage = str(getattr(exc, "failure_stage", "") or "failed").strip() or "failed"
            error_payload = {
                "success": False,
                "error": str(exc),
                "failure_stage": failure_stage,
            }
            apify_payload = getattr(exc, "apify", None)
            if apify_payload is not None:
                error_payload["apify"] = sanitize_json_compatible(apify_payload)
            retryable = getattr(exc, "retryable", None)
            if retryable is not None:
                error_payload["retryable"] = bool(retryable)
            partial_result = getattr(exc, "partial_result", None)
            if partial_result is not None:
                error_payload["partial_result"] = sanitize_json_compatible(partial_result)
            updates = {
                "status": "failed",
                "stage": failure_stage,
                "message": str(exc),
                "error": str(exc),
                "result": error_payload,
                "progress": build_job_progress(done=1, total=1),
            }
            if partial_result is not None:
                updates["partial_result"] = sanitize_json_compatible(partial_result)
            update_job(job["id"], **updates)

    thread = threading.Thread(target=runner, daemon=True, name=f"job-{job['id']}")
    thread.start()


def build_cancelled_result(message="任务已取消"):
    return {"success": False, "cancelled": True, "message": message}


def chunk_list(items, chunk_size):
    chunk_size = max(1, int(chunk_size or 1))
    return [items[index:index + chunk_size] for index in range(0, len(items), chunk_size)]


def build_target_preview(identifiers, max_items=5):
    return {
        "targets": [str(item) for item in list(identifiers or [])[:max_items]],
        "target_count": len(list(identifiers or [])),
    }


def build_visual_review_cache_prompt_descriptor(prompt_selection):
    if not isinstance(prompt_selection, dict):
        return {}
    prompt_text = str(prompt_selection.get("prompt") or "").strip()
    return {
        "platform": normalize_visual_prompt_lookup_key(prompt_selection.get("platform")),
        "provider": normalize_vision_provider_name(prompt_selection.get("provider")),
        "model": normalize_visual_prompt_lookup_key(prompt_selection.get("model")),
        "source": str(prompt_selection.get("source") or "").strip(),
        "visual_contract_source": str(prompt_selection.get("visual_contract_source") or "").strip(),
        "resolved_cover_limit": screening.coerce_positive_int(prompt_selection.get("resolved_cover_limit")) or 0,
        "prompt_key": creator_cache.stable_cache_key(prompt_text),
        "visual_runtime_contract_key": creator_cache.stable_cache_key(
            prompt_selection.get("visual_runtime_contract") or {}
        ),
    }


def build_visual_review_cache_context(
    platform,
    *,
    requested_provider="",
    routing_strategy="",
    providers=None,
    active_rulespec=None,
    active_visual_prompts=None,
):
    normalized_platform = str(platform or "").strip().lower()
    normalized_requested_provider = normalize_vision_provider_name(requested_provider)
    normalized_routing_strategy = normalize_visual_review_routing_strategy(routing_strategy)
    resolved_active_rulespec = load_active_rulespec() if active_rulespec is None else active_rulespec
    resolved_active_visual_prompts = (
        load_active_visual_prompts() if active_visual_prompts is None else active_visual_prompts
    )
    resolved_providers = list(providers or get_available_vision_providers(normalized_requested_provider))

    provider_descriptors = []
    for provider in resolved_providers:
        provider_name = normalize_vision_provider_name((provider or {}).get("name"))
        configured_model = str(resolve_vision_provider_model(provider) or "").strip()
        prompt_selection = resolve_visual_review_prompt_selection(
            provider_name,
            normalized_platform,
            model_name=configured_model,
            active_visual_prompts=resolved_active_visual_prompts,
            active_rulespec=resolved_active_rulespec,
        )
        provider_descriptors.append(
            {
                "name": provider_name,
                "api_style": str((provider or {}).get("api_style") or "").strip().lower(),
                "model": configured_model,
                "model_candidates": dedupe_non_empty_strings(resolve_vision_provider_model_candidates(provider)),
                "base_urls": dedupe_non_empty_strings(
                    (provider or {}).get("base_url_candidates") or resolve_vision_provider_base_urls(provider)
                ),
                "request_timeout_seconds": resolve_vision_provider_request_timeout(provider),
                "prompt": build_visual_review_cache_prompt_descriptor(prompt_selection),
            }
        )
    provider_descriptors.sort(
        key=lambda item: (
            str(item.get("name") or ""),
            str(item.get("model") or ""),
            str(item.get("api_style") or ""),
        )
    )

    stage_descriptors = []
    raw_stage_plan = []
    if normalized_routing_strategy == VISUAL_REVIEW_ROUTING_TIERED:
        raw_stage_plan = build_visual_review_routing_plan()
    elif normalized_routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED:
        raw_stage_plan = build_visual_review_probe_ranked_plan()
    for stage in raw_stage_plan:
        provider_name = normalize_vision_provider_name((stage or {}).get("provider"))
        configured_model = str((stage or {}).get("model") or "").strip()
        prompt_selection = resolve_visual_review_prompt_selection(
            provider_name,
            normalized_platform,
            model_name=configured_model,
            active_visual_prompts=resolved_active_visual_prompts,
            active_rulespec=resolved_active_rulespec,
        )
        stage_descriptors.append(
            {
                "stage": str((stage or {}).get("stage") or "").strip(),
                "group": str((stage or {}).get("group") or "").strip(),
                "provider": provider_name,
                "model": configured_model,
                "timeout_seconds": screening.coerce_positive_int((stage or {}).get("timeout_seconds")) or 0,
                "prompt": build_visual_review_cache_prompt_descriptor(prompt_selection),
            }
        )

    context_payload = {
        "version": "visual_review_cache_v1",
        "platform": normalized_platform,
        "requested_provider": normalized_requested_provider,
        "routing_strategy": normalized_routing_strategy or "direct",
        "active_rulespec_key": creator_cache.stable_cache_key(resolved_active_rulespec or {}),
        "active_visual_prompts_key": creator_cache.stable_cache_key(resolved_active_visual_prompts or {}),
        "providers": provider_descriptors,
        "routing_plan": stage_descriptors,
    }
    return {
        "context_key": creator_cache.stable_cache_key(context_payload),
        "context_payload": context_payload,
    }


def get_requested_visual_identifiers(platform, payload):
    candidates = []
    for key in ("identifiers", "usernames", "profiles", "urls"):
        for item in (payload.get(key) or []):
            identifier = screening.extract_platform_identifier(platform, item)
            if identifier:
                candidates.append(identifier)
    return list(dict.fromkeys(candidates))


def resolve_visual_review_targets(platform, payload):
    requested_identifiers = set(get_requested_visual_identifiers(platform, payload))
    targets = []
    for item in merge_upload_metadata_into_reviews(platform, load_profile_reviews(platform)):
        if str(item.get("status") or "").strip() != "Pass":
            continue
        identifier = screening.resolve_profile_review_identifier(platform, item)
        if requested_identifiers and identifier not in requested_identifiers:
            continue
        targets.append(item)
    return targets


def resolve_positioning_card_analysis_targets(platform, payload):
    requested_identifiers = set(get_requested_visual_identifiers(platform, payload))
    visual_results = load_visual_results(platform)
    targets = []
    for item in merge_upload_metadata_into_reviews(platform, load_profile_reviews(platform)):
        if str(item.get("status") or "").strip() != "Pass":
            continue
        identifier = screening.resolve_profile_review_identifier(platform, item)
        if not identifier:
            continue
        if requested_identifiers and identifier not in requested_identifiers:
            continue
        visual_result = visual_results.get(identifier) or {}
        if not isinstance(visual_result, dict):
            continue
        if visual_result.get("success") is False:
            continue
        if str(visual_result.get("decision") or "").strip() != "Pass":
            continue
        targets.append({
            **item,
            "_visual_result": visual_result,
        })
    return targets


def perform_visual_review(platform, payload, progress_callback=None, cancel_check=None):
    requested_provider = normalize_vision_provider_name((payload or {}).get("provider"))
    routing_strategy = resolve_visual_review_routing_strategy(payload)
    if requested_provider and requested_provider in {item["name"] for item in VISION_PROVIDER_CONFIGS}:
        routing_strategy = ""
    started_at = time.monotonic()
    preflight = build_vision_preflight(requested_provider)
    providers = get_available_vision_providers(requested_provider)
    if not providers:
        return build_vision_preflight_error_payload(requested_provider)
    active_rulespec = load_active_rulespec()
    active_visual_prompts = load_active_visual_prompts()

    targets = resolve_visual_review_targets(platform, payload)
    if not targets:
        return {
            "success": False,
            "error": "没有可复核的账号：请先完成抓取，并确保至少有一个 Prescreen=Pass 的账号。",
        }

    target_identifiers = [screening.resolve_profile_review_identifier(platform, item) for item in targets]
    results = load_visual_results(platform)
    for identifier in target_identifiers:
        if identifier:
            results.pop(identifier, None)
    creator_cache_enabled = creator_cache.creator_cache_enabled(payload)
    force_refresh_creator_cache = creator_cache.creator_cache_force_refresh(payload)
    creator_cache_db_path = creator_cache.resolve_creator_cache_db_path(payload)
    visual_cache_context = build_visual_review_cache_context(
        platform,
        requested_provider=requested_provider,
        routing_strategy=routing_strategy,
        providers=providers,
        active_rulespec=active_rulespec,
        active_visual_prompts=active_visual_prompts,
    )
    creator_cache_hit_identifiers: list[str] = []
    if creator_cache_enabled and not force_refresh_creator_cache:
        cached_visual_results = creator_cache.load_visual_cache_entries(
            platform,
            target_identifiers,
            creator_cache_db_path,
            visual_cache_context["context_key"],
        )
        for identifier, cached_visual_result in cached_visual_results.items():
            if creator_cache.is_cacheable_visual_result(cached_visual_result):
                results[identifier] = dict(cached_visual_result)
                creator_cache_hit_identifiers.append(identifier)
        if cached_visual_results:
            save_visual_results(platform, results)

    pending_targets = []
    completed_identifiers = []
    completed_identifier_set = {identifier for identifier in creator_cache_hit_identifiers if identifier}
    seen_pending_identifiers = set()
    for item in targets:
        identifier = screening.resolve_profile_review_identifier(platform, item)
        if identifier in completed_identifier_set:
            completed_identifiers.append(identifier)
            continue
        if identifier and identifier in seen_pending_identifiers:
            continue
        if identifier:
            seen_pending_identifiers.add(identifier)
        pending_targets.append(item)

    creator_cache_summary = {
        "enabled": bool(creator_cache_enabled),
        "db_path": str(creator_cache_db_path),
        "force_refresh": bool(force_refresh_creator_cache),
        "scrape_hit_count": 0,
        "visual_hit_count": len(creator_cache_hit_identifiers),
        "visual_hit_preview": creator_cache_hit_identifiers[:10],
        "visual_miss_count": len(pending_targets),
        "visual_context_key": str(visual_cache_context.get("context_key") or ""),
    }
    if not pending_targets:
        final_result = build_visual_review_partial_result(
            platform,
            results,
            targets,
            channel_race={},
        )
        final_result.update({
            "success": True,
            "message": (
                f"{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 视觉复核完成，"
                f"共处理 {len(targets)} 个账号。"
            ),
            "visual_results": results,
            "max_workers": 0,
            "selected_provider": preflight.get("preferred_provider"),
            "selected_model": "",
            "channel_race": {},
            "elapsed_seconds": round(time.monotonic() - started_at, 3),
            "creator_cache": creator_cache_summary,
        })
        return final_result

    routing_context = {}
    selected_provider_name = preflight.get("preferred_provider")
    selected_model_name = ""
    provider_candidates = get_available_vision_provider_names(requested_provider)
    probe_cover_urls = build_visual_review_candidate_cover_urls(platform, targets[0]) if targets else []
    if routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED:
        channel_race = run_probe_ranked_visual_provider_race(platform=platform, cover_urls=probe_cover_urls)
        routing_context = build_probe_ranked_runtime_context(channel_race)
        provider_candidates = dedupe_non_empty_strings([
            str(item.get("provider") or "").strip()
            for item in (channel_race.get("candidates") or [])
            if str(item.get("provider") or "").strip()
        ])
        if not channel_race.get("success"):
            return {
                "success": False,
                "error_code": "VISION_CHANNEL_RACE_FAILED",
                "error": "视觉通道赛马失败：当前优先链路都不可用。",
                "vision_preflight": preflight,
                "channel_race": channel_race,
            }
        selected_provider_name = channel_race.get("selected_provider") or selected_provider_name
        selected_model_name = str(channel_race.get("selected_model") or "").strip()
    else:
        channel_race = {}
    max_workers = resolve_visual_review_max_workers(
        payload,
        len(pending_targets) or len(targets),
        requested_provider=selected_provider_name or requested_provider,
        requested_model=selected_model_name,
        routing_strategy=routing_strategy,
    )
    item_timeout_seconds = resolve_visual_review_item_timeout_seconds(payload)
    future_poll_timeout_seconds = min(0.5, max(0.05, item_timeout_seconds / 4.0))
    if progress_callback:
            progress_callback(
                "preparing",
                "正在准备视觉复核任务",
            done=len(completed_identifiers),
            total=len(targets),
            providers=provider_candidates,
                selected_provider=selected_provider_name,
                selected_model=selected_model_name,
                routing_strategy=routing_strategy or "direct",
                max_workers=max_workers,
                item_timeout_seconds=item_timeout_seconds,
                creator_cache=creator_cache_summary,
                channel_race=snapshot_probe_ranked_channel_race(routing_context) if routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED else sanitize_json_compatible(channel_race) if channel_race else None,
                **build_target_preview(target_identifiers),
            )

    completed = len(completed_identifiers)
    target_iter = iter(pending_targets)
    future_map = {}

    def finalize_review(identifier, result=None, error_text=None):
        nonlocal completed
        if error_text is not None:
            results[identifier] = {
                "username": identifier,
                "success": False,
                "error": str(error_text),
                "reviewed_at": iso_now(),
            }
        else:
            results[identifier] = {
                "username": identifier,
                "decision": result.get("decision"),
                "reason": result.get("reason"),
                "signals": result.get("signals") or [],
                "provider": result.get("provider"),
                "model": result.get("model"),
                "configured_model": result.get("configured_model"),
                "requested_model": result.get("requested_model"),
                "response_model": result.get("response_model"),
                "effective_model": result.get("effective_model"),
                "route": result.get("route"),
                "routing_strategy": result.get("routing_strategy"),
                "trace": result.get("trace") or [],
                "judge_used": bool(result.get("judge_used")),
                "escalation_reasons": result.get("escalation_reasons") or [],
                "usage": result.get("usage") or {},
                "cover_count": result.get("cover_count"),
                "candidate_cover_count": result.get("candidate_cover_count"),
                "skipped_cover_count": result.get("skipped_cover_count"),
                "reviewed_at": result.get("reviewed_at"),
                "attempt_count": result.get("attempt_count"),
                "channel_race": result.get("channel_race") or {},
            }
            if creator_cache_enabled:
                creator_cache.persist_visual_cache_entry(
                    platform,
                    identifier,
                    results[identifier],
                    creator_cache_db_path,
                    updated_at=str(results[identifier].get("reviewed_at") or iso_now()),
                    context_key=str(visual_cache_context.get("context_key") or ""),
                    context_payload=visual_cache_context.get("context_payload") or {},
                )

        completed += 1
        save_visual_results(platform, results)
        partial_result = build_visual_review_partial_result(
            platform,
            results,
            targets,
            channel_race=snapshot_probe_ranked_channel_race(routing_context) if routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED else channel_race,
        )
        if progress_callback:
            progress_callback(
                "reviewing",
                f"已完成第 {completed}/{len(targets)} 个账号：{identifier}",
                done=completed,
                total=len(targets),
                partial_result=partial_result,
                current_identifier=identifier,
                max_workers=max_workers,
            )

    def submit_next(executor):
        if cancel_check and cancel_check():
            return False
        try:
            review_item = next(target_iter)
        except StopIteration:
            return False
        identifier = screening.resolve_profile_review_identifier(platform, review_item)
        if progress_callback:
            progress_callback(
                "reviewing",
                f"已提交视觉复核：{identifier}",
                done=completed,
                total=len(targets),
                provider_candidates=provider_candidates,
                selected_provider=selected_provider_name,
                selected_model=selected_model_name,
                routing_strategy=routing_strategy or "direct",
                current_identifier=identifier,
                max_workers=max_workers,
            )
        future = executor.submit(
            evaluate_profile_visual_review,
            platform,
            review_item,
            requested_provider,
            routing_strategy,
            routing_context,
        )
        future_map[future] = {
            "identifier": identifier,
            "review_item": review_item,
            "submitted_at": time.monotonic(),
        }
        return True

    executor = DaemonThreadPoolExecutor(max_workers=max_workers, thread_name_prefix=f"{platform}-visual")
    try:
        for _ in range(max_workers):
            if not submit_next(executor):
                break

        while future_map:
            if cancel_check and cancel_check():
                return build_cancelled_result()

            now = time.monotonic()
            expired_futures = []
            for future, metadata in list(future_map.items()):
                if future.done():
                    continue
                submitted_at = float(metadata.get("submitted_at") or 0.0)
                if submitted_at and (now - submitted_at) >= item_timeout_seconds:
                    expired_futures.append((future, metadata))
            if expired_futures:
                for future, metadata in expired_futures:
                    identifier = metadata.get("identifier") or ""
                    future_map.pop(future, None)
                    future.cancel()
                    finalize_review(
                        identifier,
                        error_text=f"视觉复核超时：{identifier} 超过 {int(item_timeout_seconds)} 秒未完成",
                    )
                    submit_next(executor)
                continue

            completed_futures, _ = wait(
                tuple(future_map.keys()),
                timeout=future_poll_timeout_seconds,
                return_when=FIRST_COMPLETED,
            )
            if not completed_futures:
                continue

            for future in completed_futures:
                metadata = future_map.pop(future)
                identifier = metadata.get("identifier") or ""
                try:
                    result = future.result()
                    finalize_review(identifier, result=result)
                except Exception as exc:
                    finalize_review(identifier, error_text=str(exc))

                submit_next(executor)
    finally:
        executor.shutdown(wait=False, cancel_futures=True)

    final_result = build_visual_review_partial_result(
        platform,
        results,
        targets,
        channel_race=snapshot_probe_ranked_channel_race(routing_context) if routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED else channel_race,
    )
    final_result.update({
        "success": True,
        "message": (
            f"{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 视觉复核完成，"
            f"共处理 {len(targets)} 个账号。"
        ),
        "visual_results": results,
        "max_workers": max_workers,
        "selected_provider": selected_provider_name,
        "selected_model": selected_model_name,
        "channel_race": snapshot_probe_ranked_channel_race(routing_context) if routing_strategy == VISUAL_REVIEW_ROUTING_PROBE_RANKED else channel_race,
        "elapsed_seconds": round(time.monotonic() - started_at, 3),
        "creator_cache": creator_cache_summary,
    })
    return final_result


def perform_positioning_card_analysis(platform, payload, progress_callback=None, cancel_check=None):
    requested_provider = normalize_vision_provider_name((payload or {}).get("provider"))
    preflight = build_vision_preflight(requested_provider)
    providers = get_available_vision_providers(requested_provider)
    if not providers:
        return build_vision_preflight_error_payload(requested_provider)

    targets = resolve_positioning_card_analysis_targets(platform, payload)
    if not targets:
        return {
            "success": False,
            "error": "没有可分析的账号：请先完成视觉复核，并确保至少有一个 Visual=Pass 的账号。",
        }

    results = load_positioning_card_results(platform)
    target_identifiers = [screening.resolve_profile_review_identifier(platform, item) for item in targets]
    max_workers = resolve_visual_review_max_workers(payload, len(targets), requested_provider=requested_provider)
    item_timeout_seconds = resolve_visual_review_item_timeout_seconds(payload)
    started_at = time.monotonic()
    future_poll_timeout_seconds = min(0.5, max(0.05, item_timeout_seconds / 4.0))
    if progress_callback:
        progress_callback(
            "preparing",
            "正在准备定位卡分析任务",
            done=0,
            total=len(targets),
            providers=[item["name"] for item in providers],
            selected_provider=preflight.get("preferred_provider") or requested_provider,
            max_workers=max_workers,
            item_timeout_seconds=item_timeout_seconds,
            **build_target_preview(target_identifiers),
        )

    completed = 0
    target_iter = iter(targets)
    future_map = {}

    def finalize_analysis(identifier, review_item, result=None, error_text=None):
        nonlocal completed
        merged_review_item = merge_upload_metadata_into_review_item(platform, review_item)
        visual_result = dict((review_item or {}).get("_visual_result") or {})
        if error_text is not None:
            results[identifier] = {
                "username": identifier,
                "profile_url": str(merged_review_item.get("profile_url") or "").strip(),
                "success": False,
                "error": str(error_text),
                "reviewed_at": iso_now(),
                "visual_status": str(visual_result.get("decision") or "").strip() or "Pass",
                "visual_reason": str(visual_result.get("reason") or "").strip(),
                "visual_reviewed_at": visual_result.get("reviewed_at"),
                "visual_contract_source": str(merged_review_item.get("visual_contract_source") or "").strip(),
            }
        else:
            results[identifier] = {
                **screening.build_positioning_card_record(
                    platform,
                    username=identifier,
                    profile_url=str(merged_review_item.get("profile_url") or "").strip(),
                    positioning_labels=result.get("positioning_labels") or [],
                    fit_recommendation=result.get("fit_recommendation"),
                    fit_summary=result.get("fit_summary"),
                    evidence_signals=result.get("evidence_signals") or [],
                    provider=result.get("provider"),
                    model=result.get("model"),
                    configured_model=result.get("configured_model"),
                    requested_model=result.get("requested_model"),
                    response_model=result.get("response_model"),
                    effective_model=result.get("effective_model"),
                    prompt_source=result.get("prompt_source"),
                    prompt_selection=result.get("prompt_selection") or {},
                    reviewed_at=result.get("reviewed_at"),
                    visual_status=str(visual_result.get("decision") or "").strip() or "Pass",
                    visual_reason=str(visual_result.get("reason") or "").strip(),
                    visual_reviewed_at=visual_result.get("reviewed_at"),
                    visual_contract_source=str(merged_review_item.get("visual_contract_source") or "").strip(),
                    usage=result.get("usage") or {},
                    cover_count=result.get("cover_count"),
                    candidate_cover_count=result.get("candidate_cover_count"),
                    skipped_cover_count=result.get("skipped_cover_count"),
                ),
                "success": True,
                "attempt_count": result.get("attempt_count"),
                "base_url": result.get("base_url"),
                "raw_text": result.get("raw_text"),
            }

        completed += 1
        save_positioning_card_results(platform, results)
        partial_result = build_positioning_card_partial_result(platform, results, targets)
        if progress_callback:
            progress_callback(
                "analyzing",
                f"已完成第 {completed}/{len(targets)} 个账号：{identifier}",
                done=completed,
                total=len(targets),
                partial_result=partial_result,
                current_identifier=identifier,
                max_workers=max_workers,
            )

    def submit_next(executor):
        if cancel_check and cancel_check():
            return False
        try:
            review_item = next(target_iter)
        except StopIteration:
            return False
        identifier = screening.resolve_profile_review_identifier(platform, review_item)
        if progress_callback:
            progress_callback(
                "analyzing",
                f"已提交定位卡分析：{identifier}",
                done=completed,
                total=len(targets),
                current_identifier=identifier,
                max_workers=max_workers,
            )
        future = executor.submit(
            evaluate_profile_positioning_card_analysis,
            platform,
            review_item,
            requested_provider,
        )
        future_map[future] = {
            "identifier": identifier,
            "review_item": review_item,
            "submitted_at": time.monotonic(),
        }
        return True

    executor = DaemonThreadPoolExecutor(max_workers=max_workers, thread_name_prefix=f"{platform}-positioning")
    try:
        for _ in range(max_workers):
            if not submit_next(executor):
                break

        while future_map:
            if cancel_check and cancel_check():
                return build_cancelled_result()

            now = time.monotonic()
            expired_futures = []
            for future, metadata in list(future_map.items()):
                if future.done():
                    continue
                submitted_at = float(metadata.get("submitted_at") or 0.0)
                if submitted_at and (now - submitted_at) >= item_timeout_seconds:
                    expired_futures.append((future, metadata))
            if expired_futures:
                for future, metadata in expired_futures:
                    identifier = metadata.get("identifier") or ""
                    review_item = metadata.get("review_item") or {}
                    future_map.pop(future, None)
                    future.cancel()
                    finalize_analysis(
                        identifier,
                        review_item,
                        error_text=f"定位卡分析超时：{identifier} 超过 {int(item_timeout_seconds)} 秒未完成",
                    )
                    submit_next(executor)
                continue

            completed_futures, _ = wait(
                tuple(future_map.keys()),
                timeout=future_poll_timeout_seconds,
                return_when=FIRST_COMPLETED,
            )
            if not completed_futures:
                continue

            for future in completed_futures:
                metadata = future_map.pop(future)
                identifier = metadata.get("identifier") or ""
                review_item = metadata.get("review_item") or {}
                try:
                    result = future.result()
                    finalize_analysis(identifier, review_item, result=result)
                except Exception as exc:
                    finalize_analysis(identifier, review_item, error_text=str(exc))

                submit_next(executor)
    finally:
        executor.shutdown(wait=False, cancel_futures=True)

    final_result = build_positioning_card_partial_result(platform, results, targets)
    final_result.update({
        "success": True,
        "message": (
            f"{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 定位卡分析完成，"
            f"共处理 {len(targets)} 个账号。"
        ),
        "positioning_card_results": results,
        "max_workers": max_workers,
        "selected_provider": preflight.get("preferred_provider") or requested_provider,
        "elapsed_seconds": round(time.monotonic() - started_at, 3),
    })
    return final_result


def get_scrape_identifiers(platform, payload):
    if platform == "tiktok":
        return [str(item) for item in (payload.get("profiles") or []) if str(item).strip()]
    if platform == "instagram":
        return [str(item) for item in (payload.get("usernames") or []) if str(item).strip()]
    if platform == "youtube":
        mode = str(payload.get("mode") or "channel").strip().lower()
        if mode == "search":
            return [str(item) for item in (payload.get("queries") or []) if str(item).strip()]
        return [str(item) for item in (payload.get("urls") or []) if str(item).strip()]
    return []


def derive_identifiers_from_upload_metadata(platform):
    metadata_lookup = load_upload_metadata(platform)
    if platform == "youtube":
        values = []
        for identifier, metadata in metadata_lookup.items():
            if isinstance(metadata, dict) and metadata.get("url"):
                values.append(str(metadata["url"]))
            else:
                values.append(screening.build_canonical_profile_url(platform, identifier))
        return values
    return [screening.build_canonical_profile_url(platform, identifier) for identifier in metadata_lookup.keys()]


def resolve_requested_identifiers(platform, payload):
    requested = get_scrape_identifiers(platform, payload)
    if requested:
        return requested
    return derive_identifiers_from_upload_metadata(platform)


def resolve_scrape_missing_retry_attempts(payload):
    value = (payload or {}).get("missing_retry_attempts")
    if value in (None, ""):
        value = (payload or {}).get("missingRetryAttempts")
    if value in (None, ""):
        value = DEFAULT_SCRAPE_MISSING_RETRY_ATTEMPTS
    try:
        return max(0, int(value))
    except Exception:
        return DEFAULT_SCRAPE_MISSING_RETRY_ATTEMPTS


def build_requested_identifier_lookup(platform, identifiers):
    lookup = {}
    for item in identifiers or []:
        requested = str(item or "").strip()
        if not requested:
            continue
        identifier = screening.extract_platform_identifier(platform, requested)
        if identifier and identifier not in lookup:
            lookup[identifier] = requested
    return lookup


def extract_missing_profile_review_identifiers(platform, profile_reviews):
    identifiers = []
    for item in profile_reviews or []:
        identifier = screening.resolve_profile_review_identifier(platform, item)
        if identifier:
            identifiers.append(identifier)
    return identifiers


def select_missing_retry_identifiers(platform, filtered_result, requested_lookup):
    retry_identifiers = []
    seen = set()
    for identifier in extract_missing_profile_review_identifiers(platform, filtered_result.get("missing_profiles") or []):
        requested = str(requested_lookup.get(identifier) or "").strip()
        if not requested or requested in seen:
            continue
        seen.add(requested)
        retry_identifiers.append(requested)
    return retry_identifiers


def _resolve_boolean_payload_flag(payload, *keys, default=False):
    for key in keys:
        if key not in (payload or {}):
            continue
        value = (payload or {}).get(key)
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            continue
        if isinstance(value, (int, float)):
            return bool(value)
        normalized = str(value).strip().lower()
        if normalized in {"1", "true", "yes", "on"}:
            return True
        if normalized in {"0", "false", "no", "off"}:
            return False
        return bool(normalized)
    return bool(default)


def build_actor_input(platform, batch, payload):
    if platform == "tiktok":
        return {
            "profiles": batch,
            "resultsPerPage": int(payload.get("limit", 20)),
            "excludePinnedPosts": _resolve_boolean_payload_flag(
                payload,
                "excludePinnedPosts",
                "exclude_pinned_posts",
                default=True,
            ),
            "shouldDownloadVideos": bool(payload.get("downloadVideos", False)),
            "shouldDownloadCovers": bool(payload.get("downloadCovers", True)),
            "shouldDownloadAvatars": bool(payload.get("downloadAvatars", False)),
            "shouldDownloadSlideshowImages": bool(payload.get("downloadSlideshow", False) or payload.get("downloadCovers", True)),
        }
    if platform == "instagram":
        return {
            "usernames": batch,
            "includeAboutSection": bool(payload.get("includeAbout", True)),
        }
    mode = str(payload.get("mode") or "channel").strip().lower()
    input_payload = {
        "maxResults": int(payload.get("limit", 10)),
        "maxResultsShorts": 0,
        "maxResultStreams": 0,
        "subtitlesLanguage": "en",
        "downloadSubtitles": bool(payload.get("downloadSubtitles", False)),
        "hasCC": bool(payload.get("hasCC", False)),
    }
    if mode == "search":
        input_payload["searchQueries"] = batch
    else:
        input_payload["startUrls"] = [{"url": item} for item in batch]
    return input_payload


def apify_request(method, url, *, token, params=None, json_payload=None):
    def _curl_get_fallback(prepared_url):
        curl_result = subprocess.run(
            [
                "curl",
                "--silent",
                "--show-error",
                "--location",
                "--max-time",
                str(APIFY_REQUEST_TIMEOUT),
                "--request",
                "GET",
                prepared_url,
                "--write-out",
                "\n__CODEX_HTTP_STATUS__:%{http_code}\n",
            ],
            capture_output=True,
            text=True,
            timeout=APIFY_REQUEST_TIMEOUT + 5,
            check=False,
        )
        if curl_result.returncode != 0:
            stderr = str(curl_result.stderr or "").strip()
            raise RuntimeError(stderr or f"curl exited with code {curl_result.returncode}")
        marker = "\n__CODEX_HTTP_STATUS__:"
        marker_index = curl_result.stdout.rfind(marker)
        if marker_index < 0:
            raise RuntimeError("curl fallback missing status marker")
        body = curl_result.stdout[:marker_index]
        status_text = (
            curl_result.stdout[marker_index + len(marker):].strip().splitlines()[0].strip()
        )
        status_code = int(status_text or "0")

        class CurlResponse:
            def __init__(self, status_code, text):
                self.status_code = status_code
                self.text = text
                self.headers = {}
                self.content = str(text or "").encode("utf-8")

            def json(self):
                return json.loads(self.text or "null")

        return CurlResponse(status_code=status_code, text=body)

    request_params = dict(params or {})
    request_params["token"] = token
    normalized_method = str(method or "GET").strip().upper()
    max_attempts = APIFY_TRANSPORT_MAX_RETRIES if normalized_method == "GET" else 1
    last_error = None
    prepared = requests.Request(normalized_method, url, params=request_params).prepare()
    for attempt in range(1, max_attempts + 1):
        try:
            return requests.request(
                normalized_method,
                url,
                params=request_params,
                json=json_payload,
                timeout=APIFY_REQUEST_TIMEOUT,
            )
        except requests.exceptions.RequestException as exc:
            last_error = exc
            if normalized_method == "GET":
                try:
                    return _curl_get_fallback(prepared.url)
                except Exception as curl_exc:
                    last_error = RuntimeError(
                        f"{exc}; curl fallback failed: {curl_exc}"
                    )
            if attempt >= max_attempts:
                break
            time.sleep(APIFY_TRANSPORT_RETRY_BACKOFF_SECONDS * attempt)
    if last_error is not None:
        raise last_error
    raise RuntimeError("Apify request failed without a captured exception")


def normalize_upstream_error_text(text, *, content_type="", max_length=180):
    normalized = re.sub(r"\s+", " ", redact_url_like_text(text)).strip()
    if not normalized:
        return ""
    lowered = normalized.lower()
    normalized_content_type = str(content_type or "").strip().lower()
    if (
        "text/html" in normalized_content_type
        or "<html" in lowered
        or "<!doctype html" in lowered
        or "<body" in lowered
        or "<head" in lowered
    ):
        title_match = re.search(r"<h1[^>]*>\s*([^<]+?)\s*</h1>", str(text or ""), re.IGNORECASE)
        title = re.sub(r"\s+", " ", str(title_match.group(1) or "").strip()) if title_match else ""
        if title:
            return f"上游返回 HTML 错误页（{title}）"
        return "上游返回 HTML 错误页"
    if len(normalized) > max_length:
        return normalized[: max_length - 1] + "…"
    return normalized


def build_upstream_http_status_summary(status_code):
    try:
        resolved = int(status_code or 0)
    except Exception:
        resolved = 0
    return UPSTREAM_HTTP_STATUS_PUBLIC_LABELS.get(resolved, "")


def extract_apify_response_error(response):
    content_type = ""
    if getattr(response, "headers", None):
        content_type = (
            response.headers.get("Content-Type")
            or response.headers.get("content-type")
            or ""
        )
    try:
        payload = response.json()
    except Exception:
        return normalize_upstream_error_text(
            getattr(response, "text", ""),
            content_type=content_type,
        ) or build_upstream_http_status_summary(getattr(response, "status_code", 0))
    if isinstance(payload, dict):
        error = payload.get("error") or {}
        if isinstance(error, dict):
            return normalize_upstream_error_text(
                error.get("message") or json.dumps(error, ensure_ascii=False),
                content_type=content_type,
            )
        return normalize_upstream_error_text(str(error or payload), content_type=content_type)
    return normalize_upstream_error_text(str(payload), content_type=content_type)


def build_vision_provider_http_error_message(response):
    status_code = int(getattr(response, "status_code", 0) or 0)
    status_summary = build_upstream_http_status_summary(status_code)
    detail = extract_apify_response_error(response)
    if detail:
        if status_summary and detail != status_summary:
            return f"HTTP {status_code} {status_summary}；{detail}"
        return f"HTTP {status_code} {detail}"
    if status_summary:
        return f"HTTP {status_code} {status_summary}"
    return f"HTTP {status_code}"


def sanitize_vision_provider_exception_message(exc):
    return normalize_upstream_error_text(str(exc)) or "视觉模型请求失败"


def build_apify_batch_context(
    *,
    actor_id,
    run_id="",
    dataset_id="",
    token="",
    selected_snapshot=None,
    estimated_batch_cost_usd=None,
    required_budget_usd=None,
    reused_guard=False,
    guard_key="",
    checked_snapshots=None,
    insufficient_snapshots=None,
    query_errors=None,
):
    return {
        "actor_id": actor_id,
        "apify_run_id": str(run_id or "").strip(),
        "apify_dataset_id": str(dataset_id or "").strip(),
        "token_masked": (
            selected_snapshot.get("masked")
            if isinstance(selected_snapshot, dict) and selected_snapshot.get("masked")
            else mask_apify_token(token)
        ),
        "estimated_batch_cost_usd": estimated_batch_cost_usd,
        "required_budget_usd": required_budget_usd,
        "remaining_monthly_usage_usd": (
            selected_snapshot.get("remaining_monthly_usage_usd")
            if isinstance(selected_snapshot, dict)
            else None
        ),
        "reused_guard": bool(reused_guard),
        "guard_key": str(guard_key or "").strip(),
        "budget_query_errors": list(query_errors or []),
        "insufficient_budget_tokens": list(insufficient_snapshots or []),
        "checked_budget_tokens": [
            {
                "token_masked": item.get("masked"),
                "remaining_monthly_usage_usd": item.get("remaining_monthly_usage_usd"),
                "checked_at": item.get("checked_at"),
            }
            for item in (checked_snapshots or [])
            if isinstance(item, dict)
        ],
    }


def start_apify_run(actor_id, input_data, token):
    actor_ref = actor_id.replace("/", "~")
    url = f"{APIFY_API_BASE}/acts/{actor_ref}/runs"
    try:
        response = apify_request("POST", url, token=token, json_payload=input_data)
    except requests.exceptions.RequestException as exc:
        raise ApifyStartError(
            503,
            f"启动 Apify 任务失败：{exc}",
        ) from exc
    if response.status_code not in (200, 201):
        raise ApifyStartError(
            response.status_code,
            f"启动 Apify 任务失败：{extract_apify_response_error(response)}",
        )
    payload = response.json() or {}
    data = payload.get("data") or {}
    run_id = data.get("id")
    dataset_id = data.get("defaultDatasetId")
    if not run_id or not dataset_id:
        raise RuntimeError("Apify API 未返回 run_id 或 dataset_id")
    return data


def start_apify_run_guarded(actor_id, input_data, token):
    guard_key = build_apify_guard_key(actor_id, input_data)
    existing_guard = get_apify_run_guard(guard_key)
    if isinstance(existing_guard, dict):
        existing_run_id = str(existing_guard.get("run_id") or "").strip()
        existing_dataset_id = str(existing_guard.get("dataset_id") or "").strip()
        if existing_run_id and existing_dataset_id:
            return {
                "run_data": {
                    "id": existing_run_id,
                    "defaultDatasetId": existing_dataset_id,
                    "status": existing_guard.get("status") or "RUNNING",
                },
                "guard_key": guard_key,
                "reused_guard": True,
            }
        raise RuntimeError(
            "检测到同一批次的本地提交保护记录，但尚未拿到可恢复的 run_id。"
            "为避免重复扣费，本次不会重复提交；如确认上次未成功提交，请手动清理 apify_run_guards.json 后重试。"
        )

    remember_apify_run_guard(
        guard_key,
        build_apify_run_guard_record(
            actor_id,
            input_data,
            token,
            status="submitting",
            request_key=guard_key,
        ),
    )

    try:
        run_data = start_apify_run(actor_id, input_data, token)
    except ApifyStartError as exc:
        if exc.retryable_with_next_token:
            clear_apify_run_guard(guard_key)
        else:
            remember_apify_run_guard(
                guard_key,
                build_apify_run_guard_record(
                    actor_id,
                    input_data,
                    token,
                    status="start_failed",
                    request_key=guard_key,
                    status_code=exc.status_code,
                    retryable_with_next_token=exc.retryable_with_next_token,
                    uncertain_submission=exc.uncertain_submission,
                    error=str(exc),
                ),
            )
        raise
    except Exception as exc:
        remember_apify_run_guard(
            guard_key,
            build_apify_run_guard_record(
                actor_id,
                input_data,
                token,
                status="start_failed",
                request_key=guard_key,
                error=str(exc),
            ),
        )
        raise RuntimeError(
            f"启动 Apify run 失败，已保留本地 guard 防止重复提交：{exc}"
        ) from exc

    remember_apify_run_guard(
        guard_key,
        build_apify_run_guard_record(
            actor_id,
            input_data,
            token,
            status=run_data.get("status") or "RUNNING",
            request_key=guard_key,
            run_id=run_data.get("id"),
            dataset_id=run_data.get("defaultDatasetId"),
        ),
    )
    return {
        "run_data": run_data,
        "guard_key": guard_key,
        "reused_guard": False,
    }


def poll_apify_run(token, run_id, cancel_check=None):
    url = f"{APIFY_API_BASE}/actor-runs/{run_id}"
    while True:
        if cancel_check and cancel_check():
            return {"cancelled": True}
        response = apify_request("GET", url, token=token)
        if response.status_code != 200:
            raise RuntimeError(f"查询 Apify run 失败：{extract_apify_response_error(response)}")
        payload = response.json() or {}
        data = payload.get("data") or {}
        status = str(data.get("status") or "").upper()
        if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
            return {"status": status, "run_data": data}
        time.sleep(APIFY_POLL_INTERVAL_SECONDS)


def download_apify_dataset_items(token, dataset_id):
    url = f"{APIFY_API_BASE}/datasets/{dataset_id}/items"
    response = apify_request("GET", url, token=token)
    if response.status_code != 200:
        raise RuntimeError(f"下载 Apify 数据集失败：{extract_apify_response_error(response)}")
    payload = response.json()
    return payload if isinstance(payload, list) else [payload]


def extract_returned_identifiers(platform, items):
    identifiers = set()
    for item in items or []:
        if platform == "tiktok":
            candidate = ((item.get("authorMeta") or {}).get("profileUrl")) or ((item.get("authorMeta") or {}).get("name"))
        elif platform == "instagram":
            candidate = item.get("url") or item.get("username")
        else:
            candidate = item.get("channelUrl") or item.get("channelName") or ((item.get("aboutChannelInfo") or {}).get("channelUrl"))
        identifier = screening.extract_platform_identifier(platform, candidate)
        if identifier:
            identifiers.add(identifier)
    return identifiers


def merge_scrape_items(platform, existing_items, new_items):
    if platform == "instagram":
        merged = {}
        for item in list(existing_items or []) + list(new_items or []):
            identifier = screening.extract_platform_identifier(platform, item.get("url") or item.get("username"))
            if not identifier:
                continue
            merged[identifier] = item
        return list(merged.values())
    return list(existing_items or []) + list(new_items or [])


def run_apify_batch(platform, batch, payload, progress_callback=None, cancel_check=None):
    actor_id = PLATFORM_ACTORS[platform]
    token_candidates = acquire_apify_token_candidates()
    if not token_candidates:
        raise RuntimeError("缺少 Apify 配置：请设置 APIFY_TOKEN 或 APIFY_API_TOKEN")

    input_data = build_actor_input(platform, batch, payload)
    estimated_batch_cost_usd = estimate_apify_batch_cost_usd(platform, batch, payload)
    required_budget_usd = apply_apify_budget_guard_band(estimated_batch_cost_usd)
    attempted_messages = []
    run_data = None
    token = None
    guard_key = None
    selected_snapshot = None
    checked_snapshots = []
    insufficient_snapshots = []
    query_errors = []
    reused_guard = False
    attempted_tokens = set()

    for attempt_index in range(1, len(token_candidates) + 1):
        available_candidates = [candidate for candidate in token_candidates if candidate not in attempted_tokens]
        if not available_candidates:
            break
        token_pick = select_apify_token_for_batch(available_candidates, required_budget_usd)
        checked_snapshots = token_pick.get("checked_snapshots") or checked_snapshots
        insufficient_snapshots = token_pick.get("insufficient_snapshots") or insufficient_snapshots
        query_errors = token_pick.get("query_errors") or query_errors
        candidate = str(token_pick.get("selected_token") or "").strip()
        selected_snapshot = token_pick.get("selected_snapshot")
        if not candidate:
            raise RuntimeError(token_pick.get("error") or "没有可用的 Apify token 可以覆盖当前批次预算。")

        if progress_callback:
            rotation_hint = "，已切换备用 token" if attempt_index > 1 else ""
            progress_callback(
                "provider_start",
                f"正在提交 Apify 任务{rotation_hint}",
                done=1,
                total=4,
                token_attempt=attempt_index,
                token_pool_size=len(token_candidates),
                token_masked=selected_snapshot.get("masked") if isinstance(selected_snapshot, dict) else mask_apify_token(candidate),
                estimated_batch_cost_usd=estimated_batch_cost_usd,
                required_budget_usd=required_budget_usd,
                **build_target_preview(batch),
            )
        try:
            guarded_start = start_apify_run_guarded(actor_id, input_data, candidate)
            run_data = guarded_start.get("run_data") or {}
            guard_key = guarded_start.get("guard_key")
            reused_guard = bool(guarded_start.get("reused_guard"))
            token = candidate
            break
        except ApifyStartError as exc:
            attempted_messages.append(str(exc))
            attempted_tokens.add(candidate)
            if not exc.retryable_with_next_token or attempt_index >= len(token_candidates):
                raise ApifyRuntimeError(
                    "start",
                    str(exc),
                    retryable=exc.retryable_with_next_token,
                    apify=build_apify_batch_context(
                        actor_id=actor_id,
                        token=candidate,
                        selected_snapshot=selected_snapshot,
                        estimated_batch_cost_usd=estimated_batch_cost_usd,
                        required_budget_usd=required_budget_usd,
                        checked_snapshots=checked_snapshots,
                        insufficient_snapshots=insufficient_snapshots,
                        query_errors=query_errors,
                    ),
                ) from exc
            continue

    if not token or not run_data:
        raise ApifyRuntimeError(
            "start",
            "所有 Apify token 启动任务失败"
            + (f"：{attempted_messages[-1]}" if attempted_messages else ""),
            retryable=True,
            apify=build_apify_batch_context(
                actor_id=actor_id,
                estimated_batch_cost_usd=estimated_batch_cost_usd,
                required_budget_usd=required_budget_usd,
                checked_snapshots=checked_snapshots,
                insufficient_snapshots=insufficient_snapshots,
                query_errors=query_errors,
            ),
        )

    run_id = run_data.get("id")
    dataset_id = run_data.get("defaultDatasetId")

    if progress_callback:
        progress_callback(
            "provider_running",
            f"Apify run {'已复用本地 guard 记录并继续等待' if reused_guard else '已创建，等待远端完成'}（run_id={run_id}）",
            done=2,
            total=4,
            apify_run_id=run_id,
            apify_dataset_id=dataset_id,
            token_masked=selected_snapshot.get("masked") if isinstance(selected_snapshot, dict) else mask_apify_token(token),
            estimated_batch_cost_usd=estimated_batch_cost_usd,
            required_budget_usd=required_budget_usd,
            reused_guard=reused_guard,
            **build_target_preview(batch),
        )
    try:
        poll_result = poll_apify_run(token, run_id, cancel_check=cancel_check)
    except Exception as exc:
        raise ApifyRuntimeError(
            "poll",
            f"查询 Apify run 失败：{exc}",
            retryable=True,
            apify=build_apify_batch_context(
                actor_id=actor_id,
                run_id=run_id,
                dataset_id=dataset_id,
                token=token,
                selected_snapshot=selected_snapshot,
                estimated_batch_cost_usd=estimated_batch_cost_usd,
                required_budget_usd=required_budget_usd,
                reused_guard=reused_guard,
                guard_key=guard_key,
                checked_snapshots=checked_snapshots,
                insufficient_snapshots=insufficient_snapshots,
                query_errors=query_errors,
            ),
        ) from exc
    if poll_result.get("cancelled"):
        return build_cancelled_result()

    final_status = poll_result.get("status")
    final_run_data = poll_result.get("run_data") or {}
    if final_status != "SUCCEEDED":
        if guard_key:
            clear_apify_run_guard(guard_key)
        raise ApifyRuntimeError(
            "poll",
            f"Apify run 结束状态异常：{final_status}",
            retryable=False,
            apify=build_apify_batch_context(
                actor_id=actor_id,
                run_id=run_id,
                dataset_id=dataset_id,
                token=token,
                selected_snapshot=selected_snapshot,
                estimated_batch_cost_usd=estimated_batch_cost_usd,
                required_budget_usd=required_budget_usd,
                reused_guard=reused_guard,
                guard_key=guard_key,
                checked_snapshots=checked_snapshots,
                insufficient_snapshots=insufficient_snapshots,
                query_errors=query_errors,
            ),
        )

    if progress_callback:
        progress_callback(
            "downloading",
            "正在下载 Apify 数据集结果",
            done=3,
            total=4,
            apify_run_id=run_id,
            apify_dataset_id=dataset_id,
            token_masked=selected_snapshot.get("masked") if isinstance(selected_snapshot, dict) else mask_apify_token(token),
            **build_target_preview(batch),
        )

    try:
        items = download_apify_dataset_items(token, dataset_id)
    except Exception as exc:
        raise ApifyRuntimeError(
            "download",
            f"下载 Apify 数据集失败：{exc}",
            retryable=True,
            apify=build_apify_batch_context(
                actor_id=actor_id,
                run_id=run_id,
                dataset_id=dataset_id,
                token=token,
                selected_snapshot=selected_snapshot,
                estimated_batch_cost_usd=estimated_batch_cost_usd,
                required_budget_usd=required_budget_usd,
                reused_guard=reused_guard,
                guard_key=guard_key,
                checked_snapshots=checked_snapshots,
                insufficient_snapshots=insufficient_snapshots,
                query_errors=query_errors,
            ),
        ) from exc
    if guard_key:
        clear_apify_run_guard(guard_key)
    usage_total_usd = None
    usage = final_run_data.get("usageTotalUsd")
    try:
        if usage is not None:
            usage_total_usd = float(usage)
    except Exception:
        usage_total_usd = None

    return {
        "success": True,
        "raw_items": items,
        "apify": {
            **build_apify_batch_context(
                actor_id=actor_id,
                run_id=run_id,
                dataset_id=dataset_id,
                token=token,
                selected_snapshot=selected_snapshot,
                estimated_batch_cost_usd=estimated_batch_cost_usd,
                required_budget_usd=required_budget_usd,
                reused_guard=reused_guard,
                guard_key=guard_key,
                checked_snapshots=checked_snapshots,
                insufficient_snapshots=insufficient_snapshots,
                query_errors=query_errors,
            ),
            "usage_total_usd": usage_total_usd,
            "status": final_status,
        },
    }


def build_partial_scrape_result(platform, raw_items, expected_identifiers):
    filtered = screening.filter_scraped_items(
        platform,
        raw_items,
        expected_profiles=expected_identifiers,
        upload_metadata_lookup=load_upload_metadata(platform),
        active_rulespec=load_active_rulespec(),
    )
    save_profile_reviews(platform, filtered.get("profile_reviews") or [])
    return {
        "platform": platform,
        "raw_count": len(raw_items or []),
        "profile_reviews": filtered.get("profile_reviews") or [],
        "successful_identifiers": filtered.get("successful_identifiers") or [],
    }


def perform_scrape(platform, payload, progress_callback=None, cancel_check=None):
    identifiers = resolve_requested_identifiers(platform, payload)
    if not identifiers:
        return {"success": False, "error": "没有可抓取的账号，请先上传名单或在 payload 中传 identifiers"}

    aggregated_items = []
    apify_runs = []
    requested_lookup = build_requested_identifier_lookup(platform, identifiers)
    creator_cache_enabled = creator_cache.creator_cache_enabled(payload)
    force_refresh_creator_cache = creator_cache.creator_cache_force_refresh(payload)
    creator_cache_db_path = creator_cache.resolve_creator_cache_db_path(payload)
    cached_scrape_entries: dict[str, list[dict[str, Any]]] = {}
    cache_hit_requested_identifiers: list[str] = []
    identifiers_to_fetch: list[str] = []
    for requested in identifiers:
        identifier = screening.extract_platform_identifier(platform, requested)
        if identifier and creator_cache_enabled and not force_refresh_creator_cache:
            if not cached_scrape_entries:
                cached_scrape_entries = creator_cache.load_scrape_cache_entries(
                    platform,
                    identifiers,
                    creator_cache_db_path,
                )
            cached_items = cached_scrape_entries.get(identifier) or []
            if cached_items:
                aggregated_items = merge_scrape_items(platform, aggregated_items, cached_items)
                cache_hit_requested_identifiers.append(requested)
                continue
        identifiers_to_fetch.append(requested)

    batches = chunk_list(identifiers_to_fetch, PLATFORM_BATCH_SIZES.get(platform, 20))
    max_missing_retry_attempts = resolve_scrape_missing_retry_attempts(payload)
    retry_history = []
    batch_failures = []
    completed_requested_identifiers = list(cache_hit_requested_identifiers)
    creator_cache_summary = {
        "enabled": bool(creator_cache_enabled),
        "db_path": str(creator_cache_db_path),
        "force_refresh": bool(force_refresh_creator_cache),
        "scrape_hit_count": len(cache_hit_requested_identifiers),
        "scrape_miss_count": len(identifiers_to_fetch),
        "scrape_hit_preview": cache_hit_requested_identifiers[:10],
        "visual_hit_count": 0,
    }

    if progress_callback:
        progress_callback(
            "preparing",
            "正在准备抓取任务",
            done=0,
            total=len(batches),
            creator_cache=creator_cache_summary,
            **build_target_preview(identifiers),
        )

    for index, batch in enumerate(batches, start=1):
        if cancel_check and cancel_check():
            return build_cancelled_result()
        try:
            batch_result = run_apify_batch(
                platform,
                batch,
                payload,
                progress_callback=progress_callback,
                cancel_check=cancel_check,
            )
        except Exception as exc:
            failure_record = {
                "batch_index": index,
                "batch_total": len(batches),
                "targets": list(batch),
                "target_preview": list(batch)[:10],
                "error": str(exc),
                "exception_type": exc.__class__.__name__,
            }
            if isinstance(exc, ApifyRuntimeError):
                failure_record["failure_stage"] = exc.failure_stage
                failure_record["retryable"] = bool(exc.retryable)
                if exc.apify:
                    failure_record["apify"] = dict(exc.apify)
            batch_failures.append(failure_record)
            partial_result = build_partial_scrape_result(
                platform,
                aggregated_items,
                completed_requested_identifiers,
            )
            if progress_callback:
                progress_callback(
                    "batch_failed",
                    f"第 {index}/{len(batches)} 批抓取失败：{exc}",
                    done=index,
                    total=len(batches),
                    partial_result=partial_result,
                    batch_index=index,
                    batch_total=len(batches),
                    error=str(exc),
                    **build_target_preview(batch),
                )
            continue
        if batch_result.get("cancelled"):
            return batch_result
        aggregated_items = merge_scrape_items(platform, aggregated_items, batch_result.get("raw_items") or [])
        apify_runs.append(batch_result.get("apify") or {})
        write_json_file(get_raw_data_path(platform), aggregated_items)
        completed_requested_identifiers.extend(batch)
        partial_result = build_partial_scrape_result(platform, aggregated_items, completed_requested_identifiers)
        if progress_callback:
            progress_callback(
                "batch_completed",
                f"第 {index}/{len(batches)} 批完成",
                done=index,
                total=len(batches),
                partial_result=partial_result,
                batch_index=index,
                batch_total=len(batches),
                **build_target_preview(batch),
            )

    filtered = screening.filter_scraped_items(
        platform,
        aggregated_items,
        expected_profiles=identifiers,
        upload_metadata_lookup=load_upload_metadata(platform),
        active_rulespec=load_active_rulespec(),
    )
    for retry_attempt in range(1, max_missing_retry_attempts + 1):
        retry_identifiers = select_missing_retry_identifiers(platform, filtered, requested_lookup)
        if not retry_identifiers:
            break
        retry_batch_size = resolve_missing_retry_batch_size(platform)
        retry_batches = chunk_list(retry_identifiers, retry_batch_size)
        previous_missing_identifiers = set(
            extract_missing_profile_review_identifiers(platform, filtered.get("missing_profiles") or [])
        )
        retry_record = {
            "attempt": retry_attempt,
            "batch_size": retry_batch_size,
            "requested_identifier_count": len(retry_identifiers),
            "requested_identifiers": retry_identifiers,
            "requested_identifier_preview": retry_identifiers[:10],
            "batch_count": len(retry_batches),
            "status": "completed",
            "errors": [],
        }
        if progress_callback:
            progress_callback(
                "retrying_missing_profiles",
                f"检测到 {len(retry_identifiers)} 个 Missing 账号，正在自动补抓（{retry_attempt}/{max_missing_retry_attempts}）",
                retry_attempt=retry_attempt,
                retry_total=max_missing_retry_attempts,
                missing_identifier_count=len(retry_identifiers),
                **build_target_preview(retry_identifiers),
            )
        for batch_index, batch in enumerate(retry_batches, start=1):
            if cancel_check and cancel_check():
                return build_cancelled_result()
            try:
                batch_result = run_apify_batch(
                    platform,
                    batch,
                    payload,
                    progress_callback=progress_callback,
                    cancel_check=cancel_check,
                )
            except Exception as exc:
                retry_record["status"] = "completed_with_errors"
                retry_record["errors"].append({
                    "batch_index": batch_index,
                    "targets": list(batch),
                    "error": str(exc),
                })
                if progress_callback:
                    progress_callback(
                        "retry_batch_failed",
                        f"Missing 自动补抓第 {batch_index}/{len(retry_batches)} 批失败：{exc}",
                        retry_attempt=retry_attempt,
                        batch_index=batch_index,
                        batch_total=len(retry_batches),
                        error=str(exc),
                        **build_target_preview(batch),
                    )
                continue
            if batch_result.get("cancelled"):
                return batch_result
            aggregated_items = merge_scrape_items(platform, aggregated_items, batch_result.get("raw_items") or [])
            apify_runs.append(batch_result.get("apify") or {})
            write_json_file(get_raw_data_path(platform), aggregated_items)
            partial_result = build_partial_scrape_result(platform, aggregated_items, identifiers)
            if progress_callback:
                progress_callback(
                    "retry_batch_completed",
                    f"Missing 自动补抓第 {batch_index}/{len(retry_batches)} 批完成",
                    partial_result=partial_result,
                    retry_attempt=retry_attempt,
                    batch_index=batch_index,
                    batch_total=len(retry_batches),
                    **build_target_preview(batch),
                )
        filtered = screening.filter_scraped_items(
            platform,
            aggregated_items,
            expected_profiles=identifiers,
            upload_metadata_lookup=load_upload_metadata(platform),
            active_rulespec=load_active_rulespec(),
        )
        remaining_missing_identifiers = set(
            extract_missing_profile_review_identifiers(platform, filtered.get("missing_profiles") or [])
        )
        retry_record["recovered_identifier_count"] = len(previous_missing_identifiers - remaining_missing_identifiers)
        retry_record["remaining_missing_count"] = len(remaining_missing_identifiers)
        retry_history.append(retry_record)
        if not remaining_missing_identifiers:
            break
    save_profile_reviews(platform, filtered.get("profile_reviews") or [])
    if creator_cache_enabled:
        creator_cache_summary["persisted_scrape_entry_count"] = creator_cache.persist_scrape_cache_entries(
            platform,
            aggregated_items,
            creator_cache_db_path,
            updated_at=iso_now(),
        )
    remaining_missing_count = len(filtered.get("missing_profiles") or [])
    retried_identifier_count = len({
        item
        for record in retry_history
        for item in (record.get("requested_identifiers") or [])
        if str(item or "").strip()
    })
    if retry_history and remaining_missing_count:
        message = (
            f"{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 抓取完成，"
            f"返回 {len(filtered.get('successful_identifiers') or [])}/{len(identifiers)} 个账号；"
            f"Missing 自动补抓 {len(retry_history)} 次后仍缺失 {remaining_missing_count} 个账号，已要求下游阻断最终导出。"
        )
    elif retry_history:
        message = (
            f"{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 抓取完成，"
            f"返回 {len(filtered.get('successful_identifiers') or [])}/{len(identifiers)} 个账号；"
            f"Missing 自动补抓 {len(retry_history)} 次后已补齐。"
        )
    else:
        message = (
            f"{UPLOAD_PLATFORM_RESPONSE_LABELS.get(platform, platform)} 抓取完成，"
            f"返回 {len(filtered.get('successful_identifiers') or [])}/{len(identifiers)} 个账号。"
        )
    result = {
        "success": True,
        "platform": platform,
        "requested_total": len(identifiers),
        "raw_count": len(aggregated_items),
        "profile_reviews": filtered.get("profile_reviews") or [],
        "successful_identifiers": filtered.get("successful_identifiers") or [],
        "profile_reviews_path": get_profile_reviews_path(platform),
        "raw_data_path": get_raw_data_path(platform),
        "apify": {
            "runs": apify_runs,
            "execution_method": "rest-batched",
            "actor_id": PLATFORM_ACTORS[platform],
            "usage_total_usd": round(
                sum(item.get("usage_total_usd") or 0 for item in apify_runs if isinstance(item, dict)),
                6,
            ) if apify_runs else None,
        },
        "retry_summary": {
            "enabled": bool(max_missing_retry_attempts > 0),
            "initial_batch_failure_count": len(batch_failures),
            "initial_batch_failures": batch_failures,
            "max_attempts": max_missing_retry_attempts,
            "attempt_count": len(retry_history),
            "retried_identifier_count": retried_identifier_count,
            "remaining_missing_count": remaining_missing_count,
            "remaining_missing_identifiers": extract_missing_profile_review_identifiers(
                platform,
                filtered.get("missing_profiles") or [],
            ),
            "history": retry_history,
        },
        "creator_cache": creator_cache_summary,
        "message": message,
    }
    return result


def merge_upload_metadata_into_review_item(platform, item):
    metadata_lookup = load_upload_metadata(platform)
    normalized = dict(item or {})
    identifier = screening.resolve_profile_review_identifier(platform, normalized)
    canonical = metadata_lookup.get(identifier) if identifier else None
    if isinstance(canonical, dict):
        normalized["upload_metadata"] = dict(canonical)
    else:
        normalized["upload_metadata"] = dict(normalized.get("upload_metadata") or {})
    if not normalized.get("profile_url") and normalized["upload_metadata"].get("url"):
        normalized["profile_url"] = normalized["upload_metadata"]["url"]
    if not normalized.get("username") and normalized["upload_metadata"].get("handle"):
        normalized["username"] = normalized["upload_metadata"]["handle"]
    return normalized


def merge_upload_metadata_into_reviews(platform, profile_reviews):
    return [merge_upload_metadata_into_review_item(platform, item) for item in (profile_reviews or []) if isinstance(item, dict)]


def list_missing_profile_reviews(platform, profile_reviews):
    missing_profiles = []
    for item in merge_upload_metadata_into_reviews(platform, profile_reviews):
        if str(item.get("status") or "").strip() != "Missing":
            continue
        identifier = screening.resolve_profile_review_identifier(platform, item)
        missing_profiles.append({
            "identifier": identifier,
            "username": str(item.get("username") or "").strip(),
            "profile_url": str(item.get("profile_url") or "").strip(),
            "reason": str(item.get("reason") or "").strip(),
        })
    return missing_profiles


def build_final_review_missing_block_payload(platform, profile_reviews):
    missing_profiles = list_missing_profile_reviews(platform, profile_reviews)
    return {
        "error": "存在名单账号未在本次抓取结果中返回，已阻止导出 final review。请先补抓或重新筛号。",
        "error_code": "FINAL_REVIEW_BLOCKED_BY_MISSING_PROFILES",
        "platform": platform,
        "missing_profile_count": len(missing_profiles),
        "missing_profiles": missing_profiles,
    }


def append_upload_metadata_to_export_row(row, review_item):
    metadata = dict(review_item.get("upload_metadata") or {})
    for export_key, metadata_key in UPLOAD_METADATA_EXPORT_FIELDS:
        row[export_key] = metadata.get(metadata_key, "")
    return row


def append_runtime_stats_to_export_row(row, review_item):
    stats = dict(review_item.get("stats") or {})
    row["runtime_avg_views"] = stats.get("avg_views", "")
    row["runtime_median_views"] = stats.get("median_views", "")
    row["runtime_video_count"] = stats.get("video_count", "")
    return row


def format_export_review_status(status):
    mapping = {
        "Pass": "Pass",
        "Reject": "Reject",
        "Missing": "Missing",
        "Not Reviewed": "Not Reviewed",
        "Error": "Error",
    }
    return mapping.get(str(status or "").strip(), str(status or "").strip())


def build_export_row_base(platform, review_item):
    merged_item = merge_upload_metadata_into_review_item(platform, review_item)
    identifier = screening.resolve_profile_review_identifier(platform, merged_item)
    row = {
        "identifier": identifier,
        "username": merged_item.get("username", ""),
        "profile_url": merged_item.get("profile_url", ""),
    }
    append_upload_metadata_to_export_row(row, merged_item)
    append_runtime_stats_to_export_row(row, merged_item)
    return row, merged_item


def build_prescreen_review_rows(platform, profile_reviews):
    rows = []
    for item in merge_upload_metadata_into_reviews(platform, profile_reviews):
        row, review_item = build_export_row_base(platform, item)
        covers = review_item.get("covers") or []
        row.update({
            "status": format_export_review_status(review_item.get("status")),
            "stage_status": format_export_review_status(review_item.get("status")),
            "reason": review_item.get("reason", ""),
            "latest_post_time": review_item.get("latest_post_time"),
            "soft_flags": "；".join(str(flag) for flag in (review_item.get("soft_flags") or []) if str(flag).strip()),
            "cover_count": len(covers),
        })
        rows.append(row)
    return rows


def build_image_review_rows(platform, profile_reviews):
    rows = []
    for item in merge_upload_metadata_into_reviews(platform, profile_reviews):
        row, review_item = build_export_row_base(platform, item)
        covers = review_item.get("covers") or []
        row.update({
            "status": format_export_review_status(review_item.get("status")),
            "stage_status": format_export_review_status(review_item.get("status")),
            "stage_reason": review_item.get("reason", ""),
            "reason": review_item.get("reason", ""),
            "latest_post_time": review_item.get("latest_post_time"),
            "soft_flags": "；".join(str(flag) for flag in (review_item.get("soft_flags") or []) if str(flag).strip()),
            "cover_count": len(covers),
        })
        for index in range(9):
            row[f"cover_{index + 1}"] = covers[index] if index < len(covers) else ""
        rows.append(row)
    return rows


def merge_visual_results(platform, requested_visual_results):
    saved = load_visual_results(platform)
    merged = {}
    for source in (saved, requested_visual_results or {}):
        if not isinstance(source, dict):
            continue
        for key, value in source.items():
            if isinstance(value, dict):
                merged[key] = value
    return merged


def merge_positioning_card_results(platform, requested_positioning_results):
    saved = load_positioning_card_results(platform)
    merged = {}
    for source in (saved, requested_positioning_results or {}):
        if not isinstance(source, dict):
            continue
        for key, value in source.items():
            if isinstance(value, dict):
                merged[key] = value
    return merged


def build_final_review_rows(platform, profile_reviews, visual_results):
    visual_lookup = {}
    for key, item in (visual_results or {}).items():
        identifier = screening.normalize_identifier(key or (item or {}).get("username"))
        if identifier and isinstance(item, dict):
            visual_lookup[identifier] = item

    rows = []
    for item in merge_upload_metadata_into_reviews(platform, profile_reviews):
        row, review_item = build_export_row_base(platform, item)
        identifier = screening.resolve_profile_review_identifier(platform, review_item)
        prescreen_status = str(review_item.get("status") or "").strip()
        prescreen_reason = str(review_item.get("reason") or "").strip()
        visual = visual_lookup.get(identifier) or {}
        visual_status = "Not Reviewed"
        visual_reason = ""
        visual_signals = ""
        final_status = prescreen_status
        final_reason = prescreen_reason
        if prescreen_status == "Pass" and visual:
            if visual.get("success") is False:
                visual_status = "Error"
                visual_reason = str(visual.get("error") or "").strip()
                final_status = "Error"
                final_reason = visual_reason or prescreen_reason
            else:
                visual_status = str(visual.get("decision") or "Pass").strip() or "Pass"
                visual_reason = str(visual.get("reason") or "").strip()
                visual_signals = "；".join(str(item) for item in (visual.get("signals") or []) if str(item).strip())
                final_status = visual_status
                final_reason = visual_reason or prescreen_reason
        row.update({
            "prescreen_status": format_export_review_status(prescreen_status),
            "prescreen_reason": prescreen_reason,
            "visual_status": format_export_review_status(visual_status),
            "visual_reason": visual_reason,
            "visual_signals": visual_signals,
            "status": format_export_review_status(final_status),
            "reason": final_reason,
            "final_status": format_export_review_status(final_status),
            "final_reason": final_reason,
        })
        rows.append(row)
    return rows


def build_positioning_card_rows(platform, profile_reviews, visual_results, positioning_results):
    visual_lookup = {}
    for key, item in (visual_results or {}).items():
        identifier = screening.normalize_identifier(key or (item or {}).get("username"))
        if identifier and isinstance(item, dict):
            visual_lookup[identifier] = item

    positioning_lookup = {}
    for key, item in (positioning_results or {}).items():
        identifier = screening.normalize_identifier(key or (item or {}).get("username"))
        if identifier and isinstance(item, dict):
            positioning_lookup[identifier] = item

    rows = []
    for item in merge_upload_metadata_into_reviews(platform, profile_reviews):
        row, review_item = build_export_row_base(platform, item)
        identifier = screening.resolve_profile_review_identifier(platform, review_item)
        visual = visual_lookup.get(identifier) or {}
        positioning = positioning_lookup.get(identifier) or {}
        row.update({
            "prescreen_status": format_export_review_status(review_item.get("status")),
            "visual_status": format_export_review_status(visual.get("decision") or visual.get("visual_status") or "Not Reviewed"),
            "visual_reason": str(visual.get("reason") or visual.get("visual_reason") or "").strip(),
            "positioning_stage_status": (
                "Error"
                if positioning.get("success") is False
                else ("Completed" if positioning else "Not Reviewed")
            ),
            "positioning_labels": "；".join(
                str(label) for label in (positioning.get("positioning_labels") or []) if str(label).strip()
            ),
            "fit_recommendation": str(positioning.get("fit_recommendation") or "").strip(),
            "fit_summary": str(positioning.get("fit_summary") or "").strip(),
            "evidence_signals": "；".join(
                str(signal) for signal in (positioning.get("evidence_signals") or []) if str(signal).strip()
            ),
            "positioning_error": str(positioning.get("error") or "").strip(),
        })
        rows.append(row)
    return rows


def workbook_bytes_from_sheets(sheet_payloads):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, rows in sheet_payloads:
            pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output


@app.route("/api/health", methods=["GET"])
def health_check():
    ensure_runtime_dirs()
    token_pool = get_apify_token_pool()
    vision_preflight = build_vision_preflight()
    return jsonify({
        "status": "ok",
        "smoke_ready": True,
        "checks": {
            "apify": "configured" if token_pool else "unconfigured",
            "apify_token_pool_size": len(token_pool),
            "vision": vision_preflight.get("status"),
            "vision_providers": vision_preflight.get("runnable_provider_names") or [],
            "vision_preflight": vision_preflight,
            "origins": BACKEND_ALLOWED_ORIGINS,
        },
    })


@app.route("/api/vision/providers/probe", methods=["POST"])
def probe_vision_provider_api():
    payload = request.get_json(silent=True) or {}
    requested_provider = normalize_vision_provider_name(payload.get("provider"))
    preflight = build_vision_preflight(requested_provider)
    providers = get_available_vision_providers(requested_provider)
    if not providers:
        return jsonify(build_vision_preflight_error_payload(requested_provider)), 400
    selected_provider = providers[0]
    try:
        probe_result = probe_vision_provider(selected_provider)
    except Exception as exc:
        return jsonify({
            "success": False,
            "error_code": "VISION_PROVIDER_PROBE_FAILED",
            "error": str(exc),
            "provider": selected_provider["name"],
            "vision_preflight": preflight,
        }), 502
    return jsonify({
        "success": True,
        "provider": selected_provider["name"],
        "probe": probe_result,
        "vision_preflight": preflight,
    })


@app.route("/api/apify/balance", methods=["GET"])
def apify_balance():
    payload, status_code = build_apify_balance_payload()
    return jsonify(payload), status_code


@app.route("/api/apify/balance/refresh", methods=["POST"])
def refresh_apify_balance():
    payload, status_code = build_apify_balance_payload(force_refresh=True)
    return jsonify(payload), status_code


@app.route("/apify/balance", methods=["GET"])
def apify_balance_dashboard():
    payload, status_code = build_apify_balance_payload()
    return render_template(
        "apify_balance_dashboard.html",
        refresh_interval_seconds=APIFY_BALANCE_REFRESH_INTERVAL_SECONDS,
        api_balance_path="/api/apify/balance",
        api_refresh_path="/api/apify/balance/refresh",
        initial_payload=payload,
        initial_status_code=status_code,
    )


@app.route("/operator", methods=["GET"])
def operator_console():
    return render_template(
        "operator_console.html",
        api_tasks_path="/api/operator/tasks",
        api_runs_path="/api/operator/runs",
        api_file_path="/api/operator/file",
        initial_runs=list_operator_runs(),
        default_env_file=".env",
        default_matching_strategy="brand-keyword-fast-path",
        default_vision_provider="reelx",
        default_platforms=["instagram", "tiktok"],
        default_max_identifiers_per_platform=1,
    )


@app.route("/api/operator/tasks", methods=["GET"])
def operator_tasks():
    try:
        payload = load_operator_task_candidates(
            env_file=request.args.get("env_file", ""),
            task_upload_url=request.args.get("task_upload_url", ""),
            employee_info_url=request.args.get("employee_info_url", ""),
        )
    except Exception as exc:  # noqa: BLE001
        error_payload, status_code = _classify_operator_config_error(exc)
        return jsonify(error_payload), status_code
    return jsonify(payload)


@app.route("/api/operator/runs", methods=["GET", "POST"])
def operator_runs_collection():
    if request.method == "GET":
        return jsonify({"success": True, "runs": list_operator_runs()})
    payload = request.get_json(silent=True) or {}
    try:
        run_payload = launch_operator_run(
            task_name=str(payload.get("task_name") or "").strip(),
            env_file=str(payload.get("env_file") or "").strip(),
            task_upload_url=str(payload.get("task_upload_url") or "").strip(),
            employee_info_url=str(payload.get("employee_info_url") or "").strip(),
            matching_strategy=str(payload.get("matching_strategy") or "brand-keyword-fast-path").strip()
            or "brand-keyword-fast-path",
            brand_keyword=str(payload.get("brand_keyword") or "").strip(),
            brand_match_include_from=_boolish(payload.get("brand_match_include_from"), default=False),
            platforms=_normalize_operator_platforms(payload.get("platforms")),
            vision_provider=str(payload.get("vision_provider") or "").strip(),
            max_identifiers_per_platform=max(0, int(payload.get("max_identifiers_per_platform") or 0)),
            mail_limit=max(0, int(payload.get("mail_limit") or 0)),
            sent_since=str(payload.get("sent_since") or "").strip(),
            reuse_existing=_boolish(payload.get("reuse_existing"), default=True),
            probe_vision_provider_only=_boolish(payload.get("probe_vision_provider_only"), default=False),
            skip_scrape=_boolish(payload.get("skip_scrape"), default=False),
            skip_visual=_boolish(payload.get("skip_visual"), default=False),
            skip_positioning_card_analysis=_boolish(payload.get("skip_positioning_card_analysis"), default=False),
        )
    except Exception as exc:  # noqa: BLE001
        error_payload, status_code = _classify_operator_config_error(exc)
        return jsonify(error_payload), status_code
    return jsonify({"success": True, "run": run_payload}), 202


@app.route("/api/operator/runs/<run_id>", methods=["GET"])
def operator_run_detail(run_id):
    run = refresh_operator_run(run_id)
    if run is None:
        return jsonify({"success": False, "error": "未找到 operator run"}), 404
    return jsonify({"success": True, "run": _serialize_operator_run(run)})


@app.route("/api/operator/file", methods=["GET"])
def operator_file_download():
    raw_path = str(request.args.get("path") or "").strip()
    if not raw_path:
        return jsonify({"success": False, "error": "缺少 path"}), 400
    resolved = Path(raw_path).expanduser().resolve()
    if not resolved.exists():
        return jsonify({"success": False, "error": "文件不存在"}), 404
    if not _path_is_within_root(resolved, BASE_DIR):
        return jsonify({"success": False, "error": "不允许下载工作区外文件"}), 403
    return send_file(resolved, as_attachment=True, download_name=resolved.name)


@app.route("/api/upload", methods=["POST"])
def upload_file():
    ensure_runtime_dirs()
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "No selected file"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)

    try:
        frames = load_canonical_upload_workbook_frames(filepath)
        if not frames:
            return build_upload_validation_error(
                "上传表没有可用数据行",
                ["请确认至少有一个 sheet 包含表头和账号数据。"],
            )
        df = pd.concat(frames, ignore_index=True)
        parsed, error_response = parse_canonical_upload_workbook(df, filename)
        if error_response:
            return error_response
        for platform in PLATFORM_ACTORS:
            save_upload_metadata(platform, parsed["metadata_by_platform"].get(platform, {}), replace=True)
        return jsonify({
            "success": True,
            "filename": f"processed_{filename}",
            "stats": parsed["stats"],
            "preview": parsed["preview"],
            "grouped_data": parsed["grouped_data"],
            "metadata_counts": {
                platform: len(parsed["metadata_by_platform"].get(platform, {}))
                for platform in PLATFORM_ACTORS
            },
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/rulespec/compile", methods=["POST"])
def compile_rulespec_api():
    payload = request.get_json(silent=True) or {}
    sop_text = payload.get("sop_text")
    persist = bool(payload.get("persist", True))
    try:
        compiled = rules_module.compile_rulespec_from_text(sop_text)
        if persist:
            persist_active_rulespec(compiled)
        return jsonify({
            "success": True,
            "rule_spec": compiled.get("rule_spec") or {},
            "field_match_report": compiled.get("field_match_report") or {},
            "missing_capabilities": compiled.get("missing_capabilities") or {},
            "compiled_at": compiled.get("compiled_at"),
            "persisted": persist,
        })
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/jobs/scrape", methods=["POST"])
def start_scrape_job():
    payload = request.get_json(silent=True) or {}
    platform = str(payload.get("platform") or "").strip().lower()
    data = payload.get("payload") or {}
    if platform not in PLATFORM_ACTORS:
        return jsonify({"success": False, "error": "平台参数无效"}), 400
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "payload 必须是对象"}), 400
    if not get_apify_token_pool():
        return jsonify({
            "success": False,
            "error_code": "MISSING_APIFY_CONFIG",
            "error": "缺少 Apify 配置：请设置 APIFY_TOKEN 或 APIFY_API_TOKEN。",
        }), 400

    job = create_job("scrape", platform=platform, message="采集任务已创建")
    start_background_job(job, lambda progress_callback, cancel_check: perform_scrape(platform, data, progress_callback, cancel_check))
    return jsonify({"success": True, "job": get_job(job["id"])})


@app.route("/api/jobs/visual-review", methods=["POST"])
def start_visual_review_job():
    payload = request.get_json(silent=True) or {}
    platform = str(payload.get("platform") or "").strip().lower()
    data = payload.get("payload") or {}
    if platform not in PLATFORM_ACTORS:
        return jsonify({"success": False, "error": "平台参数无效"}), 400
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "payload 必须是对象"}), 400
    requested_provider = normalize_vision_provider_name((data or {}).get("provider"))
    if not get_available_vision_providers(requested_provider):
        return jsonify(build_vision_preflight_error_payload(requested_provider)), 400

    job = create_job("visual-review", platform=platform, message="视觉复核任务已创建")
    start_background_job(job, lambda progress_callback, cancel_check: perform_visual_review(platform, data, progress_callback, cancel_check))
    return jsonify({"success": True, "job": get_job(job["id"])})


@app.route("/api/jobs/positioning-card-analysis", methods=["POST"])
def start_positioning_card_analysis_job():
    payload = request.get_json(silent=True) or {}
    platform = str(payload.get("platform") or "").strip().lower()
    data = payload.get("payload") or {}
    if platform not in PLATFORM_ACTORS:
        return jsonify({"success": False, "error": "平台参数无效"}), 400
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "payload 必须是对象"}), 400
    requested_provider = normalize_vision_provider_name((data or {}).get("provider"))
    if not get_available_vision_providers(requested_provider):
        return jsonify(build_vision_preflight_error_payload(requested_provider)), 400

    job = create_job("positioning-card-analysis", platform=platform, message="定位卡分析任务已创建")
    start_background_job(
        job,
        lambda progress_callback, cancel_check: perform_positioning_card_analysis(
            platform,
            data,
            progress_callback,
            cancel_check,
        ),
    )
    return jsonify({"success": True, "job": get_job(job["id"])})


@app.route("/api/jobs/<job_id>", methods=["GET"])
def get_job_status(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"success": False, "error": "未找到任务"}), 404
    return jsonify({"success": True, "job": job})


@app.route("/api/jobs/<job_id>/cancel", methods=["POST"])
def cancel_job(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({"success": False, "error": "未找到任务"}), 404
        if job["status"] in JOB_TERMINAL_STATUSES:
            return jsonify({"success": False, "error": "任务已结束，无法取消"}), 400
        job["cancel_requested"] = True
        job["status"] = "cancelling"
        job["stage"] = "cancelling"
        job["message"] = "正在取消任务"
        job["updated_at"] = iso_now()
    return jsonify({"success": True, "job": get_job(job_id)})


@app.route("/api/results/<platform>", methods=["GET"])
def get_results(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    return jsonify(load_json_payload(get_raw_data_path(platform), default=[]))


@app.route("/api/artifacts/<platform>/status", methods=["GET"])
def artifact_status(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    profile_reviews = load_profile_reviews(platform)
    visual_results = load_visual_results(platform)
    positioning_card_results = load_positioning_card_results(platform)
    missing_profiles = list_missing_profile_reviews(platform, profile_reviews)
    final_review_export_blocked = bool(missing_profiles)
    return jsonify({
        "platform": platform,
        "raw_data_path": get_raw_data_path(platform),
        "profile_reviews_path": get_profile_reviews_path(platform),
        "visual_results_path": get_visual_results_path(platform),
        "positioning_card_results_path": get_positioning_card_results_path(platform),
        "raw_count": len(load_json_payload(get_raw_data_path(platform), default=[])),
        "profile_review_count": len(profile_reviews),
        "visual_review_count": len(visual_results),
        "positioning_card_result_count": len(positioning_card_results),
        "missing_profile_count": len(missing_profiles),
        "missing_profiles_preview": missing_profiles[:20],
        "final_review_export_blocked": final_review_export_blocked,
        "saved_final_review_artifacts_available": bool(profile_reviews and visual_results and not final_review_export_blocked),
        "saved_positioning_card_artifacts_available": bool(positioning_card_results),
    })


@app.route("/api/download/<platform>/prescreen-review", methods=["GET"])
def download_prescreen_review(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    rows = build_prescreen_review_rows(platform, load_profile_reviews(platform))
    if not rows:
        return jsonify({"error": "Profile review data is empty"}), 400
    output = workbook_bytes_from_sheets([("Prescreen Review", rows)])
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{platform}_prescreen_review.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download/<platform>/image-review", methods=["GET"])
def download_image_review(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    rows = build_image_review_rows(platform, load_profile_reviews(platform))
    if not rows:
        return jsonify({"error": "Profile review data is empty"}), 400
    output = workbook_bytes_from_sheets([("Image Review", rows)])
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{platform}_image_review.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download/<platform>/test-info", methods=["GET"])
def download_test_info(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    profile_reviews = load_profile_reviews(platform)
    if not profile_reviews:
        return jsonify({"error": "No test data available to export"}), 404
    summary_rows = build_prescreen_review_rows(platform, profile_reviews)
    output = workbook_bytes_from_sheets([("Profile Reviews", summary_rows)])
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{platform}_test_info.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download/<platform>/test-info-json", methods=["GET"])
def download_test_info_json(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    payload = {
        "profile_reviews": build_prescreen_review_rows(platform, load_profile_reviews(platform)),
    }
    if not payload["profile_reviews"]:
        return jsonify({"error": "No test data available to export"}), 404
    payload["row_count"] = len(payload["profile_reviews"])
    response = app.response_class(
        response=json.dumps(payload, ensure_ascii=False, indent=2),
        mimetype="application/json",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={platform}_test_info.json"
    return response


@app.route("/api/download/<platform>/positioning-card-json", methods=["GET"])
def download_positioning_card_json(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    positioning_card_results = load_positioning_card_results(platform)
    if not positioning_card_results:
        return jsonify({"error": "No positioning card data available to export"}), 404
    records = []
    for key, value in positioning_card_results.items():
        if not isinstance(value, dict):
            continue
        identifier = screening.normalize_identifier(key or value.get("username"))
        if not identifier:
            continue
        records.append({
            "identifier": identifier,
            **sanitize_json_compatible(value),
        })
    payload = {
        "platform": platform,
        "row_count": len(records),
        "positioning_card_results": sorted(records, key=lambda item: str(item.get("identifier") or "")),
    }
    response = app.response_class(
        response=json.dumps(payload, ensure_ascii=False, indent=2),
        mimetype="application/json",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={platform}_positioning_card_results.json"
    return response


@app.route("/api/download/<platform>/positioning-card-review", methods=["GET"])
def download_positioning_card_review(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    profile_reviews = load_profile_reviews(platform)
    visual_results = load_visual_results(platform)
    positioning_card_results = load_positioning_card_results(platform)
    rows = build_positioning_card_rows(platform, profile_reviews, visual_results, positioning_card_results)
    if not rows:
        return jsonify({"error": "No positioning card rows available to export"}), 400
    output = workbook_bytes_from_sheets([("Positioning Card Review", rows)])
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{platform}_positioning_card_review.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download/<platform>/final-review", methods=["POST"])
def download_final_review(platform):
    if platform not in PLATFORM_ACTORS:
        return jsonify({"error": "Invalid platform"}), 400
    payload = request.get_json(silent=True) or {}
    profile_reviews = payload.get("profile_reviews")
    if profile_reviews is None:
        profile_reviews = load_profile_reviews(platform)
    if not isinstance(profile_reviews, list) or not profile_reviews:
        return jsonify({"error": "No profile review data available to export"}), 400
    missing_profiles = list_missing_profile_reviews(platform, profile_reviews)
    if missing_profiles:
        return jsonify(build_final_review_missing_block_payload(platform, profile_reviews)), 409
    visual_results = merge_visual_results(platform, payload.get("visual_results") or {})
    rows = build_final_review_rows(platform, profile_reviews, visual_results)
    if not rows:
        return jsonify({"error": "No final review rows available to export"}), 400
    output = workbook_bytes_from_sheets([("Final Review", rows)])
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{platform}_final_review.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    ensure_runtime_dirs()
    start_apify_balance_poller()
    app.run(
        host=BACKEND_BIND_HOST,
        port=BACKEND_PORT,
        debug=BACKEND_DEBUG,
        use_reloader=BACKEND_USE_RELOADER,
    )
